import os
from openpyxl import load_workbook

# Ajusta esta ruta a donde te queda descargado el Excel de radicados
PATH = r"data\radicados\Control Correspondencia.xlsx"  # <-- cámbialo si aplica

def norm(s: str) -> str:
    return "".join(ch.lower() for ch in (s or "").strip() if ch.isalnum())

def main():
    if not os.path.exists(PATH):
        print(f"❌ No existe: {PATH}")
        return

    wb = load_workbook(PATH, data_only=True)

    print("\n===== DEBUG RADICADOS EXCEL =====")
    print("Hojas:", wb.sheetnames)

    for sh_name in wb.sheetnames:
        sh = wb[sh_name]
        # lee primera fila con datos (asumimos header en fila 1)
        headers = []
        for c in range(1, sh.max_column + 1):
            v = sh.cell(row=1, column=c).value
            headers.append(v)

        # muestra sólo si hay algo
        if any(h is not None for h in headers):
            print(f"\n--- Hoja: {sh_name} ---")
            for i, h in enumerate(headers, start=1):
                if h is None:
                    continue
                print(f"Col {i:02d}: {h!r} | norm={norm(str(h))}")

    print("\n=================================\n")

if __name__ == "__main__":
    main()
