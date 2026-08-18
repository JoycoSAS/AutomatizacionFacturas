# -*- coding: utf-8 -*-
"""
JOYCO - Facturas Procesador
Subida remota de cierre diario V3 a repositorio único de backups.

Política aplicada:
- El cierre diario NO se sube al SharePoint principal de operación.
- Se sube únicamente al repositorio remoto de backups configurado por .env.
- La verificación remota descarga cada archivo subido y compara:
  - Excel: datos internos hoja/celda.
  - Otros archivos: SHA256 binario exacto.
- Genera validación remota local y también la sube/verifica.

Variables .env soportadas, en orden de prioridad:
- BACKUP_DRIVE_ID u ONEDRIVE_BACKUP_DRIVE_ID o SP_BACKUP2_DRIVE_ID
- BACKUP_ROOT_FOLDER u ONEDRIVE_BACKUP_FOLDER o SP_BACKUP2_FOLDER

Estructura remota esperada:
  BACKUP_ROOT_FOLDER/YYYY/TRIMESTRE_YYYY-MM-DD_A_YYYY-MM-DD/
    YYYY-MM_MesNombre/SEMANA_YYYY-MM-DD_a_YYYY-MM-DD/Diario_YYYY-MM-DD/

Compatibilidad:
- Se conserva el nombre del archivo para no romper wrappers existentes:
  scripts/subir_cierre_diario_sharepoint.py
"""

from __future__ import annotations

import argparse
import datetime as _dt
import hashlib
import json
import os
import shutil
import sys
import time
from pathlib import Path
from typing import Any, Optional, Tuple
from urllib.parse import quote
from zoneinfo import ZoneInfo

import requests
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

try:
    import config  # noqa: F401  # Carga .env del proyecto.
except Exception:
    pass

from services.m365.token import get_access_token
from trimestre_activo import cargar_trimestre_activo

VERSION_UPLOAD = "2026-08-14-UPLOAD-CIERRE-DIARIO-V5-RETENCION-IDEMPOTENTE"
GRAPH = "https://graph.microsoft.com/v1.0"
RETENCION_LOCAL_DIAS = 7
ZONA_HORARIA = ZoneInfo("America/Bogota")

DATA_DIR = ROOT / "data"
CIERRES_DIR = DATA_DIR / "cierres_diarios"
TMP_VERIFY_DIR = DATA_DIR / "_tmp_verificacion_cierre_diario_v3"

EXCLUIR_NOMBRES = {".env", ".env.local", ".env.production"}
EXCLUIR_EXT = {".tmp", ".lock"}
EXCEL_EXT = {".xlsx", ".xlsm"}


def env_first(*names: str, default: str = "") -> str:
    for name in names:
        value = str(os.getenv(name, "") or "").strip()
        if value:
            return value
    return default


def resumen_identificador(value: Any) -> str:
    texto = str(value or "").strip()
    if not texto:
        return "(vacío)"
    visible = texto[-6:] if len(texto) > 6 else texto
    return f"***{visible} (longitud={len(texto)})"


def ssl_verify() -> bool:
    return str(os.getenv("SSL_VERIFY", "true") or "true").strip().lower() not in {
        "0",
        "false",
        "no",
        "off",
    }


def headers() -> dict[str, str]:
    token = get_access_token()
    return {"Authorization": f"Bearer {token}"}


def h_json() -> dict[str, str]:
    h = headers()
    h["Content-Type"] = "application/json"
    return h


def encode_path(path: str) -> str:
    return quote(str(path).strip("/"), safe="/")


def encode_drive_id(drive_id: str) -> str:
    return quote(str(drive_id), safe="!")


def graph_get(url: str, *, ok: Tuple[int, ...] = (200,), timeout: int = 60) -> requests.Response:
    r = requests.get(url, headers=headers(), timeout=timeout, verify=ssl_verify())
    if r.status_code not in ok:
        raise RuntimeError(f"GET {r.status_code} {url} -> {r.text[:700]}")
    return r


def existe_path(drive_id: str, remote_path: str) -> bool:
    remote_path = remote_path.strip("/")
    if not remote_path:
        return True
    url = f"{GRAPH}/drives/{encode_drive_id(drive_id)}/root:/{encode_path(remote_path)}:"
    r = requests.get(url, headers=headers(), timeout=60, verify=ssl_verify())
    if r.status_code == 200:
        return True
    if r.status_code == 404:
        return False
    raise RuntimeError(f"GET {r.status_code} {url} -> {r.text[:700]}")


def crear_folder(drive_id: str, parent_path: str, folder_name: str) -> None:
    body = {
        "name": folder_name,
        "folder": {},
        "@microsoft.graph.conflictBehavior": "fail",
    }
    parent_path = parent_path.strip("/")
    if parent_path:
        url = f"{GRAPH}/drives/{encode_drive_id(drive_id)}/root:/{encode_path(parent_path)}:/children"
    else:
        url = f"{GRAPH}/drives/{encode_drive_id(drive_id)}/root/children"

    r = requests.post(url, headers=h_json(), json=body, timeout=60, verify=ssl_verify())
    if r.status_code in (200, 201, 409):
        return
    raise RuntimeError(f"POST {r.status_code} {url} -> {r.text[:700]}")


def ensure_folder_recursive(drive_id: str, folder_path: str) -> None:
    folder_path = folder_path.strip("/")
    if not folder_path:
        return

    actual = ""
    for parte in [p for p in folder_path.split("/") if p]:
        siguiente = f"{actual}/{parte}".strip("/")
        if not existe_path(drive_id, siguiente):
            crear_folder(drive_id, actual, parte)
        actual = siguiente


def graph_put_content(drive_id: str, remote_path: str, local_file: Path) -> dict[str, Any]:
    url = f"{GRAPH}/drives/{encode_drive_id(drive_id)}/root:/{encode_path(remote_path)}:/content"
    data = local_file.read_bytes()
    r = requests.put(url, headers=headers(), data=data, timeout=300, verify=ssl_verify())
    if r.status_code not in (200, 201):
        raise RuntimeError(f"PUT {r.status_code} {url} -> {r.text[:700]}")
    return r.json()


def graph_download_item_content(drive_id: str, item_id: str) -> bytes:
    url = f"{GRAPH}/drives/{encode_drive_id(drive_id)}/items/{quote(item_id, safe='')}/content"
    r = requests.get(
        url,
        headers=headers(),
        timeout=300,
        verify=ssl_verify(),
        allow_redirects=True,
    )
    if r.status_code != 200:
        raise RuntimeError(f"DOWNLOAD {r.status_code} {url} -> {r.text[:700]}")
    return r.content


def validar_drive_id(drive_id: str) -> Optional[dict[str, Any]]:
    if not drive_id:
        return None
    url = f"{GRAPH}/drives/{encode_drive_id(drive_id)}?$select=id,name,webUrl,driveType"
    r = requests.get(url, headers=headers(), timeout=60, verify=ssl_verify())
    if r.status_code == 200:
        return r.json()
    print(f"⚠️ Drive ID no válido o no accesible: {resumen_identificador(drive_id)}")
    print(f"   Respuesta Graph: {r.status_code} {r.text[:350]}")
    return None


def listar_drives_site(hostname: str, site_path: str) -> list[dict[str, Any]]:
    if not hostname or not site_path:
        return []
    site_path = site_path if site_path.startswith("/") else f"/{site_path}"
    url = f"{GRAPH}/sites/{hostname}:{site_path}:/drives?$select=id,name,webUrl,driveType"
    r = graph_get(url, timeout=60)
    return r.json().get("value", [])


def resolver_drive_backup() -> tuple[str, dict[str, Any]]:
    drive_id = env_first("BACKUP_DRIVE_ID", "ONEDRIVE_BACKUP_DRIVE_ID", "SP_BACKUP2_DRIVE_ID")
    drive = validar_drive_id(drive_id)
    if drive:
        print(
            "✅ Drive backup validado: "
            f"{drive.get('name')} | {resumen_identificador(drive.get('id'))}"
        )
        return drive["id"], drive

    hostname = env_first("BACKUP_HOSTNAME", "ONEDRIVE_BACKUP_HOSTNAME", "SP_BACKUP2_HOSTNAME")
    site_path = env_first("BACKUP_SITE_PATH", "ONEDRIVE_BACKUP_SITE_PATH", "SP_BACKUP2_SITE_PATH")
    if hostname and site_path:
        print("🔎 Buscando drive backup desde hostname/site_path...")
        drives = listar_drives_site(hostname, site_path)
        if not drives:
            raise RuntimeError("No se encontraron drives para el repositorio remoto de backups.")

        print("📚 Drives encontrados:")
        for d in drives:
            print(
                f"   - {d.get('name')} | "
                f"{resumen_identificador(d.get('id'))} | {d.get('webUrl')}"
            )

        preferidos = {"documentos", "documents", "shared documents", "onedrive"}
        elegido = None
        for d in drives:
            if (d.get("name") or "").strip().lower() in preferidos:
                elegido = d
                break
        if not elegido:
            elegido = drives[0]

        print(
            "✅ Drive backup resuelto: "
            f"{elegido.get('name')} | "
            f"{resumen_identificador(elegido.get('id'))}"
        )
        print("💡 Si este ID funciona, guárdalo en BACKUP_DRIVE_ID o SP_BACKUP2_DRIVE_ID.")
        return elegido["id"], elegido

    raise RuntimeError(
        "No hay drive de backups configurado. Define BACKUP_DRIVE_ID, "
        "ONEDRIVE_BACKUP_DRIVE_ID o SP_BACKUP2_DRIVE_ID en .env."
    )


def sha256_bytes(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def normalizar_valor_excel(v: Any) -> str:
    if v is None:
        return "<NULL>"
    if isinstance(v, (_dt.datetime, _dt.date, _dt.time)):
        return v.isoformat()
    if isinstance(v, float):
        return repr(v)
    return f"{type(v).__name__}:{str(v)}"


def digest_datos_excel(path: Path) -> tuple[str, dict[str, Any]]:
    h = hashlib.sha256()
    resumen = {
        "sheets": [],
        "non_empty_cells": 0,
        "max_rows_total": 0,
        "max_cols_total": 0,
    }

    wb = load_workbook(path, read_only=True, data_only=False, keep_links=False)
    try:
        sheet_names = list(wb.sheetnames)
        h.update(json.dumps(sheet_names, ensure_ascii=False).encode("utf-8"))

        for ws in wb.worksheets:
            sheet_info = {
                "title": ws.title,
                "max_row": int(ws.max_row or 0),
                "max_column": int(ws.max_column or 0),
                "non_empty_cells": 0,
            }
            resumen["max_rows_total"] += sheet_info["max_row"]
            resumen["max_cols_total"] += sheet_info["max_column"]
            h.update(f"\n[SHEET]{ws.title}|{ws.max_row}|{ws.max_column}".encode("utf-8"))

            for row in ws.iter_rows():
                for cell in row:
                    v = cell.value
                    if v is None:
                        continue
                    sheet_info["non_empty_cells"] += 1
                    resumen["non_empty_cells"] += 1
                    payload = f"{ws.title}|{cell.coordinate}|{normalizar_valor_excel(v)}\n"
                    h.update(payload.encode("utf-8", errors="replace"))

            resumen["sheets"].append(sheet_info)
    finally:
        wb.close()

    return h.hexdigest(), resumen


def escribir_temporal_descarga(rel: str, data: bytes) -> Path:
    TMP_VERIFY_DIR.mkdir(parents=True, exist_ok=True)
    seguro = rel.replace("/", "__").replace("\\", "__")
    p = TMP_VERIFY_DIR / f"download_{_dt.datetime.now().strftime('%Y%m%d_%H%M%S')}_{seguro}"
    p.write_bytes(data)
    return p


def verificar_archivo_subido(
    *,
    local: Path,
    rel: str,
    drive_id: str,
    item: dict[str, Any],
) -> tuple[bool, dict[str, Any]]:
    item_id = item.get("id")
    if not item_id:
        raise RuntimeError("Graph no devolvió item.id; no se puede verificar descarga exacta.")

    remote_bytes = graph_download_item_content(drive_id, item_id)

    base_result = {
        "rel": rel,
        "local": str(local),
        "remote_item_id": item_id,
        "remote_web_url": item.get("webUrl"),
        "bytes_local": local.stat().st_size,
        "bytes_remoto": len(remote_bytes),
    }

    if local.suffix.lower() in EXCEL_EXT:
        tmp = escribir_temporal_descarga(rel, remote_bytes)
        try:
            digest_local, resumen_local = digest_datos_excel(local)
            digest_remoto, resumen_remoto = digest_datos_excel(tmp)
            base_result.update(
                {
                    "tipo_verificacion": "excel_datos_hoja_celda",
                    "digest_local": digest_local,
                    "digest_remoto": digest_remoto,
                    "resumen_local": resumen_local,
                    "resumen_remoto": resumen_remoto,
                    "ok": digest_local == digest_remoto,
                }
            )
            if digest_local != digest_remoto:
                print(f"❌ Excel con datos distintos: {rel}")
                print(f"   Digest local:  {digest_local}")
                print(f"   Digest remoto: {digest_remoto}")
                return False, base_result

            print(
                f"✅ Excel verificado por DATOS: {rel} | "
                f"celdas_no_vacias={resumen_local.get('non_empty_cells')} | "
                f"hojas={len(resumen_local.get('sheets', []))}"
            )
            return True, base_result
        finally:
            try:
                tmp.unlink(missing_ok=True)
            except Exception:
                pass

    hash_local = sha256_file(local)
    hash_remoto = sha256_bytes(remote_bytes)
    base_result.update(
        {
            "tipo_verificacion": "sha256_binario_exacto",
            "sha256_local": hash_local,
            "sha256_remoto": hash_remoto,
            "ok": hash_local == hash_remoto,
        }
    )
    if hash_local != hash_remoto:
        print(f"❌ Hash distinto: {rel}")
        print(f"   SHA256 local:  {hash_local}")
        print(f"   SHA256 remoto: {hash_remoto}")
        return False, base_result

    print(f"✅ Archivo verificado por SHA256 exacto: {rel} ({local.stat().st_size} bytes)")
    return True, base_result


def subir_y_verificar_con_reintentos(
    *,
    drive_id: str,
    remote_path: str,
    local: Path,
    rel: str,
) -> tuple[bool, dict[str, Any]]:
    item = graph_put_content(drive_id, remote_path, local)

    intentos = 4
    espera = 2
    ultimo_error = None
    ultimo_resultado: dict[str, Any] = {}

    for intento in range(1, intentos + 1):
        try:
            ok, resultado = verificar_archivo_subido(local=local, rel=rel, drive_id=drive_id, item=item)
            resultado["intento_verificacion"] = intento
            if ok:
                return True, resultado
            ultimo_resultado = resultado
        except Exception as exc:
            ultimo_error = exc
            ultimo_resultado = {
                "rel": rel,
                "ok": False,
                "error": str(exc),
                "intento_verificacion": intento,
            }
            print(f"⚠️ Verificación intento {intento}/{intentos} falló para {rel}: {exc}")

        if intento < intentos:
            print(f"   Reintentando verificación en {espera}s...")
            time.sleep(espera)
            espera *= 2

    if ultimo_error:
        print(f"❌ Verificación definitiva fallida para {rel}: {ultimo_error}")
    else:
        print(f"❌ Verificación definitiva fallida para {rel}.")
    return False, ultimo_resultado


def normalizar_rel(path: Path) -> str:
    return path.as_posix().replace("\\", "/")


def debe_subir(p: Path) -> bool:
    if not p.is_file():
        return False
    if p.name in EXCLUIR_NOMBRES:
        return False
    if p.suffix.lower() in EXCLUIR_EXT:
        return False
    if p.name.startswith("validacion_remota_"):
        return False
    if p.name.startswith("estado_retencion_"):
        return False
    return True


def listar_archivos_recursivos(base: Path) -> list[Path]:
    return [p for p in sorted(base.rglob("*")) if debe_subir(p)]


def semana_lunes_domingo(fecha: _dt.date) -> tuple[_dt.date, _dt.date]:
    inicio = fecha - _dt.timedelta(days=fecha.weekday())
    fin = inicio + _dt.timedelta(days=6)
    return inicio, fin

MESES_ES = {
    1: "Enero",
    2: "Febrero",
    3: "Marzo",
    4: "Abril",
    5: "Mayo",
    6: "Junio",
    7: "Julio",
    8: "Agosto",
    9: "Septiembre",
    10: "Octubre",
    11: "Noviembre",
    12: "Diciembre",
}


def nombre_mes_dir(fecha: _dt.date) -> str:
    return f"{fecha.strftime('%Y-%m')}_{MESES_ES[int(fecha.strftime('%m'))]}"



def parse_fecha(value: Optional[str]) -> _dt.date:
    if value:
        return _dt.date.fromisoformat(value)
    return _dt.date.today()


def buscar_cierre_local(
    fecha: _dt.date,
    trimestre: dict[str, Any],
) -> Path:
    mes_nombre = nombre_mes_dir(fecha)
    fecha_s = fecha.isoformat()
    inicio, fin = semana_lunes_domingo(fecha)
    semana = f"SEMANA_{inicio.isoformat()}_a_{fin.isoformat()}"

    esperado = (
        CIERRES_DIR
        / trimestre["anio"]
        / trimestre["nombre_carpeta"]
        / mes_nombre
        / semana
        / f"Diario_{fecha_s}"
    )

    if esperado.exists() and esperado.is_dir():
        return esperado

    raise RuntimeError(
        "No se encontró el cierre diario en la jerarquía trimestral oficial. "
        f"Ruta esperada: {esperado}. "
        "No se usará automáticamente la estructura anterior para evitar "
        "crear backups fuera del trimestre activo."
    )


def remote_base_para_cierre(cierre_dir: Path, fecha: _dt.date) -> str:
    root_folder = env_first(
        "BACKUP_ROOT_FOLDER",
        "ONEDRIVE_BACKUP_FOLDER",
        "SP_BACKUP2_FOLDER",
    )
    if not root_folder:
        raise RuntimeError(
            "Falta BACKUP_ROOT_FOLDER/ONEDRIVE_BACKUP_FOLDER/"
            "SP_BACKUP2_FOLDER en .env."
        )

    try:
        rel_cierre = (
            cierre_dir.resolve()
            .relative_to(CIERRES_DIR.resolve())
            .as_posix()
        )
    except ValueError as exc:
        raise RuntimeError(
            "La carpeta del cierre diario está fuera de CIERRES_DIR y no "
            "puede convertirse en una ruta remota segura: "
            f"{cierre_dir}"
        ) from exc

    esperado_nombre = f"Diario_{fecha.isoformat()}"
    if cierre_dir.name != esperado_nombre:
        raise RuntimeError(
            "La carpeta del cierre diario no coincide con la fecha solicitada. "
            f"Esperado={esperado_nombre} | Encontrado={cierre_dir.name}"
        )

    partes = Path(rel_cierre).parts
    if len(partes) < 5 or not partes[1].startswith("TRIMESTRE_"):
        raise RuntimeError(
            "La ruta local no contiene la jerarquía trimestral oficial: "
            f"{rel_cierre}"
        )

    return f"{root_folder}/{rel_cierre}".strip("/")


def validar_cierre_local_minimo(cierre_dir: Path, fecha: _dt.date) -> None:
    fecha_s = fecha.isoformat()
    obligatorios = [
        cierre_dir / "01_Excel_Diario" / f"facturas_diario_{fecha_s}.xlsx",
        cierre_dir / "04_Manifest" / f"manifest_diario_{fecha_s}.json",
        cierre_dir / "04_Manifest" / f"resumen_diario_{fecha_s}.txt",
        cierre_dir / "05_Validaciones" / f"validacion_local_{fecha_s}.json",
    ]
    faltantes = [str(p) for p in obligatorios if not p.exists()]
    if faltantes:
        raise RuntimeError("El cierre local no está completo. Faltan: " + "; ".join(faltantes))

    validacion_path = cierre_dir / "05_Validaciones" / f"validacion_local_{fecha_s}.json"
    try:
        validacion = json.loads(validacion_path.read_text(encoding="utf-8"))
    except Exception as exc:
        raise RuntimeError(f"No se pudo leer validación local: {validacion_path} | {exc}")

    if not validacion.get("ok"):
        raise RuntimeError(f"La validación local del cierre no está OK: {validacion_path}")


def escribir_validacion_remota(
    *,
    cierre_dir: Path,
    fecha: _dt.date,
    trimestre: dict[str, Any],
    drive: dict[str, Any],
    remote_base: str,
    resultados: list[dict[str, Any]],
    ok: bool,
) -> Path:
    validaciones_dir = cierre_dir / "05_Validaciones"
    validaciones_dir.mkdir(parents=True, exist_ok=True)
    fecha_s = fecha.isoformat()
    path = validaciones_dir / f"validacion_remota_{fecha_s}.json"

    payload = {
        "tipo": "validacion_remota_cierre_diario",
        "version": VERSION_UPLOAD,
        "fecha": fecha_s,
        "trimestre": trimestre,
        "generado_en": _dt.datetime.now().isoformat(timespec="seconds"),
        "cierre_local": str(cierre_dir),
        "drive": {
            "id": drive.get("id"),
            "name": drive.get("name"),
            "driveType": drive.get("driveType"),
            "webUrl": drive.get("webUrl"),
        },
        "remote_base": remote_base,
        "total_archivos_verificados": len(resultados),
        "ok": bool(ok),
        "resultados": resultados,
    }
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return path


def digest_publicacion_desde_resultados(
    resultados: list[dict[str, Any]],
) -> str:
    """
    Huella estable del contenido publicado.

    Excel:
        digest lógico de datos/celdas.
    Otros:
        SHA256 binario local.

    No usa IDs, URLs, timestamps ni datos variables de Graph.
    """
    normalizados: list[dict[str, str]] = []

    for resultado in resultados:
        if not isinstance(resultado, dict):
            raise RuntimeError(
                "Resultado de publicación inválido."
            )

        if resultado.get("ok") is not True:
            raise RuntimeError(
                "No se puede calcular digest con resultados no OK."
            )

        rel = str(resultado.get("rel") or "").strip()

        if not rel:
            raise RuntimeError(
                "Resultado sin ruta relativa."
            )

        tipo = str(
            resultado.get("tipo_verificacion") or ""
        ).strip()

        digest_excel = resultado.get("digest_local")
        sha_binario = str(
            resultado.get("sha256_local") or ""
        ).strip().lower()

        if tipo == "excel_datos_hoja_celda":
            if isinstance(digest_excel, str):
                huella = digest_excel.strip().lower()
            elif isinstance(digest_excel, dict):
                huella = str(
                    digest_excel.get("sha256_datos") or ""
                ).strip().lower()
            else:
                huella = ""

            if not huella:
                raise RuntimeError(
                    f"Excel sin digest local: {rel}"
                )
        else:
            huella = sha_binario

            if not huella:
                raise RuntimeError(
                    f"Archivo sin SHA256 local: {rel}"
                )

        normalizados.append(
            {
                "rel": rel,
                "tipo": tipo,
                "huella": huella,
            }
        )

    normalizados.sort(key=lambda x: x["rel"])

    serializado = json.dumps(
        normalizados,
        ensure_ascii=False,
        sort_keys=True,
        separators=(",", ":"),
    ).encode("utf-8")

    return sha256_bytes(serializado)


def digest_publicacion_desde_validacion(
    validacion_path: Path,
) -> str:
    data = json.loads(
        validacion_path.read_text(encoding="utf-8")
    )

    if data.get("ok") is not True:
        raise RuntimeError(
            "La validación remota anterior no está OK."
        )

    resultados = data.get("resultados")

    if not isinstance(resultados, list) or not resultados:
        raise RuntimeError(
            "La validación remota anterior no tiene resultados."
        )

    return digest_publicacion_desde_resultados(
        resultados
    )


def crear_estado_retencion_temporal(
    fecha: _dt.date,
    remote_base: str,
    trimestre: dict[str, Any],
    digest_publicacion: str,
    total_archivos_publicacion: int,
    digest_publicacion_previo: str | None = None,
) -> tuple[Path, dict[str, Any]]:
    fecha_s = fecha.isoformat()

    estado_local = (
        buscar_cierre_local(fecha, trimestre)
        / "05_Validaciones"
        / f"estado_retencion_diario_{fecha_s}.json"
    )

    # Reutiliza la fecha original únicamente si el estado existente
    # corresponde al mismo cierre y al mismo contenido publicado.
    if estado_local.is_file():
        try:
            existente = json.loads(
                estado_local.read_text(encoding="utf-8")
            )

            digest_existente = str(
                existente.get("digest_publicacion") or ""
            ).strip().lower()

            mismo_contenido = (
                digest_existente == digest_publicacion
                if digest_existente
                else (
                    bool(digest_publicacion_previo)
                    and digest_publicacion_previo
                    == digest_publicacion
                )
            )

            vigente = bool(
                existente.get("tipo")
                == "estado_retencion_local_cierre_diario"
                and existente.get("ok") is True
                and existente.get(
                    "validacion_remota_publicada_y_verificada"
                ) is True
                and existente.get("fecha") == fecha_s
                and existente.get("remote_base") == remote_base
                and int(
                    existente.get("retencion_local_dias") or -1
                ) == RETENCION_LOCAL_DIAS
                and bool(existente.get("validado_en"))
                and bool(
                    existente.get(
                        "eliminacion_local_permitida_desde"
                    )
                )
                and mismo_contenido
            )

            if vigente:
                payload = dict(existente)

                # Migración transparente de estados históricos que
                # todavía no tenían digest_publicacion.
                payload["digest_publicacion"] = digest_publicacion
                payload["total_archivos_publicacion"] = (
                    total_archivos_publicacion
                )

                TMP_VERIFY_DIR.mkdir(
                    parents=True,
                    exist_ok=True,
                )

                path_tmp = (
                    TMP_VERIFY_DIR
                    / f"estado_retencion_diario_{fecha_s}.json"
                )

                path_tmp.write_text(
                    json.dumps(
                        payload,
                        ensure_ascii=False,
                        indent=2,
                    ),
                    encoding="utf-8",
                )

                return path_tmp, payload

        except Exception:
            pass

    ahora = _dt.datetime.now(ZONA_HORARIA)

    payload = {
        "tipo": "estado_retencion_local_cierre_diario",
        "version": VERSION_UPLOAD,
        "trimestre": {
            "periodo_activo": trimestre["periodo_activo"],
            "nombre_carpeta": trimestre["nombre_carpeta"],
            "fecha_inicio": trimestre["fecha_inicio"],
            "fecha_fin": trimestre["fecha_fin"],
            "ruta_relativa": trimestre["ruta_relativa"],
        },
        "fecha": fecha_s,
        "validacion_remota_publicada_y_verificada": True,
        "validado_en": ahora.isoformat(timespec="seconds"),
        "retencion_local_dias": RETENCION_LOCAL_DIAS,
        "eliminacion_local_permitida_desde": (
            ahora + _dt.timedelta(days=RETENCION_LOCAL_DIAS)
        ).isoformat(timespec="seconds"),
        "remote_base": remote_base,
        "digest_publicacion": digest_publicacion,
        "total_archivos_publicacion": total_archivos_publicacion,
        "ok": True,
    }

    TMP_VERIFY_DIR.mkdir(parents=True, exist_ok=True)

    path_tmp = (
        TMP_VERIFY_DIR
        / f"estado_retencion_diario_{fecha_s}.json"
    )

    path_tmp.write_text(
        json.dumps(
            payload,
            ensure_ascii=False,
            indent=2,
        ),
        encoding="utf-8",
    )

    return path_tmp, payload


def limpiar_tmp() -> None:
    try:
        if TMP_VERIFY_DIR.exists():
            shutil.rmtree(TMP_VERIFY_DIR, ignore_errors=True)
    except Exception:
        pass


def main() -> int:
    parser = argparse.ArgumentParser(description="Sube cierre diario V5 al repositorio remoto único de backups.")
    parser.add_argument("--fecha", default="", help="Fecha del cierre en formato YYYY-MM-DD. Default: hoy.")
    parser.add_argument("--dry-run", action="store_true", help="Solo muestra qué subiría, sin llamar Graph.")
    args = parser.parse_args()

    fecha = parse_fecha(args.fecha)
    fecha_s = fecha.isoformat()

    print("=" * 100)
    print("SUBIDA CIERRE DIARIO V5 A REPOSITORIO DE BACKUPS - FACTURAS JOYCO")
    print("=" * 100)
    print(f"Versión: {VERSION_UPLOAD}")
    print(f"Root: {ROOT}")
    print(f"Fecha cierre: {fecha_s}")
    print(f"Dry-run: {args.dry_run}")
    print("-" * 100)

    try:
        trimestre = cargar_trimestre_activo(ROOT, fecha)
        cierre_dir = buscar_cierre_local(fecha, trimestre)
        validar_cierre_local_minimo(cierre_dir, fecha)
    except Exception as exc:
        print(f"❌ Cierre local no válido: {exc}")
        return 1

    remote_base = remote_base_para_cierre(cierre_dir, fecha)
    archivos = listar_archivos_recursivos(cierre_dir)
    if not archivos:
        print("❌ No hay archivos válidos para subir.")
        return 1

    estado_local = (
        cierre_dir
        / "05_Validaciones"
        / f"estado_retencion_diario_{fecha_s}.json"
    )
    print(f"📁 Cierre local: {cierre_dir}")
    print(f"📆 Trimestre activo: {trimestre['nombre_carpeta']}")
    print(
        "📆 Rango trimestre: "
        f"{trimestre['fecha_inicio']} a {trimestre['fecha_fin']}"
    )
    print(f"☁️ Ruta remota base: {remote_base}")
    print(f"📦 Archivos detectados: {len(archivos)}")
    for p in archivos:
        rel = p.relative_to(cierre_dir).as_posix()
        print(f"   - {rel} ({p.stat().st_size} bytes)")

    if args.dry_run:
        print("-" * 100)
        print("DRY-RUN: no se llamó a Microsoft Graph ni se subió nada.")
        print("Si la ruta remota base es correcta, ejecuta sin --dry-run.")
        print("=" * 100)
        return 0

    digest_publicacion_previo = None
    validacion_remota_previa = (
        cierre_dir
        / "05_Validaciones"
        / f"validacion_remota_{fecha_s}.json"
    )

    if estado_local.is_file() and validacion_remota_previa.is_file():
        try:
            digest_publicacion_previo = (
                digest_publicacion_desde_validacion(
                    validacion_remota_previa
                )
            )
        except Exception as exc:
            print(
                "⚠️ No se pudo calcular digest histórico "
                f"de publicación: {exc}"
            )

    try:
        drive_id, drive = resolver_drive_backup()
        print(f"📌 Repositorio destino: {drive.get('name')} | {drive.get('webUrl')}")
        ensure_folder_recursive(drive_id, remote_base)
        print("✅ Carpeta remota base verificada/creada.")
    except Exception as exc:
        print(f"❌ Error preparando repositorio remoto de backups: {exc}")
        return 1

    resultados: list[dict[str, Any]] = []
    ok_todos = True

    for local in archivos:
        rel = local.relative_to(cierre_dir).as_posix()
        remote_path = f"{remote_base}/{rel}".strip("/")
        remote_dir = "/".join(remote_path.split("/")[:-1])
        try:
            ensure_folder_recursive(drive_id, remote_dir)
            print("☁️ Subiendo:")
            print(f"   Local:  {local}")
            print(f"   Remoto: {remote_path}")
            ok, resultado = subir_y_verificar_con_reintentos(
                drive_id=drive_id,
                remote_path=remote_path,
                local=local,
                rel=rel,
            )
            resultado["remote_path"] = remote_path
            resultados.append(resultado)
            if not ok:
                ok_todos = False
        except Exception as exc:
            ok_todos = False
            print(f"❌ Error subiendo/verificando {rel}: {exc}")
            resultados.append(
                {
                    "rel": rel,
                    "remote_path": remote_path,
                    "ok": False,
                    "error": str(exc),
                }
            )

    digest_publicacion_actual = None

    if ok_todos:
        try:
            digest_publicacion_actual = (
                digest_publicacion_desde_resultados(
                    resultados
                )
            )
        except Exception as exc:
            ok_todos = False
            print(
                "❌ No se pudo calcular digest estable "
                f"de publicación: {exc}"
            )

    validacion_remota = escribir_validacion_remota(
        cierre_dir=cierre_dir,
        fecha=fecha,
        trimestre=trimestre,
        drive=drive,
        remote_base=remote_base,
        resultados=resultados,
        ok=ok_todos,
    )

    # Sube la validación remota como evidencia final.
    validacion_publicada = False
    try:
        rel = validacion_remota.relative_to(cierre_dir).as_posix()
        remote_path = f"{remote_base}/{rel}".strip("/")
        print("☁️ Subiendo validación remota final:")
        print(f"   Local:  {validacion_remota}")
        print(f"   Remoto: {remote_path}")
        ok_validacion, resultado_validacion = subir_y_verificar_con_reintentos(
            drive_id=drive_id,
            remote_path=remote_path,
            local=validacion_remota,
            rel=rel,
        )
        validacion_publicada = bool(ok_validacion)
        if not ok_validacion:
            ok_todos = False
        resultados.append(resultado_validacion)
    except Exception as exc:
        ok_todos = False
        print(f"❌ Error subiendo/verificando validación remota final: {exc}")

    if ok_todos and validacion_publicada:
        try:
            if not digest_publicacion_actual:
                raise RuntimeError(
                    "No existe digest de publicación validado."
                )

            estado_tmp, estado_payload = crear_estado_retencion_temporal(
                fecha,
                remote_base,
                trimestre,
                digest_publicacion_actual,
                len(
                    json.loads(
                        validacion_remota.read_text(
                            encoding="utf-8"
                        )
                    ).get("resultados") or []
                ),
                digest_publicacion_previo,
            )
            rel_estado = f"05_Validaciones/{estado_tmp.name}"
            remote_estado = f"{remote_base}/{rel_estado}".strip("/")

            print("☁️ Subiendo estado de retención:")
            print(f"   Local temporal: {estado_tmp}")
            print(f"   Remoto:         {remote_estado}")

            ok_estado, _resultado_estado = subir_y_verificar_con_reintentos(
                drive_id=drive_id,
                remote_path=remote_estado,
                local=estado_tmp,
                rel=rel_estado,
            )

            if not ok_estado:
                raise RuntimeError(
                    "No se pudo verificar el estado remoto de retención."
                )

            estado_local.parent.mkdir(
                parents=True,
                exist_ok=True,
            )

            temporal_local = estado_local.with_suffix(
                estado_local.suffix + ".tmp"
            )

            temporal_local.write_bytes(
                estado_tmp.read_bytes()
            )

            temporal_local.replace(
                estado_local
            )

            print(
                "✅ Retención local habilitada hasta: "
                f"{estado_payload['eliminacion_local_permitida_desde']}"
            )
        except Exception as exc:
            ok_todos = False
            print(
                "❌ Error registrando retención diaria: "
                f"{type(exc).__name__}: {exc}"
            )

    limpiar_tmp()

    print("-" * 100)
    print(f"📄 Validación remota local: {validacion_remota}")
    if ok_todos:
        print("✅ Subida de cierre diario V5 terminada correctamente en el repositorio único de backups.")
        print("✅ Verificación aplicada archivo por archivo después de descargar desde Graph.")
        print(
            f"✅ Retención local: política de {RETENCION_LOCAL_DIAS} días; "
            "el vencimiento original se conserva si el contenido no cambia."
        )
        print("=" * 100)
        return 0

    print("❌ Subida de cierre diario V5 terminó con errores.")
    print("⚠️ No borres ni archives localmente hasta revisar el error.")
    print("=" * 100)
    return 1


if __name__ == "__main__":
    raise SystemExit(main())
