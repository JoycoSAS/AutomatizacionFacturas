# -*- coding: utf-8 -*-
"""
JOYCO - Facturas Procesador
Sincronización segura de la jerarquía trimestral completa en un único destino.

Estructura oficial:

Backups_Facturas_Produccion/
└── YYYY/
    └── TRIMESTRE_YYYY-MM-DD_A_YYYY-MM-DD/
        ├── YYYY-MM_Mes/
        │   ├── SEMANA_YYYY-MM-DD_a_YYYY-MM-DD/
        │   │   ├── Diario_YYYY-MM-DD/
        │   │   └── Semanal/
        │   └── Mensual/
        └── Trimestral/
            ├── 01_Excel_Cierre/
            ├── 02_Manifest/
            ├── 03_Validaciones/
            └── 04_Resumen/

Principios:
- Existe un único destino oficial: Backups_Facturas_Produccion.
- Se sincroniza el contenedor completo del trimestre.
- Cada archivo permanece en una sola ubicación dentro de la jerarquía.
- No se copian diarios dentro de Semanal, semanas dentro de Mensual ni meses
  dentro de Trimestral.
- Los archivos remotos idénticos se verifican y se omiten; no se vuelven a
  subir ni generan versiones innecesarias.
- Los archivos faltantes se crean y los diferentes se actualizan.
- Nunca se eliminan archivos remotos desde este script.
- No se reemplaza data/facturas.xlsx ni el Excel activo de SharePoint.
- No se actualiza cierre_trimestral_state.json.
- No se ejecutan autoeliminaciones; esa fase se implementa y prueba aparte.

Modos:
- --dry-run:
  valida únicamente la jerarquía local y muestra el destino calculado.
  No usa Microsoft Graph.
- --plan-remoto:
  consulta el destino en modo de solo lectura y clasifica archivos como
  idénticos, faltantes, diferentes o extras. No crea ni modifica nada.
- --upload-cierre --confirmar SINCRONIZAR_TRIMESTRE_COMPLETO:
  sincroniza y verifica la jerarquía trimestral completa en el único destino.
"""

from __future__ import annotations

import argparse
import datetime
import hashlib
import json
import os
import re
import sys
import time
from pathlib import Path
from typing import Any, Optional, Tuple
from urllib.parse import quote
from zoneinfo import ZoneInfo

import requests
from dotenv import load_dotenv
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

try:
    import config  # noqa: F401
except Exception:
    pass

from services.m365.token import get_access_token
from trimestre_activo import cargar_trimestre_activo

DATA_DIR = ROOT / "data"
STATE_PATH = DATA_DIR / "state" / "cierre_trimestral_state.json"
FACTURAS_PATH = DATA_DIR / "facturas.xlsx"
CIERRES_DIR = DATA_DIR / "cierres_diarios"
TMP_VERIFY_DIR = DATA_DIR / "_tmp_verificacion_jerarquia_trimestral"

VERSION_UPLOAD = (
    "2026-08-13-UPLOAD-CIERRE-TRIMESTRAL-V7-RETENCION-30D"
)
GRAPH = "https://graph.microsoft.com/v1.0"
RETENCION_LOCAL_DIAS = 30
ZONA_HORARIA = ZoneInfo("America/Bogota")
CONFIRMACION_UPLOAD = "SINCRONIZAR_TRIMESTRE_COMPLETO"
ESTADO_PREPARADO = "PREPARADO_PENDIENTE_VALIDACION_REMOTA"
ROOT_OFICIAL = "Backups_Facturas_Produccion"

load_dotenv(ROOT / ".env")

BACKUP_DRIVE_ID = (
    os.getenv("BACKUP_DRIVE_ID")
    or os.getenv("SP_BACKUP2_DRIVE_ID")
    or ""
).strip()
BACKUP_ROOT_FOLDER = (
    os.getenv("BACKUP_ROOT_FOLDER")
    or os.getenv("SP_BACKUP2_FOLDER")
    or ""
).strip().strip("/")

EXCLUIR_NOMBRES = {
    ".env",
    "Thumbs.db",
    "desktop.ini",
}
EXCLUIR_EXT = {".tmp", ".lock", ".pyc"}
EXCEL_EXT = {".xlsx", ".xlsm"}
PREFIJO_VALIDACION_REMOTA = "validacion_remota_cierre_trimestral_"
PREFIJOS_CONTROL_NO_PUBLICAR = (
    PREFIJO_VALIDACION_REMOTA,
    "estado_retencion_trimestral_",
    "estado_finalizacion_cierre_trimestral_",
    "RESUMEN_FINALIZACION_CIERRE_TRIMESTRAL_",
)
DIRECTORIOS_TRIMESTRALES = {
    "01_Excel_Cierre",
    "02_Manifest",
    "03_Validaciones",
    "04_Resumen",
}
DIRECTORIOS_DUPLICACION_PROHIBIDOS = {
    "03_Soporte_Diarios",
    "03_Soporte_Semanas",
    "03_Respaldos_Mensuales_Validados",
    "07_Paquete_Mensual",
}
RE_MES = re.compile(r"^(\d{4})-(\d{2})_.+$")
RE_SEMANA = re.compile(
    r"^SEMANA_(\d{4}-\d{2}-\d{2})_[aA]_(\d{4}-\d{2}-\d{2})$"
)
RE_DIARIO = re.compile(r"^Diario_(\d{4}-\d{2}-\d{2})$")


def ssl_verify() -> bool:
    return (os.getenv("SSL_VERIFY") or "true").strip().lower() not in {
        "0",
        "false",
        "no",
        "off",
    }


def headers() -> dict[str, str]:
    return {"Authorization": f"Bearer {get_access_token()}"}


def headers_json() -> dict[str, str]:
    resultado = headers()
    resultado["Content-Type"] = "application/json"
    return resultado


def encode_path(path: str) -> str:
    return quote(str(path).strip("/"), safe="/")


def encode_drive_id(drive_id: str) -> str:
    return quote(str(drive_id), safe="!")


def resumen_drive(drive_id: str) -> str:
    valor = str(drive_id or "").strip()
    if not valor:
        return "(vacío)"
    sufijo = valor[-6:] if len(valor) >= 6 else valor
    return f"***{sufijo} (longitud={len(valor)})"


def url_graph_segura(url: str) -> str:
    valor = str(url or "")

    def ocultar(match: re.Match[str]) -> str:
        identificador = match.group(2)
        sufijo = identificador[-6:] if len(identificador) >= 6 else identificador
        return f"{match.group(1)}***{sufijo}"

    valor = re.sub(r"(/drives/)([^/?]+)", ocultar, valor)
    valor = re.sub(r"(/items/)([^/?]+)", ocultar, valor)
    return valor


def sha256_bytes(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as archivo:
        for bloque in iter(lambda: archivo.read(1024 * 1024), b""):
            digest.update(bloque)
    return digest.hexdigest()


def normalizar_valor_excel(valor: Any) -> str:
    if valor is None:
        return "<NULL>"
    if isinstance(valor, (datetime.datetime, datetime.date, datetime.time)):
        return valor.isoformat()
    if isinstance(valor, float):
        return repr(valor)
    return f"{type(valor).__name__}:{valor}"


def digest_datos_excel(path: Path) -> Tuple[str, dict[str, Any]]:
    digest = hashlib.sha256()
    resumen: dict[str, Any] = {
        "sheets": [],
        "non_empty_cells": 0,
        "max_rows_total": 0,
        "max_cols_total": 0,
    }

    libro = load_workbook(
        path,
        read_only=True,
        data_only=False,
        keep_links=False,
    )
    try:
        nombres = list(libro.sheetnames)
        digest.update(json.dumps(nombres, ensure_ascii=False).encode("utf-8"))

        for hoja in libro.worksheets:
            info = {
                "title": hoja.title,
                "max_row": int(hoja.max_row or 0),
                "max_column": int(hoja.max_column or 0),
                "non_empty_cells": 0,
            }
            resumen["max_rows_total"] += info["max_row"]
            resumen["max_cols_total"] += info["max_column"]
            digest.update(
                f"\n[SHEET]{hoja.title}|{hoja.max_row}|{hoja.max_column}".encode(
                    "utf-8"
                )
            )

            for fila in hoja.iter_rows():
                for celda in fila:
                    if celda.value is None:
                        continue
                    info["non_empty_cells"] += 1
                    resumen["non_empty_cells"] += 1
                    payload = (
                        f"{hoja.title}|{celda.coordinate}|"
                        f"{normalizar_valor_excel(celda.value)}\n"
                    )
                    digest.update(payload.encode("utf-8", errors="replace"))

            resumen["sheets"].append(info)
    finally:
        libro.close()

    return digest.hexdigest(), resumen


def leer_json(path: Path, descripcion: str) -> dict[str, Any]:
    if not path.exists() or not path.is_file():
        raise RuntimeError(f"No existe {descripcion}: {path}")
    try:
        datos = json.loads(path.read_text(encoding="utf-8-sig"))
    except Exception as exc:
        raise RuntimeError(
            f"No fue posible leer {descripcion}: {path}: {exc}"
        ) from exc
    if not isinstance(datos, dict):
        raise RuntimeError(f"{descripcion} no contiene un objeto JSON: {path}")
    return datos


def parse_fecha_iso(value: Any, campo: str) -> datetime.date:
    try:
        return datetime.date.fromisoformat(str(value or "").strip())
    except Exception as exc:
        raise RuntimeError(
            f"Fecha inválida en {campo}: {value!r}. Usa YYYY-MM-DD."
        ) from exc


def validar_path_exacto(registrado: Any, esperado: Path, descripcion: str) -> None:
    texto = str(registrado or "").strip()
    if not texto:
        raise RuntimeError(f"La preparación no registra {descripcion}.")
    actual = Path(texto)
    if actual.resolve() != esperado.resolve():
        raise RuntimeError(
            f"{descripcion} no coincide con la ruta oficial. "
            f"Registrado={actual} | Esperado={esperado}"
        )


def debe_incluir_directorio(path: Path) -> bool:
    nombre = path.name
    if nombre == "__pycache__":
        return False
    if nombre.startswith(".tmp_") or nombre.startswith("_tmp"):
        return False
    if nombre.startswith(".migrando-"):
        return False
    return True


def debe_incluir_archivo(path: Path) -> bool:
    if not path.is_file():
        return False
    if path.name in EXCLUIR_NOMBRES:
        return False
    if path.name.startswith("~$"):
        return False
    if path.suffix.lower() in EXCLUIR_EXT:
        return False
    if any(not debe_incluir_directorio(Path(parte)) for parte in path.parts):
        return False
    if path.name.startswith(PREFIJOS_CONTROL_NO_PUBLICAR):
        return False
    return True


def listar_directorios_locales(base: Path) -> list[Path]:
    resultado = [
        path
        for path in base.rglob("*")
        if path.is_dir()
        and all(
            debe_incluir_directorio(Path(parte))
            for parte in path.relative_to(base).parts
        )
    ]
    return sorted(
        resultado,
        key=lambda p: (
            len(p.relative_to(base).parts),
            p.as_posix().lower(),
        ),
    )


def listar_archivos_locales(base: Path) -> list[Path]:
    return sorted(
        [path for path in base.rglob("*") if debe_incluir_archivo(path)],
        key=lambda p: p.as_posix().lower(),
    )


def digest_archivos(archivos: list[Path], base: Path) -> tuple[str, int]:
    digest = hashlib.sha256()
    total_bytes = 0
    for archivo in archivos:
        relativa = archivo.relative_to(base).as_posix()
        tamano = archivo.stat().st_size
        total_bytes += tamano
        digest.update(
            f"{relativa}|{tamano}|{sha256_file(archivo)}\n".encode("utf-8")
        )
    return digest.hexdigest(), total_bytes


def digest_jerarquia_inferior(
    carpeta_periodo: Path,
    inicio: datetime.date,
    fin: datetime.date,
) -> tuple[str, int, int, int]:
    digest = hashlib.sha256()
    total_archivos = 0
    total_bytes = 0
    meses_detectados = 0

    for mes_dir in sorted(
        carpeta_periodo.iterdir(),
        key=lambda p: p.name.lower(),
    ):
        if not mes_dir.is_dir() or mes_dir.name == "Trimestral":
            continue

        match = RE_MES.match(mes_dir.name)
        if not match:
            continue

        try:
            mes_fecha = datetime.date(
                int(match.group(1)),
                int(match.group(2)),
                1,
            )
        except ValueError:
            continue

        if (mes_fecha.year, mes_fecha.month) < (inicio.year, inicio.month):
            continue
        if (mes_fecha.year, mes_fecha.month) > (fin.year, fin.month):
            continue

        meses_detectados += 1
        archivos_mes = sorted(
            [p for p in mes_dir.rglob("*") if p.is_file()],
            key=lambda p: p.as_posix().lower(),
        )

        for archivo in archivos_mes:
            relativa = archivo.relative_to(carpeta_periodo).as_posix()
            tamano = archivo.stat().st_size
            digest.update(
                f"{relativa}|{tamano}|{sha256_file(archivo)}\n".encode("utf-8")
            )
            total_archivos += 1
            total_bytes += tamano

    if meses_detectados <= 0:
        raise RuntimeError(
            "La carpeta del trimestre no contiene carpetas mensuales."
        )

    return digest.hexdigest(), total_archivos, total_bytes, meses_detectados


def validar_excel_limpio(path: Path) -> dict[str, Any]:
    if not path.exists() or not path.is_file():
        raise RuntimeError(f"No existe el Excel limpio candidato: {path}")

    libro = load_workbook(
        path,
        read_only=False,
        data_only=False,
        keep_links=False,
    )
    try:
        if "Facturas" not in libro.sheetnames:
            raise RuntimeError("El candidato limpio no contiene la hoja Facturas.")
        hoja = libro["Facturas"]
        tabla = (
            hoja.tables["TblFacturas"].ref
            if "TblFacturas" in hoja.tables
            else None
        )
        info = {
            "filas": int(hoja.max_row or 0),
            "filas_datos": max(int(hoja.max_row or 0) - 1, 0),
            "columnas": int(hoja.max_column or 0),
            "tabla": tabla,
            "sha256": sha256_file(path),
        }
        if info["filas"] != 1 or info["columnas"] != 19 or tabla != "A1:S1":
            raise RuntimeError(
                "El Excel limpio candidato no conserva la estructura esperada: "
                f"{info}"
            )
        return info
    finally:
        libro.close()


def validar_estructura_jerarquica(
    carpeta_periodo: Path,
    fecha_inicio: datetime.date,
    fecha_fin: datetime.date,
) -> dict[str, Any]:
    if not carpeta_periodo.exists() or not carpeta_periodo.is_dir():
        raise RuntimeError(
            f"No existe la carpeta local del trimestre: {carpeta_periodo}"
        )

    prohibidos = [
        path
        for path in carpeta_periodo.rglob("*")
        if path.is_dir() and path.name in DIRECTORIOS_DUPLICACION_PROHIBIDOS
    ]
    if prohibidos:
        raise RuntimeError(
            "Se detectaron carpetas que duplican cierres inferiores y no "
            "pertenecen a la estructura aprobada: "
            + "; ".join(str(path) for path in prohibidos)
        )

    top_dirs = [
        path
        for path in carpeta_periodo.iterdir()
        if path.is_dir() and debe_incluir_directorio(path)
    ]
    trimestrales = [path for path in top_dirs if path.name == "Trimestral"]
    meses = [path for path in top_dirs if RE_MES.match(path.name)]
    desconocidos = [
        path
        for path in top_dirs
        if path.name != "Trimestral" and not RE_MES.match(path.name)
    ]

    if len(trimestrales) != 1:
        raise RuntimeError(
            "Debe existir exactamente una carpeta Trimestral en la raíz "
            f"del periodo. Encontradas={len(trimestrales)}"
        )
    if not meses:
        raise RuntimeError("El trimestre no contiene carpetas mensuales.")
    if desconocidos:
        raise RuntimeError(
            "La raíz del trimestre contiene directorios no aprobados: "
            + "; ".join(path.name for path in desconocidos)
        )

    carpeta_trimestral = trimestrales[0]
    subdirs_trimestral = {
        path.name
        for path in carpeta_trimestral.iterdir()
        if path.is_dir() and debe_incluir_directorio(path)
    }
    if subdirs_trimestral != DIRECTORIOS_TRIMESTRALES:
        raise RuntimeError(
            "La carpeta Trimestral no contiene exactamente la estructura "
            f"aprobada. Encontrados={sorted(subdirs_trimestral)} | "
            f"Esperados={sorted(DIRECTORIOS_TRIMESTRALES)}"
        )

    if any(RE_MES.match(path.name) for path in carpeta_trimestral.rglob("*") if path.is_dir()):
        raise RuntimeError("Se detectaron carpetas mensuales dentro de Trimestral.")
    if any(RE_SEMANA.match(path.name) for path in carpeta_trimestral.rglob("*") if path.is_dir()):
        raise RuntimeError("Se detectaron semanas dentro de Trimestral.")
    if any(RE_DIARIO.match(path.name) for path in carpeta_trimestral.rglob("*") if path.is_dir()):
        raise RuntimeError("Se detectaron diarios dentro de Trimestral.")

    resumen = {
        "meses": 0,
        "semanas": 0,
        "diarios": 0,
        "cierres_semanales": 0,
        "cierres_mensuales": 0,
        "cierres_trimestrales": 1,
        "advertencias": [],
    }

    for mes_dir in sorted(meses, key=lambda p: p.name.lower()):
        match_mes = RE_MES.match(mes_dir.name)
        assert match_mes is not None
        mes_clave = f"{match_mes.group(1)}-{match_mes.group(2)}"
        resumen["meses"] += 1

        hijos_dir = [
            path
            for path in mes_dir.iterdir()
            if path.is_dir() and debe_incluir_directorio(path)
        ]
        semanas = [path for path in hijos_dir if RE_SEMANA.match(path.name)]
        mensuales = [path for path in hijos_dir if path.name == "Mensual"]
        desconocidos_mes = [
            path
            for path in hijos_dir
            if path.name != "Mensual" and not RE_SEMANA.match(path.name)
        ]
        if desconocidos_mes:
            raise RuntimeError(
                f"El mes {mes_dir.name} contiene directorios no aprobados: "
                + "; ".join(path.name for path in desconocidos_mes)
            )
        if len(mensuales) > 1:
            raise RuntimeError(
                f"El mes {mes_dir.name} contiene más de una carpeta Mensual."
            )
        if mensuales:
            resumen["cierres_mensuales"] += 1
        else:
            resumen["advertencias"].append(
                f"{mes_dir.name} no contiene cierre Mensual."
            )

        for semana_dir in sorted(semanas, key=lambda p: p.name.lower()):
            match_semana = RE_SEMANA.match(semana_dir.name)
            assert match_semana is not None
            inicio_semana = parse_fecha_iso(match_semana.group(1), "inicio semana")
            fin_semana = parse_fecha_iso(match_semana.group(2), "fin semana")
            if inicio_semana > fin_semana:
                raise RuntimeError(
                    f"Semana con rango invertido: {semana_dir.name}"
                )
            if fin_semana.strftime("%Y-%m") != mes_clave:
                raise RuntimeError(
                    "La semana debe pertenecer al mes de su fecha final. "
                    f"Semana={semana_dir.name} | Mes={mes_dir.name}"
                )

            resumen["semanas"] += 1
            hijos_semana = [
                path
                for path in semana_dir.iterdir()
                if path.is_dir() and debe_incluir_directorio(path)
            ]
            diarios = [path for path in hijos_semana if RE_DIARIO.match(path.name)]
            semanales = [path for path in hijos_semana if path.name == "Semanal"]
            desconocidos_semana = [
                path
                for path in hijos_semana
                if path.name != "Semanal" and not RE_DIARIO.match(path.name)
            ]
            if desconocidos_semana:
                raise RuntimeError(
                    f"La semana {semana_dir.name} contiene directorios no aprobados: "
                    + "; ".join(path.name for path in desconocidos_semana)
                )
            if len(semanales) > 1:
                raise RuntimeError(
                    f"La semana {semana_dir.name} contiene más de una carpeta Semanal."
                )
            if semanales:
                resumen["cierres_semanales"] += 1
                diarios_dentro_semanal = [
                    path
                    for path in semanales[0].rglob("*")
                    if path.is_dir() and RE_DIARIO.match(path.name)
                ]
                if diarios_dentro_semanal:
                    raise RuntimeError(
                        "Se detectaron diarios copiados dentro de Semanal: "
                        + "; ".join(str(path) for path in diarios_dentro_semanal)
                    )
            else:
                resumen["advertencias"].append(
                    f"{semana_dir.name} no contiene cierre Semanal."
                )

            for diario_dir in diarios:
                match_diario = RE_DIARIO.match(diario_dir.name)
                assert match_diario is not None
                fecha_diario = parse_fecha_iso(match_diario.group(1), "fecha diario")
                if not (inicio_semana <= fecha_diario <= fin_semana):
                    raise RuntimeError(
                        "El diario está fuera del rango de su semana. "
                        f"Diario={diario_dir.name} | Semana={semana_dir.name}"
                    )
                if not (fecha_inicio <= fecha_diario <= fecha_fin):
                    raise RuntimeError(
                        "El diario está fuera del rango del trimestre. "
                        f"Diario={diario_dir.name} | "
                        f"Trimestre={fecha_inicio} a {fecha_fin}"
                    )
                resumen["diarios"] += 1

    return resumen


def resolver_preparacion_trimestral() -> dict[str, Any]:
    estado_activo = leer_json(STATE_PATH, "el estado trimestral activo")
    if str(estado_activo.get("estado") or "").strip().upper() != "ACTIVO":
        raise RuntimeError("El estado trimestral no está ACTIVO.")

    fecha_inicio = parse_fecha_iso(
        estado_activo.get("fecha_inicio_periodo_activo"),
        "fecha_inicio_periodo_activo",
    )
    fecha_fin = parse_fecha_iso(
        estado_activo.get("proximo_cierre_estimado"),
        "proximo_cierre_estimado",
    )
    trimestre_inicio = cargar_trimestre_activo(ROOT, fecha_inicio)
    trimestre_fin = cargar_trimestre_activo(ROOT, fecha_fin)

    if trimestre_inicio["ruta_relativa"] != trimestre_fin["ruta_relativa"]:
        raise RuntimeError(
            "El helper trimestral devolvió rutas distintas para el periodo activo."
        )

    carpeta_periodo = CIERRES_DIR / Path(trimestre_fin["ruta_relativa"])
    estructura = validar_estructura_jerarquica(
        carpeta_periodo,
        fecha_inicio,
        fecha_fin,
    )
    carpeta_trimestral = carpeta_periodo / "Trimestral"
    carpeta_excel = carpeta_trimestral / "01_Excel_Cierre"
    carpeta_manifest = carpeta_trimestral / "02_Manifest"
    carpeta_validaciones = carpeta_trimestral / "03_Validaciones"
    carpeta_resumen = carpeta_trimestral / "04_Resumen"

    estados_preparacion = sorted(
        carpeta_validaciones.glob("estado_preparacion_cierre_trimestral_*.json")
    )
    if len(estados_preparacion) != 1:
        raise RuntimeError(
            "Se esperaba exactamente un estado de preparación y se encontraron "
            f"{len(estados_preparacion)}."
        )
    estado_preparacion_path = estados_preparacion[0]
    preparacion = leer_json(
        estado_preparacion_path,
        "el estado de preparación trimestral",
    )

    if preparacion.get("estado") != ESTADO_PREPARADO:
        raise RuntimeError(
            f"Estado de preparación inesperado: {preparacion.get('estado')!r}"
        )
    politica_preparacion = str(preparacion.get("politica_subida") or "").strip()
    if politica_preparacion not in {
        "SOLO_CARPETA_TRIMESTRAL",
        "JERARQUIA_COMPLETA_UN_DESTINO",
    }:
        raise RuntimeError(
            "Política de preparación no reconocida: "
            f"{politica_preparacion!r}"
        )

    if str(preparacion.get("periodo") or "").strip() != str(
        trimestre_fin["periodo_activo"]
    ):
        raise RuntimeError("El periodo preparado no coincide con el activo.")
    if str(preparacion.get("nombre_carpeta_trimestre") or "").strip() != str(
        trimestre_fin["nombre_carpeta"]
    ):
        raise RuntimeError("La carpeta preparada no coincide con el estado.")
    if str(preparacion.get("fecha_inicio") or "").strip() != fecha_inicio.isoformat():
        raise RuntimeError("La fecha inicial de la preparación es distinta.")
    if str(preparacion.get("fecha_fin") or "").strip() != fecha_fin.isoformat():
        raise RuntimeError("La fecha final de la preparación es distinta.")

    validar_path_exacto(
        preparacion.get("carpeta_periodo"),
        carpeta_periodo,
        "la carpeta del periodo",
    )
    validar_path_exacto(
        preparacion.get("carpeta_cierre_trimestral"),
        carpeta_trimestral,
        "la carpeta de cierre trimestral",
    )

    excel_historico = carpeta_excel / (
        f"facturas_{fecha_inicio.isoformat()}_a_{fecha_fin.isoformat()}.xlsx"
    )
    manifestes = sorted(carpeta_manifest.glob("manifest_cierre_trimestral_*.json"))
    inventarios = sorted(
        carpeta_manifest.glob("inventario_jerarquia_trimestral_*.json")
    )
    validaciones_locales = sorted(
        carpeta_validaciones.glob("validacion_local_cierre_trimestral_*.json")
    )
    resumenes = sorted(
        carpeta_resumen.glob("RESUMEN_CIERRE_TRIMESTRAL_*.txt")
    )

    for descripcion, encontrados in {
        "manifest": manifestes,
        "inventario": inventarios,
        "validación local": validaciones_locales,
        "resumen": resumenes,
    }.items():
        if len(encontrados) != 1:
            raise RuntimeError(
                f"Se esperaba exactamente un {descripcion} y se encontraron "
                f"{len(encontrados)}."
            )

    manifest = manifestes[0]
    inventario = inventarios[0]
    validacion_local = validaciones_locales[0]
    resumen = resumenes[0]

    validar_path_exacto(
        preparacion.get("excel_historico"),
        excel_historico,
        "el Excel histórico",
    )
    validar_path_exacto(preparacion.get("manifest"), manifest, "el manifest")
    validar_path_exacto(
        preparacion.get("inventario_jerarquia"),
        inventario,
        "el inventario",
    )
    validar_path_exacto(
        preparacion.get("validacion_local"),
        validacion_local,
        "la validación local",
    )

    carpeta_control = (
        DATA_DIR
        / "state"
        / "preparaciones_trimestrales"
        / trimestre_fin["nombre_carpeta"]
    )
    candidato = Path(str(preparacion.get("excel_limpio_candidato") or ""))
    respaldo_estado = Path(str(preparacion.get("respaldo_estado_activo") or ""))
    if candidato.resolve().parent != carpeta_control.resolve():
        raise RuntimeError(
            "El Excel limpio candidato no está fuera del backup en la carpeta "
            f"de control esperada: {carpeta_control}"
        )
    if respaldo_estado.resolve().parent != carpeta_control.resolve():
        raise RuntimeError(
            "El respaldo del estado no está en la carpeta de control esperada."
        )

    requeridos = [
        excel_historico,
        manifest,
        inventario,
        validacion_local,
        estado_preparacion_path,
        resumen,
        candidato,
        respaldo_estado,
        FACTURAS_PATH,
    ]
    faltantes = [str(path) for path in requeridos if not path.is_file()]
    if faltantes:
        raise RuntimeError(
            "La preparación trimestral está incompleta. Faltan: "
            + "; ".join(faltantes)
        )

    hash_historico = str(preparacion.get("excel_historico_sha256") or "")
    hash_activo = str(preparacion.get("excel_activo_sha256_al_preparar") or "")
    hash_candidato = str(
        preparacion.get("excel_limpio_candidato_sha256") or ""
    )
    hash_jerarquia_preparada = str(
        preparacion.get("jerarquia_sha256_al_preparar") or ""
    )
    if not all(
        [hash_historico, hash_activo, hash_candidato, hash_jerarquia_preparada]
    ):
        raise RuntimeError("La preparación no contiene todos los hashes obligatorios.")
    if sha256_file(excel_historico) != hash_historico:
        raise RuntimeError("El Excel histórico cambió después de prepararse.")
    if sha256_file(FACTURAS_PATH) != hash_activo:
        raise RuntimeError("data/facturas.xlsx cambió después de la preparación.")
    if hash_historico != hash_activo:
        raise RuntimeError("El Excel histórico y el activo preparado no coinciden.")
    if sha256_file(candidato) != hash_candidato:
        raise RuntimeError("El candidato limpio cambió después de prepararse.")

    candidato_info = validar_excel_limpio(candidato)
    manifest_datos = leer_json(manifest, "el manifest trimestral")
    inventario_datos = leer_json(inventario, "el inventario trimestral")
    validacion_datos = leer_json(validacion_local, "la validación local")
    if manifest_datos.get("estado") != ESTADO_PREPARADO:
        raise RuntimeError("El manifest no conserva el estado preparado.")
    if validacion_datos.get("ok") is not True:
        raise RuntimeError("La validación local trimestral no está OK.")
    if validacion_datos.get("cierres_inferiores_copiados") is not False:
        raise RuntimeError("La validación indica cierres inferiores copiados.")
    if validacion_datos.get("zip_creado") is not False:
        raise RuntimeError("La validación trimestral indica que creó un ZIP.")

    (
        hash_jerarquia_actual,
        archivos_jerarquia_inferior,
        bytes_jerarquia_inferior,
        meses_detectados,
    ) = digest_jerarquia_inferior(carpeta_periodo, fecha_inicio, fecha_fin)
    if hash_jerarquia_actual != hash_jerarquia_preparada:
        raise RuntimeError(
            "La jerarquía de meses, semanas y diarios cambió después de preparar."
        )
    if (
        str(inventario_datos.get("sha256_contenido_jerarquia") or "")
        != hash_jerarquia_actual
    ):
        raise RuntimeError(
            "El inventario no coincide con la jerarquía inferior preparada."
        )

    archivos = listar_archivos_locales(carpeta_periodo)
    directorios = listar_directorios_locales(carpeta_periodo)
    requeridos_base = {
        excel_historico.resolve(),
        manifest.resolve(),
        inventario.resolve(),
        validacion_local.resolve(),
        estado_preparacion_path.resolve(),
        resumen.resolve(),
    }
    presentes = {path.resolve() for path in archivos}
    faltan_base = sorted(str(path) for path in requeridos_base - presentes)
    if faltan_base:
        raise RuntimeError(
            "La jerarquía completa no incluye los seis archivos trimestrales "
            "obligatorios: "
            + "; ".join(faltan_base)
        )

    digest_completo, bytes_completos = digest_archivos(
        archivos,
        carpeta_periodo,
    )

    return {
        "anio": trimestre_fin["anio"],
        "periodo": trimestre_fin["nombre_carpeta"],
        "periodo_activo": trimestre_fin["periodo_activo"],
        "fecha_inicio": fecha_inicio.isoformat(),
        "fecha_fin": fecha_fin.isoformat(),
        "ruta_relativa_trimestre": trimestre_fin["ruta_relativa"],
        "carpeta_periodo": carpeta_periodo,
        "carpeta_trimestral": carpeta_trimestral,
        "carpeta_validaciones": carpeta_validaciones,
        "excel_historico": excel_historico,
        "manifest": manifest,
        "inventario": inventario,
        "validacion_local": validacion_local,
        "estado_preparacion": estado_preparacion_path,
        "resumen": resumen,
        "candidato": candidato,
        "candidato_info": candidato_info,
        "politica_preparacion": politica_preparacion,
        "estructura": estructura,
        "jerarquia_inferior_sha256": hash_jerarquia_actual,
        "archivos_jerarquia_inferior": archivos_jerarquia_inferior,
        "bytes_jerarquia_inferior": bytes_jerarquia_inferior,
        "meses_detectados": meses_detectados,
        "archivos": archivos,
        "directorios": directorios,
        "digest_completo": digest_completo,
        "bytes_completos": bytes_completos,
    }


def validar_configuracion_remota() -> None:
    if not BACKUP_DRIVE_ID:
        raise RuntimeError(
            "Falta BACKUP_DRIVE_ID en .env. "
            "No se usará un segundo destino como reemplazo automático."
        )
    if not BACKUP_ROOT_FOLDER:
        raise RuntimeError("Falta BACKUP_ROOT_FOLDER en .env.")
    if BACKUP_ROOT_FOLDER != ROOT_OFICIAL:
        raise RuntimeError(
            "BACKUP_ROOT_FOLDER no coincide con el único destino oficial. "
            f"Configurado={BACKUP_ROOT_FOLDER!r} | Esperado={ROOT_OFICIAL!r}"
        )


def calcular_ruta_remota(contexto: dict[str, Any]) -> str:
    ruta_relativa = str(contexto["ruta_relativa_trimestre"]).strip("/")
    partes = Path(ruta_relativa).parts
    if (
        len(partes) != 2
        or partes[0] != contexto["anio"]
        or partes[1] != contexto["periodo"]
    ):
        raise RuntimeError(
            "La ruta relativa trimestral no coincide con el periodo: "
            f"{ruta_relativa}"
        )

    remoto = f"{BACKUP_ROOT_FOLDER}/{ruta_relativa}".strip("/")
    if remoto.endswith("/Trimestral"):
        raise RuntimeError(
            "La ruta remota termina en Trimestral; debe apuntar al contenedor "
            "completo del trimestre."
        )
    if "03_Cierres_Trimestrales" in Path(remoto).parts:
        raise RuntimeError(
            "La ruta remota usa la jerarquía anterior retirada: " + remoto
        )
    if not remoto.startswith(f"{ROOT_OFICIAL}/"):
        raise RuntimeError("La ruta remota está fuera del destino oficial.")
    return remoto


def validar_drive() -> dict[str, Any]:
    url = (
        f"{GRAPH}/drives/{encode_drive_id(BACKUP_DRIVE_ID)}"
        "?$select=id,name,driveType,webUrl"
    )
    respuesta = requests.get(
        url,
        headers=headers(),
        timeout=60,
        verify=ssl_verify(),
    )
    if respuesta.status_code != 200:
        raise RuntimeError(
            f"GET {respuesta.status_code} {url_graph_segura(url)} -> "
            f"{respuesta.text[:500]}"
        )
    return respuesta.json()


def obtener_item_por_path(drive_id: str, remote_path: str) -> Optional[dict[str, Any]]:
    url = (
        f"{GRAPH}/drives/{encode_drive_id(drive_id)}"
        f"/root:/{encode_path(remote_path)}:"
        "?$select=id,name,size,file,folder,eTag,parentReference"
    )
    respuesta = requests.get(
        url,
        headers=headers(),
        timeout=60,
        verify=ssl_verify(),
    )
    if respuesta.status_code == 200:
        return respuesta.json()
    if respuesta.status_code == 404:
        return None
    raise RuntimeError(
        f"GET {respuesta.status_code} {url_graph_segura(url)} -> "
        f"{respuesta.text[:500]}"
    )


def existe_path(drive_id: str, remote_path: str) -> bool:
    if not remote_path.strip("/"):
        return True
    return obtener_item_por_path(drive_id, remote_path) is not None


def crear_folder(drive_id: str, parent_path: str, folder_name: str) -> None:
    body = {
        "name": folder_name,
        "folder": {},
        "@microsoft.graph.conflictBehavior": "fail",
    }
    if parent_path.strip("/"):
        url = (
            f"{GRAPH}/drives/{encode_drive_id(drive_id)}"
            f"/root:/{encode_path(parent_path)}:/children"
        )
    else:
        url = f"{GRAPH}/drives/{encode_drive_id(drive_id)}/root/children"

    respuesta = requests.post(
        url,
        headers=headers_json(),
        json=body,
        timeout=60,
        verify=ssl_verify(),
    )
    if respuesta.status_code in (200, 201, 409):
        return
    raise RuntimeError(
        f"POST {respuesta.status_code} {url_graph_segura(url)} -> "
        f"{respuesta.text[:500]}"
    )


def ensure_folder_recursive(drive_id: str, folder_path: str) -> None:
    actual = ""
    for parte in [segmento for segmento in folder_path.strip("/").split("/") if segmento]:
        siguiente = f"{actual}/{parte}".strip("/")
        if not existe_path(drive_id, siguiente):
            crear_folder(drive_id, actual, parte)
        actual = siguiente


def listar_hijos(drive_id: str, item_id: str) -> list[dict[str, Any]]:
    url: Optional[str] = (
        f"{GRAPH}/drives/{encode_drive_id(drive_id)}"
        f"/items/{quote(item_id, safe='')}/children"
        "?$select=id,name,size,file,folder,eTag,parentReference&$top=200"
    )
    resultado: list[dict[str, Any]] = []
    while url:
        respuesta = requests.get(
            url,
            headers=headers(),
            timeout=120,
            verify=ssl_verify(),
        )
        if respuesta.status_code != 200:
            raise RuntimeError(
                f"GET {respuesta.status_code} {url_graph_segura(url)} -> "
                f"{respuesta.text[:500]}"
            )
        datos = respuesta.json()
        resultado.extend(datos.get("value", []))
        url = datos.get("@odata.nextLink")
    return resultado


def inventariar_remoto(
    drive_id: str,
    remote_base: str,
) -> dict[str, Any]:
    base_item = obtener_item_por_path(drive_id, remote_base)
    if base_item is None:
        return {
            "base_existe": False,
            "directorios": set(),
            "archivos": {},
        }
    if not isinstance(base_item.get("folder"), dict):
        raise RuntimeError(
            "La ruta remota del trimestre existe, pero no es una carpeta: "
            + remote_base
        )

    directorios: set[str] = set()
    archivos: dict[str, dict[str, Any]] = {}
    pendientes: list[tuple[str, str]] = [(str(base_item["id"]), "")]

    while pendientes:
        item_id, prefijo = pendientes.pop()
        for item in listar_hijos(drive_id, item_id):
            nombre = str(item.get("name") or "").strip()
            if not nombre:
                continue
            relativa = f"{prefijo}/{nombre}".strip("/")
            if isinstance(item.get("folder"), dict):
                directorios.add(relativa)
                pendientes.append((str(item["id"]), relativa))
            elif isinstance(item.get("file"), dict):
                archivos[relativa] = item

    return {
        "base_existe": True,
        "directorios": directorios,
        "archivos": archivos,
    }


def graph_download_item_content(drive_id: str, item_id: str) -> bytes:
    url = (
        f"{GRAPH}/drives/{encode_drive_id(drive_id)}"
        f"/items/{quote(item_id, safe='')}/content"
    )
    respuesta = requests.get(
        url,
        headers=headers(),
        timeout=300,
        verify=ssl_verify(),
        allow_redirects=True,
    )
    if respuesta.status_code != 200:
        raise RuntimeError(
            f"DOWNLOAD {respuesta.status_code} {url_graph_segura(url)} -> "
            f"{respuesta.text[:500]}"
        )
    return respuesta.content


def escribir_temporal_descarga(relativa: str, data: bytes) -> Path:
    TMP_VERIFY_DIR.mkdir(parents=True, exist_ok=True)
    seguro = relativa.replace("/", "__").replace("\\", "__")
    path = TMP_VERIFY_DIR / f"remoto_{seguro}"
    path.write_bytes(data)
    return path


def comparar_local_remoto(
    local: Path,
    relativa: str,
    item_remoto: dict[str, Any],
) -> tuple[bool, str]:
    item_id = str(item_remoto.get("id") or "").strip()
    if not item_id:
        raise RuntimeError(f"El archivo remoto no tiene item.id: {relativa}")

    if local.suffix.lower() not in EXCEL_EXT:
        tamano_remoto = int(item_remoto.get("size") or -1)
        if tamano_remoto != local.stat().st_size:
            return False, "TAMAÑO_DISTINTO"

    contenido = graph_download_item_content(BACKUP_DRIVE_ID, item_id)

    if local.suffix.lower() in EXCEL_EXT:
        temporal = escribir_temporal_descarga(relativa, contenido)
        try:
            digest_local, _ = digest_datos_excel(local)
            digest_remoto, _ = digest_datos_excel(temporal)
            return (
                digest_local == digest_remoto,
                "DATOS_EXCEL_IGUALES"
                if digest_local == digest_remoto
                else "DATOS_EXCEL_DISTINTOS",
            )
        finally:
            temporal.unlink(missing_ok=True)

    hash_local = sha256_file(local)
    hash_remoto = sha256_bytes(contenido)
    return (
        hash_local == hash_remoto,
        "SHA256_IGUAL" if hash_local == hash_remoto else "SHA256_DISTINTO",
    )


def construir_plan_remoto(
    contexto: dict[str, Any],
    remote_base: str,
) -> dict[str, Any]:
    inventario = inventariar_remoto(BACKUP_DRIVE_ID, remote_base)

    dirs_locales = {
        path.relative_to(contexto["carpeta_periodo"]).as_posix()
        for path in contexto["directorios"]
    }
    archivos_locales = {
        path.relative_to(contexto["carpeta_periodo"]).as_posix(): path
        for path in contexto["archivos"]
    }
    dirs_remotos: set[str] = inventario["directorios"]
    archivos_remotos: dict[str, dict[str, Any]] = inventario["archivos"]

    conflictos = []
    for relativa in sorted(dirs_locales & set(archivos_remotos)):
        conflictos.append(
            f"La ruta local es carpeta pero la ruta remota es archivo: {relativa}"
        )
    for relativa in sorted(set(archivos_locales) & dirs_remotos):
        conflictos.append(
            f"La ruta local es archivo pero la ruta remota es carpeta: {relativa}"
        )

    faltan_directorios = sorted(
        dirs_locales - dirs_remotos,
        key=lambda ruta: (len(Path(ruta).parts), ruta.lower()),
    )
    identicos: list[str] = []
    faltantes: list[str] = []
    diferentes: list[str] = []
    detalles: dict[str, str] = {}

    existentes = [rel for rel in archivos_locales if rel in archivos_remotos]
    total_existentes = len(existentes)
    revisados = 0

    for relativa, local in archivos_locales.items():
        remoto = archivos_remotos.get(relativa)
        if remoto is None:
            faltantes.append(relativa)
            detalles[relativa] = "NO_EXISTE_REMOTO"
            continue

        iguales, motivo = comparar_local_remoto(local, relativa, remoto)
        detalles[relativa] = motivo
        if iguales:
            identicos.append(relativa)
        else:
            diferentes.append(relativa)

        revisados += 1
        if total_existentes >= 100 and revisados % 100 == 0:
            print(
                f"   Comparación remota: {revisados}/{total_existentes} "
                "archivos existentes revisados..."
            )

    gestionados = set(archivos_locales)
    prefijo_evidencia = (
        "Trimestral/03_Validaciones/" + PREFIJO_VALIDACION_REMOTA
    )
    extras = sorted(
        relativa
        for relativa in set(archivos_remotos) - gestionados
        if not relativa.startswith(prefijo_evidencia)
    )

    return {
        "base_existe": inventario["base_existe"],
        "faltan_directorios": faltan_directorios,
        "identicos": sorted(identicos),
        "faltantes": sorted(faltantes),
        "diferentes": sorted(diferentes),
        "extras": extras,
        "conflictos": conflictos,
        "detalles": detalles,
        "archivos_remotos": archivos_remotos,
        "directorios_remotos": dirs_remotos,
    }


def graph_put_content(drive_id: str, remote_path: str, local_file: Path) -> dict[str, Any]:
    url = (
        f"{GRAPH}/drives/{encode_drive_id(drive_id)}"
        f"/root:/{encode_path(remote_path)}:/content"
    )
    respuesta = requests.put(
        url,
        headers=headers(),
        data=local_file.read_bytes(),
        timeout=300,
        verify=ssl_verify(),
    )
    if respuesta.status_code not in (200, 201):
        raise RuntimeError(
            f"PUT {respuesta.status_code} {url_graph_segura(url)} -> "
            f"{respuesta.text[:500]}"
        )
    return respuesta.json()


def verificar_archivo_subido(
    local: Path,
    relativa: str,
    item: dict[str, Any],
) -> bool:
    item_id = str(item.get("id") or "").strip()
    if not item_id:
        raise RuntimeError("Graph no devolvió item.id para la verificación.")

    contenido = graph_download_item_content(BACKUP_DRIVE_ID, item_id)
    if local.suffix.lower() in EXCEL_EXT:
        temporal = escribir_temporal_descarga(relativa, contenido)
        try:
            digest_local, resumen_local = digest_datos_excel(local)
            digest_remoto, _ = digest_datos_excel(temporal)
            if digest_local != digest_remoto:
                print(f"❌ Excel remoto con datos distintos: {relativa}")
                return False
            print(
                f"✅ Excel verificado por datos: {relativa} | "
                f"celdas_no_vacias={resumen_local['non_empty_cells']}"
            )
            return True
        finally:
            temporal.unlink(missing_ok=True)

    if sha256_file(local) != sha256_bytes(contenido):
        print(f"❌ SHA256 remoto distinto: {relativa}")
        return False
    print(f"✅ Archivo verificado por SHA256: {relativa}")
    return True


def subir_y_verificar(
    remote_path: str,
    local: Path,
    relativa: str,
) -> bool:
    item: Optional[dict[str, Any]] = None
    ultimo_error: Optional[Exception] = None
    espera = 2

    for intento in range(1, 4):
        try:
            item = graph_put_content(BACKUP_DRIVE_ID, remote_path, local)
            break
        except Exception as exc:
            ultimo_error = exc
            print(f"⚠️ Subida intento {intento}/3 falló para {relativa}: {exc}")
            try:
                item_existente = obtener_item_por_path(
                    BACKUP_DRIVE_ID,
                    remote_path,
                )
                if item_existente:
                    item = item_existente
                    print(
                        "✅ El archivo existe pese al error del PUT; se "
                        "verificará su contenido."
                    )
                    break
            except Exception as consulta_exc:
                print(
                    "⚠️ No fue posible comprobar aún la ruta remota: "
                    f"{consulta_exc}"
                )
            if intento < 3:
                time.sleep(espera)
                espera *= 2

    if item is None:
        print(f"❌ No fue posible subir {relativa}: {ultimo_error}")
        return False

    espera = 2
    for intento in range(1, 5):
        try:
            if verificar_archivo_subido(local, relativa, item):
                return True
        except Exception as exc:
            print(
                f"⚠️ Verificación intento {intento}/4 falló para "
                f"{relativa}: {exc}"
            )
        if intento < 4:
            time.sleep(espera)
            espera *= 2

    print(f"❌ Verificación definitiva fallida: {relativa}")
    return False


def ruta_evidencia(contexto: dict[str, Any]) -> Path:
    periodo_seguro = re.sub(
        r"[^A-Za-z0-9._-]+",
        "_",
        contexto["periodo_activo"],
    ).strip("._-")
    return contexto["carpeta_validaciones"] / (
        f"{PREFIJO_VALIDACION_REMOTA}{periodo_seguro}.json"
    )


def evidencia_vigente(
    path: Path,
    contexto: dict[str, Any],
    remote_base: str,
) -> bool:
    if not path.is_file():
        return False
    try:
        datos = leer_json(path, "la evidencia remota existente")
    except Exception:
        return False
    return bool(
        datos.get("tipo") == "VALIDACION_REMOTA_ESTRUCTURA_TRIMESTRAL"
        and datos.get("ok") is True
        and datos.get("estado")
        == "JERARQUIA_COMPLETA_VALIDADA_EN_DESTINO_UNICO"
        and datos.get("destino_unico") is True
        and datos.get("remote_base") == remote_base
        and datos.get("digest_jerarquia_completa")
        == contexto["digest_completo"]
        and int(datos.get("total_archivos_base") or -1)
        == len(contexto["archivos"])
    )


def preparar_evidencia(
    contexto: dict[str, Any],
    remote_base: str,
    plan: dict[str, Any],
    creados: int,
    actualizados: int,
    omitidos: int,
) -> tuple[Path, bool]:
    path = ruta_evidencia(contexto)
    if evidencia_vigente(path, contexto, remote_base):
        return path, True

    datos = {
        "tipo": "VALIDACION_REMOTA_ESTRUCTURA_TRIMESTRAL",
        "version_script": VERSION_UPLOAD,
        "generado_en": datetime.datetime.now().astimezone().isoformat(
            timespec="seconds"
        ),
        "ok": True,
        "estado": "JERARQUIA_COMPLETA_VALIDADA_EN_DESTINO_UNICO",
        "destino_unico": True,
        "periodo": contexto["periodo"],
        "periodo_activo": contexto["periodo_activo"],
        "fecha_inicio": contexto["fecha_inicio"],
        "fecha_fin": contexto["fecha_fin"],
        "carpeta_local_publicada": str(contexto["carpeta_periodo"]),
        "remote_base": remote_base,
        "drive": resumen_drive(BACKUP_DRIVE_ID),
        "digest_jerarquia_completa": contexto["digest_completo"],
        "total_archivos_base": len(contexto["archivos"]),
        "total_directorios": len(contexto["directorios"]),
        "total_bytes": contexto["bytes_completos"],
        "estructura": contexto["estructura"],
        "sincronizacion": {
            "archivos_creados": creados,
            "archivos_actualizados": actualizados,
            "archivos_omitidos_por_ser_identicos": omitidos,
            "archivos_extras_remotos_conservados": len(plan["extras"]),
            "directorios_creados": len(plan["faltan_directorios"]),
            "eliminaciones_remotas": 0,
        },
        "controles": {
            "unico_destino": True,
            "jerarquia_completa": True,
            "cada_archivo_en_una_sola_ubicacion": True,
            "diarios_copiados_dentro_de_semanal": False,
            "semanas_copiadas_dentro_de_mensual": False,
            "meses_copiados_dentro_de_trimestral": False,
            "excel_limpio_candidato_subido": False,
            "excel_activo_reemplazado": False,
            "estado_trimestral_actualizado": False,
            "autoeliminaciones_ejecutadas": False,
            "archivos_remotos_eliminados": False,
        },
    }

    temporal = path.with_suffix(path.suffix + ".tmp")
    temporal.write_text(
        json.dumps(datos, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    temporal.replace(path)
    return path, False


def ruta_estado_retencion(
    contexto: dict[str, Any],
) -> Path:
    periodo_seguro = re.sub(
        r"[^A-Za-z0-9._-]+",
        "_",
        contexto["periodo_activo"],
    ).strip("._-")

    return contexto["carpeta_validaciones"] / (
        f"estado_retencion_trimestral_{periodo_seguro}.json"
    )


def crear_estado_retencion_temporal(
    contexto: dict[str, Any],
    remote_base: str,
) -> tuple[Path, Path, dict[str, Any]]:
    estado_local = ruta_estado_retencion(contexto)

    # Si ya existe un estado valido para exactamente esta misma
    # jerarquia, se conserva su fecha original de retencion.
    if estado_local.is_file():
        try:
            existente = leer_json(
                estado_local,
                "el estado local de retencion trimestral",
            )

            vigente = bool(
                existente.get("tipo")
                == "estado_retencion_local_cierre_trimestral"
                and existente.get("ok") is True
                and existente.get(
                    "validacion_remota_publicada_y_verificada"
                ) is True
                and existente.get("remote_base") == remote_base
                and existente.get("digest_jerarquia_completa")
                == contexto["digest_completo"]
                and int(existente.get("total_archivos_base") or -1)
                == len(contexto["archivos"])
                and int(existente.get("retencion_local_dias") or -1)
                == RETENCION_LOCAL_DIAS
                and bool(existente.get("validado_en"))
                and bool(
                    existente.get(
                        "eliminacion_local_permitida_desde"
                    )
                )
            )

            if vigente:
                TMP_VERIFY_DIR.mkdir(
                    parents=True,
                    exist_ok=True,
                )
                temporal = TMP_VERIFY_DIR / estado_local.name
                temporal.write_bytes(estado_local.read_bytes())

                return temporal, estado_local, existente

        except Exception:
            pass

    ahora = datetime.datetime.now(ZONA_HORARIA)

    payload = {
        "tipo": "estado_retencion_local_cierre_trimestral",
        "version": VERSION_UPLOAD,
        "periodo": contexto["periodo"],
        "periodo_activo": contexto["periodo_activo"],
        "fecha_inicio": contexto["fecha_inicio"],
        "fecha_fin": contexto["fecha_fin"],
        "validacion_remota_publicada_y_verificada": True,
        "validado_en": ahora.isoformat(timespec="seconds"),
        "retencion_local_dias": RETENCION_LOCAL_DIAS,
        "eliminacion_local_permitida_desde": (
            ahora + datetime.timedelta(days=RETENCION_LOCAL_DIAS)
        ).isoformat(timespec="seconds"),
        "remote_base": remote_base,
        "digest_jerarquia_completa": contexto["digest_completo"],
        "total_archivos_base": len(contexto["archivos"]),
        "ok": True,
    }

    TMP_VERIFY_DIR.mkdir(parents=True, exist_ok=True)
    temporal = TMP_VERIFY_DIR / estado_local.name

    temporal.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    return temporal, estado_local, payload


def limpiar_temporales() -> None:
    try:
        if TMP_VERIFY_DIR.exists():
            for path in TMP_VERIFY_DIR.iterdir():
                if path.is_file():
                    path.unlink(missing_ok=True)
            TMP_VERIFY_DIR.rmdir()
    except Exception:
        pass


def imprimir_diagnostico_local(
    contexto: dict[str, Any],
    remote_base: str,
) -> None:
    print("Jerarquía trimestral local validada.")
    print(f"Periodo operativo: {contexto['periodo_activo']}")
    print(f"Rango: {contexto['fecha_inicio']} a {contexto['fecha_fin']}")
    print(f"Carpeta completa que se sincronizará: {contexto['carpeta_periodo']}")
    print(f"Destino único: {remote_base}")
    print(f"Drive: {resumen_drive(BACKUP_DRIVE_ID)}")
    print(f"Directorios locales: {len(contexto['directorios'])}")
    print(f"Archivos base: {len(contexto['archivos'])}")
    print(f"Bytes totales: {contexto['bytes_completos']}")
    print(f"SHA256 lógico de la jerarquía completa: {contexto['digest_completo']}")
    estructura = contexto["estructura"]
    print(
        "Niveles detectados: "
        f"meses={estructura['meses']} | "
        f"semanas={estructura['semanas']} | "
        f"diarios={estructura['diarios']} | "
        f"cierres_semanales={estructura['cierres_semanales']} | "
        f"cierres_mensuales={estructura['cierres_mensuales']} | "
        f"cierres_trimestrales={estructura['cierres_trimestrales']}"
    )
    if estructura["advertencias"]:
        print("Advertencias de completitud histórica:")
        for advertencia in estructura["advertencias"]:
            print(f"   - {advertencia}")
    if contexto["politica_preparacion"] == "SOLO_CARPETA_TRIMESTRAL":
        print(
            "Nota: la preparación local conserva el nombre de política "
            "anterior, pero su digest de la jerarquía inferior fue revalidado."
        )
    print("-" * 100)
    print("CONTROLES:")
    print(f"- Único destino oficial: {ROOT_OFICIAL}")
    print("- Se sincroniza el trimestre completo, no solo Trimestral.")
    print("- Los archivos idénticos se omiten y no se vuelven a subir.")
    print("- No se eliminan archivos remotos.")
    print("- No se sube el candidato limpio.")
    print("- No se reemplaza el Excel activo local ni remoto.")
    print("- No se actualiza cierre_trimestral_state.json.")
    print("- No se ejecutan autoeliminaciones en este script.")


def imprimir_plan(plan: dict[str, Any]) -> None:
    print("-" * 100)
    print("PLAN REMOTO DE SOLO LECTURA")
    print(f"Carpeta trimestral remota existente: {plan['base_existe']}")
    print(f"Directorios faltantes: {len(plan['faltan_directorios'])}")
    print(f"Archivos idénticos que se omitirán: {len(plan['identicos'])}")
    print(f"Archivos faltantes que se crearían: {len(plan['faltantes'])}")
    print(f"Archivos diferentes que se actualizarían: {len(plan['diferentes'])}")
    print(f"Archivos extras remotos que se conservarán: {len(plan['extras'])}")
    print(f"Conflictos de tipo archivo/carpeta: {len(plan['conflictos'])}")

    for titulo, valores in (
        ("CONFLICTOS", plan["conflictos"]),
        ("PRIMEROS ARCHIVOS FALTANTES", plan["faltantes"][:20]),
        ("PRIMEROS ARCHIVOS DIFERENTES", plan["diferentes"][:20]),
        ("PRIMEROS EXTRAS REMOTOS", plan["extras"][:20]),
    ):
        if valores:
            print(f"{titulo}:")
            for valor in valores:
                print(f"   - {valor}")


def ejecutar_sincronizacion(
    contexto: dict[str, Any],
    remote_base: str,
    plan: dict[str, Any],
) -> int:
    if plan["conflictos"]:
        print("Sincronización bloqueada por conflictos archivo/carpeta.")
        return 1

    print("-" * 100)
    print("Preparando estructura remota faltante...")
    ensure_folder_recursive(BACKUP_DRIVE_ID, remote_base)
    for relativa in plan["faltan_directorios"]:
        ensure_folder_recursive(
            BACKUP_DRIVE_ID,
            f"{remote_base}/{relativa}".strip("/"),
        )

    archivos_por_relativa = {
        path.relative_to(contexto["carpeta_periodo"]).as_posix(): path
        for path in contexto["archivos"]
    }
    pendientes = [
        ("CREAR", relativa) for relativa in plan["faltantes"]
    ] + [
        ("ACTUALIZAR", relativa) for relativa in plan["diferentes"]
    ]

    creados = 0
    actualizados = 0
    errores: list[str] = []

    for indice, (accion, relativa) in enumerate(pendientes, start=1):
        local = archivos_por_relativa[relativa]
        remote_path = f"{remote_base}/{relativa}".strip("/")
        try:
            print(
                f"[{indice}/{len(pendientes)}] {accion}: {relativa}"
            )
            if not subir_y_verificar(remote_path, local, relativa):
                errores.append(relativa)
            elif accion == "CREAR":
                creados += 1
            else:
                actualizados += 1
        except Exception as exc:
            print(f"❌ Error procesando {relativa}: {exc}")
            errores.append(relativa)

    if errores:
        print("-" * 100)
        print(f"Sincronización incompleta. Archivos con error: {len(errores)}")
        for relativa in errores[:50]:
            print(f"   - {relativa}")
        print("No se generó evidencia de éxito.")
        return 1

    evidencia_path, reutilizada = preparar_evidencia(
        contexto,
        remote_base,
        plan,
        creados,
        actualizados,
        len(plan["identicos"]),
    )
    relativa_evidencia = evidencia_path.relative_to(
        contexto["carpeta_periodo"]
    ).as_posix()
    remote_evidencia = f"{remote_base}/{relativa_evidencia}".strip("/")

    print("-" * 100)
    if reutilizada:
        print("Evidencia local vigente reutilizada sin reescribirla.")
        item = obtener_item_por_path(BACKUP_DRIVE_ID, remote_evidencia)
        if item is not None:
            iguales, _ = comparar_local_remoto(
                evidencia_path,
                relativa_evidencia,
                item,
            )
            if iguales:
                print("Evidencia remota ya era idéntica; no se volvió a subir.")
            elif not subir_y_verificar(
                remote_evidencia,
                evidencia_path,
                relativa_evidencia,
            ):
                return 1
        elif not subir_y_verificar(
            remote_evidencia,
            evidencia_path,
            relativa_evidencia,
        ):
            return 1
    else:
        print(f"Publicando evidencia final: {relativa_evidencia}")
        if not subir_y_verificar(
            remote_evidencia,
            evidencia_path,
            relativa_evidencia,
        ):
            return 1

    try:
        (
            estado_tmp,
            estado_local,
            estado_payload,
        ) = crear_estado_retencion_temporal(
            contexto,
            remote_base,
        )

        relativa_estado = estado_local.relative_to(
            contexto["carpeta_periodo"]
        ).as_posix()

        remote_estado = (
            f"{remote_base}/{relativa_estado}"
        ).strip("/")

        print("-" * 100)
        print("Publicando estado de retencion trimestral:")
        print(f"   Local temporal: {estado_tmp}")
        print(f"   Remoto:         {remote_estado}")

        item_estado = obtener_item_por_path(
            BACKUP_DRIVE_ID,
            remote_estado,
        )

        if item_estado is not None:
            iguales, _ = comparar_local_remoto(
                estado_tmp,
                relativa_estado,
                item_estado,
            )

            if iguales:
                print(
                    "Estado remoto de retencion ya era "
                    "identico; no se volvio a subir."
                )
            elif not subir_y_verificar(
                remote_estado,
                estado_tmp,
                relativa_estado,
            ):
                raise RuntimeError(
                    "No se pudo publicar/verificar el estado "
                    "remoto de retencion trimestral."
                )

        elif not subir_y_verificar(
            remote_estado,
            estado_tmp,
            relativa_estado,
        ):
            raise RuntimeError(
                "No se pudo publicar/verificar el estado "
                "remoto de retencion trimestral."
            )

        estado_local.parent.mkdir(
            parents=True,
            exist_ok=True,
        )

        temporal_local = estado_local.with_suffix(
            estado_local.suffix + ".tmp"
        )
        temporal_local.write_bytes(
            estado_tmp.read_bytes()
        )
        temporal_local.replace(estado_local)

        estado_tmp.unlink(missing_ok=True)

        print(
            "Retencion local trimestral habilitada hasta: "
            f"{estado_payload['eliminacion_local_permitida_desde']}"
        )

    except Exception as exc:
        print(
            "ERROR registrando retencion trimestral: "
            f"{type(exc).__name__}: {exc}"
        )
        return 1

    print("-" * 100)
    print("✅ JERARQUÍA TRIMESTRAL COMPLETA SINCRONIZADA Y VERIFICADA.")
    print(f"Archivos creados: {creados}")
    print(f"Archivos actualizados: {actualizados}")
    print(f"Archivos idénticos omitidos: {len(plan['identicos'])}")
    print(f"Extras remotos conservados sin eliminación: {len(plan['extras'])}")
    print("Destino utilizado: uno solo.")
    print("El Excel activo y el estado trimestral permanecen intactos.")
    print(
        f"Retencion local trimestral: politica de "
        f"{RETENCION_LOCAL_DIAS} dias; el vencimiento original "
        "se conserva si la jerarquia no cambia."
    )
    print("Las autoeliminaciones siguen pendientes y bloqueadas.")
    return 0


def main() -> int:
    parser = argparse.ArgumentParser(
        description=(
            "Valida, compara o sincroniza la jerarquía trimestral completa "
            "en Backups_Facturas_Produccion."
        )
    )
    modos = parser.add_mutually_exclusive_group()
    modos.add_argument(
        "--dry-run",
        action="store_true",
        help="Valida localmente; no usa Graph.",
    )
    modos.add_argument(
        "--plan-remoto",
        action="store_true",
        help="Compara con Graph en solo lectura; no modifica nada.",
    )
    modos.add_argument(
        "--upload-cierre",
        action="store_true",
        help="Sincroniza la jerarquía trimestral completa.",
    )
    parser.add_argument(
        "--confirmar",
        default=None,
        help="Confirmación obligatoria para la sincronización real.",
    )
    args = parser.parse_args()

    if args.upload_cierre:
        modo = "SINCRONIZAR JERARQUÍA COMPLETA"
    elif args.plan_remoto:
        modo = "PLAN REMOTO SOLO LECTURA"
    else:
        modo = "DRY RUN LOCAL"

    print("=" * 100)
    print(f"BACKUP TRIMESTRAL - UN DESTINO - {modo}")
    print("=" * 100)
    print(f"Versión: {VERSION_UPLOAD}")
    print(f"Root: {ROOT}")
    print("-" * 100)

    try:
        contexto = resolver_preparacion_trimestral()
        remote_base = calcular_ruta_remota(contexto)
        imprimir_diagnostico_local(contexto, remote_base)

        if not args.plan_remoto and not args.upload_cierre:
            print("-" * 100)
            print("DRY-RUN local finalizado.")
            print("No se usó Graph ni se modificó ningún archivo.")
            print("=" * 100)
            return 0

        validar_configuracion_remota()
        print("-" * 100)
        print("Validando el único drive de backup...")
        drive = validar_drive()
        print(
            f"Drive validado: {drive.get('name')} | "
            f"tipo={drive.get('driveType')} | "
            f"{resumen_drive(str(drive.get('id') or ''))}"
        )
        print("Inventariando y comparando el destino en modo seguro...")
        plan = construir_plan_remoto(contexto, remote_base)
        imprimir_plan(plan)

        if args.plan_remoto:
            print("-" * 100)
            print("PLAN REMOTO FINALIZADO.")
            print("No se creó, subió, actualizó ni eliminó ningún archivo.")
            print("=" * 100)
            return 1 if plan["conflictos"] else 0

        if args.confirmar != CONFIRMACION_UPLOAD:
            print("-" * 100)
            print("Sincronización real bloqueada por falta de confirmación.")
            print(
                "Usa --upload-cierre --confirmar "
                f"{CONFIRMACION_UPLOAD}"
            )
            print("=" * 100)
            return 1

        resultado = ejecutar_sincronizacion(contexto, remote_base, plan)
        print("=" * 100)
        return resultado
    except Exception as exc:
        print(f"❌ Error en el backup trimestral: {type(exc).__name__}: {exc}")
        print("No se modificó el Excel activo ni el estado trimestral.")
        print("=" * 100)
        return 1
    finally:
        limpiar_temporales()


if __name__ == "__main__":
    raise SystemExit(main())
