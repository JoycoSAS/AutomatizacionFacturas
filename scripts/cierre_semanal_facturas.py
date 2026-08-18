# -*- coding: utf-8 -*-
"""
JOYCO - Facturas Procesador
Cierre semanal local seguro V4

Politica:
- El cierre semanal soporta modo normal y modo de reconstruccion controlada.
- En modo normal usa Excel operativo + auditorias; en reconstruccion concatena cierres diarios fisicos validados.
- Los cierres diarios permanecen como carpetas hermanas de Semanal y no se copian dentro del cierre semanal.
- Incluye logs de producción local y VPS correspondientes al rango semanal.
- No sube a OneDrive/SharePoint. La subida se hace con scripts/subir_cierre_semanal_sharepoint.py.
- No incluye .env real ni secretos.

Estructura generada:
  data/cierres_diarios/YYYY/TRIMESTRE_YYYY-MM-DD_A_YYYY-MM-DD/
    YYYY-MM_Mes/SEMANA_YYYY-MM-DD_a_YYYY-MM-DD/Semanal/
      01_Excel_Semanal/facturas_semanal_YYYY-MM-DD_a_YYYY-MM-DD.xlsx
      02_Auditorias_Semana/audit_*.csv
      04_Manifest_Semanal/manifest_semanal_YYYY-MM-DD_a_YYYY-MM-DD.json
      04_Manifest_Semanal/resumen_semanal_YYYY-MM-DD_a_YYYY-MM-DD.txt
      05_Validaciones/validacion_local_semanal_YYYY-MM-DD_a_YYYY-MM-DD.json
"""

from __future__ import annotations

import argparse
import csv
import datetime as _dt
import hashlib
import json
import os
import platform
import re
import shutil
import socket
import sys
import time
import unicodedata
from pathlib import Path
from typing import Any, Optional, Sequence, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

try:
    import config  # noqa: F401
except Exception:
    pass

from trimestre_activo import cargar_trimestre_activo

VERSION_CIERRE = "2026-08-13-CIERRE-SEMANAL-V4-DESDE-DIARIOS-RECONSTRUCCION"

DATA_DIR = ROOT / "data"
AUDIT_DIR = DATA_DIR / "audit"
LOGS_DIR = ROOT / "logs"
DATA_LOGS_DIR = DATA_DIR / "logs"

# En local, los logs pueden permanecer dentro del repositorio.
# En el VPS se almacenan fuera del código para separar aplicación y operación.
VPS_LOG_ROOT = Path(
    os.getenv(
        "FACTURAS_LOG_ROOT",
        "/var/log/joyco/facturas-procesador",
    )
)

CIERRES_DIR = DATA_DIR / "cierres_diarios"
LOCKS_DIR = DATA_DIR / "locks"

FACTURAS_PATH = Path(os.getenv("ARCHIVO_EXCEL_LOCAL", str(DATA_DIR / "facturas.xlsx")))
if not FACTURAS_PATH.is_absolute():
    FACTURAS_PATH = ROOT / FACTURAS_PATH

CIERRE_SEMANAL_DESDE_DIARIOS = (
    str(os.getenv("CIERRE_SEMANAL_DESDE_DIARIOS") or "")
    .strip()
    .lower()
    in {"1", "true", "si", "sí", "yes", "on"}
)

LOCK_FILE = LOCKS_DIR / "cierre_semanal_facturas.lock"
LOCK_TTL_SECONDS = int(os.getenv("LOCK_TTL_SECONDS", "3600") or "3600")

MESES_ES = {
    1: "Enero",
    2: "Febrero",
    3: "Marzo",
    4: "Abril",
    5: "Mayo",
    6: "Junio",
    7: "Julio",
    8: "Agosto",
    9: "Septiembre",
    10: "Octubre",
    11: "Noviembre",
    12: "Diciembre",
}

PALABRAS_SENSIBLES = (
    "SECRET",
    "PASSWORD",
    "PASS",
    "TOKEN",
    "KEY",
    "CLIENT_SECRET",
    "PRIVATE",
    "CERT",
)

NOMBRES_EXCLUIDOS = {
    ".env",
    ".env.local",
    ".env.production",
    "token_cache.json",
    "msal_cache.bin",
}

COLUMNAS_FECHA_PROCESO_CANDIDATAS = [
    "fecha_procesamiento",
    "fecha procesamiento",
    "fecha proceso",
    "procesado_en",
    "procesado en",
    "fecha_registro",
    "fecha registro",
    "fecha de registro",
    "fecha_carga",
    "fecha carga",
    "creado_en",
    "created_at",
]

COLS_NUMERO = [
    "número de factura",
    "numero de factura",
    "numerofactura",
    "numero factura",
    "factura",
    "numero",
]
COLS_RADICADO = ["radicado"]
COLS_CUFE = ["cufe", "cude", "cufe/cude", "uuid"]
COLS_ARCHIVO = ["archivo", "archivo origen", "nombre archivo", "pdf", "xml"]
COLS_ARCHIVO_AUDIT_FALLBACK = [
    "pdf_name",
    "pdf name",
    "archivo_pdf",
    "archivo pdf",
    "nombre_pdf",
    "nombre pdf",
    "attachment_name",
    "attachment name",
]
COLS_ASUNTO_AUDIT = [
    "subject",
    "subj",
    "asunto",
    "email_subject",
    "email subject",
    "mail_subject",
    "mail subject",
]

PATRON_RADICADO_ASUNTO = re.compile(
    r"(?i)\bradicado"
    r"(?:\s*(?:n[.°ºo]*|no\.?))?"
    r"\s*[-:#]?\s*(\d{4,})\b"
)

EXT_LOGS = {".log", ".txt", ".csv", ".json"}
EXCEL_SHEET_NAME = "Facturas"
EXCEL_TABLE_NAME = "TblFacturasSemanal"


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Genera cierre semanal local seguro V4.")
    parser.add_argument(
        "--fecha",
        default=_dt.date.today().isoformat(),
        help="Fecha dentro de la semana a cerrar, formato YYYY-MM-DD. Default: hoy.",
    )
    parser.add_argument(
        "--inicio",
        default="",
        help="Inicio de semana/rango en formato YYYY-MM-DD. Opcional.",
    )
    parser.add_argument(
        "--fin",
        default="",
        help="Fin de semana/rango en formato YYYY-MM-DD. Opcional.",
    )
    parser.add_argument("--dry-run", action="store_true", help="Diagnostica sin crear/copiar archivos.")
    parser.add_argument(
        "--permitir-vacio",
        action="store_true",
        help="Permite generar cierre aunque no haya filas semanales.",
    )
    return parser.parse_args()


def validar_fecha(fecha: str) -> _dt.date:
    try:
        return _dt.date.fromisoformat(str(fecha).strip())
    except Exception as exc:
        raise RuntimeError(f"Fecha invalida {fecha!r}. Usa formato YYYY-MM-DD.") from exc


def rango_semana_lunes_domingo(fecha: _dt.date) -> tuple[_dt.date, _dt.date]:
    inicio = fecha - _dt.timedelta(days=fecha.weekday())
    fin = inicio + _dt.timedelta(days=6)
    return inicio, fin


def rango_desde_args(args: argparse.Namespace) -> tuple[_dt.date, _dt.date]:
    if args.inicio or args.fin:
        if not args.inicio or not args.fin:
            raise RuntimeError("Si usas --inicio o --fin debes indicar ambos.")
        inicio = validar_fecha(args.inicio)
        fin = validar_fecha(args.fin)
        if fin < inicio:
            raise RuntimeError("El --fin no puede ser menor que --inicio.")
        return inicio, fin
    return rango_semana_lunes_domingo(validar_fecha(args.fecha))


def fechas_en_rango(inicio: _dt.date, fin: _dt.date) -> list[_dt.date]:
    dias = []
    actual = inicio
    while actual <= fin:
        dias.append(actual)
        actual += _dt.timedelta(days=1)
    return dias


def mes_carpeta(fecha: _dt.date) -> str:
    return f"{fecha:%Y-%m}_{MESES_ES.get(fecha.month, fecha.strftime('%B'))}"


def semana_nombre(inicio: _dt.date, fin: _dt.date) -> str:
    return f"SEMANA_{inicio.isoformat()}_a_{fin.isoformat()}"


def rel(path: Path) -> str:
    try:
        return str(path.resolve().relative_to(ROOT.resolve())).replace("\\", "/")
    except Exception:
        return str(path).replace("\\", "/")


def norm(value: Any) -> str:
    s = "" if value is None else str(value)
    s = s.strip().replace("\xa0", " ")
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = re.sub(r"\s+", " ", s)
    return s.lower()


def norm_key(value: Any) -> str:
    s = norm(value)
    return re.sub(r"[^a-z0-9]+", "", s)


def clean_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (_dt.datetime, _dt.date)):
        return value.isoformat()
    return str(value).strip()


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def iso_mtime(path: Path) -> str:
    try:
        return _dt.datetime.fromtimestamp(path.stat().st_mtime).isoformat(timespec="seconds")
    except Exception:
        return ""


def acquire_lock() -> Tuple[bool, str]:
    LOCKS_DIR.mkdir(parents=True, exist_ok=True)
    now_ts = time.time()
    if LOCK_FILE.exists():
        age = now_ts - LOCK_FILE.stat().st_mtime
        if age < LOCK_TTL_SECONDS:
            return False, f"Lock activo: {LOCK_FILE} | edad={age:.0f}s | ttl={LOCK_TTL_SECONDS}s"
        try:
            LOCK_FILE.unlink()
        except Exception as exc:
            return False, f"No se pudo eliminar lock vencido: {LOCK_FILE} | {exc}"
    payload = {
        "script": "cierre_semanal_facturas.py",
        "version": VERSION_CIERRE,
        "pid": os.getpid(),
        "started_at": _dt.datetime.now().isoformat(timespec="seconds"),
        "host": socket.gethostname(),
    }
    LOCK_FILE.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
    return True, "Lock creado"


def release_lock() -> None:
    try:
        LOCK_FILE.unlink(missing_ok=True)
    except Exception:
        pass


def es_sensible(clave: str) -> bool:
    clave_upper = str(clave or "").strip().upper()
    return any(p in clave_upper for p in PALABRAS_SENSIBLES)


def redactar_env_line(linea: str) -> str:
    if not linea.strip() or linea.lstrip().startswith("#") or "=" not in linea:
        return linea
    clave, valor = linea.split("=", 1)
    if es_sensible(clave):
        return f"{clave}=***REDACTADO***"
    if re.search(r"[A-Za-z0-9_\-]{32,}", valor) and clave.strip().upper() not in {
        "SP_DRIVE_ID",
        "SP_DRIVE_ID_RADICADOS",
        "SP_BACKUP2_DRIVE_ID",
        "BACKUP_DRIVE_ID",
        "ONEDRIVE_BACKUP_DRIVE_ID",
    }:
        return f"{clave}=***REDACTADO***"
    return linea


def info_archivo(destino: Path, origen: Optional[Path] = None, categoria: str = "") -> dict[str, Any]:
    st = destino.stat()
    return {
        "categoria": categoria,
        "origen": rel(origen) if origen else "generado",
        "destino": rel(destino),
        "nombre": destino.name,
        "bytes": st.st_size,
        "sha256": sha256_file(destino),
        "mtime_origen": iso_mtime(origen) if origen and origen.exists() else "",
        "mtime_destino": iso_mtime(destino),
    }


def copiar_archivo(origen: Path, destino: Path, categoria: str = "") -> Optional[dict[str, Any]]:
    if not origen.exists() or not origen.is_file() or origen.name in NOMBRES_EXCLUIDOS:
        return None
    destino.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(origen, destino)
    return info_archivo(destino, origen=origen, categoria=categoria)


def leer_csv_dicts(path: Path) -> list[dict[str, str]]:
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            with path.open("r", encoding=encoding, newline="") as f:
                reader = csv.DictReader(f)
                return [dict(r or {}) for r in reader]
        except UnicodeDecodeError:
            continue
        except Exception:
            return []
    return []


def recolectar_audits(inicio: _dt.date, fin: _dt.date) -> list[Path]:
    rutas: list[Path] = []
    fechas = [d.isoformat() for d in fechas_en_rango(inicio, fin)]
    for fecha in fechas:
        patrones = [
            f"audit_detalle_{fecha}.csv",
            f"audit_runs_{fecha}.csv",
            f"audit_*_{fecha}.csv",
            f"*{fecha}*.csv",
        ]
        for base in (AUDIT_DIR, DATA_DIR, ROOT):
            if not base.exists():
                continue
            for patron in patrones:
                for p in base.glob(patron):
                    if p.is_file() and "audit" in p.name.lower():
                        rutas.append(p)
    vistos = set()
    out = []
    for p in rutas:
        try:
            key = str(p.resolve()).lower()
        except Exception:
            key = str(p).lower()
        if key not in vistos:
            vistos.add(key)
            out.append(p)
    return sorted(out)


def log_corresponde_rango(path: Path, inicio: _dt.date, fin: _dt.date) -> bool:
    name = path.name.lower()
    for d in fechas_en_rango(inicio, fin):
        fecha = d.isoformat()
        if fecha in name or fecha.replace("-", "") in name:
            return True
    try:
        mdate = _dt.datetime.fromtimestamp(path.stat().st_mtime).date()
        return inicio <= mdate <= fin
    except Exception:
        return False


def bases_logs_disponibles() -> list[tuple[str, Path]]:
    """
    Devuelve las ubicaciones de logs compatibles con local y VPS.

    Los paths se deduplican por ruta resuelta para evitar copiar dos veces
    el mismo directorio cuando data/logs es un enlace simbólico.
    """
    candidatos = [
        ("app_logs", LOGS_DIR),
        ("data_logs", DATA_LOGS_DIR),
        ("runtime", VPS_LOG_ROOT / "runtime"),
        ("runs", VPS_LOG_ROOT / "runs"),
        ("cron", VPS_LOG_ROOT / "cron"),
    ]

    vistos: set[str] = set()
    bases: list[tuple[str, Path]] = []

    for etiqueta, base in candidatos:
        if not base.exists() or not base.is_dir():
            continue

        try:
            key = str(base.resolve()).casefold()
        except Exception:
            key = str(base).casefold()

        if key in vistos:
            continue

        vistos.add(key)
        bases.append((etiqueta, base))

    return bases


def recolectar_logs(inicio: _dt.date, fin: _dt.date) -> list[Path]:
    rutas: list[Path] = []

    for _etiqueta, base in bases_logs_disponibles():
        for p in base.rglob("*"):
            if not p.is_file():
                continue
            if p.suffix.lower() not in EXT_LOGS:
                continue
            if log_corresponde_rango(p, inicio, fin):
                rutas.append(p)

    vistos: set[str] = set()
    out: list[Path] = []

    for p in rutas:
        try:
            key = str(p.resolve()).casefold()
        except Exception:
            key = str(p).casefold()

        if key not in vistos:
            vistos.add(key)
            out.append(p)

    return sorted(out, key=lambda p: str(p).casefold())



def find_col(headers: Sequence[Any], candidates: Sequence[str]) -> Optional[int]:
    candidate_norms = {norm(c) for c in candidates}
    candidate_keys = {norm_key(c) for c in candidates}
    for idx, header in enumerate(headers):
        hn = norm(header)
        hk = norm_key(header)
        if hn in candidate_norms or hk in candidate_keys:
            return idx
    return None


def fecha_valor_en_rango(value: Any, inicio: _dt.date, fin: _dt.date) -> bool:
    if value is None:
        return False
    if isinstance(value, _dt.datetime):
        return inicio <= value.date() <= fin
    if isinstance(value, _dt.date):
        return inicio <= value <= fin
    s = str(value).strip()
    if not s:
        return False
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            d = _dt.datetime.strptime(s[:10], fmt).date()
            return inicio <= d <= fin
        except Exception:
            pass
    # Soporta valores ISO con fecha dentro del texto.
    for d in fechas_en_rango(inicio, fin):
        if d.isoformat() in s:
            return True
    return False


def extraer_valor_row(row: dict[str, Any], candidates: Sequence[str]) -> str:
    by_norm = {norm(k): v for k, v in row.items()}
    by_key = {norm_key(k): v for k, v in row.items()}
    for c in candidates:
        if norm(c) in by_norm:
            return clean_value(by_norm[norm(c)])
        if norm_key(c) in by_key:
            return clean_value(by_key[norm_key(c)])
    return ""


def extraer_primer_valor_no_vacio(
    row: dict[str, Any],
    candidates: Sequence[str],
) -> str:
    """
    Busca el primer valor no vacío entre varios nombres posibles de columna.

    A diferencia de extraer_valor_row(), no se detiene cuando una columna
    existe pero viene vacía. Esto permite leer auditorías que conservan
    columnas antiguas vacías y guardan el dato real en otra columna.
    """
    by_norm = {norm(k): v for k, v in row.items()}
    by_key = {norm_key(k): v for k, v in row.items()}

    for candidate in candidates:
        values: list[Any] = []

        candidate_norm = norm(candidate)
        candidate_key = norm_key(candidate)

        if candidate_norm in by_norm:
            values.append(by_norm[candidate_norm])

        if candidate_key in by_key:
            values.append(by_key[candidate_key])

        for value in values:
            cleaned = clean_value(value)
            if cleaned:
                return cleaned

    return ""


def extraer_radicado_desde_asunto(asunto: str) -> str:
    if not asunto:
        return ""

    match = PATRON_RADICADO_ASUNTO.search(str(asunto))
    return match.group(1) if match else ""


def extraer_candidatos_desde_audits(
    audit_files: list[Path],
) -> dict[str, set[str]]:
    """
    Extrae identificadores de las auditorías de la semana.

    Los registros normales suelen coincidir por CUFE. Los registros mínimos
    pueden no tener CUFE, pero conservan número, radicado y archivo. En esos
    casos:

    - el radicado puede venir dentro del asunto del correo;
    - el archivo puede venir en pdf_name u otra columna equivalente.

    Nunca se habilita una coincidencia únicamente por número. Las
    combinaciones número+radicado, número+archivo o radicado+archivo reducen
    el riesgo de incluir registros ajenos al rango semanal.
    """
    candidatos: dict[str, set[str]] = {
        "cufe": set(),
        "numero": set(),
        "radicado": set(),
        "archivo": set(),
        "numero_radicado": set(),
        "numero_archivo": set(),
        "radicado_archivo": set(),
    }

    for audit in audit_files:
        for row in leer_csv_dicts(audit):
            cufe = extraer_primer_valor_no_vacio(row, COLS_CUFE)
            numero = extraer_primer_valor_no_vacio(row, COLS_NUMERO)
            radicado = extraer_primer_valor_no_vacio(row, COLS_RADICADO)
            archivo = extraer_primer_valor_no_vacio(row, COLS_ARCHIVO)

            if not archivo:
                archivo = extraer_primer_valor_no_vacio(
                    row,
                    COLS_ARCHIVO_AUDIT_FALLBACK,
                )

            if not radicado:
                asunto = extraer_primer_valor_no_vacio(
                    row,
                    COLS_ASUNTO_AUDIT,
                )
                radicado = extraer_radicado_desde_asunto(asunto)

            cufe_key = norm_key(cufe)
            numero_key = norm_key(numero)
            radicado_key = norm_key(radicado)
            archivo_key = norm_key(archivo)

            if cufe_key:
                candidatos["cufe"].add(cufe_key)
            if numero_key:
                candidatos["numero"].add(numero_key)
            if radicado_key:
                candidatos["radicado"].add(radicado_key)
            if archivo_key:
                candidatos["archivo"].add(archivo_key)

            if numero_key and radicado_key:
                candidatos["numero_radicado"].add(
                    f"{numero_key}|{radicado_key}"
                )
            if numero_key and archivo_key:
                candidatos["numero_archivo"].add(
                    f"{numero_key}|{archivo_key}"
                )
            if radicado_key and archivo_key:
                candidatos["radicado_archivo"].add(
                    f"{radicado_key}|{archivo_key}"
                )

    return candidatos



def fila_match_candidatos(row_values: Sequence[Any], headers: Sequence[Any], candidatos: dict[str, set[str]]) -> bool:
    idx_cufe = find_col(headers, COLS_CUFE)
    idx_numero = find_col(headers, COLS_NUMERO)
    idx_radicado = find_col(headers, COLS_RADICADO)
    idx_archivo = find_col(headers, COLS_ARCHIVO)

    cufe = norm_key(row_values[idx_cufe]) if idx_cufe is not None and idx_cufe < len(row_values) else ""
    numero = norm_key(row_values[idx_numero]) if idx_numero is not None and idx_numero < len(row_values) else ""
    radicado = norm_key(row_values[idx_radicado]) if idx_radicado is not None and idx_radicado < len(row_values) else ""
    archivo = norm_key(row_values[idx_archivo]) if idx_archivo is not None and idx_archivo < len(row_values) else ""

    if cufe and cufe in candidatos["cufe"]:
        return True
    if numero and radicado and f"{numero}|{radicado}" in candidatos["numero_radicado"]:
        return True
    if numero and archivo and f"{numero}|{archivo}" in candidatos["numero_archivo"]:
        return True
    if radicado and archivo and f"{radicado}|{archivo}" in candidatos["radicado_archivo"]:
        return True
    if numero and radicado and numero in candidatos["numero"] and radicado in candidatos["radicado"]:
        return True
    if numero and archivo and numero in candidatos["numero"] and archivo in candidatos["archivo"]:
        return True
    return False



def obtener_filas_semanales_desde_diarios(
    inicio: _dt.date,
    fin: _dt.date,
) -> tuple[list[Any], list[list[Any]], dict[str, Any]]:
    """
    Reconstruye el Excel semanal concatenando únicamente los Excel
    diarios físicos existentes dentro del rango solicitado.

    Este modo se usa para reconstrucción histórica controlada.
    No vuelve a inferir fechas desde auditorías ni desde facturas.xlsx.
    """
    headers_base: Optional[list[Any]] = None
    rows_semana: list[list[Any]] = []
    diarios_usados: list[dict[str, Any]] = []

    fecha = inicio

    while fecha <= fin:
        fecha_str = fecha.isoformat()

        patron = (
            f"**/Diario_{fecha_str}/"
            f"01_Excel_Diario/"
            f"facturas_diario_{fecha_str}.xlsx"
        )

        encontrados = sorted(
            p for p in CIERRES_DIR.glob(patron)
            if p.is_file()
        )

        if len(encontrados) > 1:
            raise RuntimeError(
                "Se encontraron varios cierres diarios para "
                f"{fecha_str}: {[rel(x) for x in encontrados]}"
            )

        if not encontrados:
            fecha += _dt.timedelta(days=1)
            continue

        excel_diario = encontrados[0]

        wb = load_workbook(
            excel_diario,
            data_only=True,
            read_only=True,
        )

        try:
            ws = (
                wb[EXCEL_SHEET_NAME]
                if EXCEL_SHEET_NAME in wb.sheetnames
                else wb[wb.sheetnames[0]]
            )

            rows_iter = ws.iter_rows(values_only=True)

            try:
                raw_headers = next(rows_iter)
            except StopIteration:
                raise RuntimeError(
                    f"Excel diario vacío/sin encabezados: {excel_diario}"
                )

            headers = list(raw_headers or [])

            if headers_base is None:
                headers_base = headers
            elif headers != headers_base:
                raise RuntimeError(
                    "Encabezados incompatibles entre cierres diarios. "
                    f"Fecha problemática: {fecha_str}"
                )

            filas_dia = 0

            for values in rows_iter:
                row = list(values or [])

                if len(row) < len(headers):
                    row.extend(
                        [None] * (len(headers) - len(row))
                    )
                elif len(row) > len(headers):
                    row = row[:len(headers)]

                if any(v not in (None, "") for v in row):
                    rows_semana.append(row)
                    filas_dia += 1

            diarios_usados.append(
                {
                    "fecha": fecha_str,
                    "excel": rel(excel_diario),
                    "filas": filas_dia,
                }
            )

        finally:
            wb.close()

        fecha += _dt.timedelta(days=1)

    if headers_base is None:
        raise RuntimeError(
            "No se encontró ningún cierre diario físico dentro "
            f"del rango {inicio.isoformat()} a {fin.isoformat()}."
        )

    meta: dict[str, Any] = {
        "metodo_seleccion": "cierres_diarios_v5_concatenados",
        "fuente": "cierres_diarios_fisicos",
        "rango_inicio": inicio.isoformat(),
        "rango_fin": fin.isoformat(),
        "diarios_usados": diarios_usados,
        "cantidad_diarios_usados": len(diarios_usados),
        "filas_semanales": len(rows_semana),
        "advertencias": [],
    }

    print(
        "Reconstruccion semanal desde diarios: "
        f"diarios={len(diarios_usados)} | "
        f"filas={len(rows_semana)}"
    )

    for item in diarios_usados:
        print(
            f"   {item['fecha']} | "
            f"filas={item['filas']} | "
            f"{item['excel']}"
        )

    return headers_base, rows_semana, meta


def obtener_filas_semanales_desde_excel(
    inicio: _dt.date,
    fin: _dt.date,
    audit_files: list[Path],
) -> tuple[list[Any], list[list[Any]], dict[str, Any]]:
    if not FACTURAS_PATH.exists():
        raise RuntimeError(f"No existe Excel operativo local: {FACTURAS_PATH}")

    print(f"Leyendo Excel operativo: {FACTURAS_PATH}")
    wb = load_workbook(FACTURAS_PATH, data_only=True, read_only=True)
    try:
        ws = wb[EXCEL_SHEET_NAME] if EXCEL_SHEET_NAME in wb.sheetnames else wb[wb.sheetnames[0]]
        print(f"Hoja usada: {ws.title} | max_row={ws.max_row} | max_column={ws.max_column}")
        rows_iter = ws.iter_rows(values_only=True)
        try:
            headers_raw = next(rows_iter)
        except StopIteration:
            raise RuntimeError("El Excel operativo no tiene filas.")
        headers = list(headers_raw or [])
        total_cols = len(headers)
        rows_all: list[list[Any]] = []
        leidas = 0
        for values in rows_iter:
            leidas += 1
            row = list(values or [])
            if total_cols > 0:
                if len(row) < total_cols:
                    row.extend([None] * (total_cols - len(row)))
                elif len(row) > total_cols:
                    row = row[:total_cols]
            if any(v not in (None, "") for v in row):
                rows_all.append(row)
            if leidas % 1000 == 0:
                print(f"   ... filas leidas desde Excel: {leidas} | filas con datos: {len(rows_all)}")
        print(f"Lectura Excel terminada: filas leidas={leidas} | filas con datos={len(rows_all)}")

        meta: dict[str, Any] = {
            "excel_operativo": rel(FACTURAS_PATH),
            "hoja_usada": ws.title,
            "filas_operativas_total": len(rows_all),
            "metodo_seleccion": "",
            "advertencias": [],
        }

        idx_fecha = find_col(headers, COLUMNAS_FECHA_PROCESO_CANDIDATAS)
        if idx_fecha is not None:
            rows_fecha = [
                row for row in rows_all
                if idx_fecha < len(row) and fecha_valor_en_rango(row[idx_fecha], inicio, fin)
            ]
            meta["metodo_seleccion"] = "columna_fecha_procesamiento_rango_semanal"
            meta["columna_fecha_procesamiento"] = headers[idx_fecha]
            return headers, rows_fecha, meta

        candidatos = extraer_candidatos_desde_audits(audit_files)
        total_candidates = sum(len(v) for v in candidatos.values())
        meta["audit_files_usados"] = [rel(p) for p in audit_files]
        meta["candidatos_extraidos"] = {k: len(v) for k, v in candidatos.items()}

        if total_candidates <= 0:
            meta["metodo_seleccion"] = "sin_candidatos_auditoria_semana"
            meta["advertencias"].append(
                "No se encontro columna de fecha de procesamiento ni candidatos suficientes en auditorias de la semana. "
                "Se genera Excel semanal solo con encabezados."
            )
            return headers, [], meta

        print(f"Cruzando Excel operativo contra auditorias de la semana. Candidatos={total_candidates}")
        rows_match = []
        for idx, row in enumerate(rows_all, start=1):
            if fila_match_candidatos(row, headers, candidatos):
                rows_match.append(row)
            if idx % 1000 == 0:
                print(f"   ... cruce auditoria semanal: {idx}/{len(rows_all)} | matches={len(rows_match)}")
        meta["metodo_seleccion"] = "auditorias_semana_vs_excel_operativo"
        return headers, rows_match, meta
    finally:
        wb.close()


def ajustar_ancho_columnas(ws) -> None:
    for column_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells[:80]:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 45)


def crear_excel_semanal(destino: Path, headers: list[Any], rows: list[list[Any]]) -> dict[str, Any]:
    destino.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = EXCEL_SHEET_NAME
    header_clean = ["" if h is None else str(h).strip() for h in headers]
    ws.append(header_clean)
    for row in rows:
        normalized = list(row)
        if len(normalized) < len(header_clean):
            normalized += [None] * (len(header_clean) - len(normalized))
        ws.append(normalized[: len(header_clean)])

    max_col = max(1, len(header_clean))
    max_row = max(1, len(rows) + 1)
    last_col = get_column_letter(max_col)
    table_ref = f"A1:{last_col}{max_row}"
    tab = Table(displayName=EXCEL_TABLE_NAME, ref=table_ref)
    tab.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium4",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tab)
    ws.freeze_panes = "A2"
    ajustar_ancho_columnas(ws)
    wb.save(destino)
    wb.close()
    return {
        "ruta": rel(destino),
        "filas_datos": len(rows),
        "columnas": len(header_clean),
        "tabla": EXCEL_TABLE_NAME,
        "bytes": destino.stat().st_size,
        "sha256": sha256_file(destino),
    }


def copiar_auditorias(audit_files: list[Path], destino_dir: Path) -> list[dict[str, Any]]:
    items = []
    destino_dir.mkdir(parents=True, exist_ok=True)
    for origen in audit_files:
        info = copiar_archivo(origen, destino_dir / origen.name, categoria="auditoria_semana")
        if info:
            items.append(info)
    return items


def copiar_logs(
    log_files: list[Path],
    destino_dir: Path,
) -> list[dict[str, Any]]:
    items: list[dict[str, Any]] = []
    destino_dir.mkdir(parents=True, exist_ok=True)
    bases = bases_logs_disponibles()

    for origen in log_files:
        rel_log: Optional[Path] = None

        for etiqueta, base in bases:
            try:
                rel_origen = origen.resolve().relative_to(base.resolve())
                rel_log = Path(etiqueta) / rel_origen
                break
            except Exception:
                continue

        if rel_log is None:
            rel_log = Path("otros") / origen.name

        destino = destino_dir / rel_log
        info = copiar_archivo(
            origen,
            destino,
            categoria="log_semana",
        )

        if info:
            items.append(info)

    return items



def crear_snapshot_env_redactado(destino_dir: Path, *args, **kwargs) -> Optional[dict[str, Any]]:
    """
    Crea snapshot redactado del .env para soporte técnico semanal.
    Asegura la carpeta destino antes de escribir.
    """
    env_path = ROOT / ".env"
    if not env_path.exists():
        return None

    destino_dir.mkdir(parents=True, exist_ok=True)

    semana_inicio = ""
    semana_fin = ""

    if len(args) >= 1:
        semana_inicio = str(args[0] or "").strip()
    if len(args) >= 2:
        semana_fin = str(args[1] or "").strip()

    semana_inicio = str(kwargs.get("semana_inicio", semana_inicio) or "").strip()
    semana_fin = str(kwargs.get("semana_fin", semana_fin) or "").strip()

    if semana_inicio and semana_fin:
        nombre = f"snapshot_env_redactado_semanal_{semana_inicio}_a_{semana_fin}.txt"
    else:
        nombre = "snapshot_env_redactado_semanal.txt"

    destino = destino_dir / nombre
    destino.parent.mkdir(parents=True, exist_ok=True)

    lineas = env_path.read_text(encoding="utf-8", errors="replace").splitlines()
    destino.write_text(
        "\n".join(redactar_env_line(x) for x in lineas) + "\n",
        encoding="utf-8",
    )

    return info_archivo(destino, origen=env_path, categoria="config_redactada")


def validar_excel_generado(path: Path, expected_columns: int) -> dict[str, Any]:
    out: dict[str, Any] = {
        "archivo": rel(path),
        "existe": path.exists(),
        "abre_ok": False,
        "hojas": [],
        "filas_datos": 0,
        "columnas": 0,
        "errores": [],
    }
    if not path.exists():
        out["errores"].append("No existe Excel semanal.")
        return out
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            out["abre_ok"] = True
            out["hojas"] = list(wb.sheetnames)
            ws = wb[EXCEL_SHEET_NAME] if EXCEL_SHEET_NAME in wb.sheetnames else wb[wb.sheetnames[0]]
            out["columnas"] = int(ws.max_column or 0)
            out["filas_datos"] = max(int(ws.max_row or 0) - 1, 0)
            if expected_columns and out["columnas"] != expected_columns:
                out["errores"].append(f"Columnas distintas. Esperadas={expected_columns}, detectadas={out['columnas']}")
        finally:
            wb.close()
    except Exception as exc:
        out["errores"].append(f"No se pudo abrir Excel semanal: {type(exc).__name__}: {exc}")
    return out


def generar_resumen_txt(path: Path, manifest: dict[str, Any]) -> None:
    lines = [
        "CIERRE SEMANAL LOCAL SEGURO V4 - FACTURAS JOYCO",
        "=" * 90,
        f"Version: {manifest['version']}",
        f"Rango semana: {manifest['semana_inicio']} a {manifest['semana_fin']}",
        f"Generado: {manifest['generado_en']}",
        f"Root: {manifest['root']}",
        f"Carpeta cierre: {manifest['carpeta_cierre']}",
        "",
        "Resumen:",
        f"- Filas en Excel semanal: {manifest['excel_semanal']['filas_datos']}",
        f"- Columnas en Excel semanal: {manifest['excel_semanal']['columnas']}",
        f"- Metodo seleccion: {manifest['seleccion_filas'].get('metodo_seleccion')}",
        f"- Auditorias copiadas: {manifest['conteos']['auditorias']}",
        f"- Logs copiados: {manifest['conteos']['logs']}",
        f"- Total archivos evidencia: {manifest['total_archivos']}",
        f"- Total bytes evidencia: {manifest['total_bytes']}",
        "",
        "Regla aplicada:",
        "- La fuente utilizada por el cierre semanal queda registrada en fuente_datos segun el modo ejecutado.",
        "- En modo reconstruccion consolida cierres diarios fisicos validados; "
        "en modo normal usa Excel principal + auditorias.",
        "- Los cierres diarios permanecen en sus carpetas hermanas y no se duplican dentro de Semanal.",
        "- No se incluye .env real ni secretos.",
        "- No se sube nada desde este script; la subida remota se hace en el paso de OneDrive.",
        "",
    ]
    advertencias = manifest.get("advertencias") or []
    if advertencias:
        lines.append("Advertencias:")
        for adv in advertencias:
            lines.append(f"- {adv}")
        lines.append("")
    lines.append("=" * 90)
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def listar_archivos_para_manifest(base: Path, excluir: Optional[set[Path]] = None) -> list[dict[str, Any]]:
    excluir_resolved = {p.resolve() for p in (excluir or set()) if p.exists()}
    items = []
    for p in sorted(base.rglob("*")):
        if not p.is_file():
            continue
        try:
            if p.resolve() in excluir_resolved:
                continue
        except Exception:
            pass
        items.append(info_archivo(p, categoria="evidencia"))
    return items


def main() -> int:
    args = _parse_args()
    inicio, fin = rango_desde_args(args)

    # La semana se organiza según la fecha en que termina. Esto conserva la
    # estructura aprobada para semanas que cruzan de mes o de trimestre; por
    # ejemplo, SEMANA_2026-07-27_a_2026-08-02 queda dentro de agosto y del
    # trimestre activo que contiene el 2026-08-02.
    trimestre = cargar_trimestre_activo(ROOT, fin)
    anio = trimestre["anio"]
    mes_nombre = mes_carpeta(fin)
    semana = semana_nombre(inicio, fin)
    rango = f"{inicio.isoformat()}_a_{fin.isoformat()}"

    semana_dir = (
        CIERRES_DIR
        / Path(trimestre["ruta_relativa"])
        / mes_nombre
        / semana
    )
    cierre_dir = semana_dir / "Semanal"

    excel_dir = cierre_dir / "01_Excel_Semanal"
    auditoria_dir = cierre_dir / "02_Auditorias_Semana"
    manifest_dir = cierre_dir / "04_Manifest_Semanal"
    validaciones_dir = cierre_dir / "05_Validaciones"
    logs_dir = cierre_dir / "06_Logs_Semana"

    excel_path = excel_dir / f"facturas_semanal_{rango}.xlsx"
    manifest_path = manifest_dir / f"manifest_semanal_{rango}.json"
    resumen_path = manifest_dir / f"resumen_semanal_{rango}.txt"
    validacion_path = validaciones_dir / f"validacion_local_semanal_{rango}.json"

    print("=" * 100)
    print("CIERRE SEMANAL LOCAL SEGURO V4 - FACTURAS JOYCO")
    print("=" * 100)
    print(f"Version: {VERSION_CIERRE}")
    print(f"Root: {ROOT}")
    print(f"Trimestre activo: {trimestre['nombre_carpeta']}")
    print(
        "Rango trimestre: "
        f"{trimestre['fecha_inicio']} a {trimestre['fecha_fin']}"
    )
    print(f"Semana: {inicio.isoformat()} a {fin.isoformat()}")
    print(f"Mes carpeta: {mes_nombre}")
    print(f"Carpeta local destino: {cierre_dir}")
    print(f"Dry-run: {args.dry_run}")
    print("-" * 100)

    ok_lock, msg_lock = acquire_lock()
    if not ok_lock:
        print(f"ERROR: {msg_lock}")
        print("No se ejecuta el cierre para evitar cruces.")
        return 2

    try:
        audit_files = recolectar_audits(inicio, fin)
        log_files = recolectar_logs(inicio, fin)
        if CIERRE_SEMANAL_DESDE_DIARIOS:
            headers, filas_semanales, meta_seleccion = (
                obtener_filas_semanales_desde_diarios(
                    inicio,
                    fin,
                )
            )
        else:
            headers, filas_semanales, meta_seleccion = (
                obtener_filas_semanales_desde_excel(
                    inicio,
                    fin,
                    audit_files,
                )
            )

        advertencias: list[str] = []
        advertencias.extend(meta_seleccion.get("advertencias", []))
        if not audit_files:
            advertencias.append("No se encontraron archivos de auditoria dentro del rango semanal.")
        if not log_files:
            advertencias.append("No se encontraron logs especificos de la semana.")
        if not filas_semanales:
            advertencias.append("El Excel semanal no tendra filas de datos para esta semana.")

        print(f"Auditorias detectadas: {len(audit_files)}")
        print(f"Logs detectados: {len(log_files)}")
        print(f"Metodo seleccion filas: {meta_seleccion.get('metodo_seleccion')}")
        print(f"Filas semanales detectadas: {len(filas_semanales)}")

        if args.dry_run:
            print("-" * 100)
            print("DRY-RUN: no se generaron archivos.")
            if advertencias:
                print("Advertencias:")
                for adv in advertencias:
                    print(f"  - {adv}")
            return 0

        if cierre_dir.exists():
            print(f"Ya existe carpeta de cierre semanal. Se regenerara: {cierre_dir}")
            shutil.rmtree(cierre_dir, ignore_errors=True)

        for d in (excel_dir, auditoria_dir, manifest_dir, validaciones_dir, logs_dir):
            d.mkdir(parents=True, exist_ok=True)

        excel_info = crear_excel_semanal(excel_path, list(headers), filas_semanales)
        auditorias_info = copiar_auditorias(audit_files, auditoria_dir)
        logs_info = copiar_logs(log_files, logs_dir)
        env_info = crear_snapshot_env_redactado(manifest_dir, inicio.isoformat(), fin.isoformat())

        validacion_excel = validar_excel_generado(excel_path, expected_columns=len(headers))
        validacion = {
            "tipo": "validacion_local_cierre_semanal",
            "version": VERSION_CIERRE,
            "trimestre": {
                "periodo_activo": trimestre["periodo_activo"],
                "nombre_carpeta": trimestre["nombre_carpeta"],
                "fecha_inicio": trimestre["fecha_inicio"],
                "fecha_fin": trimestre["fecha_fin"],
                "state_path": trimestre["state_path"],
            },
            "semana_inicio": inicio.isoformat(),
            "semana_fin": fin.isoformat(),
            "generado_en": _dt.datetime.now().isoformat(timespec="seconds"),
            "excel": validacion_excel,
            "auditorias_detectadas": len(audit_files),
            "auditorias_copiadas": len(auditorias_info),
            "logs_detectados": len(log_files),
            "logs_copiados": len(logs_info),
            "ok": bool(validacion_excel.get("abre_ok")) and not validacion_excel.get("errores"),
            "advertencias": advertencias,
        }
        validacion_path.write_text(json.dumps(validacion, ensure_ascii=False, indent=2), encoding="utf-8")

        manifest = {
            "tipo": "cierre_semanal_local_seguro_v4",
            "version": VERSION_CIERRE,
            "trimestre": {
                "periodo_activo": trimestre["periodo_activo"],
                "nombre_carpeta": trimestre["nombre_carpeta"],
                "fecha_inicio": trimestre["fecha_inicio"],
                "fecha_fin": trimestre["fecha_fin"],
                "ruta_relativa": trimestre["ruta_relativa"],
                "state_path": trimestre["state_path"],
            },
            "anio": anio,
            "mes_carpeta": mes_nombre,
            "semana_inicio": inicio.isoformat(),
            "semana_fin": fin.isoformat(),
            "semana_nombre": semana,
            "generado_en": _dt.datetime.now().isoformat(timespec="seconds"),
            "root": str(ROOT),
            "host": socket.gethostname(),
            "platform": platform.platform(),
            "python": sys.version.replace("\n", " "),
            "carpeta_cierre": rel(cierre_dir),
            "excel_semanal": excel_info,
            "seleccion_filas": meta_seleccion,
            "fuente_datos": {
                "modo": (
                    "cierres_diarios_fisicos"
                    if CIERRE_SEMANAL_DESDE_DIARIOS
                    else "excel_principal_mas_auditorias"
                ),
                "principal": (
                    None
                    if CIERRE_SEMANAL_DESDE_DIARIOS
                    else rel(FACTURAS_PATH)
                ),
                "cierres_diarios": (
                    meta_seleccion.get("diarios_usados", [])
                    if CIERRE_SEMANAL_DESDE_DIARIOS
                    else []
                ),
                "auditorias_semana": [rel(p) for p in audit_files],
                "nota": (
                    "Reconstruccion semanal desde cierres diarios fisicos validados; "
                    "los diarios permanecen como carpetas hermanas y no se duplican."
                    if CIERRE_SEMANAL_DESDE_DIARIOS
                    else
                    "Modo normal desde Excel principal y auditorias; "
                    "los cierres diarios no se duplican dentro de Semanal."
                ),
            },
            "conteos": {
                "filas_semanales": len(filas_semanales),
                "auditorias": len(auditorias_info),
                "logs": len(logs_info),
                "config_redactada": 1 if env_info else 0,
            },
            "validacion_local": rel(validacion_path),
            "advertencias": advertencias,
            "nota": "Cierre semanal V4: fuente determinada por el modo de ejecucion, sin duplicar evidencias diarias.",
        }

        archivos_pre = listar_archivos_para_manifest(cierre_dir, excluir={manifest_path})
        manifest["archivos"] = archivos_pre
        manifest["total_archivos"] = len(archivos_pre)
        manifest["total_bytes"] = sum(int(x.get("bytes", 0) or 0) for x in archivos_pre)
        generar_resumen_txt(resumen_path, manifest)

        archivos = listar_archivos_para_manifest(cierre_dir, excluir={manifest_path})
        manifest["archivos"] = archivos
        manifest["total_archivos"] = len(archivos)
        manifest["total_bytes"] = sum(int(x.get("bytes", 0) or 0) for x in archivos)
        manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")

        print(f"Excel semanal: {excel_path}")
        print(f"Filas Excel semanal: {excel_info['filas_datos']}")
        print(f"Auditorias copiadas: {len(auditorias_info)}")
        print(f"Logs copiados: {len(logs_info)}")
        print(f"Manifest: {manifest_path}")
        print(f"Resumen: {resumen_path}")
        print(f"Validacion local: {validacion_path}")
        print(f"Total archivos evidencia: {manifest['total_archivos']}")
        print(f"Total bytes evidencia: {manifest['total_bytes']}")

        if advertencias:
            print("Advertencias:")
            for adv in advertencias:
                print(f"  - {adv}")

        if not args.permitir_vacio and len(filas_semanales) == 0:
            print("Cierre generado sin filas semanales. Codigo 3 para revision controlada.")
            print("Si la semana realmente no tuvo facturas nuevas, usa --permitir-vacio.")
            return 3

        print("=" * 100)
        print("Cierre semanal local V4 generado correctamente.")
        print("Siguiente paso: subir/validar en OneDrive con scripts\\subir_cierre_semanal_sharepoint.py")
        print("=" * 100)
        return 0

    except Exception as exc:
        print(f"ERROR generando cierre semanal local V4: {type(exc).__name__}: {exc}")
        print("No subas ni archives nada hasta revisar el error.")
        return 1
    finally:
        release_lock()


if __name__ == "__main__":
    raise SystemExit(main())
