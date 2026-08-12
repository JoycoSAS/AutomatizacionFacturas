# -*- coding: utf-8 -*-
"""
JOYCO - Facturas Procesador
Cierre diario local seguro V3

Política V3:
- El cierre diario NO copia el histórico completo de facturas.xlsx.
- Genera un Excel diario con las facturas registradas/procesadas en la fecha del cierre.
- Incluye auditorías del día, logs de producción local/VPS, manifest, resumen y validación local.
- No sube a OneDrive/SharePoint. La subida se hará con el script remoto correspondiente.
- No incluye .env real ni secretos.

Estructura generada:
  data/cierres_diarios/YYYY/TRIMESTRE_YYYY-MM-DD_A_YYYY-MM-DD/
    YYYY-MM_MesNombre/SEMANA_YYYY-MM-DD_a_YYYY-MM-DD/Diario_YYYY-MM-DD/
    01_Excel_Diario/facturas_diario_YYYY-MM-DD.xlsx
    02_Auditoria/audit_*.csv
    03_Logs/logs_produccion_YYYY-MM-DD/
    04_Manifest/manifest_diario_YYYY-MM-DD.json
    04_Manifest/resumen_diario_YYYY-MM-DD.txt
    05_Validaciones/validacion_local_YYYY-MM-DD.json
"""

from __future__ import annotations

import argparse
import csv
import datetime as _dt
import hashlib
import json
import os
import platform
import re
import shutil
import socket
import sys
import time
import unicodedata
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

try:
    import config  # noqa: F401  # Carga .env/configuración del proyecto sin imprimir secretos.
except Exception:
    pass

from trimestre_activo import cargar_trimestre_activo

VERSION_CIERRE = "2026-08-12-CIERRE-DIARIO-SEGURO-V3-FILTRO-FECHA-AUDITORIA"

DATA_DIR = ROOT / "data"
AUDIT_DIR = DATA_DIR / "audit"
LOGS_DIR = ROOT / "logs"
DATA_LOGS_DIR = DATA_DIR / "logs"

# En local, los logs pueden permanecer dentro del repositorio.
# En el VPS se almacenan fuera del código para separar aplicación y operación.
VPS_LOG_ROOT = Path(
    os.getenv(
        "FACTURAS_LOG_ROOT",
        "/var/log/joyco/facturas-procesador",
    )
)

CIERRES_DIR = DATA_DIR / "cierres_diarios"
LOCKS_DIR = DATA_DIR / "locks"

FACTURAS_PATH = Path(os.getenv("ARCHIVO_EXCEL_LOCAL", str(DATA_DIR / "facturas.xlsx")))
if not FACTURAS_PATH.is_absolute():
    FACTURAS_PATH = ROOT / FACTURAS_PATH

LOCK_FILE = LOCKS_DIR / "cierre_diario_seguro.lock"
LOCK_TTL_SECONDS = int(os.getenv("LOCK_TTL_SECONDS", "3600") or "3600")

PALABRAS_SENSIBLES = (
    "SECRET",
    "PASSWORD",
    "PASS",
    "TOKEN",
    "KEY",
    "CLIENT_SECRET",
    "PRIVATE",
    "CERT",
)

NOMBRES_EXCLUIDOS = {
    ".env",
    ".env.local",
    ".env.production",
    "token_cache.json",
    "msal_cache.bin",
}

COLUMNAS_FECHA_PROCESO_CANDIDATAS = [
    "fecha_procesamiento",
    "fecha procesamiento",
    "fecha proceso",
    "procesado_en",
    "procesado en",
    "fecha_registro",
    "fecha registro",
    "fecha de registro",
    "fecha_carga",
    "fecha carga",
    "creado_en",
    "created_at",
]

COLS_NUMERO = [
    "número de factura",
    "numero de factura",
    "numerofactura",
    "numero factura",
    "factura",
    "numero",
]
COLS_RADICADO = ["radicado"]
COLS_CUFE = ["cufe", "cude", "cufe/cude", "uuid"]
COLS_ARCHIVO = ["archivo", "archivo origen", "nombre archivo", "pdf", "xml"]
COLS_ARCHIVO_AUDIT_FALLBACK = [
    "pdf_name",
    "pdf name",
    "archivo_pdf",
    "archivo pdf",
    "nombre_pdf",
    "nombre pdf",
    "attachment_name",
    "attachment name",
]
COLS_ASUNTO_AUDIT = [
    "subject",
    "subj",
    "asunto",
    "email_subject",
    "email subject",
    "mail_subject",
    "mail subject",
]
COLS_FECHA_AUDIT = [
    "fecha_hora",
    "fecha hora",
    "fecha_hora_proceso",
    "fecha hora proceso",
    "timestamp",
    "procesado_en",
    "procesado en",
    "processed_at",
    "processed at",
]
COLS_INICIO_AUDIT = [
    "inicio",
    "fecha_inicio",
    "fecha inicio",
    "started_at",
    "started at",
]
COLS_FIN_AUDIT = [
    "fin",
    "fecha_fin",
    "fecha fin",
    "finished_at",
    "finished at",
]
COLS_CONCEPTO = ["concepto"]

PATRON_RADICADO_ASUNTO = re.compile(
    r"(?i)\bradicado"
    r"(?:\s*(?:n[.°ºo]*|no\.?))?"
    r"\s*[-:#]?\s*(\d{4,})\b"
)

EXT_LOGS = {".log", ".txt", ".csv", ".json"}
EXCEL_SHEET_NAME = "Facturas"
EXCEL_TABLE_NAME = "TblFacturasDiario"


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Genera cierre diario local seguro V3.")
    parser.add_argument(
        "--fecha",
        default=_dt.datetime.now().strftime("%Y-%m-%d"),
        help="Fecha del cierre en formato YYYY-MM-DD. Default: hoy.",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Solo diagnostica qué se generaría. No crea/copia archivos.",
    )
    parser.add_argument(
        "--permitir-vacio",
        action="store_true",
        help="Permite generar cierre aunque el Excel diario quede sin filas. Útil si no hubo facturas nuevas.",
    )
    return parser.parse_args()


def validar_fecha(fecha: str) -> _dt.date:
    try:
        return _dt.datetime.strptime(fecha, "%Y-%m-%d").date()
    except ValueError as exc:
        raise RuntimeError(f"Fecha inválida {fecha!r}. Usa formato YYYY-MM-DD.") from exc


def rango_semana_lunes_domingo(fecha: _dt.date) -> tuple[_dt.date, _dt.date]:
    inicio = fecha - _dt.timedelta(days=fecha.weekday())
    fin = inicio + _dt.timedelta(days=6)
    return inicio, fin

MESES_ES = {
    1: "Enero",
    2: "Febrero",
    3: "Marzo",
    4: "Abril",
    5: "Mayo",
    6: "Junio",
    7: "Julio",
    8: "Agosto",
    9: "Septiembre",
    10: "Octubre",
    11: "Noviembre",
    12: "Diciembre",
}


def nombre_mes_dir(fecha: _dt.date) -> str:
    return f"{fecha.strftime('%Y-%m')}_{MESES_ES[int(fecha.strftime('%m'))]}"



def rel(path: Path) -> str:
    try:
        return str(path.resolve().relative_to(ROOT.resolve())).replace("\\", "/")
    except Exception:
        return str(path).replace("\\", "/")


def norm(value: Any) -> str:
    s = "" if value is None else str(value)
    s = s.strip().replace("\xa0", " ")
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = re.sub(r"\s+", " ", s)
    return s.lower()


def norm_key(value: Any) -> str:
    s = norm(value)
    return re.sub(r"[^a-z0-9]+", "", s)


def clean_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (_dt.datetime, _dt.date)):
        return value.isoformat()
    return str(value).strip()


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def iso_mtime(path: Path) -> str:
    try:
        return _dt.datetime.fromtimestamp(path.stat().st_mtime).isoformat(timespec="seconds")
    except Exception:
        return ""


def acquire_lock() -> Tuple[bool, str]:
    LOCKS_DIR.mkdir(parents=True, exist_ok=True)
    now_ts = time.time()

    if LOCK_FILE.exists():
        age = now_ts - LOCK_FILE.stat().st_mtime
        if age < LOCK_TTL_SECONDS:
            return False, f"Lock activo: {LOCK_FILE} | edad={age:.0f}s | ttl={LOCK_TTL_SECONDS}s"
        try:
            LOCK_FILE.unlink()
        except Exception as exc:
            return False, f"No se pudo eliminar lock vencido: {LOCK_FILE} | {exc}"

    payload = {
        "script": "cierre_diario_seguro.py",
        "version": VERSION_CIERRE,
        "pid": os.getpid(),
        "started_at": _dt.datetime.now().isoformat(timespec="seconds"),
        "host": socket.gethostname(),
    }
    LOCK_FILE.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
    return True, "Lock creado"


def release_lock() -> None:
    try:
        LOCK_FILE.unlink(missing_ok=True)
    except Exception:
        pass


def es_sensible(clave: str) -> bool:
    clave_upper = str(clave or "").strip().upper()
    return any(palabra in clave_upper for palabra in PALABRAS_SENSIBLES)


def redactar_env_line(linea: str) -> str:
    if not linea.strip() or linea.lstrip().startswith("#") or "=" not in linea:
        return linea
    clave, valor = linea.split("=", 1)
    if es_sensible(clave):
        return f"{clave}=***REDACTADO***"
    if re.search(r"[A-Za-z0-9_\-]{32,}", valor) and clave.strip().upper() not in {
        "SP_DRIVE_ID",
        "SP_DRIVE_ID_RADICADOS",
        "SP_BACKUP2_DRIVE_ID",
        "ONEDRIVE_BACKUP_DRIVE_ID",
        "BACKUP_DRIVE_ID",
    }:
        return f"{clave}=***REDACTADO***"
    return linea


def copiar_archivo(origen: Path, destino: Path) -> Optional[Dict[str, Any]]:
    if not origen.exists() or not origen.is_file() or origen.name in NOMBRES_EXCLUIDOS:
        return None
    destino.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(origen, destino)
    return info_archivo(destino, origen=origen)


def info_archivo(destino: Path, origen: Optional[Path] = None, categoria: str = "") -> Dict[str, Any]:
    st = destino.stat()
    return {
        "categoria": categoria,
        "origen": rel(origen) if origen else "generado",
        "destino": rel(destino),
        "nombre": destino.name,
        "bytes": st.st_size,
        "sha256": sha256_file(destino),
        "mtime_origen": iso_mtime(origen) if origen and origen.exists() else "",
        "mtime_destino": iso_mtime(destino),
    }


def leer_csv_dicts(path: Path) -> list[dict[str, str]]:
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            with path.open("r", encoding=encoding, newline="") as f:
                reader = csv.DictReader(f)
                return [dict(r or {}) for r in reader]
        except UnicodeDecodeError:
            continue
        except Exception:
            return []
    return []


def _fecha_en_nombre_archivo(path: Path, fecha: str) -> bool:
    fecha_compacta = fecha.replace("-", "")
    name = path.name.lower()
    return fecha in name or fecha_compacta in name


def _parse_datetime_audit(value: Any) -> Optional[_dt.datetime]:
    if value is None:
        return None
    if isinstance(value, _dt.datetime):
        return value
    if isinstance(value, _dt.date):
        return _dt.datetime.combine(value, _dt.time.min)

    s = str(value).strip()
    if not s:
        return None

    candidate = s.replace("Z", "+00:00")
    try:
        parsed = _dt.datetime.fromisoformat(candidate)
        if parsed.tzinfo is not None:
            parsed = parsed.replace(tzinfo=None)
        return parsed
    except Exception:
        pass

    for fmt in (
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%dT%H:%M:%S",
        "%d/%m/%Y %H:%M:%S",
        "%d-%m-%Y %H:%M:%S",
        "%Y-%m-%d",
        "%d/%m/%Y",
        "%d-%m-%Y",
    ):
        try:
            return _dt.datetime.strptime(s[:19], fmt)
        except Exception:
            continue
    return None


def _intervalo_audit_incluye_fecha(row: dict[str, Any], fecha: str) -> bool:
    inicio_raw = extraer_primer_valor_no_vacio(row, COLS_INICIO_AUDIT)
    fin_raw = extraer_primer_valor_no_vacio(row, COLS_FIN_AUDIT)

    inicio = _parse_datetime_audit(inicio_raw)
    fin = _parse_datetime_audit(fin_raw)

    if inicio is None and fin is None:
        return False

    target = validar_fecha(fecha)
    inicio_dia = _dt.datetime.combine(target, _dt.time.min)
    fin_dia = _dt.datetime.combine(target, _dt.time.max)

    if inicio is None:
        inicio = fin
    if fin is None:
        fin = inicio

    assert inicio is not None and fin is not None
    if fin < inicio:
        inicio, fin = fin, inicio

    return inicio <= fin_dia and fin >= inicio_dia


def _fila_audit_corresponde_fecha(
    row: dict[str, Any],
    fecha: str,
    *,
    archivo: Optional[Path] = None,
    permitir_fallback_nombre: bool = True,
) -> tuple[bool, str]:
    """
    Decide si una fila de auditoría pertenece a la fecha solicitada.

    Prioridad:
    1. fecha_hora (o equivalente) de la propia fila;
    2. intervalo inicio/fin para filas de resumen de ejecución;
    3. compatibilidad histórica por nombre de archivo solo si la fila
       no contiene ninguna fecha utilizable.
    """
    fecha_hora = extraer_primer_valor_no_vacio(row, COLS_FECHA_AUDIT)
    if fecha_hora:
        if fecha_valor_coincide(fecha_hora, fecha):
            return True, "fecha_hora"
        return False, "otra_fecha"

    inicio = extraer_primer_valor_no_vacio(row, COLS_INICIO_AUDIT)
    fin = extraer_primer_valor_no_vacio(row, COLS_FIN_AUDIT)
    if inicio or fin:
        if _intervalo_audit_incluye_fecha(row, fecha):
            return True, "intervalo_run"
        return False, "otro_intervalo"

    if permitir_fallback_nombre and archivo is not None and _fecha_en_nombre_archivo(archivo, fecha):
        return True, "fallback_nombre"

    return False, "sin_fecha"


def _rutas_audit_disponibles() -> list[Path]:
    rutas: list[Path] = []
    patrones = (
        "audit_detalle_*.csv",
        "audit_runs_*.csv",
        "audit_*.csv",
    )

    for base in (AUDIT_DIR, DATA_DIR, ROOT):
        if not base.exists():
            continue
        for patron in patrones:
            for p in base.glob(patron):
                if p.is_file() and "audit" in p.name.lower():
                    rutas.append(p)

    vistos: set[str] = set()
    out: list[Path] = []
    for p in rutas:
        try:
            key = str(p.resolve()).casefold()
        except Exception:
            key = str(p).casefold()
        if key not in vistos:
            vistos.add(key)
            out.append(p)

    return sorted(out, key=lambda p: str(p).casefold())


def recolectar_audits(fecha: str) -> list[Path]:
    """
    Localiza auditorías relevantes por contenido, no solo por el nombre.

    Esto permite reconstruir correctamente días históricos cuando una
    ejecución larga comenzó un día y terminó al siguiente, dejando filas
    del día anterior dentro de un CSV cuyo nombre corresponde al día de
    finalización.
    """
    relevantes: list[Path] = []

    for p in _rutas_audit_disponibles():
        rows = leer_csv_dicts(p)

        if not rows:
            if _fecha_en_nombre_archivo(p, fecha):
                relevantes.append(p)
            continue

        for row in rows:
            coincide, _motivo = _fila_audit_corresponde_fecha(
                row,
                fecha,
                archivo=p,
                permitir_fallback_nombre=True,
            )
            if coincide:
                relevantes.append(p)
                break

    vistos: set[str] = set()
    out: list[Path] = []
    for p in relevantes:
        try:
            key = str(p.resolve()).casefold()
        except Exception:
            key = str(p).casefold()
        if key not in vistos:
            vistos.add(key)
            out.append(p)

    return sorted(out, key=lambda p: str(p).casefold())


def log_corresponde_fecha(path: Path, fecha: str) -> bool:
    name = path.name.lower()
    fecha_compacta = fecha.replace("-", "")
    if fecha in name or fecha_compacta in name:
        return True
    try:
        mdate = _dt.datetime.fromtimestamp(path.stat().st_mtime).date().isoformat()
        return mdate == fecha
    except Exception:
        return False


def bases_logs_disponibles() -> list[tuple[str, Path]]:
    """
    Devuelve las ubicaciones de logs compatibles con local y VPS.

    Los paths se deduplican por ruta resuelta para evitar copiar dos veces
    el mismo directorio cuando data/logs es un enlace simbólico.
    """
    candidatos = [
        ("app_logs", LOGS_DIR),
        ("data_logs", DATA_LOGS_DIR),
        ("runtime", VPS_LOG_ROOT / "runtime"),
        ("runs", VPS_LOG_ROOT / "runs"),
        ("cron", VPS_LOG_ROOT / "cron"),
    ]

    vistos: set[str] = set()
    bases: list[tuple[str, Path]] = []

    for etiqueta, base in candidatos:
        if not base.exists() or not base.is_dir():
            continue

        try:
            key = str(base.resolve()).casefold()
        except Exception:
            key = str(base).casefold()

        if key in vistos:
            continue

        vistos.add(key)
        bases.append((etiqueta, base))

    return bases


def recolectar_logs(fecha: str) -> list[Path]:
    rutas: list[Path] = []

    for _etiqueta, base in bases_logs_disponibles():
        for p in base.rglob("*"):
            if not p.is_file():
                continue
            if p.suffix.lower() not in EXT_LOGS:
                continue
            if log_corresponde_fecha(p, fecha):
                rutas.append(p)

    vistos: set[str] = set()
    out: list[Path] = []

    for p in rutas:
        try:
            key = str(p.resolve()).casefold()
        except Exception:
            key = str(p).casefold()

        if key not in vistos:
            vistos.add(key)
            out.append(p)

    return sorted(out, key=lambda p: str(p).casefold())


def find_col(headers: Sequence[Any], candidates: Sequence[str]) -> Optional[int]:
    candidate_norms = {norm(c) for c in candidates}
    candidate_keys = {norm_key(c) for c in candidates}
    for idx, header in enumerate(headers):
        hn = norm(header)
        hk = norm_key(header)
        if hn in candidate_norms or hk in candidate_keys:
            return idx
    return None


def all_header_map(headers: Sequence[Any]) -> dict[str, int]:
    out: dict[str, int] = {}
    for i, h in enumerate(headers):
        if str(h or "").strip():
            out[norm(h)] = i
            out[norm_key(h)] = i
    return out


def fecha_valor_coincide(value: Any, fecha: str) -> bool:
    if value is None:
        return False
    if isinstance(value, _dt.datetime):
        return value.date().isoformat() == fecha
    if isinstance(value, _dt.date):
        return value.isoformat() == fecha
    s = str(value).strip()
    if not s:
        return False
    if fecha in s:
        return True
    # Soporta ISO datetime o dd/mm/yyyy de forma básica.
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return _dt.datetime.strptime(s[:10], fmt).date().isoformat() == fecha
        except Exception:
            pass
    return False


def extraer_valor_row(row: dict[str, Any], candidates: Sequence[str]) -> str:
    by_norm = {norm(k): v for k, v in row.items()}
    by_key = {norm_key(k): v for k, v in row.items()}
    for c in candidates:
        if norm(c) in by_norm:
            return clean_value(by_norm[norm(c)])
        if norm_key(c) in by_key:
            return clean_value(by_key[norm_key(c)])
    return ""


def extraer_primer_valor_no_vacio(
    row: dict[str, Any],
    candidates: Sequence[str],
) -> str:
    """
    Busca el primer valor no vacío entre varios nombres posibles de columna.

    A diferencia de extraer_valor_row(), no se detiene cuando una columna
    existe pero viene vacía. Esto es necesario para auditorías que conservan
    columnas antiguas vacías y guardan el dato real en una columna nueva.
    """
    by_norm = {norm(k): v for k, v in row.items()}
    by_key = {norm_key(k): v for k, v in row.items()}

    for candidate in candidates:
        values: list[Any] = []

        candidate_norm = norm(candidate)
        candidate_key = norm_key(candidate)

        if candidate_norm in by_norm:
            values.append(by_norm[candidate_norm])

        if candidate_key in by_key:
            values.append(by_key[candidate_key])

        for value in values:
            cleaned = clean_value(value)
            if cleaned:
                return cleaned

    return ""


def extraer_radicado_desde_asunto(asunto: str) -> str:
    if not asunto:
        return ""

    match = PATRON_RADICADO_ASUNTO.search(str(asunto))
    return match.group(1) if match else ""


def extraer_candidatos_desde_audits(
    audit_files: list[Path],
    fecha: str,
) -> tuple[dict[str, set[str]], dict[str, Any]]:
    """
    Extrae identificadores únicamente de las filas de auditoría que
    corresponden realmente a la fecha del cierre.

    La fecha interna de la fila (fecha_hora o equivalente) tiene prioridad
    sobre el nombre físico del CSV. Esto evita que una ejecución que cruza
    medianoche arrastre al cierre del día siguiente todas las facturas
    procesadas el día anterior.

    Para auditorías antiguas sin fecha por fila se mantiene compatibilidad
    usando el nombre del archivo como fallback.
    """
    candidatos: dict[str, set[str]] = {
        "cufe": set(),
        "numero": set(),
        "radicado": set(),
        "archivo": set(),
        "numero_radicado": set(),
        "numero_archivo": set(),
        "radicado_archivo": set(),
    }

    meta: dict[str, Any] = {
        "fecha_objetivo": fecha,
        "archivos_revisados": [rel(p) for p in audit_files],
        "filas_auditoria_leidas": 0,
        "filas_aceptadas_fecha": 0,
        "filas_descartadas_otra_fecha": 0,
        "filas_aceptadas_intervalo_run": 0,
        "filas_aceptadas_fallback_nombre": 0,
        "filas_sin_fecha_ignoradas": 0,
    }

    for audit in audit_files:
        for row in leer_csv_dicts(audit):
            meta["filas_auditoria_leidas"] += 1

            coincide, motivo = _fila_audit_corresponde_fecha(
                row,
                fecha,
                archivo=audit,
                permitir_fallback_nombre=True,
            )

            if not coincide:
                if motivo in {"otra_fecha", "otro_intervalo"}:
                    meta["filas_descartadas_otra_fecha"] += 1
                else:
                    meta["filas_sin_fecha_ignoradas"] += 1
                continue

            if motivo == "intervalo_run":
                meta["filas_aceptadas_intervalo_run"] += 1
            elif motivo == "fallback_nombre":
                meta["filas_aceptadas_fallback_nombre"] += 1
            else:
                meta["filas_aceptadas_fecha"] += 1

            cufe = extraer_primer_valor_no_vacio(row, COLS_CUFE)
            numero = extraer_primer_valor_no_vacio(row, COLS_NUMERO)
            radicado = extraer_primer_valor_no_vacio(row, COLS_RADICADO)
            archivo = extraer_primer_valor_no_vacio(row, COLS_ARCHIVO)

            if not archivo:
                archivo = extraer_primer_valor_no_vacio(
                    row,
                    COLS_ARCHIVO_AUDIT_FALLBACK,
                )

            if not radicado:
                asunto = extraer_primer_valor_no_vacio(
                    row,
                    COLS_ASUNTO_AUDIT,
                )
                radicado = extraer_radicado_desde_asunto(asunto)

            cufe_key = norm_key(cufe)
            numero_key = norm_key(numero)
            radicado_key = norm_key(radicado)
            archivo_key = norm_key(archivo)

            if cufe_key:
                candidatos["cufe"].add(cufe_key)
            if numero_key:
                candidatos["numero"].add(numero_key)
            if radicado_key:
                candidatos["radicado"].add(radicado_key)
            if archivo_key:
                candidatos["archivo"].add(archivo_key)

            if numero_key and radicado_key:
                candidatos["numero_radicado"].add(
                    f"{numero_key}|{radicado_key}"
                )
            if numero_key and archivo_key:
                candidatos["numero_archivo"].add(
                    f"{numero_key}|{archivo_key}"
                )
            if radicado_key and archivo_key:
                candidatos["radicado_archivo"].add(
                    f"{radicado_key}|{archivo_key}"
                )

    return candidatos, meta


def fila_match_candidatos(
    row_values: Sequence[Any],
    headers: Sequence[Any],
    candidatos: dict[str, set[str]],
) -> bool:
    idx_cufe = find_col(headers, COLS_CUFE)
    idx_numero = find_col(headers, COLS_NUMERO)
    idx_radicado = find_col(headers, COLS_RADICADO)
    idx_archivo = find_col(headers, COLS_ARCHIVO)

    cufe = norm_key(row_values[idx_cufe]) if idx_cufe is not None and idx_cufe < len(row_values) else ""
    numero = norm_key(row_values[idx_numero]) if idx_numero is not None and idx_numero < len(row_values) else ""
    radicado = norm_key(row_values[idx_radicado]) if idx_radicado is not None and idx_radicado < len(row_values) else ""
    archivo = norm_key(row_values[idx_archivo]) if idx_archivo is not None and idx_archivo < len(row_values) else ""

    if cufe and cufe in candidatos["cufe"]:
        return True
    if numero and radicado and f"{numero}|{radicado}" in candidatos["numero_radicado"]:
        return True
    if numero and archivo and f"{numero}|{archivo}" in candidatos["numero_archivo"]:
        return True
    if radicado and archivo and f"{radicado}|{archivo}" in candidatos["radicado_archivo"]:
        return True

    # Fallback más flexible, solo cuando hay combinación para reducir falsos positivos.
    if numero and radicado and numero in candidatos["numero"] and radicado in candidatos["radicado"]:
        return True
    if numero and archivo and numero in candidatos["numero"] and archivo in candidatos["archivo"]:
        return True

    return False


def obtener_filas_diarias_desde_excel(
    fecha: str,
    audit_files: list[Path],
) -> tuple[list[Any], list[list[Any]], dict[str, Any]]:
    if not FACTURAS_PATH.exists():
        raise RuntimeError(f"No existe Excel operativo local: {FACTURAS_PATH}")

    print(f"?? Leyendo Excel operativo: {FACTURAS_PATH}")

    wb = load_workbook(FACTURAS_PATH, data_only=True, read_only=True)
    try:
        if EXCEL_SHEET_NAME in wb.sheetnames:
            ws = wb[EXCEL_SHEET_NAME]
        else:
            ws = wb[wb.sheetnames[0]]

        print(f"?? Hoja usada: {ws.title} | max_row={ws.max_row} | max_column={ws.max_column}")

        rows_iter = ws.iter_rows(values_only=True)

        try:
            headers_raw = next(rows_iter)
        except StopIteration:
            raise RuntimeError("El Excel operativo no tiene filas.")

        headers = list(headers_raw or [])
        total_cols = len(headers)

        rows_all: list[list[Any]] = []
        leidas = 0

        for values in rows_iter:
            leidas += 1

            row = list(values or [])

            if total_cols > 0:
                if len(row) < total_cols:
                    row.extend([None] * (total_cols - len(row)))
                elif len(row) > total_cols:
                    row = row[:total_cols]

            if any(v not in (None, "") for v in row):
                rows_all.append(row)

            if leidas % 1000 == 0:
                print(f"   ... filas le?das desde Excel: {leidas} | filas con datos: {len(rows_all)}")

        print(f"? Lectura Excel terminada: filas le?das={leidas} | filas con datos={len(rows_all)}")

        meta: dict[str, Any] = {
            "excel_operativo": rel(FACTURAS_PATH),
            "hoja_usada": ws.title,
            "filas_operativas_total": len(rows_all),
            "metodo_seleccion": "",
            "advertencias": [],
        }

        idx_fecha = find_col(headers, COLUMNAS_FECHA_PROCESO_CANDIDATAS)
        if idx_fecha is not None:
            rows_fecha = [
                row for row in rows_all
                if idx_fecha < len(row) and fecha_valor_coincide(row[idx_fecha], fecha)
            ]
            meta["metodo_seleccion"] = "columna_fecha_procesamiento"
            meta["columna_fecha_procesamiento"] = headers[idx_fecha]
            return headers, rows_fecha, meta

        candidatos, meta_auditoria = extraer_candidatos_desde_audits(
            audit_files,
            fecha,
        )
        total_candidates = sum(len(v) for v in candidatos.values())

        meta["audit_files_usados"] = [rel(p) for p in audit_files]
        meta["candidatos_extraidos"] = {k: len(v) for k, v in candidatos.items()}
        meta["filtro_auditoria_por_fecha"] = meta_auditoria

        if total_candidates <= 0:
            meta["metodo_seleccion"] = "sin_candidatos_auditoria"
            meta["advertencias"].append(
                "No se encontr? columna de fecha de procesamiento ni candidatos suficientes en auditor?a. "
                "Se genera Excel diario solo con encabezados."
            )
            return headers, [], meta

        print(f"?? Cruzando Excel operativo contra auditor?a del d?a. Candidatos={total_candidates}")
        print(
            "?? Filtro auditoria por fecha: "
            f"leidas={meta_auditoria['filas_auditoria_leidas']} | "
            f"aceptadas_fecha={meta_auditoria['filas_aceptadas_fecha']} | "
            f"aceptadas_intervalo={meta_auditoria['filas_aceptadas_intervalo_run']} | "
            f"fallback_nombre={meta_auditoria['filas_aceptadas_fallback_nombre']} | "
            f"descartadas_otra_fecha={meta_auditoria['filas_descartadas_otra_fecha']} | "
            f"sin_fecha_ignoradas={meta_auditoria['filas_sin_fecha_ignoradas']}"
        )

        rows_match = []
        for idx, row in enumerate(rows_all, start=1):
            if fila_match_candidatos(row, headers, candidatos):
                rows_match.append(row)

            if idx % 1000 == 0:
                print(f"   ... cruce auditor?a: {idx}/{len(rows_all)} | matches={len(rows_match)}")

        meta["metodo_seleccion"] = "auditoria_del_dia_vs_excel_operativo"
        return headers, rows_match, meta

    finally:
        wb.close()

def ajustar_ancho_columnas(ws) -> None:
    for column_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells[:80]:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 45)


def crear_excel_diario(destino: Path, headers: list[Any], rows: list[list[Any]]) -> dict[str, Any]:
    destino.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = EXCEL_SHEET_NAME

    header_clean = ["" if h is None else str(h).strip() for h in headers]
    ws.append(header_clean)

    for row in rows:
        normalized = list(row)
        if len(normalized) < len(header_clean):
            normalized += [None] * (len(header_clean) - len(normalized))
        ws.append(normalized[: len(header_clean)])

    max_col = max(1, len(header_clean))
    max_row = max(1, len(rows) + 1)
    last_col = get_column_letter(max_col)
    table_ref = f"A1:{last_col}{max_row}"

    # Excel permite tabla de solo encabezado. Mantiene estructura aunque no haya registros del día.
    tab = Table(displayName=EXCEL_TABLE_NAME, ref=table_ref)
    tab.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tab)
    ws.freeze_panes = "A2"
    ajustar_ancho_columnas(ws)

    wb.save(destino)
    wb.close()

    return {
        "ruta": rel(destino),
        "filas_datos": len(rows),
        "columnas": len(header_clean),
        "tabla": EXCEL_TABLE_NAME,
        "bytes": destino.stat().st_size,
        "sha256": sha256_file(destino),
    }


def _leer_csv_con_campos(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            with path.open("r", encoding=encoding, newline="") as f:
                reader = csv.DictReader(f)
                fieldnames = list(reader.fieldnames or [])
                rows = [dict(r or {}) for r in reader]
                return fieldnames, rows
        except UnicodeDecodeError:
            continue
        except Exception:
            return [], []
    return [], []


def _nombre_audit_diario(origen: Path, fecha: str) -> str:
    name = origen.name
    lower = name.lower()
    if lower.startswith("audit_detalle_"):
        return f"audit_detalle_{fecha}.csv"
    if lower.startswith("audit_runs_"):
        return f"audit_runs_{fecha}.csv"
    return name


def copiar_auditorias(
    audit_files: list[Path],
    destino_dir: Path,
    fecha: str,
) -> list[dict[str, Any]]:
    """
    Copia evidencia de auditoría filtrada a la fecha del cierre.

    Si el CSV de origen mezcla días, solo se escriben en el cierre las filas
    que pertenecen a la fecha solicitada. Los archivos detalle/runs se
    renombran con la fecha real del cierre para evitar confusión histórica.
    """
    items: list[dict[str, Any]] = []
    destino_dir.mkdir(parents=True, exist_ok=True)

    acumulados: dict[str, dict[str, Any]] = {}

    for origen in audit_files:
        fieldnames, rows = _leer_csv_con_campos(origen)
        if not fieldnames:
            continue

        filtradas: list[dict[str, str]] = []
        for row in rows:
            coincide, _motivo = _fila_audit_corresponde_fecha(
                row,
                fecha,
                archivo=origen,
                permitir_fallback_nombre=True,
            )
            if coincide:
                filtradas.append(row)

        if not filtradas:
            continue

        nombre_destino = _nombre_audit_diario(origen, fecha)
        bucket = acumulados.setdefault(
            nombre_destino,
            {
                "fieldnames": [],
                "rows": [],
                "origenes": [],
            },
        )

        for field in fieldnames:
            if field not in bucket["fieldnames"]:
                bucket["fieldnames"].append(field)

        bucket["rows"].extend(filtradas)
        bucket["origenes"].append(origen)

    for nombre_destino, bucket in sorted(acumulados.items()):
        destino = destino_dir / nombre_destino
        fieldnames = list(bucket["fieldnames"])
        rows = list(bucket["rows"])

        unicas: list[dict[str, str]] = []
        vistos: set[tuple[str, ...]] = set()
        for row in rows:
            key = tuple(clean_value(row.get(field, "")) for field in fieldnames)
            if key in vistos:
                continue
            vistos.add(key)
            unicas.append(row)

        with destino.open("w", encoding="utf-8-sig", newline="") as f:
            writer = csv.DictWriter(
                f,
                fieldnames=fieldnames,
                extrasaction="ignore",
            )
            writer.writeheader()
            writer.writerows(unicas)

        origenes = bucket["origenes"]
        info = info_archivo(
            destino,
            origen=origenes[0] if origenes else None,
            categoria="auditoria",
        )
        info["origenes_fuente"] = [rel(p) for p in origenes]
        info["filas_filtradas_fecha"] = len(unicas)
        items.append(info)

    return items


def copiar_logs(
    log_files: list[Path],
    destino_dir: Path,
) -> list[dict[str, Any]]:
    items: list[dict[str, Any]] = []
    destino_dir.mkdir(parents=True, exist_ok=True)
    bases = bases_logs_disponibles()

    for origen in log_files:
        rel_log: Optional[Path] = None

        for etiqueta, base in bases:
            try:
                rel_origen = origen.resolve().relative_to(base.resolve())
                rel_log = Path(etiqueta) / rel_origen
                break
            except Exception:
                continue

        if rel_log is None:
            rel_log = Path("otros") / origen.name

        destino = destino_dir / rel_log
        info = copiar_archivo(origen, destino)

        if info:
            info["categoria"] = "log_produccion"
            items.append(info)

    return items


def crear_snapshot_env_redactado(destino_dir: Path, fecha: str) -> Optional[dict[str, Any]]:
    env_path = ROOT / ".env"
    if not env_path.exists():
        return None
    destino_dir.mkdir(parents=True, exist_ok=True)
    destino = destino_dir / f"snapshot_env_redactado_{fecha}.txt"
    lineas = env_path.read_text(encoding="utf-8", errors="replace").splitlines()
    destino.write_text("\n".join(redactar_env_line(x) for x in lineas) + "\n", encoding="utf-8")
    info = info_archivo(destino, origen=env_path, categoria="config_redactada")
    return info


def validar_excel_generado(path: Path, expected_columns: int) -> dict[str, Any]:
    out: dict[str, Any] = {
        "archivo": rel(path),
        "existe": path.exists(),
        "abre_ok": False,
        "hojas": [],
        "filas_datos": 0,
        "columnas": 0,
        "errores": [],
    }
    if not path.exists():
        out["errores"].append("No existe Excel diario.")
        return out

    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            out["abre_ok"] = True
            out["hojas"] = list(wb.sheetnames)
            ws = wb[EXCEL_SHEET_NAME] if EXCEL_SHEET_NAME in wb.sheetnames else wb[wb.sheetnames[0]]
            out["columnas"] = int(ws.max_column or 0)
            out["filas_datos"] = max(int(ws.max_row or 0) - 1, 0)
            if expected_columns and out["columnas"] != expected_columns:
                out["errores"].append(
                    f"Columnas distintas. Esperadas={expected_columns}, detectadas={out['columnas']}"
                )
        finally:
            wb.close()
    except Exception as exc:
        out["errores"].append(f"No se pudo abrir Excel diario: {type(exc).__name__}: {exc}")

    return out


def generar_resumen_txt(path: Path, manifest: dict[str, Any]) -> None:
    lines = [
        "CIERRE DIARIO LOCAL SEGURO V3 - FACTURAS JOYCO",
        "=" * 90,
        f"Versión: {manifest['version']}",
        f"Fecha cierre: {manifest['fecha']}",
        f"Trimestre: {manifest['trimestre']['nombre_carpeta']}",
        f"Rango trimestre: {manifest['trimestre']['fecha_inicio']} a {manifest['trimestre']['fecha_fin']}",
        f"Generado: {manifest['generado_en']}",
        f"Root: {manifest['root']}",
        f"Carpeta cierre: {manifest['carpeta_cierre']}",
        "",
        "Resumen:",
        f"- Filas en Excel diario: {manifest['excel_diario']['filas_datos']}",
        f"- Columnas en Excel diario: {manifest['excel_diario']['columnas']}",
        f"- Método selección: {manifest['seleccion_filas'].get('metodo_seleccion')}",
        f"- Auditorías copiadas: {manifest['conteos']['auditorias']}",
        f"- Logs copiados: {manifest['conteos']['logs']}",
        f"- Total archivos evidencia: {manifest['total_archivos']}",
        f"- Total bytes evidencia: {manifest['total_bytes']}",
        "",
        "Regla aplicada:",
        "- Este cierre diario contiene únicamente lo registrado/procesado en la fecha del cierre, no el histórico completo.",
        "- Si no hubo registros del día, el Excel diario queda solo con encabezados y el manifest deja advertencia.",
        "- No se incluye .env real ni secretos.",
        "- No se sube nada desde este script; la subida remota se hace en el paso de OneDrive.",
        "",
    ]
    advertencias = manifest.get("advertencias") or []
    if advertencias:
        lines.append("Advertencias:")
        for adv in advertencias:
            lines.append(f"- {adv}")
        lines.append("")
    lines.append("=" * 90)
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def listar_archivos_para_manifest(base: Path, excluir: Optional[set[Path]] = None) -> list[dict[str, Any]]:
    excluir_resolved = {p.resolve() for p in (excluir or set()) if p.exists()}
    items: list[dict[str, Any]] = []
    for p in sorted(base.rglob("*")):
        if not p.is_file():
            continue
        try:
            if p.resolve() in excluir_resolved:
                continue
        except Exception:
            pass
        items.append(info_archivo(p, categoria="evidencia"))
    return items


def main() -> int:
    args = _parse_args()
    fecha_date = validar_fecha(args.fecha)
    fecha = fecha_date.isoformat()
    mes = fecha_date.strftime("%Y-%m")
    anio = fecha_date.strftime("%Y")
    mes_nombre = nombre_mes_dir(fecha_date)
    semana_inicio, semana_fin = rango_semana_lunes_domingo(fecha_date)
    semana_nombre = f"SEMANA_{semana_inicio.isoformat()}_a_{semana_fin.isoformat()}"
    trimestre = cargar_trimestre_activo(ROOT, fecha_date)

    cierre_dia_dir = (
        CIERRES_DIR
        / trimestre["anio"]
        / trimestre["nombre_carpeta"]
        / mes_nombre
        / semana_nombre
        / f"Diario_{fecha}"
    )
    excel_dir = cierre_dia_dir / "01_Excel_Diario"
    auditoria_dir = cierre_dia_dir / "02_Auditoria"
    logs_dir = cierre_dia_dir / "03_Logs" / f"logs_produccion_{fecha}"
    manifest_dir = cierre_dia_dir / "04_Manifest"
    validaciones_dir = cierre_dia_dir / "05_Validaciones"

    excel_diario_path = excel_dir / f"facturas_diario_{fecha}.xlsx"
    manifest_path = manifest_dir / f"manifest_diario_{fecha}.json"
    resumen_path = manifest_dir / f"resumen_diario_{fecha}.txt"
    validacion_path = validaciones_dir / f"validacion_local_{fecha}.json"

    print("=" * 100)
    print("CIERRE DIARIO LOCAL SEGURO V3 - FACTURAS JOYCO")
    print("=" * 100)
    print(f"Versión: {VERSION_CIERRE}")
    print(f"Root: {ROOT}")
    print(f"Fecha: {fecha}")
    print(f"Trimestre activo: {trimestre['nombre_carpeta']}")
    print(
        "Rango trimestre: "
        f"{trimestre['fecha_inicio']} a {trimestre['fecha_fin']}"
    )
    print(f"Mes carpeta: {mes_nombre}")
    print(f"Semana: {semana_inicio.isoformat()} a {semana_fin.isoformat()}")
    print(f"Carpeta local destino: {cierre_dia_dir}")
    print(f"Dry-run: {args.dry_run}")
    print("-" * 100)

    ok_lock, msg_lock = acquire_lock()
    if not ok_lock:
        print(f"❌ {msg_lock}")
        print("⚠️ No se ejecuta el cierre para evitar cruces.")
        return 2

    try:
        audit_files = recolectar_audits(fecha)
        log_files = recolectar_logs(fecha)
        headers, filas_diarias, meta_seleccion = obtener_filas_diarias_desde_excel(fecha, audit_files)

        advertencias: list[str] = []
        advertencias.extend(meta_seleccion.get("advertencias", []))
        if not audit_files:
            advertencias.append("No se encontraron archivos de auditoría del día.")
        if not log_files:
            advertencias.append("No se encontraron logs específicos del día.")
        if not filas_diarias:
            advertencias.append("El Excel diario no tendrá filas de datos para esta fecha.")

        print(f"📌 Auditorías detectadas: {len(audit_files)}")
        print(f"📌 Logs detectados: {len(log_files)}")
        print(f"📌 Método selección filas: {meta_seleccion.get('metodo_seleccion')}")
        print(f"📌 Filas diarias detectadas: {len(filas_diarias)}")

        if args.dry_run:
            print("-" * 100)
            print("DRY-RUN: no se generaron archivos.")
            if advertencias:
                print("⚠️ Advertencias:")
                for adv in advertencias:
                    print(f"  - {adv}")
            return 0

        if cierre_dia_dir.exists():
            print(f"⚠️ Ya existe carpeta de cierre diario. Se regenerará: {cierre_dia_dir}")
            shutil.rmtree(cierre_dia_dir, ignore_errors=True)

        for d in (excel_dir, auditoria_dir, logs_dir, manifest_dir, validaciones_dir):
            d.mkdir(parents=True, exist_ok=True)

        excel_info = crear_excel_diario(excel_diario_path, list(headers), filas_diarias)
        auditorias_info = copiar_auditorias(audit_files, auditoria_dir, fecha)
        logs_info = copiar_logs(log_files, logs_dir)
        env_info = crear_snapshot_env_redactado(manifest_dir, fecha)

        validacion_excel = validar_excel_generado(excel_diario_path, expected_columns=len(headers))
        validacion = {
            "tipo": "validacion_local_cierre_diario",
            "version": VERSION_CIERRE,
            "fecha": fecha,
            "trimestre": trimestre,
            "generado_en": _dt.datetime.now().isoformat(timespec="seconds"),
            "excel": validacion_excel,
            "auditorias_detectadas": len(audit_files),
            "auditorias_copiadas": len(auditorias_info),
            "logs_detectados": len(log_files),
            "logs_copiados": len(logs_info),
            "ok": bool(validacion_excel.get("abre_ok")) and not validacion_excel.get("errores"),
            "advertencias": advertencias,
        }
        validacion_path.write_text(json.dumps(validacion, ensure_ascii=False, indent=2), encoding="utf-8")

        manifest = {
            "tipo": "cierre_diario_local_seguro_v3",
            "version": VERSION_CIERRE,
            "fecha": fecha,
            "trimestre": trimestre,
            "anio": anio,
            "mes": mes,
            "mes_nombre": mes_nombre,
            "semana_inicio": semana_inicio.isoformat(),
            "semana_fin": semana_fin.isoformat(),
            "semana_nombre": semana_nombre,
            "generado_en": _dt.datetime.now().isoformat(timespec="seconds"),
            "root": str(ROOT),
            "host": socket.gethostname(),
            "platform": platform.platform(),
            "python": sys.version.replace("\n", " "),
            "carpeta_cierre": rel(cierre_dia_dir),
            "excel_diario": excel_info,
            "seleccion_filas": meta_seleccion,
            "conteos": {
                "filas_diarias": len(filas_diarias),
                "auditorias": len(auditorias_info),
                "logs": len(logs_info),
                "config_redactada": 1 if env_info else 0,
            },
            "validacion_local": rel(validacion_path),
            "advertencias": advertencias,
            "nota": (
                "Cierre diario V3: evidencia solo del día, no histórico completo, "
                "almacenada dentro del trimestre operativo activo."
            ),
        }

        # Totales preliminares antes de generar el resumen.
        # Esto evita KeyError si el resumen consulta total_archivos/total_bytes.
        archivos_evidencia_pre_resumen = listar_archivos_para_manifest(cierre_dia_dir, excluir={manifest_path})
        manifest["archivos"] = archivos_evidencia_pre_resumen
        manifest["total_archivos"] = len(archivos_evidencia_pre_resumen)
        manifest["total_bytes"] = sum(
            int(x.get("bytes", 0) or 0)
            for x in archivos_evidencia_pre_resumen
        )
        generar_resumen_txt(resumen_path, manifest)

        # El manifest se escribe después del resumen/validación para registrar la evidencia final.
        archivos_evidencia = listar_archivos_para_manifest(cierre_dia_dir, excluir={manifest_path})
        manifest["archivos"] = archivos_evidencia
        manifest["total_archivos"] = len(archivos_evidencia)
        manifest["total_bytes"] = sum(int(x.get("bytes", 0) or 0) for x in archivos_evidencia)
        manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")

        print(f"✅ Excel diario: {excel_diario_path}")
        print(f"✅ Filas Excel diario: {excel_info['filas_datos']}")
        print(f"✅ Auditorías copiadas: {len(auditorias_info)}")
        print(f"✅ Logs copiados: {len(logs_info)}")
        print(f"✅ Manifest: {manifest_path}")
        print(f"✅ Resumen: {resumen_path}")
        print(f"✅ Validación local: {validacion_path}")
        print(f"✅ Total archivos evidencia: {manifest['total_archivos']}")
        print(f"✅ Total bytes evidencia: {manifest['total_bytes']}")

        if advertencias:
            print("⚠️ Advertencias:")
            for adv in advertencias:
                print(f"  - {adv}")

        if not args.permitir_vacio and len(filas_diarias) == 0:
            print("⚠️ Cierre generado sin filas diarias. Se devuelve código 3 para revisión controlada.")
            print("   Si el día realmente no tuvo facturas nuevas, usa --permitir-vacio.")
            return 3

        print("=" * 100)
        print("✅ Cierre diario local V3 generado correctamente.")
        print("Siguiente paso: subir/validar en OneDrive con el uploader trimestral-compatible.")
        print("=" * 100)
        return 0

    except Exception as exc:
        print(f"❌ Error generando cierre diario local V3: {type(exc).__name__}: {exc}")
        print("⚠️ No subas ni archives nada hasta revisar el error.")
        return 1
    finally:
        release_lock()


if __name__ == "__main__":
    raise SystemExit(main())
