# -*- coding: utf-8 -*-
"""
JOYCO - Facturas Procesador
Subida remota de cierre mensual V4 a repositorio unico de backups.

Politica:
- El cierre mensual NO se sube al SharePoint principal de operacion.
- Se sube unicamente al repositorio remoto de backups configurado por .env:
  BACKUP_DRIVE_ID / ONEDRIVE_BACKUP_DRIVE_ID / SP_BACKUP2_DRIVE_ID
  BACKUP_ROOT_FOLDER / ONEDRIVE_BACKUP_FOLDER / SP_BACKUP2_FOLDER
- La verificacion remota descarga cada archivo subido y compara:
  - Excel: datos internos hoja/celda.
  - Los demas archivos: SHA256 exacto.
- Reintenta errores temporales de Graph sobre la misma ruta remota.
- Permite reanudar una ejecucion incompleta y procesar solo los archivos fallidos.
- Registra los 20 dias de retencion local solo despues de validar toda la subida.
- Conserva obligatoriamente la jerarquia Trimestre/Mes/Mensual.
- Excluye estructuras obsoletas que duplicaban semanas o empaquetaban el cierre en ZIP.
- Bloquea rutas antiguas o cierres que no pertenezcan por completo al
  trimestre operativo activo.

Estructura remota esperada:
  BACKUP_ROOT_FOLDER/YYYY/TRIMESTRE_YYYY-MM-DD_A_YYYY-MM-DD/
    YYYY-MM_MesNombre/Mensual/
"""

from __future__ import annotations

import argparse
import datetime as _dt
import hashlib
import json
import os
import shutil
import sys
import time
from pathlib import Path
from typing import Any, Optional, Tuple
from urllib.parse import quote
from zoneinfo import ZoneInfo

try:
    import truststore
except ImportError:
    truststore = None
else:
    truststore.inject_into_ssl()

import requests
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

try:
    import config  # noqa: F401
except Exception:
    pass

from services.m365.token import get_access_token
from trimestre_activo import cargar_trimestre_activo

VERSION_UPLOAD = "2026-08-05-UPLOAD-CIERRE-MENSUAL-V4-SIN-DUPLICAR-SEMANAS-NI-ZIP"
GRAPH = "https://graph.microsoft.com/v1.0"

HTTP_REINTENTABLES_SUBIDA = {408, 429, 500, 502, 503, 504}
INTENTOS_SUBIDA = 4
ESPERA_BASE_SUBIDA_SEGUNDOS = 2
ESPERA_MAX_SUBIDA_SEGUNDOS = 60
RETENCION_LOCAL_DIAS = 20
ZONA_HORARIA = ZoneInfo("America/Bogota")

DATA_DIR = ROOT / "data"
CIERRES_DIR = DATA_DIR / "cierres_diarios"
TMP_VERIFY_DIR = DATA_DIR / "_tmp_verificacion_cierre_mensual_v3"

EXCLUIR_NOMBRES = {".env", ".env.local", ".env.production"}
EXCLUIR_EXT = {".tmp", ".lock"}
EXCEL_EXT = {".xlsx", ".xlsm"}
DIRECTORIOS_OBSOLETOS = {"03_Soporte_Semanas", "07_Paquete_Mensual"}

MESES_ES = {
    1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril", 5: "Mayo", 6: "Junio",
    7: "Julio", 8: "Agosto", 9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre",
}


def resumen_identificador(value: Any) -> str:
    texto = str(value or "").strip()
    if not texto:
        return "(vacio)"
    visible = texto[-6:] if len(texto) > 6 else texto
    return f"***{visible} (longitud={len(texto)})"


def ssl_verify() -> bool:
    return str(os.getenv("SSL_VERIFY", "true")).strip().lower() not in {"0", "false", "no", "off"}


def env_first(*names: str, default: str = "") -> str:
    for name in names:
        value = (os.getenv(name) or "").strip()
        if value:
            return value
    return default


def encode_path(path: str) -> str:
    return quote(str(path).strip("/"), safe="/")


def encode_drive_id(drive_id: str) -> str:
    return quote(str(drive_id), safe="!")


def headers() -> dict[str, str]:
    token = get_access_token()
    return {"Authorization": f"Bearer {token}"}


def h_json() -> dict[str, str]:
    h = headers()
    h["Content-Type"] = "application/json"
    return h


def graph_get(
    url: str, *, ok: Tuple[int, ...] = (200,), timeout: int = 60
) -> requests.Response:
    r = requests.get(url, headers=headers(), timeout=timeout, verify=ssl_verify())
    if r.status_code not in ok:
        raise RuntimeError(f"GET {r.status_code} {url} -> {r.text[:700]}")
    return r


def espera_reintento_subida(
    respuesta: Optional[requests.Response], intento: int
) -> int:
    if respuesta is not None:
        retry_after = str(respuesta.headers.get("Retry-After", "") or "").strip()
        if retry_after:
            try:
                return max(1, min(int(float(retry_after)), ESPERA_MAX_SUBIDA_SEGUNDOS))
            except (TypeError, ValueError):
                pass
    espera = ESPERA_BASE_SUBIDA_SEGUNDOS * (2 ** max(0, intento - 1))
    return min(espera, ESPERA_MAX_SUBIDA_SEGUNDOS)


def graph_put_content(drive_id: str, remote_path: str, local_file: Path) -> dict[str, Any]:
    url = f"{GRAPH}/drives/{encode_drive_id(drive_id)}/root:/{encode_path(remote_path)}:/content"
    data = local_file.read_bytes()
    ultimo_error = ""
    for intento in range(1, INTENTOS_SUBIDA + 1):
        respuesta: Optional[requests.Response] = None
        try:
            respuesta = requests.put(
                url, headers=headers(), data=data, timeout=300, verify=ssl_verify()
            )
        except (requests.Timeout, requests.ConnectionError) as exc:
            ultimo_error = (
                f"PUT interrumpido para {remote_path} en intento "
                f"{intento}/{INTENTOS_SUBIDA}: {type(exc).__name__}: {exc}"
            )
            if intento >= INTENTOS_SUBIDA:
                raise RuntimeError(ultimo_error) from exc
            espera = espera_reintento_subida(None, intento)
            print(ultimo_error)
            print(f"Reintentando subida en {espera}s sobre la misma ruta remota...")
            time.sleep(espera)
            continue
        except requests.RequestException as exc:
            raise RuntimeError(
                f"PUT no reintentable para {remote_path}: {type(exc).__name__}: {exc}"
            ) from exc

        if respuesta.status_code in (200, 201):
            item = respuesta.json()
            item["_joyco_intento_subida"] = intento
            item["_joyco_codigo_http_subida"] = respuesta.status_code
            if intento > 1:
                print(
                    f"Subida recuperada en intento {intento}/{INTENTOS_SUBIDA}: "
                    f"{remote_path}"
                )
            return item

        ultimo_error = (
            f"PUT {respuesta.status_code} para {remote_path} -> {respuesta.text[:700]}"
        )
        if respuesta.status_code not in HTTP_REINTENTABLES_SUBIDA:
            raise RuntimeError(ultimo_error)
        if intento >= INTENTOS_SUBIDA:
            raise RuntimeError(
                f"{ultimo_error} | reintentos agotados: {INTENTOS_SUBIDA}"
            )
        espera = espera_reintento_subida(respuesta, intento)
        print(
            f"Subida intento {intento}/{INTENTOS_SUBIDA} fallo temporalmente "
            f"con HTTP {respuesta.status_code}: {remote_path}"
        )
        print(f"Reintentando subida en {espera}s sobre la misma ruta remota...")
        time.sleep(espera)
    raise RuntimeError(ultimo_error or f"PUT fallido para {remote_path}")


def graph_download_item_content(drive_id: str, item_id: str) -> bytes:
    url = f"{GRAPH}/drives/{encode_drive_id(drive_id)}/items/{quote(item_id, safe='')}/content"
    r = requests.get(
        url, headers=headers(), timeout=300, verify=ssl_verify(), allow_redirects=True
    )
    if r.status_code != 200:
        raise RuntimeError(f"DOWNLOAD {r.status_code} {url} -> {r.text[:700]}")
    return r.content


def validar_drive_id(drive_id: str) -> Optional[dict[str, Any]]:
    if not drive_id:
        return None
    url = f"{GRAPH}/drives/{encode_drive_id(drive_id)}?$select=id,name,webUrl,driveType"
    r = requests.get(url, headers=headers(), timeout=60, verify=ssl_verify())
    if r.status_code == 200:
        return r.json()
    print(
        "Drive ID no valido o no accesible: "
        f"{resumen_identificador(drive_id)}"
    )
    print(f"Respuesta Graph: {r.status_code} {r.text[:350]}")
    return None


def listar_drives_site(hostname: str, site_path: str) -> list[dict[str, Any]]:
    if not hostname or not site_path:
        return []
    site_path = site_path if site_path.startswith("/") else f"/{site_path}"
    url = f"{GRAPH}/sites/{hostname}:{site_path}:/drives?$select=id,name,webUrl,driveType"
    return graph_get(url, timeout=60).json().get("value", [])


def resolver_drive_backup() -> tuple[str, dict[str, Any]]:
    drive_id = env_first(
        "BACKUP_DRIVE_ID", "ONEDRIVE_BACKUP_DRIVE_ID", "SP_BACKUP2_DRIVE_ID"
    )
    drive = validar_drive_id(drive_id)
    if drive:
        print(
            "Drive backup validado: "
            f"{drive.get('name')} | "
            f"{resumen_identificador(drive.get('id'))}"
        )
        return str(drive["id"]), drive

    hostname = env_first(
        "BACKUP_HOSTNAME", "ONEDRIVE_BACKUP_HOSTNAME", "SP_BACKUP2_HOSTNAME"
    )
    site_path = env_first(
        "BACKUP_SITE_PATH", "ONEDRIVE_BACKUP_SITE_PATH", "SP_BACKUP2_SITE_PATH"
    )
    if hostname and site_path:
        drives = listar_drives_site(hostname, site_path)
        if not drives:
            raise RuntimeError("No se encontraron drives para el repositorio de backups.")
        preferidos = {"documentos", "documents", "shared documents", "onedrive"}
        elegido = next(
            (d for d in drives if str(d.get("name", "")).strip().lower() in preferidos),
            drives[0],
        )
        print(
            "Drive backup resuelto: "
            f"{elegido.get('name')} | "
            f"{resumen_identificador(elegido.get('id'))}"
        )
        return str(elegido["id"]), elegido
    raise RuntimeError(
        "No hay drive de backups configurado. Define BACKUP_DRIVE_ID, "
        "ONEDRIVE_BACKUP_DRIVE_ID o SP_BACKUP2_DRIVE_ID en .env."
    )


def existe_path(drive_id: str, remote_path: str) -> bool:
    remote_path = remote_path.strip("/")
    if not remote_path:
        return True
    url = f"{GRAPH}/drives/{encode_drive_id(drive_id)}/root:/{encode_path(remote_path)}:"
    ultimo_error = ""
    for intento in range(1, INTENTOS_SUBIDA + 1):
        respuesta: Optional[requests.Response] = None
        try:
            respuesta = requests.get(
                url, headers=headers(), timeout=60, verify=ssl_verify()
            )
        except (requests.Timeout, requests.ConnectionError) as exc:
            ultimo_error = (
                f"GET interrumpido para {remote_path} en intento "
                f"{intento}/{INTENTOS_SUBIDA}: {type(exc).__name__}: {exc}"
            )
            if intento >= INTENTOS_SUBIDA:
                raise RuntimeError(ultimo_error) from exc
            espera = espera_reintento_subida(None, intento)
            print(ultimo_error)
            print(f"Reintentando consulta de ruta en {espera}s...")
            time.sleep(espera)
            continue

        if respuesta.status_code == 200:
            return True
        if respuesta.status_code == 404:
            return False

        ultimo_error = (
            f"GET {respuesta.status_code} para {remote_path} -> "
            f"{respuesta.text[:700]}"
        )
        if respuesta.status_code not in HTTP_REINTENTABLES_SUBIDA:
            raise RuntimeError(ultimo_error)
        if intento >= INTENTOS_SUBIDA:
            raise RuntimeError(
                f"{ultimo_error} | reintentos agotados: {INTENTOS_SUBIDA}"
            )
        espera = espera_reintento_subida(respuesta, intento)
        print(
            f"Consulta de ruta intento {intento}/{INTENTOS_SUBIDA} "
            f"fallo temporalmente con HTTP {respuesta.status_code}: {remote_path}"
        )
        print(f"Reintentando consulta de ruta en {espera}s...")
        time.sleep(espera)
    raise RuntimeError(ultimo_error or f"GET fallido para {remote_path}")


def crear_folder(drive_id: str, parent_path: str, folder_name: str) -> None:
    parent_path = parent_path.strip("/")
    if parent_path:
        url = f"{GRAPH}/drives/{encode_drive_id(drive_id)}/root:/{encode_path(parent_path)}:/children"
    else:
        url = f"{GRAPH}/drives/{encode_drive_id(drive_id)}/root/children"
    body = {"name": folder_name, "folder": {}, "@microsoft.graph.conflictBehavior": "fail"}
    r = requests.post(url, headers=h_json(), json=body, timeout=60, verify=ssl_verify())
    if r.status_code not in (200, 201, 409):
        raise RuntimeError(f"POST {r.status_code} {url} -> {r.text[:700]}")


def ensure_folder_recursive(drive_id: str, folder_path: str) -> None:
    parts = [p for p in folder_path.strip("/").split("/") if p]
    current = ""
    for part in parts:
        parent = current
        current = f"{current}/{part}".strip("/")
        if not existe_path(drive_id, current):
            crear_folder(drive_id, parent, part)


def sha256_bytes(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def normalizar_valor_excel(value: Any) -> str:
    if value is None:
        return "<NULL>"
    if isinstance(value, (_dt.datetime, _dt.date, _dt.time)):
        return value.isoformat()
    if isinstance(value, float):
        return repr(value)
    return f"{type(value).__name__}:{value}"


def excel_digest(path: Path) -> dict[str, Any]:
    wb = load_workbook(path, read_only=True, data_only=False, keep_links=False)
    try:
        digest = hashlib.sha256()
        hojas: list[dict[str, Any]] = []
        celdas = 0
        digest.update(json.dumps(list(wb.sheetnames), ensure_ascii=False).encode("utf-8"))
        for ws in wb.worksheets:
            info = {
                "title": ws.title,
                "max_row": int(ws.max_row or 0),
                "max_column": int(ws.max_column or 0),
                "non_empty_cells": 0,
            }
            digest.update(
                f"\n[SHEET]{ws.title}|{ws.max_row}|{ws.max_column}".encode("utf-8")
            )
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is None:
                        continue
                    info["non_empty_cells"] += 1
                    celdas += 1
                    payload = (
                        f"{ws.title}|{cell.coordinate}|"
                        f"{normalizar_valor_excel(cell.value)}\n"
                    )
                    digest.update(payload.encode("utf-8", errors="replace"))
            hojas.append(info)
        return {
            "sha256_datos": digest.hexdigest(),
            "hojas": hojas,
            "celdas_no_vacias": celdas,
        }
    finally:
        wb.close()


def verificar_archivo_subido(
    *, local: Path, item: dict[str, Any], rel_path: str, drive_id: str
) -> tuple[bool, dict[str, Any]]:
    item_id = item.get("id")
    if not item_id:
        raise RuntimeError(f"Graph no devolvio item.id para {rel_path}")
    remote_bytes = graph_download_item_content(drive_id, str(item_id))
    base = {
        "archivo": rel_path,
        "local": str(local),
        "remote_item_id": item_id,
        "remote_web_url": item.get("webUrl"),
        "bytes_local": local.stat().st_size,
        "bytes_remoto": len(remote_bytes),
    }
    if local.suffix.lower() in EXCEL_EXT:
        TMP_VERIFY_DIR.mkdir(parents=True, exist_ok=True)
        tmp = TMP_VERIFY_DIR / rel_path.replace("/", "__").replace("\\", "__")
        tmp.write_bytes(remote_bytes)
        local_digest = excel_digest(local)
        remote_digest = excel_digest(tmp)
        ok = local_digest["sha256_datos"] == remote_digest["sha256_datos"]
        tmp.unlink(missing_ok=True)
        base.update({
            "tipo": "excel_datos_hoja_celda",
            "ok": ok,
            "local_digest": local_digest,
            "remoto_digest": remote_digest,
        })
        if ok:
            print(
                f"Excel verificado por DATOS: {rel_path} | "
                f"celdas_no_vacias={remote_digest['celdas_no_vacias']} | "
                f"hojas={len(remote_digest['hojas'])}"
            )
        return ok, base
    local_sha = sha256_file(local)
    remote_sha = sha256_bytes(remote_bytes)
    ok = local_sha == remote_sha
    base.update({
        "tipo": "sha256_binario_exacto",
        "ok": ok,
        "sha256_local": local_sha,
        "sha256_remoto": remote_sha,
    })
    if ok:
        print(f"Archivo verificado por SHA256 exacto: {rel_path} ({local.stat().st_size} bytes)")
    return ok, base


def subir_y_verificar_con_reintentos(
    *, drive_id: str, remote_path: str, local: Path, rel_path: str
) -> tuple[bool, dict[str, Any]]:
    item = graph_put_content(drive_id, remote_path, local)
    intento_subida = int(item.pop("_joyco_intento_subida", 1) or 1)
    codigo_http = int(item.pop("_joyco_codigo_http_subida", 0) or 0)
    ultimo: dict[str, Any] = {}
    for intento in range(1, 5):
        try:
            ok, resultado = verificar_archivo_subido(
                local=local, item=item, rel_path=rel_path, drive_id=drive_id
            )
            resultado.update({
                "intento_subida": intento_subida,
                "codigo_http_subida": codigo_http,
                "intento_verificacion": intento,
            })
            ultimo = resultado
            if ok:
                return True, resultado
        except Exception as exc:
            ultimo = {
                "archivo": rel_path,
                "ok": False,
                "error": f"{type(exc).__name__}: {exc}",
                "intento_subida": intento_subida,
                "codigo_http_subida": codigo_http,
                "intento_verificacion": intento,
            }
            print(f"Verificacion intento {intento}/4 fallo para {rel_path}: {exc}")
        if intento < 4:
            espera = 2 ** intento
            print(f"Reintentando verificacion en {espera}s...")
            time.sleep(espera)
    return False, ultimo


def parse_fecha(fecha: str) -> _dt.date:
    if fecha:
        return _dt.date.fromisoformat(fecha)
    hoy = _dt.datetime.now(ZONA_HORARIA).date()
    return hoy.replace(day=1) - _dt.timedelta(days=1)


def ultimo_dia_mes(fecha: _dt.date) -> _dt.date:
    if fecha.month == 12:
        siguiente = _dt.date(fecha.year + 1, 1, 1)
    else:
        siguiente = _dt.date(fecha.year, fecha.month + 1, 1)
    return siguiente - _dt.timedelta(days=1)


def rango_mes(fecha: _dt.date) -> tuple[_dt.date, _dt.date]:
    inicio = _dt.date(fecha.year, fecha.month, 1)
    return inicio, ultimo_dia_mes(fecha)


def rango_desde_args(args: argparse.Namespace) -> tuple[_dt.date, _dt.date]:
    if args.inicio or args.fin:
        if not args.inicio or not args.fin:
            raise RuntimeError("Si usas --inicio o --fin debes enviar ambos.")
        inicio = _dt.date.fromisoformat(args.inicio)
        fin = _dt.date.fromisoformat(args.fin)
        if fin < inicio:
            raise RuntimeError("El --fin no puede ser menor que --inicio.")
    else:
        inicio, fin = rango_mes(parse_fecha(args.fecha))
    if inicio.day != 1 or fin != ultimo_dia_mes(inicio):
        raise RuntimeError("La subida mensual exige un mes calendario completo.")
    if (inicio.year, inicio.month) != (fin.year, fin.month):
        raise RuntimeError("El inicio y el fin deben pertenecer al mismo mes.")
    primer_dia_actual = _dt.datetime.now(ZONA_HORARIA).date().replace(day=1)
    if fin >= primer_dia_actual:
        raise RuntimeError("No se permite subir el cierre de un mes abierto o futuro.")
    return inicio, fin


def mes_carpeta(fecha: _dt.date) -> str:
    return f"{fecha:%Y-%m}_{MESES_ES.get(fecha.month, fecha.strftime('%B'))}"


def periodo_nombre(inicio: _dt.date, fin: _dt.date) -> str:
    if inicio.day == 1 and fin == ultimo_dia_mes(inicio):
        return inicio.strftime("%Y-%m")
    return f"{inicio.isoformat()}_a_{fin.isoformat()}"


def buscar_cierre_mensual_local(
    inicio: _dt.date,
    fin: _dt.date,
    trimestre: dict[str, Any],
) -> Path:
    mes = mes_carpeta(inicio)
    esperado = (
        CIERRES_DIR
        / Path(trimestre["ruta_relativa"])
        / mes
        / "Mensual"
    )

    if esperado.exists() and esperado.is_dir():
        return esperado

    raise RuntimeError(
        "No se encontro el cierre mensual en la jerarquia trimestral oficial. "
        f"Periodo={periodo_nombre(inicio, fin)} | Ruta esperada={esperado}. "
        "No se usara automaticamente la estructura anterior para evitar "
        "crear backups fuera del trimestre activo."
    )


def debe_subir(path: Path) -> bool:
    if not path.is_file():
        return False
    if path.name in EXCLUIR_NOMBRES:
        return False
    if path.suffix.lower() in EXCLUIR_EXT:
        return False
    if any(part.startswith("_tmp") for part in path.parts):
        return False
    if any(part in DIRECTORIOS_OBSOLETOS for part in path.parts):
        return False
    if path.name.startswith("validacion_remota_mensual_"):
        return False
    if path.name.startswith("estado_retencion_mensual_"):
        return False
    return True


def listar_archivos(base: Path) -> list[Path]:
    return [p for p in sorted(base.rglob("*")) if debe_subir(p)]


def cargar_reintento_fallidos(
    *,
    validacion_path: Path,
    cierre_dir: Path,
    archivos: list[Path],
    periodo: str,
    remote_base: str,
    trimestre: dict[str, Any],
) -> tuple[list[Path], dict[str, dict[str, Any]]]:
    if not validacion_path.exists():
        raise RuntimeError(
            "No existe la validacion remota anterior necesaria para reintentar: "
            f"{validacion_path}"
        )
    try:
        validacion = json.loads(validacion_path.read_text(encoding="utf-8"))
    except Exception as exc:
        raise RuntimeError(
            f"No se pudo leer la validacion remota anterior: {validacion_path} | {exc}"
        ) from exc

    if str(validacion.get("periodo", "")) != periodo:
        raise RuntimeError("La validacion remota anterior corresponde a otro periodo.")
    if str(validacion.get("remote_base", "")).strip("/") != remote_base.strip("/"):
        raise RuntimeError("La validacion remota anterior corresponde a otro destino.")

    trimestre_anterior = validacion.get("trimestre")
    if not isinstance(trimestre_anterior, dict):
        raise RuntimeError(
            "La validacion remota anterior no registra el trimestre operativo."
        )
    if (
        str(trimestre_anterior.get("nombre_carpeta", "")).strip()
        != trimestre["nombre_carpeta"]
    ):
        raise RuntimeError(
            "La validacion remota anterior corresponde a otro trimestre."
        )
    if int(validacion.get("total_archivos_esperados", -1)) != len(archivos):
        raise RuntimeError(
            "La cantidad actual de archivos no coincide con la validacion remota anterior."
        )

    locales_por_ruta = {
        p.relative_to(cierre_dir).as_posix(): p
        for p in archivos
    }
    resultados_anteriores = validacion.get("resultados", [])
    if not isinstance(resultados_anteriores, list):
        raise RuntimeError("La validacion remota anterior no contiene resultados validos.")

    resultados_por_ruta: dict[str, dict[str, Any]] = {}
    for resultado in resultados_anteriores:
        if not isinstance(resultado, dict):
            raise RuntimeError("La validacion remota anterior contiene un resultado invalido.")
        rel_path = str(resultado.get("archivo", "")).strip()
        if not rel_path or rel_path in resultados_por_ruta:
            raise RuntimeError(
                "La validacion remota anterior contiene rutas vacias o duplicadas."
            )
        resultados_por_ruta[rel_path] = resultado

    if set(resultados_por_ruta) != set(locales_por_ruta):
        raise RuntimeError(
            "Las rutas actuales no coinciden con las registradas en la validacion anterior."
        )

    pendientes = [
        locales_por_ruta[rel_path]
        for rel_path, resultado in resultados_por_ruta.items()
        if resultado.get("ok") is not True
    ]
    exitos_previos = {
        rel_path: resultado
        for rel_path, resultado in resultados_por_ruta.items()
        if resultado.get("ok") is True
    }
    return pendientes, exitos_previos


def validar_cierre_local_minimo(
    cierre_dir: Path,
    inicio: _dt.date,
    fin: _dt.date,
    trimestre: dict[str, Any],
) -> None:
    periodo = periodo_nombre(inicio, fin)
    manifest_path = (
        cierre_dir
        / "04_Manifest_Mensual"
        / f"manifest_mensual_{periodo}.json"
    )
    validacion_path = (
        cierre_dir
        / "05_Validaciones"
        / f"validacion_local_mensual_{periodo}.json"
    )
    obligatorios = [
        cierre_dir
        / "01_Excel_Mensual"
        / f"facturas_mensual_{periodo}.xlsx",
        manifest_path,
        cierre_dir
        / "04_Manifest_Mensual"
        / f"resumen_mensual_{periodo}.txt",
        validacion_path,
    ]
    faltantes = [str(p) for p in obligatorios if not p.exists()]
    if faltantes:
        raise RuntimeError(
            "El cierre mensual local no esta completo. Faltan: "
            + "; ".join(faltantes)
        )

    try:
        validacion = json.loads(
            validacion_path.read_text(encoding="utf-8")
        )
    except Exception as exc:
        raise RuntimeError(
            f"No se pudo leer validacion local: {validacion_path} | {exc}"
        ) from exc

    if not validacion.get("ok"):
        raise RuntimeError(
            "La validacion local del cierre mensual no esta OK: "
            f"{validacion_path}"
        )

    trimestre_validacion = validacion.get("trimestre")
    if not isinstance(trimestre_validacion, dict):
        raise RuntimeError(
            "La validacion local mensual no registra el trimestre operativo."
        )
    if (
        str(trimestre_validacion.get("nombre_carpeta", "")).strip()
        != trimestre["nombre_carpeta"]
    ):
        raise RuntimeError(
            "La validacion local mensual corresponde a otro trimestre."
        )

    try:
        manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    except Exception as exc:
        raise RuntimeError(
            f"No se pudo leer manifest mensual: {manifest_path} | {exc}"
        ) from exc

    trimestre_manifest = manifest.get("trimestre")
    if not isinstance(trimestre_manifest, dict):
        raise RuntimeError(
            "El manifest mensual no registra el trimestre operativo."
        )
    if (
        str(trimestre_manifest.get("nombre_carpeta", "")).strip()
        != trimestre["nombre_carpeta"]
    ):
        raise RuntimeError(
            "El manifest mensual corresponde a otro trimestre."
        )

    for item in manifest.get("archivos", []):
        rel_path = str(item.get("ruta_relativa", ""))
        archivo = cierre_dir / rel_path
        if not rel_path or not archivo.exists() or not archivo.is_file():
            raise RuntimeError(
                f"Archivo del manifest no existe: {rel_path}"
            )
        if archivo.stat().st_size != int(item.get("bytes", -1)):
            raise RuntimeError(
                f"Tamano distinto al manifest: {rel_path}"
            )
        if sha256_file(archivo) != item.get("sha256"):
            raise RuntimeError(
                f"SHA256 distinto al manifest: {rel_path}"
            )


def remote_base_para_cierre(
    cierre_dir: Path,
    inicio: _dt.date,
    fin: _dt.date,
    trimestre: dict[str, Any],
) -> str:
    root_folder = env_first(
        "BACKUP_ROOT_FOLDER",
        "ONEDRIVE_BACKUP_FOLDER",
        "SP_BACKUP2_FOLDER",
    )
    if not root_folder:
        raise RuntimeError(
            "Falta BACKUP_ROOT_FOLDER/ONEDRIVE_BACKUP_FOLDER/"
            "SP_BACKUP2_FOLDER en .env."
        )

    esperado = (
        CIERRES_DIR
        / Path(trimestre["ruta_relativa"])
        / mes_carpeta(inicio)
        / "Mensual"
    )

    try:
        cierre_resuelto = cierre_dir.resolve()
        esperado_resuelto = esperado.resolve()
        rel_cierre = cierre_resuelto.relative_to(
            CIERRES_DIR.resolve()
        ).as_posix()
    except ValueError as exc:
        raise RuntimeError(
            "La carpeta del cierre mensual esta fuera de CIERRES_DIR y no "
            f"puede convertirse en una ruta remota segura: {cierre_dir}"
        ) from exc

    if cierre_resuelto != esperado_resuelto:
        raise RuntimeError(
            "La carpeta mensual no coincide con la jerarquia trimestral "
            f"oficial. Detectada={cierre_dir} | Esperada={esperado}"
        )

    if not (
        trimestre["fecha_inicio"] <= inicio.isoformat()
        and fin.isoformat() <= trimestre["fecha_fin"]
    ):
        raise RuntimeError(
            "El mes solicitado no pertenece completamente al trimestre activo."
        )

    return f"{root_folder}/{rel_cierre}".strip("/")


def escribir_validacion_remota(
    cierre_dir: Path,
    inicio: _dt.date,
    fin: _dt.date,
    trimestre: dict[str, Any],
    drive: dict[str, Any],
    remote_base: str,
    resultados: list[dict[str, Any]],
    total_esperados: int,
) -> Path:
    validaciones_dir = cierre_dir / "05_Validaciones"
    validaciones_dir.mkdir(parents=True, exist_ok=True)
    periodo = periodo_nombre(inicio, fin)
    path = (
        validaciones_dir
        / f"validacion_remota_mensual_{periodo}.json"
    )
    payload = {
        "tipo": "validacion_remota_cierre_mensual",
        "version": VERSION_UPLOAD,
        "trimestre": {
            "periodo_activo": trimestre["periodo_activo"],
            "nombre_carpeta": trimestre["nombre_carpeta"],
            "fecha_inicio": trimestre["fecha_inicio"],
            "fecha_fin": trimestre["fecha_fin"],
            "ruta_relativa": trimestre["ruta_relativa"],
            "state_path": trimestre["state_path"],
        },
        "periodo": periodo,
        "fecha_inicio": inicio.isoformat(),
        "fecha_fin": fin.isoformat(),
        "generado_en": _dt.datetime.now(
            ZONA_HORARIA
        ).isoformat(timespec="seconds"),
        "cierre_local": str(cierre_dir),
        "drive": {
            "id": drive.get("id"),
            "name": drive.get("name"),
            "webUrl": drive.get("webUrl"),
            "driveType": drive.get("driveType"),
        },
        "remote_base": remote_base,
        "total_archivos_esperados": total_esperados,
        "total_resultados": len(resultados),
        "total_archivos_verificados": sum(
            1 for x in resultados if x.get("ok")
        ),
        "total_archivos_fallidos": sum(
            1 for x in resultados if not x.get("ok")
        ),
        "ok": (
            len(resultados) == total_esperados
            and total_esperados > 0
            and all(x.get("ok") for x in resultados)
        ),
        "resultados": resultados,
    }
    path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return path


def crear_estado_retencion_temporal(
    inicio: _dt.date,
    fin: _dt.date,
    remote_base: str,
    trimestre: dict[str, Any],
) -> tuple[Path, dict[str, Any]]:
    ahora = _dt.datetime.now(ZONA_HORARIA)
    periodo = periodo_nombre(inicio, fin)
    payload = {
        "tipo": "estado_retencion_local_cierre_mensual",
        "version": VERSION_UPLOAD,
        "trimestre": {
            "periodo_activo": trimestre["periodo_activo"],
            "nombre_carpeta": trimestre["nombre_carpeta"],
            "fecha_inicio": trimestre["fecha_inicio"],
            "fecha_fin": trimestre["fecha_fin"],
            "ruta_relativa": trimestre["ruta_relativa"],
        },
        "periodo": periodo,
        "validacion_remota_publicada_y_verificada": True,
        "validado_en": ahora.isoformat(timespec="seconds"),
        "retencion_local_dias": RETENCION_LOCAL_DIAS,
        "eliminacion_local_permitida_desde": (
            ahora + _dt.timedelta(days=RETENCION_LOCAL_DIAS)
        ).isoformat(timespec="seconds"),
        "remote_base": remote_base,
        "ok": True,
    }
    TMP_VERIFY_DIR.mkdir(parents=True, exist_ok=True)
    path = (
        TMP_VERIFY_DIR
        / f"estado_retencion_mensual_{periodo}.json"
    )
    path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return path, payload


def main() -> int:
    parser = argparse.ArgumentParser(description="Sube cierre mensual V4 al repositorio remoto unico de backups.")
    parser.add_argument("--fecha", default="", help="Fecha dentro del mes cerrado. Default: mes anterior.")
    parser.add_argument("--inicio", default="", help="Primer dia del mes, formato YYYY-MM-DD.")
    parser.add_argument("--fin", default="", help="Ultimo dia del mismo mes, formato YYYY-MM-DD.")
    parser.add_argument("--dry-run", action="store_true", help="Muestra archivos/ruta sin subir.")
    parser.add_argument(
        "--reintentar-fallidos",
        action="store_true",
        help=(
            "Lee la validacion remota anterior, conserva los resultados correctos "
            "y procesa unicamente los archivos fallidos."
        ),
    )
    args = parser.parse_args()

    try:
        inicio, fin = rango_desde_args(args)
    except Exception as exc:
        print(f"Periodo mensual no valido: {exc}")
        return 2
    periodo = periodo_nombre(inicio, fin)

    try:
        trimestre_inicio = cargar_trimestre_activo(ROOT, inicio)
        trimestre_fin = cargar_trimestre_activo(ROOT, fin)
    except Exception as exc:
        print(f"Trimestre operativo no valido para el mes: {exc}")
        return 2

    if (
        trimestre_inicio["nombre_carpeta"]
        != trimestre_fin["nombre_carpeta"]
    ):
        print(
            "Trimestre operativo no valido para el mes: el inicio y el fin "
            "quedaron en trimestres distintos."
        )
        return 2

    trimestre = trimestre_inicio

    print("=" * 100)
    print("SUBIDA CIERRE MENSUAL V4 A REPOSITORIO DE BACKUPS - FACTURAS JOYCO")
    print("=" * 100)
    print(f"Version: {VERSION_UPLOAD}")
    print(f"Root: {ROOT}")
    print(f"Periodo: {periodo}")
    print(f"Rango: {inicio.isoformat()} a {fin.isoformat()}")
    print(f"Trimestre activo: {trimestre['nombre_carpeta']}")
    print(
        "Rango trimestre: "
        f"{trimestre['fecha_inicio']} a {trimestre['fecha_fin']}"
    )
    print(f"Dry-run: {args.dry_run}")
    print("-" * 100)

    try:
        cierre_dir = buscar_cierre_mensual_local(
            inicio,
            fin,
            trimestre,
        )
        validar_cierre_local_minimo(
            cierre_dir,
            inicio,
            fin,
            trimestre,
        )
    except Exception as exc:
        print(f"Cierre mensual local no valido: {exc}")
        return 1

    try:
        remote_base = remote_base_para_cierre(
            cierre_dir,
            inicio,
            fin,
            trimestre,
        )
    except Exception as exc:
        print(f"No se pudo calcular ruta remota: {exc}")
        return 1

    todos_los_archivos = listar_archivos(cierre_dir)
    if not todos_los_archivos:
        print("No hay archivos validos para subir.")
        return 1
    print(f"Cierre local: {cierre_dir}")
    print(f"Ruta remota base: {remote_base}")
    print(f"Archivos detectados: {len(todos_los_archivos)}")
    for p in todos_los_archivos:
        print(f"   - {p.relative_to(cierre_dir).as_posix()} ({p.stat().st_size} bytes)")

    validacion_anterior = (
        cierre_dir
        / "05_Validaciones"
        / f"validacion_remota_mensual_{periodo}.json"
    )
    archivos = todos_los_archivos
    resultados_por_ruta: dict[str, dict[str, Any]] = {}
    if args.reintentar_fallidos:
        try:
            archivos, resultados_por_ruta = cargar_reintento_fallidos(
                validacion_path=validacion_anterior,
                cierre_dir=cierre_dir,
                archivos=todos_los_archivos,
                periodo=periodo,
                remote_base=remote_base,
                trimestre=trimestre,
            )
        except Exception as exc:
            print(f"No se puede reintentar de forma segura: {exc}")
            return 1
        print("Modo: reintentar unicamente archivos fallidos.")
        print(f"Resultados correctos conservados: {len(resultados_por_ruta)}")
        print(f"Archivos pendientes de reintento: {len(archivos)}")
        for p in archivos:
            print(f"   * {p.relative_to(cierre_dir).as_posix()}")

    if args.dry_run:
        print("-" * 100)
        print("DRY-RUN: no se llamo a Microsoft Graph ni se subio nada.")
        print("Si la ruta remota base es correcta, ejecuta sin --dry-run.")
        print("=" * 100)
        return 0

    try:
        drive_id, drive = resolver_drive_backup()
        print(f"Repositorio destino: {drive.get('name')} | {drive.get('webUrl')}")
        ensure_folder_recursive(drive_id, remote_base)
        print("Carpeta remota base verificada/creada.")
    except Exception as exc:
        print(f"Error preparando repositorio remoto de backups: {exc}")
        return 1

    ok_todos = True
    estado_local = (
        cierre_dir / "05_Validaciones" / f"estado_retencion_mensual_{periodo}.json"
    )
    estado_local.unlink(missing_ok=True)
    try:
        if TMP_VERIFY_DIR.exists():
            shutil.rmtree(TMP_VERIFY_DIR, ignore_errors=True)
        TMP_VERIFY_DIR.mkdir(parents=True, exist_ok=True)
    except Exception:
        pass

    for local in archivos:
        rel_path = local.relative_to(cierre_dir).as_posix()
        remote_path = f"{remote_base}/{rel_path}".strip("/")
        try:
            remote_dir = "/".join(remote_path.split("/")[:-1])
            ensure_folder_recursive(drive_id, remote_dir)
            print("Subiendo:")
            print(f"   Local:  {local}")
            print(f"   Remoto: {remote_path}")
            ok, resultado = subir_y_verificar_con_reintentos(
                drive_id=drive_id,
                remote_path=remote_path,
                local=local,
                rel_path=rel_path,
            )
            resultado["remote_path"] = remote_path
            resultados_por_ruta[rel_path] = resultado
            if not ok:
                ok_todos = False
        except Exception as exc:
            ok_todos = False
            print(f"ERROR subiendo/verificando {rel_path}: {type(exc).__name__}: {exc}")
            resultados_por_ruta[rel_path] = {
                "archivo": rel_path,
                "remote_path": remote_path,
                "ok": False,
                "error": f"{type(exc).__name__}: {exc}",
            }

    rutas_esperadas = [
        p.relative_to(cierre_dir).as_posix()
        for p in todos_los_archivos
    ]
    resultados = [
        resultados_por_ruta[rel_path]
        for rel_path in rutas_esperadas
        if rel_path in resultados_por_ruta
    ]
    if len(resultados) != len(todos_los_archivos):
        ok_todos = False
        print(
            "ERROR: no se obtuvo un resultado para cada archivo del cierre "
            f"({len(resultados)}/{len(todos_los_archivos)})."
        )

    validacion_remota = escribir_validacion_remota(
        cierre_dir,
        inicio,
        fin,
        trimestre,
        drive,
        remote_base,
        resultados,
        len(todos_los_archivos),
    )
    payload_validacion = json.loads(validacion_remota.read_text(encoding="utf-8"))
    ok_todos = ok_todos and bool(payload_validacion.get("ok"))
    rel_val = validacion_remota.relative_to(cierre_dir).as_posix()
    remote_val = f"{remote_base}/{rel_val}".strip("/")
    validacion_publicada = False
    try:
        print("Subiendo validacion remota final:")
        print(f"   Local:  {validacion_remota}")
        print(f"   Remoto: {remote_val}")
        validacion_publicada, _resultado_validacion = subir_y_verificar_con_reintentos(
            drive_id=drive_id,
            remote_path=remote_val,
            local=validacion_remota,
            rel_path=rel_val,
        )
        if not validacion_publicada:
            ok_todos = False
    except Exception as exc:
        ok_todos = False
        print(f"ERROR subiendo/verificando validacion remota final: {type(exc).__name__}: {exc}")

    if ok_todos and validacion_publicada:
        try:
            estado_tmp, estado_payload = crear_estado_retencion_temporal(
                inicio,
                fin,
                remote_base,
                trimestre,
            )
            rel_estado = f"05_Validaciones/{estado_tmp.name}"
            remote_estado = f"{remote_base}/{rel_estado}".strip("/")
            ok_estado, _resultado_estado = subir_y_verificar_con_reintentos(
                drive_id=drive_id,
                remote_path=remote_estado,
                local=estado_tmp,
                rel_path=rel_estado,
            )
            if not ok_estado:
                raise RuntimeError("No se pudo verificar el estado remoto de retencion.")
            estado_local.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(estado_tmp, estado_local)
            print(
                "Retencion local habilitada hasta: "
                f"{estado_payload['eliminacion_local_permitida_desde']}"
            )
        except Exception as exc:
            ok_todos = False
            estado_local.unlink(missing_ok=True)
            print(f"ERROR registrando retencion mensual: {type(exc).__name__}: {exc}")

    shutil.rmtree(TMP_VERIFY_DIR, ignore_errors=True)
    print("-" * 100)
    print(f"Validacion remota local: {validacion_remota}")
    if ok_todos:
        print("Subida de cierre mensual V4 terminada correctamente en el repositorio unico de backups.")
        print("Verificacion aplicada archivo por archivo despues de descargar desde Graph.")
        print(f"Retencion local: {RETENCION_LOCAL_DIAS} dias desde la validacion remota.")
        print("=" * 100)
        return 0
    print("Subida de cierre mensual V4 termino con errores.")
    print("No borres ni archives localmente hasta revisar el error.")
    print("=" * 100)
    return 1


if __name__ == "__main__":
    raise SystemExit(main())
