# -*- coding: utf-8 -*-
"""
JOYCO - Facturas Procesador
Preparación segura del cierre trimestral sin duplicar cierres inferiores.

Estructura oficial generada:

    data/cierres_diarios/<AÑO>/TRIMESTRE_<INICIO>_A_<FIN>/
        <MES>/...
        Trimestral/
            01_Excel_Cierre/
            02_Manifest/
            03_Validaciones/
            04_Resumen/

Reglas principales:
- Los cierres diarios, semanales y mensuales permanecen en sus carpetas.
- No se copian evidencias mensuales dentro de Trimestral.
- No se crea ZIP trimestral.
- Dentro del backup trimestral se conserva un solo Excel histórico.
- El Excel limpio candidato y el respaldo del estado se guardan fuera del
  backup, dentro de data/state/preparaciones_trimestrales/.
- Usa una ruta de staging corta compatible con Windows.
- data/facturas.xlsx y cierre_trimestral_state.json no se modifican durante
  esta fase.
- La preparación queda bloqueada para finalización hasta que el uploader
  trimestral publique y verifique exclusivamente la carpeta Trimestral.

Modos:
- --dry-run: valida e inventaría; no modifica archivos.
- --preparar --confirmar PREPARAR_CIERRE_TRIMESTRAL: crea la preparación.
- --real y --local: comandos antiguos bloqueados.
"""

from __future__ import annotations

import argparse
import calendar
import hashlib
import json
import os
import re
import shutil
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Dict, Optional

from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

from trimestre_activo import cargar_trimestre_activo

ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "data"
STATE_PATH = DATA_DIR / "state" / "cierre_trimestral_state.json"
FACTURAS_PATH = DATA_DIR / "facturas.xlsx"
CIERRES_DIR = DATA_DIR / "cierres_diarios"
PREPARACIONES_DIR = DATA_DIR / "state" / "preparaciones_trimestrales"

VERSION = "2026-08-06-CIERRE-TRIMESTRAL-V8-STAGING-CORTO-WINDOWS"
CONFIRMACION_PREPARAR = "PREPARAR_CIERRE_TRIMESTRAL"
ESTADO_PREPARADO = "PREPARADO_PENDIENTE_VALIDACION_REMOTA"

HEADERS_ESPERADOS = [
    "Radicado",
    "ProyectoProceso",
    "Archivo",
    "Empresa emisora",
    "CUFE",
    "Ciudad emisora",
    "Código ciudad",
    "NIT",
    "Cliente",
    "Número de factura",
    "Año",
    "Mes",
    "Día",
    "Tipo de contribuyente",
    "Actividad económica",
    "DESCRIPCIÓN",
    "Concepto",
    "VALOR",
    "Estado_calidad",
]


class ErrorCierreTrimestral(RuntimeError):
    """Error controlado del cierre trimestral."""


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as archivo:
        for chunk in iter(lambda: archivo.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def escribir_json_atomico(path: Path, datos: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporal = path.with_suffix(path.suffix + ".tmp")
    temporal.write_text(
        json.dumps(datos, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    os.replace(temporal, path)


def escribir_texto_atomico(path: Path, contenido: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporal = path.with_suffix(path.suffix + ".tmp")
    temporal.write_text(contenido, encoding="utf-8")
    os.replace(temporal, path)


def parse_fecha(valor: Any, campo: str) -> date:
    try:
        return date.fromisoformat(str(valor or "").strip())
    except Exception as exc:
        raise ErrorCierreTrimestral(
            f"Fecha inválida en {campo}: {valor!r}. Usa YYYY-MM-DD."
        ) from exc


def add_months(fecha: date, meses: int) -> date:
    total = fecha.month - 1 + meses
    anio = fecha.year + total // 12
    mes = total % 12 + 1
    dia = min(fecha.day, calendar.monthrange(anio, mes)[1])
    return date(anio, mes, dia)


def calcular_siguiente_periodo(fecha_fin_actual: date) -> Dict[str, str]:
    inicio = fecha_fin_actual + timedelta(days=1)
    fin = add_months(inicio, 3) - timedelta(days=1)
    return {
        "periodo_activo": (
            f"{inicio.year}-CICLO_{inicio:%Y%m%d}_A_{fin:%Y%m%d}"
        ),
        "fecha_inicio_periodo_activo": inicio.isoformat(),
        "proximo_cierre_estimado": fin.isoformat(),
    }


def nombre_seguro(valor: str) -> str:
    limpio = re.sub(r"[^A-Za-z0-9._-]+", "_", valor.strip()).strip("._-")
    if not limpio:
        raise ErrorCierreTrimestral(
            f"No fue posible construir un nombre seguro desde: {valor!r}"
        )
    return limpio


def cargar_estado() -> dict:
    if not STATE_PATH.exists() or not STATE_PATH.is_file():
        raise ErrorCierreTrimestral(
            f"No existe el estado trimestral: {STATE_PATH}"
        )

    try:
        estado = json.loads(STATE_PATH.read_text(encoding="utf-8-sig"))
    except Exception as exc:
        raise ErrorCierreTrimestral(
            f"No fue posible leer el estado trimestral: {exc}"
        ) from exc

    if not isinstance(estado, dict):
        raise ErrorCierreTrimestral(
            "El estado trimestral no contiene un objeto JSON."
        )

    requeridos = (
        "periodo_activo",
        "fecha_inicio_periodo_activo",
        "proximo_cierre_estimado",
        "estado",
    )
    faltantes = [c for c in requeridos if not str(estado.get(c) or "").strip()]
    if faltantes:
        raise ErrorCierreTrimestral(
            "Estado trimestral incompleto. Faltan: " + ", ".join(faltantes)
        )

    if str(estado["estado"]).strip().upper() != "ACTIVO":
        raise ErrorCierreTrimestral(
            f"El trimestre no está ACTIVO: {estado.get('estado')!r}"
        )

    inicio = parse_fecha(
        estado["fecha_inicio_periodo_activo"],
        "fecha_inicio_periodo_activo",
    )
    fin = parse_fecha(
        estado["proximo_cierre_estimado"],
        "proximo_cierre_estimado",
    )
    if inicio > fin:
        raise ErrorCierreTrimestral(
            "La fecha inicial del trimestre es posterior a la fecha final."
        )

    return estado


def diagnosticar_excel(path: Path = FACTURAS_PATH) -> dict:
    if not path.exists() or not path.is_file():
        raise ErrorCierreTrimestral(f"No existe el Excel: {path}")

    wb = load_workbook(
        path,
        read_only=False,
        data_only=False,
        keep_links=False,
    )
    try:
        if "Facturas" not in wb.sheetnames:
            raise ErrorCierreTrimestral(
                "El Excel no contiene la hoja Facturas."
            )
        ws = wb["Facturas"]
        headers = [cell.value for cell in ws[1]]
        if headers != HEADERS_ESPERADOS:
            raise ErrorCierreTrimestral(
                "Los encabezados del Excel no coinciden con la estructura esperada."
            )
        if int(ws.max_column or 0) != len(HEADERS_ESPERADOS):
            raise ErrorCierreTrimestral(
                f"Columnas inválidas: {ws.max_column}. Esperadas: 19."
            )
        if "TblFacturas" not in ws.tables:
            raise ErrorCierreTrimestral(
                "El Excel no contiene la tabla requerida TblFacturas."
            )
        tabla_ref = ws.tables["TblFacturas"].ref
        return {
            "archivo": str(path),
            "hojas": list(wb.sheetnames),
            "filas": int(ws.max_row or 0),
            "filas_datos": max(int(ws.max_row or 0) - 1, 0),
            "columnas": int(ws.max_column or 0),
            "tbl_facturas_ref": tabla_ref,
            "bytes": path.stat().st_size,
            "sha256": sha256_file(path),
        }
    finally:
        wb.close()


def crear_excel_limpio(origen: Path, destino: Path) -> None:
    wb = load_workbook(
        origen,
        read_only=False,
        data_only=False,
        keep_links=False,
    )
    try:
        if "Facturas" not in wb.sheetnames:
            raise ErrorCierreTrimestral(
                "No se puede crear el candidato limpio: falta Facturas."
            )
        ws = wb["Facturas"]
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)
        ws.tables.clear()
        tabla = Table(displayName="TblFacturas", ref="A1:S1")
        tabla.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(tabla)
        destino.parent.mkdir(parents=True, exist_ok=True)
        wb.save(destino)
    finally:
        wb.close()


def datos_periodo(estado: dict) -> dict:
    periodo_activo = str(estado["periodo_activo"]).strip()
    inicio = parse_fecha(
        estado["fecha_inicio_periodo_activo"],
        "fecha_inicio_periodo_activo",
    )
    fin = parse_fecha(
        estado["proximo_cierre_estimado"],
        "proximo_cierre_estimado",
    )

    trimestre_inicio = cargar_trimestre_activo(ROOT, inicio)
    trimestre_fin = cargar_trimestre_activo(ROOT, fin)
    if trimestre_inicio["ruta_relativa"] != trimestre_fin["ruta_relativa"]:
        raise ErrorCierreTrimestral(
            "El helper trimestral devolvió rutas diferentes para el mismo periodo."
        )

    base = CIERRES_DIR / Path(trimestre_fin["ruta_relativa"])
    carpeta_trimestral = base / "Trimestral"
    seguro = nombre_seguro(periodo_activo)
    nombre_excel = f"facturas_{inicio.isoformat()}_a_{fin.isoformat()}.xlsx"
    preparacion = PREPARACIONES_DIR / trimestre_fin["nombre_carpeta"]

    return {
        "periodo": periodo_activo,
        "periodo_seguro": seguro,
        "fecha_inicio": inicio.isoformat(),
        "fecha_fin": fin.isoformat(),
        "inicio_date": inicio,
        "fin_date": fin,
        "anio": trimestre_fin["anio"],
        "nombre_trimestre": trimestre_fin["nombre_carpeta"],
        "ruta_relativa_trimestre": trimestre_fin["ruta_relativa"],
        "carpeta_base": base,
        "carpeta_trimestral": carpeta_trimestral,
        "carpeta_excel": carpeta_trimestral / "01_Excel_Cierre",
        "carpeta_manifest": carpeta_trimestral / "02_Manifest",
        "carpeta_validaciones": carpeta_trimestral / "03_Validaciones",
        "carpeta_resumen": carpeta_trimestral / "04_Resumen",
        "nombre_excel": nombre_excel,
        "excel_historico": carpeta_trimestral / "01_Excel_Cierre" / nombre_excel,
        "carpeta_preparacion": preparacion,
        "excel_candidato": preparacion / (
            f"facturas_NUEVO_CANDIDATO_{seguro}.xlsx"
        ),
        "respaldo_estado": preparacion / (
            f"cierre_trimestral_state_ANTES_{seguro}.json"
        ),
        "state_path": Path(trimestre_fin["state_path"]),
    }


def meses_esperados(inicio: date, fin: date) -> list[str]:
    salida: list[str] = []
    actual = date(inicio.year, inicio.month, 1)
    limite = date(fin.year, fin.month, 1)
    while actual <= limite:
        salida.append(actual.strftime("%Y-%m"))
        actual = add_months(actual, 1)
    return salida


def inventariar_jerarquia_trimestral(periodo: dict) -> dict:
    base: Path = periodo["carpeta_base"]
    if not base.exists() or not base.is_dir():
        raise ErrorCierreTrimestral(
            f"No existe la carpeta oficial del trimestre: {base}"
        )

    esperados = meses_esperados(periodo["inicio_date"], periodo["fin_date"])
    digest = hashlib.sha256()
    meses: list[dict] = []
    archivos_inventario: list[dict] = []
    total_bytes = 0
    total_semanas = 0
    total_diarios = 0
    total_semanales = 0
    total_mensuales = 0

    for mes_dir in sorted(base.iterdir(), key=lambda p: p.name.casefold()):
        if not mes_dir.is_dir() or mes_dir.name == "Trimestral":
            continue
        if mes_dir.name.startswith(".tmp_"):
            continue

        match = re.match(r"^(\d{4}-\d{2})_", mes_dir.name)
        if not match or match.group(1) not in esperados:
            continue

        semanas = sorted(
            [
                p
                for p in mes_dir.iterdir()
                if p.is_dir() and p.name.startswith("SEMANA_")
            ],
            key=lambda p: p.name.casefold(),
        )
        diarios = [
            p
            for semana in semanas
            for p in semana.iterdir()
            if p.is_dir() and p.name.startswith("Diario_")
        ]
        semanales = [
            semana / "Semanal"
            for semana in semanas
            if (semana / "Semanal").is_dir()
        ]
        mensual = mes_dir / "Mensual"

        archivos_mes = sorted(
            [p for p in mes_dir.rglob("*") if p.is_file()],
            key=lambda p: p.as_posix().casefold(),
        )
        bytes_mes = 0
        for archivo in archivos_mes:
            if archivo.is_symlink():
                raise ErrorCierreTrimestral(
                    f"No se permiten enlaces simbólicos en la evidencia: {archivo}"
                )
            relativa = archivo.relative_to(base).as_posix()
            tamano = archivo.stat().st_size
            hash_archivo = sha256_file(archivo)
            digest.update(
                f"{relativa}|{tamano}|{hash_archivo}\n".encode("utf-8")
            )
            archivos_inventario.append(
                {
                    "ruta_relativa": relativa,
                    "bytes": tamano,
                    "sha256": hash_archivo,
                }
            )
            bytes_mes += tamano
            total_bytes += tamano

        meses.append(
            {
                "mes": match.group(1),
                "carpeta": mes_dir.name,
                "ruta": str(mes_dir),
                "semanas": len(semanas),
                "cierres_diarios": len(diarios),
                "cierres_semanales": len(semanales),
                "cierre_mensual_presente": mensual.is_dir(),
                "archivos": len(archivos_mes),
                "bytes": bytes_mes,
            }
        )
        total_semanas += len(semanas)
        total_diarios += len(diarios)
        total_semanales += len(semanales)
        total_mensuales += int(mensual.is_dir())

    if not meses:
        raise ErrorCierreTrimestral(
            "La carpeta trimestral no contiene carpetas mensuales del periodo."
        )

    presentes = [item["mes"] for item in meses]
    return {
        "tipo": "INVENTARIO_JERARQUIA_TRIMESTRAL",
        "version_script": VERSION,
        "generado_en": datetime.now().isoformat(timespec="seconds"),
        "carpeta_trimestre": str(base),
        "nombre_trimestre": periodo["nombre_trimestre"],
        "fecha_inicio": periodo["fecha_inicio"],
        "fecha_fin": periodo["fecha_fin"],
        "meses_esperados": esperados,
        "meses_presentes": presentes,
        "meses_sin_carpeta": [m for m in esperados if m not in presentes],
        "total_meses_presentes": len(meses),
        "total_semanas": total_semanas,
        "total_cierres_diarios": total_diarios,
        "total_cierres_semanales": total_semanales,
        "total_cierres_mensuales": total_mensuales,
        "total_archivos": len(archivos_inventario),
        "total_bytes": total_bytes,
        "sha256_contenido_jerarquia": digest.hexdigest(),
        "meses": meses,
        "archivos": archivos_inventario,
        "politica": {
            "copias_de_cierres_inferiores": False,
            "incluye_carpeta_trimestral": False,
            "verificacion": "ruta relativa + bytes + SHA256 por archivo",
        },
    }


def validar_fecha_preparacion(estado: dict, confirmar: Optional[str]) -> None:
    if confirmar != CONFIRMACION_PREPARAR:
        raise ErrorCierreTrimestral(
            "Preparación real bloqueada. Usa --preparar --confirmar "
            + CONFIRMACION_PREPARAR
        )
    fecha_fin = parse_fecha(
        estado["proximo_cierre_estimado"],
        "proximo_cierre_estimado",
    )
    if date.today() < fecha_fin:
        raise ErrorCierreTrimestral(
            f"No se permite preparar antes de {fecha_fin.isoformat()}."
        )


def estado_preparacion_path(periodo: dict) -> Path:
    return periodo["carpeta_validaciones"] / (
        "estado_preparacion_cierre_trimestral_"
        f"{periodo['periodo_seguro']}.json"
    )


def validar_preparacion_existente(
    periodo: dict,
    info_excel: dict,
    inventario: dict,
) -> Optional[dict]:
    paquete: Path = periodo["carpeta_trimestral"]
    preparacion: Path = periodo["carpeta_preparacion"]

    if not paquete.exists() and not preparacion.exists():
        return None
    if not paquete.is_dir() or not preparacion.is_dir():
        raise ErrorCierreTrimestral(
            "Existe una preparación trimestral incompleta. Revisión manual requerida."
        )

    estado_path = estado_preparacion_path(periodo)
    if not estado_path.exists():
        raise ErrorCierreTrimestral(
            f"No existe el estado de preparación esperado: {estado_path}"
        )

    datos = json.loads(estado_path.read_text(encoding="utf-8-sig"))
    if datos.get("estado") != ESTADO_PREPARADO:
        raise ErrorCierreTrimestral(
            f"Estado de preparación inesperado: {datos.get('estado')!r}"
        )
    if datos.get("excel_activo_sha256_al_preparar") != info_excel["sha256"]:
        raise ErrorCierreTrimestral(
            "data/facturas.xlsx cambió después de la preparación existente."
        )
    if (
        datos.get("jerarquia_sha256_al_preparar")
        != inventario["sha256_contenido_jerarquia"]
    ):
        raise ErrorCierreTrimestral(
            "La jerarquía de meses/semanas/diarios cambió después de preparar."
        )

    historico = Path(str(datos.get("excel_historico") or ""))
    candidato = Path(str(datos.get("excel_limpio_candidato") or ""))
    if not historico.exists() or sha256_file(historico) != info_excel["sha256"]:
        raise ErrorCierreTrimestral(
            "El Excel histórico existente no coincide con el Excel preparado."
        )
    if not candidato.exists():
        raise ErrorCierreTrimestral(
            f"No existe el candidato limpio registrado: {candidato}"
        )
    return datos


def construir_manifest(
    *,
    periodo: dict,
    estado: dict,
    info_excel: dict,
    info_candidato: dict,
    inventario: dict,
    inventario_path_final: Path,
    generado_en: str,
) -> dict:
    siguiente = calcular_siguiente_periodo(periodo["fin_date"])
    return {
        "tipo": "PREPARACION_CIERRE_TRIMESTRAL",
        "version_script": VERSION,
        "generado_en": generado_en,
        "estado": ESTADO_PREPARADO,
        "finalizacion_autorizada": False,
        "periodo": periodo["periodo"],
        "nombre_carpeta_trimestre": periodo["nombre_trimestre"],
        "ruta_relativa_trimestre": periodo["ruta_relativa_trimestre"],
        "fecha_inicio": periodo["fecha_inicio"],
        "fecha_fin": periodo["fecha_fin"],
        "carpeta_periodo": str(periodo["carpeta_base"]),
        "carpeta_cierre_trimestral": str(periodo["carpeta_trimestral"]),
        "excel_historico": str(periodo["excel_historico"]),
        "excel_historico_bytes": info_excel["bytes"],
        "excel_historico_sha256": info_excel["sha256"],
        "excel_info_al_preparar": info_excel,
        "inventario_jerarquia": str(inventario_path_final),
        "inventario_jerarquia_sha256": sha256_file(inventario_path_final),
        "jerarquia_trimestral": inventario,
        "excel_limpio_candidato": str(periodo["excel_candidato"]),
        "excel_limpio_candidato_sha256": info_candidato["sha256"],
        "excel_limpio_candidato_info": info_candidato,
        "respaldo_estado_activo": str(periodo["respaldo_estado"]),
        "estado_trimestral_al_preparar": estado,
        "siguiente_periodo_previsto": siguiente,
        "estructura": {
            "01_Excel_Cierre": "un único Excel histórico",
            "02_Manifest": "manifest e inventario",
            "03_Validaciones": "validación local y estado de preparación",
            "04_Resumen": "resumen legible",
        },
        "controles": {
            "cierres_inferiores_copiados": False,
            "zip_trimestral_creado": False,
            "excel_historico_duplicado": False,
            "excel_candidato_fuera_del_backup": True,
            "excel_activo_no_reemplazado": True,
            "estado_trimestral_no_actualizado": True,
            "remotos_no_modificados": True,
            "requiere_subida_solo_carpeta_trimestral": True,
            "requiere_finalizacion_separada": True,
        },
    }


def construir_resumen(manifest: dict) -> str:
    inv = manifest["jerarquia_trimestral"]
    siguiente = manifest["siguiente_periodo_previsto"]
    lineas = [
        "PREPARACIÓN SEGURA DEL CIERRE TRIMESTRAL",
        "=" * 80,
        f"Versión: {manifest['version_script']}",
        f"Generado en: {manifest['generado_en']}",
        f"Estado: {manifest['estado']}",
        "",
        "Periodo:",
        f"- Inicio: {manifest['fecha_inicio']}",
        f"- Fin: {manifest['fecha_fin']}",
        f"- Carpeta: {manifest['carpeta_cierre_trimestral']}",
        "",
        "Inventario de la jerarquía existente:",
        f"- Meses presentes: {inv['total_meses_presentes']}",
        f"- Semanas: {inv['total_semanas']}",
        f"- Cierres diarios: {inv['total_cierres_diarios']}",
        f"- Cierres semanales: {inv['total_cierres_semanales']}",
        f"- Cierres mensuales: {inv['total_cierres_mensuales']}",
        f"- Archivos verificados: {inv['total_archivos']}",
        f"- Bytes: {inv['total_bytes']}",
        f"- SHA256 jerarquía: {inv['sha256_contenido_jerarquia']}",
        "",
        "Excel histórico:",
        f"- Ruta: {manifest['excel_historico']}",
        f"- Filas de datos: {manifest['excel_info_al_preparar']['filas_datos']}",
        f"- SHA256: {manifest['excel_historico_sha256']}",
        "",
        "Candidato limpio fuera del backup:",
        f"- Ruta: {manifest['excel_limpio_candidato']}",
        f"- Tabla: {manifest['excel_limpio_candidato_info']['tbl_facturas_ref']}",
        f"- SHA256: {manifest['excel_limpio_candidato_sha256']}",
        "",
        "Siguiente ciclo previsto:",
        f"- Periodo: {siguiente['periodo_activo']}",
        f"- Inicio: {siguiente['fecha_inicio_periodo_activo']}",
        f"- Fin: {siguiente['proximo_cierre_estimado']}",
        "",
        "CONTROLES:",
        "- No se copiaron diarios, semanales ni mensuales dentro de Trimestral.",
        "- No se generó ZIP.",
        "- El backup contiene un solo Excel histórico.",
        "- data/facturas.xlsx no fue reemplazado.",
        "- cierre_trimestral_state.json no fue actualizado.",
        "- No se modificaron destinos remotos.",
        "- La finalización permanece bloqueada.",
    ]
    if inv.get("meses_sin_carpeta"):
        lineas.extend(
            [
                "",
                "ADVERTENCIA:",
                "- Meses sin carpeta local: "
                + ", ".join(inv["meses_sin_carpeta"]),
            ]
        )
    return "\n".join(lineas) + "\n"


def imprimir_plan(
    periodo: dict,
    info_excel: dict,
    inventario: dict,
    modo: str,
) -> None:
    print("Estado trimestral: ACTIVO")
    print(f"Periodo: {periodo['fecha_inicio']} a {periodo['fecha_fin']}")
    print(f"Carpeta base: {periodo['carpeta_base']}")
    print(f"Carpeta trimestral a crear: {periodo['carpeta_trimestral']}")
    print("-" * 100)
    print("Jerarquía existente inventariada:")
    print(f"Meses presentes: {inventario['total_meses_presentes']}")
    print(f"Semanas: {inventario['total_semanas']}")
    print(f"Cierres diarios: {inventario['total_cierres_diarios']}")
    print(f"Cierres semanales: {inventario['total_cierres_semanales']}")
    print(f"Cierres mensuales: {inventario['total_cierres_mensuales']}")
    print(f"Archivos verificados: {inventario['total_archivos']}")
    print(f"Bytes: {inventario['total_bytes']}")
    print(f"SHA256 jerarquía: {inventario['sha256_contenido_jerarquia']}")
    if inventario.get("meses_sin_carpeta"):
        print(
            "Advertencia - meses sin carpeta: "
            + ", ".join(inventario["meses_sin_carpeta"])
        )
    print("-" * 100)
    print("Excel activo validado:")
    print(f"Ruta: {FACTURAS_PATH}")
    print(f"Filas de datos: {info_excel['filas_datos']}")
    print(f"Tabla: {info_excel['tbl_facturas_ref']}")
    print(f"SHA256: {info_excel['sha256']}")
    print("-" * 100)
    print("PLAN:")
    print("1. Crear Trimestral/01_Excel_Cierre con un solo Excel histórico.")
    print("2. Crear Trimestral/02_Manifest con manifest e inventario SHA256.")
    print("3. Crear Trimestral/03_Validaciones con controles locales.")
    print("4. Crear Trimestral/04_Resumen con el resumen del cierre.")
    print("5. No copiar diarios, semanales ni mensuales.")
    print("6. No crear ZIP.")
    print("7. Guardar el candidato limpio fuera del backup.")
    if modo == "DRY RUN":
        print("-" * 100)
        print("DRY-RUN: no se creará ni modificará ningún archivo.")


def publicar_dos_carpetas(
    stage_paquete: Path,
    final_paquete: Path,
    stage_control: Path,
    final_control: Path,
) -> None:
    if final_paquete.exists() or final_control.exists():
        raise ErrorCierreTrimestral(
            "Los destinos finales aparecieron durante la preparación."
        )

    paquete_publicado = False
    try:
        os.replace(stage_paquete, final_paquete)
        paquete_publicado = True
        final_control.parent.mkdir(parents=True, exist_ok=True)
        os.replace(stage_control, final_control)
    except Exception:
        if paquete_publicado and final_paquete.exists() and not stage_paquete.exists():
            try:
                os.replace(final_paquete, stage_paquete)
            except Exception as rollback_exc:
                raise ErrorCierreTrimestral(
                    "Falló la publicación del control y también el rollback "
                    f"del paquete trimestral: {rollback_exc}"
                )
        raise


def ejecutar_preparacion(
    periodo: dict,
    estado: dict,
    info_excel: dict,
    inventario: dict,
) -> int:
    existente = validar_preparacion_existente(
        periodo,
        info_excel,
        inventario,
    )
    if existente is not None:
        print("-" * 100)
        print("La preparación trimestral ya existe y coincide.")
        print(f"Estado: {existente.get('estado')}")
        print(f"Excel histórico: {existente.get('excel_historico')}")
        print("No se creó ni modificó ningún archivo.")
        print("=" * 100)
        return 0

    base: Path = periodo["carpeta_base"]
    if not base.exists() or not base.is_dir():
        raise ErrorCierreTrimestral(f"No existe la carpeta base: {base}")

    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    # El staging se crea en una ruta corta para evitar el límite clásico
    # de 260 caracteres de Windows. Permanece en el mismo volumen para
    # conservar la publicación atómica mediante os.replace().
    stage_root = DATA_DIR / f".tmp_ct_{stamp}"
    stage_paquete = stage_root / "Trimestral"
    stage_control = PREPARACIONES_DIR / f".tmp_{periodo['nombre_trimestre']}_{stamp}"

    if stage_root.exists():
        shutil.rmtree(stage_root)
    if stage_control.exists():
        shutil.rmtree(stage_control)

    stage_excel = stage_paquete / "01_Excel_Cierre"
    stage_manifest = stage_paquete / "02_Manifest"
    stage_validaciones = stage_paquete / "03_Validaciones"
    stage_resumen = stage_paquete / "04_Resumen"
    for carpeta in (
        stage_excel,
        stage_manifest,
        stage_validaciones,
        stage_resumen,
        stage_control,
    ):
        carpeta.mkdir(parents=True, exist_ok=False)

    hash_excel_inicial = info_excel["sha256"]
    hash_jerarquia_inicial = inventario["sha256_contenido_jerarquia"]
    generado_en = datetime.now().isoformat(timespec="seconds")

    try:
        historico_stage = stage_excel / periodo["nombre_excel"]
        shutil.copy2(FACTURAS_PATH, historico_stage)
        if sha256_file(historico_stage) != hash_excel_inicial:
            raise ErrorCierreTrimestral(
                "El Excel histórico copiado no coincide con el Excel activo."
            )

        candidato_stage = stage_control / periodo["excel_candidato"].name
        crear_excel_limpio(FACTURAS_PATH, candidato_stage)
        info_candidato = diagnosticar_excel(candidato_stage)
        if (
            info_candidato["filas"] != 1
            or info_candidato["columnas"] != 19
            or info_candidato["tbl_facturas_ref"] != "A1:S1"
        ):
            raise ErrorCierreTrimestral(
                f"El candidato limpio es inválido: {info_candidato}"
            )

        respaldo_estado_stage = stage_control / periodo["respaldo_estado"].name
        shutil.copy2(STATE_PATH, respaldo_estado_stage)
        if sha256_file(respaldo_estado_stage) != sha256_file(STATE_PATH):
            raise ErrorCierreTrimestral(
                "El respaldo del estado no coincide con el archivo original."
            )

        inventario_name = (
            "inventario_jerarquia_trimestral_"
            f"{periodo['periodo_seguro']}.json"
        )
        inventario_stage = stage_manifest / inventario_name
        escribir_json_atomico(inventario_stage, inventario)

        manifest_name = (
            "manifest_cierre_trimestral_"
            f"{periodo['periodo_seguro']}.json"
        )
        manifest_stage = stage_manifest / manifest_name

        # El manifest usa las rutas finales, aunque todavía se esté trabajando
        # en staging.
        info_candidato_final = dict(info_candidato)
        info_candidato_final["archivo"] = str(periodo["excel_candidato"])
        inventario_final = periodo["carpeta_manifest"] / inventario_name
        manifest = construir_manifest(
            periodo=periodo,
            estado=estado,
            info_excel=info_excel,
            info_candidato=info_candidato_final,
            inventario=inventario,
            inventario_path_final=inventario_stage,
            generado_en=generado_en,
        )
        manifest["inventario_jerarquia"] = str(inventario_final)
        escribir_json_atomico(manifest_stage, manifest)

        validacion_name = (
            "validacion_local_cierre_trimestral_"
            f"{periodo['periodo_seguro']}.json"
        )
        validacion_stage = stage_validaciones / validacion_name
        validacion = {
            "tipo": "VALIDACION_LOCAL_CIERRE_TRIMESTRAL",
            "version_script": VERSION,
            "generado_en": generado_en,
            "ok": True,
            "periodo": periodo["periodo"],
            "fecha_inicio": periodo["fecha_inicio"],
            "fecha_fin": periodo["fecha_fin"],
            "carpeta_trimestral": str(periodo["carpeta_trimestral"]),
            "excel_historico": str(periodo["excel_historico"]),
            "excel_historico_sha256": hash_excel_inicial,
            "jerarquia_sha256": hash_jerarquia_inicial,
            "archivos_jerarquia_verificados": inventario["total_archivos"],
            "cierres_inferiores_copiados": False,
            "zip_creado": False,
            "excel_historico_duplicado": False,
        }
        escribir_json_atomico(validacion_stage, validacion)

        estado_name = (
            "estado_preparacion_cierre_trimestral_"
            f"{periodo['periodo_seguro']}.json"
        )
        estado_stage = stage_validaciones / estado_name
        estado_preparacion = {
            "tipo": "ESTADO_PREPARACION_CIERRE_TRIMESTRAL",
            "version_script": VERSION,
            "generado_en": generado_en,
            "estado": ESTADO_PREPARADO,
            "finalizacion_autorizada": False,
            "periodo": periodo["periodo"],
            "nombre_carpeta_trimestre": periodo["nombre_trimestre"],
            "fecha_inicio": periodo["fecha_inicio"],
            "fecha_fin": periodo["fecha_fin"],
            "carpeta_periodo": str(periodo["carpeta_base"]),
            "carpeta_cierre_trimestral": str(periodo["carpeta_trimestral"]),
            "manifest": str(periodo["carpeta_manifest"] / manifest_name),
            "inventario_jerarquia": str(inventario_final),
            "validacion_local": str(
                periodo["carpeta_validaciones"] / validacion_name
            ),
            "jerarquia_sha256_al_preparar": hash_jerarquia_inicial,
            "excel_activo": str(FACTURAS_PATH),
            "excel_activo_sha256_al_preparar": hash_excel_inicial,
            "excel_historico": str(periodo["excel_historico"]),
            "excel_historico_sha256": hash_excel_inicial,
            "excel_limpio_candidato": str(periodo["excel_candidato"]),
            "excel_limpio_candidato_sha256": info_candidato["sha256"],
            "respaldo_estado_activo": str(periodo["respaldo_estado"]),
            "validacion_remota_requerida": str(
                periodo["carpeta_validaciones"]
                / (
                    "validacion_remota_cierre_trimestral_"
                    f"{periodo['periodo_seguro']}.json"
                )
            ),
            "politica_subida": "SOLO_CARPETA_TRIMESTRAL",
        }
        escribir_json_atomico(estado_stage, estado_preparacion)

        resumen_name = (
            "RESUMEN_CIERRE_TRIMESTRAL_"
            f"{periodo['periodo_seguro']}.txt"
        )
        escribir_texto_atomico(
            stage_resumen / resumen_name,
            construir_resumen(manifest),
        )

        # Verificación final antes de publicar.
        if sha256_file(FACTURAS_PATH) != hash_excel_inicial:
            raise ErrorCierreTrimestral(
                "data/facturas.xlsx cambió durante la preparación."
            )
        inventario_control = inventariar_jerarquia_trimestral(periodo)
        if (
            inventario_control["sha256_contenido_jerarquia"]
            != hash_jerarquia_inicial
        ):
            raise ErrorCierreTrimestral(
                "La jerarquía cambió durante la preparación."
            )
        if cargar_estado() != estado:
            raise ErrorCierreTrimestral(
                "El estado trimestral cambió durante la preparación."
            )

        publicar_dos_carpetas(
            stage_paquete,
            periodo["carpeta_trimestral"],
            stage_control,
            periodo["carpeta_preparacion"],
        )
        try:
            stage_root.rmdir()
        except OSError:
            pass

        print("-" * 100)
        print("PREPARACIÓN TRIMESTRAL COMPLETADA.")
        print(f"Carpeta: {periodo['carpeta_trimestral']}")
        print(f"Excel histórico: {periodo['excel_historico']}")
        print(f"Candidato limpio: {periodo['excel_candidato']}")
        print("No se copiaron cierres diarios, semanales ni mensuales.")
        print("No se creó ZIP.")
        print("data/facturas.xlsx y el estado activo permanecen intactos.")
        print("La finalización continúa bloqueada.")
        print("=" * 100)
        return 0

    except Exception:
        if stage_root.exists():
            shutil.rmtree(stage_root, ignore_errors=True)
        if stage_control.exists():
            shutil.rmtree(stage_control, ignore_errors=True)
        raise


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Prepara el cierre trimestral sin duplicar cierres inferiores."
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Diagnostica sin modificar archivos.",
    )
    parser.add_argument(
        "--preparar",
        action="store_true",
        help="Prepara localmente el cierre sin finalizarlo.",
    )
    parser.add_argument(
        "--confirmar",
        default=None,
        help="Confirmación obligatoria para --preparar.",
    )
    parser.add_argument("--real", action="store_true", help=argparse.SUPPRESS)
    parser.add_argument("--local", action="store_true", help=argparse.SUPPRESS)
    args = parser.parse_args()

    if args.real or args.local:
        print("Los modos --real y --local pertenecen al flujo antiguo y están bloqueados.")
        return 1
    if args.dry_run and args.preparar:
        print("Usa solo un modo: --dry-run o --preparar.")
        return 1

    modo = "PREPARAR" if args.preparar else "DRY RUN"
    print("=" * 100)
    print(f"CIERRE TRIMESTRAL FACTURAS - {modo}")
    print("=" * 100)
    print(f"Versión: {VERSION}")
    print(f"Root: {ROOT}")
    print("Estructura: Meses existentes + carpeta independiente Trimestral")
    print("-" * 100)

    try:
        estado = cargar_estado()
        if args.preparar:
            validar_fecha_preparacion(estado, args.confirmar)
        periodo = datos_periodo(estado)
        inventario = inventariar_jerarquia_trimestral(periodo)
        info_excel = diagnosticar_excel()
        imprimir_plan(periodo, info_excel, inventario, modo)

        if args.preparar:
            return ejecutar_preparacion(
                periodo,
                estado,
                info_excel,
                inventario,
            )

        print("-" * 100)
        print("DRY-RUN finalizado. No se modificó ningún archivo.")
        print("=" * 100)
        return 0
    except Exception as exc:
        print(f"Error en cierre trimestral: {exc}")
        print("No se debe continuar con subida ni finalización.")
        print("=" * 100)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
