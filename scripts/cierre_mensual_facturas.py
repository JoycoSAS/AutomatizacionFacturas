# -*- coding: utf-8 -*-
"""
JOYCO - Facturas Procesador
Cierre mensual local seguro V5 sin duplicar cierres semanales ni paquetes ZIP.

Política:
- En modo normal el cierre mensual usa Excel principal + auditorias.
- En reconstruccion controlada puede consolidarse desde cierres diarios fisicos validados.
- Genera un Excel mensual cruzando el Excel operativo principal contra las auditorías reales del mes.
- NO copia manifiestos, validaciones ni archivos de los cierres semanales.
- NO genera un ZIP del cierre mensual, porque duplicaría la misma evidencia.
- Las carpetas Semanal y Diario permanecen independientes dentro del mes.
- No sube a OneDrive/SharePoint. La subida se hace con scripts/subir_cierre_mensual_sharepoint.py.
- No incluye .env real ni secretos.
- Solo permite meses calendario ya terminados.
- Genera en staging y publica atómicamente, sin destruir un cierre válido ante fallos.
- Consulta el trimestre operativo activo para mantener una sola jerarquía.
- Bloquea el cierre si el mes no pertenece completamente al trimestre activo.

Estructura generada:
data/cierres_diarios/YYYY/TRIMESTRE_YYYY-MM-DD_A_YYYY-MM-DD/
    YYYY-MM_Mes/Mensual/
        01_Excel_Mensual/facturas_mensual_YYYY-MM.xlsx
        02_Auditorias_Mes/audit_*.csv
        04_Manifest_Mensual/manifest_mensual_YYYY-MM.json
        04_Manifest_Mensual/resumen_mensual_YYYY-MM.txt
        05_Validaciones/validacion_local_mensual_YYYY-MM.json
        06_Logs_Mes/...
"""

from __future__ import annotations

import argparse
import csv
import datetime as _dt
import hashlib
import json
import os
import platform
import re
import shutil
import socket
import sys
import time
import unicodedata
from pathlib import Path
from typing import Any, Optional, Sequence
from zoneinfo import ZoneInfo

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.utils import get_column_letter
except Exception as exc:  # pragma: no cover
    raise RuntimeError("Este script requiere openpyxl instalado.") from exc

ROOT = Path(__file__).resolve().parents[1]
try:
    sys.path.insert(0, str(ROOT))
    import config  # noqa: F401
except Exception:
    pass

from trimestre_activo import cargar_trimestre_activo

VERSION_CIERRE = "2026-08-13-CIERRE-MENSUAL-V5-DESDE-DIARIOS-RECONSTRUCCION"

DATA_DIR = ROOT / "data"
AUDIT_DIR = DATA_DIR / "audit"
LOGS_DIR = ROOT / "logs"
DATA_LOGS_DIR = DATA_DIR / "logs"

# En local, los logs pueden permanecer dentro del repositorio. En el VPS se
# guardan fuera del codigo para separar aplicacion y operacion.
VPS_LOG_ROOT = Path(
    os.getenv("FACTURAS_LOG_ROOT", "/var/log/joyco/facturas-procesador")
)

CIERRES_DIR = DATA_DIR / "cierres_diarios"
LOCKS_DIR = DATA_DIR / "locks"

FACTURAS_PATH = Path(os.getenv("ARCHIVO_EXCEL_LOCAL", str(DATA_DIR / "facturas.xlsx")))
if not FACTURAS_PATH.is_absolute():
    FACTURAS_PATH = ROOT / FACTURAS_PATH

CIERRE_MENSUAL_DESDE_DIARIOS = (
    str(os.getenv("CIERRE_MENSUAL_DESDE_DIARIOS") or "")
    .strip()
    .lower()
    in {"1", "true", "si", "sí", "yes", "on"}
)

LOCK_FILE = LOCKS_DIR / "cierre_mensual_facturas.lock"
LOCK_TTL_SECONDS = int(os.getenv("LOCK_TTL_SECONDS", "3600") or "3600")

MESES_ES = {
    1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril", 5: "Mayo", 6: "Junio",
    7: "Julio", 8: "Agosto", 9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre",
}

PALABRAS_SENSIBLES = (
    "SECRET", "PASSWORD", "PASS", "TOKEN", "KEY", "CLIENT_SECRET",
    "CONTRASENA", "CONTRASEÑA", "PWD", "AUTH", "PRIVATE", "CERT",
)
NOMBRES_EXCLUIDOS = {
    ".env", ".env.local", ".env.production", "token_cache.json", "msal_cache.bin",
}

COLUMNAS_FECHA_PROCESO_CANDIDATAS = [
    "fecha_procesamiento", "fecha procesamiento", "fecha proceso",
    "procesado_en", "procesado en", "fecha_registro", "fecha registro",
    "fecha de registro", "fecha_carga", "fecha carga", "creado_en", "created_at",
]
COLS_NUMERO = [
    "número de factura", "numero de factura", "numerofactura",
    "numero factura", "factura", "numero",
]
COLS_RADICADO = ["radicado"]
COLS_CUFE = ["cufe", "cude", "cufe/cude", "uuid"]
COLS_ARCHIVO = ["archivo", "archivo origen", "nombre archivo", "pdf", "xml"]
COLS_ARCHIVO_AUDIT_FALLBACK = [
    "pdf_name", "pdf name", "archivo_pdf", "archivo pdf", "nombre_pdf",
    "nombre pdf", "attachment_name", "attachment name",
]
COLS_ASUNTO_AUDIT = [
    "subject", "subj", "asunto", "email_subject", "email subject",
    "mail_subject", "mail subject",
]
PATRON_RADICADO_ASUNTO = re.compile(
    r"(?i)\bradicado(?:\s*(?:n[.°ºo]*|no\.?))?\s*[-:#]?\s*(\d{4,})\b"
)

EXT_LOGS = {".log", ".txt", ".csv", ".json"}
EXCEL_SHEET_NAME = "Facturas"
EXCEL_TABLE_NAME = "TblFacturasMensual"
ZONA_HORARIA = ZoneInfo("America/Bogota")


def _parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Genera cierre mensual local seguro V5 sin duplicados.")
    parser.add_argument(
        "--fecha", default="",
        help="Fecha dentro del mes cerrado, formato YYYY-MM-DD. Default: mes anterior.",
    )
    parser.add_argument("--inicio", default="", help="Primer dia del mes, formato YYYY-MM-DD.")
    parser.add_argument("--fin", default="", help="Ultimo dia del mismo mes, formato YYYY-MM-DD.")
    parser.add_argument("--dry-run", action="store_true", help="Calcula y muestra resumen sin generar archivos.")
    parser.add_argument("--permitir-vacio", action="store_true", help="Permite generar cierre aunque no haya filas mensuales.")
    parser.add_argument(
        "--reemplazar", action="store_true",
        help="Reemplaza atomicamente un cierre existente; restaura el anterior si algo falla.",
    )
    return parser.parse_args()


def validar_fecha(fecha: str) -> _dt.date:
    try:
        return _dt.date.fromisoformat(fecha)
    except Exception as exc:
        raise RuntimeError(f"Fecha invalida {fecha!r}. Usa formato YYYY-MM-DD.") from exc


def ultimo_dia_mes(fecha: _dt.date) -> _dt.date:
    if fecha.month == 12:
        siguiente = _dt.date(fecha.year + 1, 1, 1)
    else:
        siguiente = _dt.date(fecha.year, fecha.month + 1, 1)
    return siguiente - _dt.timedelta(days=1)


def rango_mes(fecha: _dt.date) -> tuple[_dt.date, _dt.date]:
    inicio = _dt.date(fecha.year, fecha.month, 1)
    return inicio, ultimo_dia_mes(fecha)


def hoy_bogota() -> _dt.date:
    return _dt.datetime.now(ZONA_HORARIA).date()


def fecha_mes_anterior() -> _dt.date:
    return hoy_bogota().replace(day=1) - _dt.timedelta(days=1)


def validar_mes_cerrado(inicio: _dt.date, fin: _dt.date) -> None:
    if inicio.day != 1 or fin != ultimo_dia_mes(inicio):
        raise RuntimeError(
            "El cierre mensual exige un mes calendario completo: primer dia a ultimo dia."
        )
    if (inicio.year, inicio.month) != (fin.year, fin.month):
        raise RuntimeError("El inicio y el fin deben pertenecer al mismo mes.")
    primer_dia_mes_actual = hoy_bogota().replace(day=1)
    if fin >= primer_dia_mes_actual:
        raise RuntimeError(
            f"No se permite cerrar un mes abierto o futuro. Mes actual en Bogota: "
            f"{primer_dia_mes_actual:%Y-%m}."
        )


def rango_desde_args(args: argparse.Namespace) -> tuple[_dt.date, _dt.date]:
    if args.inicio or args.fin:
        if not args.inicio or not args.fin:
            raise RuntimeError("Si usas --inicio o --fin debes enviar ambos.")
        inicio = validar_fecha(args.inicio)
        fin = validar_fecha(args.fin)
        if fin < inicio:
            raise RuntimeError("El --fin no puede ser menor que --inicio.")
    else:
        fecha = validar_fecha(args.fecha) if args.fecha else fecha_mes_anterior()
        inicio, fin = rango_mes(fecha)
    validar_mes_cerrado(inicio, fin)
    return inicio, fin


def fechas_en_rango(inicio: _dt.date, fin: _dt.date) -> list[_dt.date]:
    out: list[_dt.date] = []
    d = inicio
    while d <= fin:
        out.append(d)
        d += _dt.timedelta(days=1)
    return out


def mes_carpeta(fecha: _dt.date) -> str:
    return f"{fecha:%Y-%m}_{MESES_ES.get(fecha.month, fecha.strftime('%B'))}"


def periodo_nombre(inicio: _dt.date, fin: _dt.date) -> str:
    if inicio.day == 1 and fin == ultimo_dia_mes(inicio):
        return inicio.strftime("%Y-%m")
    return f"{inicio.isoformat()}_a_{fin.isoformat()}"



def rel(path: Path) -> str:
    try:
        return str(path.resolve().relative_to(ROOT.resolve())).replace("\\", "/")
    except Exception:
        return str(path).replace("\\", "/")


def norm(value: Any) -> str:
    s = "" if value is None else str(value)
    s = s.strip().replace("\xa0", " ")
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", s).lower()


def norm_key(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", norm(value))


def clean_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (_dt.datetime, _dt.date)):
        return value.isoformat()
    return str(value).strip()


def find_col(headers: Sequence[Any], candidates: Sequence[str]) -> Optional[int]:
    candidate_norms = {norm(c) for c in candidates}
    candidate_keys = {norm_key(c) for c in candidates}
    for idx, header in enumerate(headers):
        if norm(header) in candidate_norms or norm_key(header) in candidate_keys:
            return idx
    return None


def fecha_valor_en_rango(value: Any, inicio: _dt.date, fin: _dt.date) -> bool:
    if value is None:
        return False
    if isinstance(value, _dt.datetime):
        return inicio <= value.date() <= fin
    if isinstance(value, _dt.date):
        return inicio <= value <= fin
    s = str(value).strip()
    if not s:
        return False
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return inicio <= _dt.datetime.strptime(s[:10], fmt).date() <= fin
        except Exception:
            pass
    return any(d.isoformat() in s for d in fechas_en_rango(inicio, fin))


def extraer_primer_valor_no_vacio(
    row: dict[str, Any], candidates: Sequence[str]
) -> str:
    by_norm = {norm(k): v for k, v in row.items()}
    by_key = {norm_key(k): v for k, v in row.items()}
    for candidate in candidates:
        values: list[Any] = []
        if norm(candidate) in by_norm:
            values.append(by_norm[norm(candidate)])
        if norm_key(candidate) in by_key:
            values.append(by_key[norm_key(candidate)])
        for value in values:
            cleaned = clean_value(value)
            if cleaned:
                return cleaned
    return ""


def extraer_radicado_desde_asunto(asunto: str) -> str:
    if not asunto:
        return ""
    match = PATRON_RADICADO_ASUNTO.search(str(asunto))
    return match.group(1) if match else ""


def crear_lock() -> tuple[bool, str]:
    LOCKS_DIR.mkdir(parents=True, exist_ok=True)
    if LOCK_FILE.exists():
        try:
            stat = LOCK_FILE.stat()
            age = _dt.datetime.now().timestamp() - stat.st_mtime
            if age < LOCK_TTL_SECONDS:
                return False, f"Existe lock activo: {LOCK_FILE} | edad={age:.0f}s"
            LOCK_FILE.unlink(missing_ok=True)
        except Exception as exc:
            return False, f"No se pudo eliminar lock vencido: {LOCK_FILE} | {exc}"
    payload = {
        "script": "cierre_mensual_facturas.py",
        "version": VERSION_CIERRE,
        "pid": os.getpid(),
        "started_at": _dt.datetime.now().isoformat(timespec="seconds"),
        "host": socket.gethostname(),
    }
    LOCK_FILE.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return True, "lock creado"


def liberar_lock() -> None:
    try:
        LOCK_FILE.unlink(missing_ok=True)
    except Exception:
        pass


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def info_archivo(path: Path, origen: Optional[Path] = None, categoria: str = "evidencia") -> dict[str, Any]:
    return {
        "categoria": categoria,
        "archivo": rel(path),
        "origen": rel(origen) if origen else None,
        "nombre": path.name,
        "bytes": path.stat().st_size if path.exists() else 0,
        "sha256": sha256_file(path) if path.exists() and path.is_file() else None,
    }


def read_csv_rows(path: Path) -> list[dict[str, str]]:
    for enc in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            with path.open("r", encoding=enc, newline="") as f:
                sample = f.read(4096)
                f.seek(0)
                try:
                    dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
                except Exception:
                    dialect = csv.excel
                reader = csv.DictReader(f, dialect=dialect)
                return [{k or "": v or "" for k, v in row.items()} for row in reader]
        except UnicodeDecodeError:
            continue
        except Exception:
            return []
    return []


def recolectar_audits(inicio: _dt.date, fin: _dt.date) -> list[Path]:
    rutas: list[Path] = []
    fechas = [d.isoformat() for d in fechas_en_rango(inicio, fin)]
    for fecha in fechas:
        patrones = [
            f"audit_detalle_{fecha}.csv",
            f"audit_runs_{fecha}.csv",
            f"audit_*_{fecha}.csv",
            f"*{fecha}*.csv",
        ]
        for base in (AUDIT_DIR, DATA_DIR, ROOT):
            if not base.exists():
                continue
            for patron in patrones:
                for p in base.glob(patron):
                    if p.is_file() and "audit" in p.name.lower():
                        rutas.append(p)
    vistos = set()
    out = []
    for p in rutas:
        try:
            key = str(p.resolve()).lower()
        except Exception:
            key = str(p).lower()
        if key not in vistos:
            vistos.add(key)
            out.append(p)
    return sorted(out)


def log_corresponde_rango(path: Path, inicio: _dt.date, fin: _dt.date) -> bool:
    name = path.name.lower()
    for d in fechas_en_rango(inicio, fin):
        fecha = d.isoformat()
        if fecha in name or fecha.replace("-", "") in name:
            return True
    try:
        mtime = _dt.date.fromtimestamp(path.stat().st_mtime)
        return inicio <= mtime <= fin
    except Exception:
        return False


def bases_logs_disponibles() -> list[tuple[str, Path]]:
    candidatos = [
        ("app_logs", LOGS_DIR),
        ("data_logs", DATA_LOGS_DIR),
        ("runtime", VPS_LOG_ROOT / "runtime"),
        ("runs", VPS_LOG_ROOT / "runs"),
        ("cron", VPS_LOG_ROOT / "cron"),
    ]
    vistos: set[str] = set()
    bases: list[tuple[str, Path]] = []
    for etiqueta, base in candidatos:
        if not base.exists() or not base.is_dir():
            continue
        try:
            key = str(base.resolve()).casefold()
        except Exception:
            key = str(base).casefold()
        if key in vistos:
            continue
        vistos.add(key)
        bases.append((etiqueta, base))
    return bases


def recolectar_logs(inicio: _dt.date, fin: _dt.date) -> list[Path]:
    rutas: list[Path] = []
    for _etiqueta, base in bases_logs_disponibles():
        for p in base.rglob("*"):
            if not p.is_file():
                continue
            if p.suffix.lower() not in EXT_LOGS:
                continue
            if log_corresponde_rango(p, inicio, fin):
                rutas.append(p)
    vistos = set()
    out = []
    for p in rutas:
        try:
            key = str(p.resolve()).lower()
        except Exception:
            key = str(p).lower()
        if key not in vistos:
            vistos.add(key)
            out.append(p)
    return sorted(out, key=lambda p: str(p).casefold())


def extraer_candidatos_desde_audits(
    audit_files: list[Path],
) -> dict[str, set[str]]:
    """Extrae identificadores fuertes, incluidos los registros MINIMA."""
    candidatos: dict[str, set[str]] = {
        "cufe": set(),
        "numero": set(),
        "radicado": set(),
        "archivo": set(),
        "numero_radicado": set(),
        "numero_archivo": set(),
        "radicado_archivo": set(),
    }
    for audit in audit_files:
        for row in read_csv_rows(audit):
            cufe = extraer_primer_valor_no_vacio(row, COLS_CUFE)
            numero = extraer_primer_valor_no_vacio(row, COLS_NUMERO)
            radicado = extraer_primer_valor_no_vacio(row, COLS_RADICADO)
            archivo = extraer_primer_valor_no_vacio(row, COLS_ARCHIVO)
            if not archivo:
                archivo = extraer_primer_valor_no_vacio(
                    row, COLS_ARCHIVO_AUDIT_FALLBACK
                )
            if not radicado:
                asunto = extraer_primer_valor_no_vacio(row, COLS_ASUNTO_AUDIT)
                radicado = extraer_radicado_desde_asunto(asunto)

            cufe_key = norm_key(cufe)
            numero_key = norm_key(numero)
            radicado_key = norm_key(radicado)
            archivo_key = norm_key(archivo)
            if cufe_key:
                candidatos["cufe"].add(cufe_key)
            if numero_key:
                candidatos["numero"].add(numero_key)
            if radicado_key:
                candidatos["radicado"].add(radicado_key)
            if archivo_key:
                candidatos["archivo"].add(archivo_key)
            if numero_key and radicado_key:
                candidatos["numero_radicado"].add(f"{numero_key}|{radicado_key}")
            if numero_key and archivo_key:
                candidatos["numero_archivo"].add(f"{numero_key}|{archivo_key}")
            if radicado_key and archivo_key:
                candidatos["radicado_archivo"].add(f"{radicado_key}|{archivo_key}")
    return candidatos


def fila_match_candidatos(
    row_values: Sequence[Any],
    headers: Sequence[Any],
    candidatos: dict[str, set[str]],
) -> bool:
    idx_cufe = find_col(headers, COLS_CUFE)
    idx_numero = find_col(headers, COLS_NUMERO)
    idx_radicado = find_col(headers, COLS_RADICADO)
    idx_archivo = find_col(headers, COLS_ARCHIVO)

    def valor(idx: Optional[int]) -> str:
        return norm_key(row_values[idx]) if idx is not None and idx < len(row_values) else ""

    cufe = valor(idx_cufe)
    numero = valor(idx_numero)
    radicado = valor(idx_radicado)
    archivo = valor(idx_archivo)

    if cufe and cufe in candidatos["cufe"]:
        return True
    if numero and radicado and f"{numero}|{radicado}" in candidatos["numero_radicado"]:
        return True
    if numero and archivo and f"{numero}|{archivo}" in candidatos["numero_archivo"]:
        return True
    if radicado and archivo and f"{radicado}|{archivo}" in candidatos["radicado_archivo"]:
        return True
    if numero and radicado and numero in candidatos["numero"] and radicado in candidatos["radicado"]:
        return True
    if numero and archivo and numero in candidatos["numero"] and archivo in candidatos["archivo"]:
        return True
    return False



def obtener_filas_mensuales_desde_diarios(
    inicio: _dt.date,
    fin: _dt.date,
) -> tuple[list[Any], list[list[Any]], dict[str, Any]]:
    """
    Reconstruye el Excel mensual concatenando los Excel diarios
    fisicos existentes dentro del rango mensual solicitado.

    No infiere fechas desde auditorias ni desde facturas.xlsx.
    """
    headers_base: Optional[list[Any]] = None
    filas_mes: list[list[Any]] = []
    diarios_usados: list[dict[str, Any]] = []

    fecha = inicio

    while fecha <= fin:
        fecha_str = fecha.isoformat()

        patron = (
            f"**/Diario_{fecha_str}/"
            f"01_Excel_Diario/"
            f"facturas_diario_{fecha_str}.xlsx"
        )

        encontrados = sorted(
            p
            for p in CIERRES_DIR.glob(patron)
            if p.is_file()
        )

        if len(encontrados) > 1:
            raise RuntimeError(
                "Se encontraron varios cierres diarios para "
                f"{fecha_str}: {[rel(x) for x in encontrados]}"
            )

        if not encontrados:
            fecha += _dt.timedelta(days=1)
            continue

        excel_diario = encontrados[0]

        wb = load_workbook(
            excel_diario,
            read_only=True,
            data_only=True,
        )

        try:
            ws = (
                wb[EXCEL_SHEET_NAME]
                if EXCEL_SHEET_NAME in wb.sheetnames
                else wb[wb.sheetnames[0]]
            )

            rows_iter = ws.iter_rows(values_only=True)
            raw_headers = next(rows_iter, None)

            if raw_headers is None:
                raise RuntimeError(
                    f"Excel diario sin encabezados: {excel_diario}"
                )

            headers = list(raw_headers or [])

            if not headers:
                raise RuntimeError(
                    f"Excel diario con encabezados vacios: {excel_diario}"
                )

            if headers_base is None:
                headers_base = headers
            elif headers != headers_base:
                raise RuntimeError(
                    "Encabezados incompatibles entre cierres diarios. "
                    f"Fecha: {fecha_str}"
                )

            filas_dia = 0

            for values in rows_iter:
                row = list(values or [])

                if len(row) < len(headers):
                    row.extend(
                        [None] * (len(headers) - len(row))
                    )
                elif len(row) > len(headers):
                    row = row[:len(headers)]

                if any(v not in (None, "") for v in row):
                    filas_mes.append(row)
                    filas_dia += 1

            diarios_usados.append(
                {
                    "fecha": fecha_str,
                    "excel": rel(excel_diario),
                    "filas": filas_dia,
                }
            )

        finally:
            wb.close()

        fecha += _dt.timedelta(days=1)

    if headers_base is None:
        raise RuntimeError(
            "No se encontro ningun cierre diario fisico dentro "
            f"del rango {inicio.isoformat()} a {fin.isoformat()}."
        )

    meta: dict[str, Any] = {
        "metodo_seleccion": "cierres_diarios_v5_mes_concatenados",
        "fuente": "cierres_diarios_fisicos",
        "rango_inicio": inicio.isoformat(),
        "rango_fin": fin.isoformat(),
        "cantidad_diarios_usados": len(diarios_usados),
        "diarios_usados": diarios_usados,
        "filas_mensuales": len(filas_mes),
        "advertencias": [],
    }

    print(
        "Reconstruccion mensual desde diarios: "
        f"diarios={len(diarios_usados)} | "
        f"filas={len(filas_mes)}"
    )

    for item in diarios_usados:
        print(
            f"   {item['fecha']} | "
            f"filas={item['filas']} | "
            f"{item['excel']}"
        )

    return headers_base, filas_mes, meta


def obtener_filas_mensuales_desde_excel(
    inicio: _dt.date,
    fin: _dt.date,
    audit_files: list[Path],
) -> tuple[list[Any], list[list[Any]], dict[str, Any]]:
    if not FACTURAS_PATH.exists():
        raise RuntimeError(f"No existe Excel operativo: {FACTURAS_PATH}")

    print(f"Leyendo Excel operativo: {FACTURAS_PATH}")
    wb = load_workbook(FACTURAS_PATH, read_only=True, data_only=True)
    try:
        ws = wb[EXCEL_SHEET_NAME] if EXCEL_SHEET_NAME in wb.sheetnames else wb[wb.sheetnames[0]]
        print(f"Hoja usada: {ws.title} | max_row={ws.max_row} | max_column={ws.max_column}")
        rows_iter = ws.iter_rows(values_only=True)
        headers = list(next(rows_iter, []) or [])
        if not headers:
            raise RuntimeError("El Excel operativo no tiene encabezados.")
        rows_all: list[list[Any]] = []
        total = 0
        for row_tuple in rows_iter:
            total += 1
            row = list(row_tuple or [])
            if len(row) < len(headers):
                row.extend([None] * (len(headers) - len(row)))
            row = row[:len(headers)]
            if any(v not in (None, "") for v in row):
                rows_all.append(row)
            if total % 1000 == 0:
                print(f"   ... filas leidas: {total} | con datos={len(rows_all)}")
        print(f"Lectura Excel terminada: filas leidas={total} | con datos={len(rows_all)}")

        meta: dict[str, Any] = {
            "excel_operativo": rel(FACTURAS_PATH),
            "hoja_usada": ws.title,
            "filas_operativas_total": len(rows_all),
            "metodo_seleccion": "",
            "advertencias": [],
        }

        idx_fecha = find_col(headers, COLUMNAS_FECHA_PROCESO_CANDIDATAS)
        if idx_fecha is not None:
            filas_fecha = [
                row for row in rows_all
                if idx_fecha < len(row) and fecha_valor_en_rango(row[idx_fecha], inicio, fin)
            ]
            meta["metodo_seleccion"] = "columna_fecha_procesamiento_rango_mensual"
            meta["columna_fecha_procesamiento"] = headers[idx_fecha]
            return headers, filas_fecha, meta

        candidatos = extraer_candidatos_desde_audits(audit_files)
        total_candidates = sum(len(values) for values in candidatos.values())
        meta["audit_files_usados"] = [rel(p) for p in audit_files]
        meta["candidatos_extraidos"] = {
            key: len(values) for key, values in candidatos.items()
        }
        if total_candidates <= 0:
            meta["metodo_seleccion"] = "sin_candidatos_auditoria_mes"
            meta["advertencias"].append(
                "No se encontro columna de fecha de procesamiento ni candidatos "
                "suficientes en auditorias del mes."
            )
            return headers, [], meta

        print(f"Cruzando Excel contra auditorias del mes. Candidatos={total_candidates}")
        rows_match: list[list[Any]] = []
        for idx, row in enumerate(rows_all, start=1):
            if fila_match_candidatos(row, headers, candidatos):
                rows_match.append(row)
            if idx % 1000 == 0:
                print(f"   ... cruce mensual: {idx}/{len(rows_all)} | matches={len(rows_match)}")
        meta["metodo_seleccion"] = "auditorias_mes_vs_excel_operativo"
        return headers, rows_match, meta
    finally:
        wb.close()


def ajustar_ancho_columnas(ws) -> None:
    for col_cells in ws.columns:
        max_len = 0
        col_letter = col_cells[0].column_letter
        for cell in col_cells[:200]:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 45)


def crear_excel_mensual(destino: Path, headers: list[Any], rows: list[list[Any]]) -> dict[str, Any]:
    destino.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = EXCEL_SHEET_NAME
    clean_headers = [str(h) if h is not None and str(h).strip() else f"Columna_{i+1}" for i, h in enumerate(headers)]
    ws.append(clean_headers)
    for row in rows:
        ws.append(row[:len(clean_headers)] + [None] * max(0, len(clean_headers) - len(row)))
    if clean_headers:
        end_col = ws.cell(row=1, column=len(clean_headers)).column_letter
        end_row = max(1, len(rows) + 1)
        table = Table(displayName=EXCEL_TABLE_NAME, ref=f"A1:{end_col}{end_row}")
        style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        table.tableStyleInfo = style
        ws.add_table(table)
    ws.freeze_panes = "A2"
    ajustar_ancho_columnas(ws)
    wb.save(destino)
    wb.close()
    return {
        "archivo": rel(destino),
        "filas_datos": len(rows),
        "columnas": len(clean_headers),
        "bytes": destino.stat().st_size,
        "sha256": sha256_file(destino),
    }


def copiar_archivo(origen: Path, destino: Path, categoria: str) -> Optional[dict[str, Any]]:
    if not origen.exists() or not origen.is_file() or origen.name in NOMBRES_EXCLUIDOS:
        return None
    destino.parent.mkdir(parents=True, exist_ok=True)
    if destino.exists():
        contador = 2
        while True:
            candidato = destino.with_name(
                f"{destino.stem}_{contador}{destino.suffix}"
            )
            if not candidato.exists():
                destino = candidato
                break
            contador += 1
    shutil.copy2(origen, destino)
    return info_archivo(destino, origen=origen, categoria=categoria)


def copiar_auditorias(audit_files: list[Path], destino_dir: Path) -> list[dict[str, Any]]:
    items: list[dict[str, Any]] = []
    for p in audit_files:
        info = copiar_archivo(p, destino_dir / p.name, "auditoria_mes")
        if info:
            items.append(info)
    return items


def copiar_logs(log_files: list[Path], destino_dir: Path) -> list[dict[str, Any]]:
    items: list[dict[str, Any]] = []
    bases = bases_logs_disponibles()
    for p in log_files:
        rel_log: Optional[Path] = None
        for etiqueta, base in bases:
            try:
                rel_log = Path(etiqueta) / p.resolve().relative_to(base.resolve())
                break
            except Exception:
                continue
        if rel_log is None:
            rel_log = Path("otros") / p.name
        info = copiar_archivo(p, destino_dir / rel_log, "log_mes")
        if info:
            items.append(info)
    return items



def redactar_env_line(line: str) -> str:
    if not line or line.strip().startswith("#") or "=" not in line:
        return line
    key, value = line.split("=", 1)
    key_up = key.upper()
    if any(p in key_up for p in PALABRAS_SENSIBLES):
        if value.strip():
            return f"{key}=***REDACTADO***"
    return line


def crear_snapshot_env_redactado(destino_dir: Path, periodo: str) -> Optional[dict[str, Any]]:
    env_path = ROOT / ".env"
    if not env_path.exists():
        return None
    destino_dir.mkdir(parents=True, exist_ok=True)
    destino = destino_dir / f"snapshot_env_redactado_mensual_{periodo}.txt"
    lineas = env_path.read_text(encoding="utf-8", errors="replace").splitlines()
    destino.write_text("\n".join(redactar_env_line(x) for x in lineas) + "\n", encoding="utf-8")
    return info_archivo(destino, origen=env_path, categoria="config_redactada")


def validar_excel_generado(
    path: Path, expected_columns: int, expected_rows: int
) -> dict[str, Any]:
    out: dict[str, Any] = {"archivo": rel(path), "existe": path.exists(), "abre_ok": False, "hojas": [], "filas_datos": 0, "columnas": 0, "errores": []}
    if not path.exists():
        out["errores"].append("No existe Excel mensual.")
        return out
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            out["hojas"] = wb.sheetnames
            ws = wb[EXCEL_SHEET_NAME] if EXCEL_SHEET_NAME in wb.sheetnames else wb[wb.sheetnames[0]]
            out["abre_ok"] = True
            out["filas_datos"] = max(0, ws.max_row - 1)
            out["columnas"] = ws.max_column
            if expected_columns and ws.max_column != expected_columns:
                out["errores"].append(f"Columnas esperadas={expected_columns}, encontradas={ws.max_column}.")
            if out["filas_datos"] != expected_rows:
                out["errores"].append(
                    f"Filas esperadas={expected_rows}, encontradas={out['filas_datos']}."
                )
            out["bytes"] = path.stat().st_size
            out["sha256"] = sha256_file(path)
        finally:
            wb.close()
    except Exception as exc:
        out["errores"].append(f"No se pudo abrir Excel mensual: {type(exc).__name__}: {exc}")
    return out


def listar_archivos_para_manifest(base: Path, excluir: Optional[set[Path]] = None) -> list[dict[str, Any]]:
    excluir_resolved = {p.resolve() for p in (excluir or set()) if p.exists()}
    items = []
    for p in sorted(base.rglob("*")):
        if not p.is_file():
            continue
        try:
            if p.resolve() in excluir_resolved:
                continue
        except Exception:
            pass
        items.append({
            "categoria": "evidencia",
            "ruta_relativa": p.relative_to(base).as_posix(),
            "nombre": p.name,
            "bytes": p.stat().st_size,
            "sha256": sha256_file(p),
        })
    return items


def validar_archivos_manifest(base: Path, archivos: list[dict[str, Any]]) -> dict[str, Any]:
    errores: list[str] = []
    verificados = 0
    for item in archivos:
        ruta = base / str(item.get("ruta_relativa", ""))
        if not ruta.exists() or not ruta.is_file():
            errores.append(f"No existe: {item.get('ruta_relativa')}")
            continue
        bytes_real = ruta.stat().st_size
        hash_real = sha256_file(ruta)
        if bytes_real != int(item.get("bytes", -1)):
            errores.append(f"Tamano distinto: {item.get('ruta_relativa')}")
            continue
        if hash_real != item.get("sha256"):
            errores.append(f"SHA256 distinto: {item.get('ruta_relativa')}")
            continue
        verificados += 1
    return {
        "total_manifest": len(archivos),
        "total_verificados": verificados,
        "errores": errores,
        "ok": not errores and verificados == len(archivos),
    }



def publicar_atomico(staging_dir: Path, cierre_dir: Path, reemplazar: bool) -> None:
    anterior: Optional[Path] = None
    if cierre_dir.exists():
        if not reemplazar:
            raise RuntimeError(
                f"Ya existe un cierre mensual en {cierre_dir}. "
                "Usa --reemplazar solo despues de revisar el existente."
            )
        sello = _dt.datetime.now(ZONA_HORARIA).strftime("%Y%m%d_%H%M%S")
        anterior = cierre_dir.with_name(f"Mensual.anterior_{sello}")
        cierre_dir.rename(anterior)
    try:
        staging_dir.rename(cierre_dir)
    except Exception:
        if anterior and anterior.exists() and not cierre_dir.exists():
            anterior.rename(cierre_dir)
        raise
    if anterior and anterior.exists():
        shutil.rmtree(anterior)


def generar_resumen_txt(path: Path, manifest: dict[str, Any]) -> None:
    lines = [
        "CIERRE MENSUAL LOCAL SEGURO V5 - FACTURAS JOYCO",
        "=" * 90,
        f"Version: {manifest['version']}",
        f"Periodo: {manifest['periodo']}",
        f"Rango: {manifest['fecha_inicio']} a {manifest['fecha_fin']}",
        (
            "Trimestre: "
            f"{manifest['trimestre']['nombre_carpeta']} "
            f"({manifest['trimestre']['fecha_inicio']} a "
            f"{manifest['trimestre']['fecha_fin']})"
        ),
        f"Generado: {manifest['generado_en']}",
        f"Root: {manifest['root']}",
        f"Carpeta cierre: {manifest['carpeta_cierre']}",
        "",
        "Resumen:",
        f"- Filas en Excel mensual: {manifest['excel_mensual']['filas_datos']}",
        f"- Columnas en Excel mensual: {manifest['excel_mensual']['columnas']}",
        f"- Metodo seleccion: {manifest['seleccion_filas'].get('metodo_seleccion')}",
        f"- Auditorias copiadas: {manifest['conteos']['auditorias']}",
        f"- Logs copiados: {manifest['conteos']['logs']}",
        f"- Total archivos evidencia: {manifest['total_archivos']}",
        f"- Total bytes evidencia: {manifest['total_bytes']}",
        "",
        "Regla aplicada:",
        "- La fuente utilizada por el cierre mensual queda registrada en fuente_datos segun el modo ejecutado.",
        "- En reconstruccion controlada consolida los cierres diarios fisicos previamente validados.",
        "- Los cierres semanales y diarios permanecen en sus carpetas independientes; no se copian dentro de Mensual.",
        "- No se incluye .env real ni secretos.",
        "=" * 90,
    ]
    if manifest.get("advertencias"):
        lines.extend(["", "Advertencias:"])
        lines.extend([f"- {x}" for x in manifest["advertencias"]])
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> int:
    args = _parse_args()
    try:
        inicio, fin = rango_desde_args(args)
    except Exception as exc:
        print(f"Periodo mensual no valido: {exc}")
        return 2

    try:
        trimestre_inicio = cargar_trimestre_activo(ROOT, inicio)
        trimestre_fin = cargar_trimestre_activo(ROOT, fin)
    except Exception as exc:
        print(f"Trimestre operativo no valido para el mes: {exc}")
        return 2

    if (
        trimestre_inicio["nombre_carpeta"]
        != trimestre_fin["nombre_carpeta"]
    ):
        print(
            "Trimestre operativo no valido para el mes: el inicio y el fin "
            "quedaron en trimestres distintos."
        )
        return 2

    trimestre = trimestre_inicio
    anio = inicio.strftime("%Y")
    mes_nombre = mes_carpeta(inicio)
    periodo = periodo_nombre(inicio, fin)

    mes_dir = (
        CIERRES_DIR
        / Path(trimestre["ruta_relativa"])
        / mes_nombre
    )
    cierre_dir = mes_dir / "Mensual"
    staging_dir = mes_dir / f".Mensual.generando_{os.getpid()}"

    print("=" * 100)
    print("CIERRE MENSUAL LOCAL SEGURO V5 - FACTURAS JOYCO")
    print("=" * 100)
    print(f"Version: {VERSION_CIERRE}")
    print(f"Root: {ROOT}")
    print(f"Periodo: {periodo}")
    print(f"Rango: {inicio.isoformat()} a {fin.isoformat()}")
    print(f"Trimestre activo: {trimestre['nombre_carpeta']}")
    print(
        "Rango trimestre: "
        f"{trimestre['fecha_inicio']} a {trimestre['fecha_fin']}"
    )
    print(f"Mes carpeta: {mes_nombre}")
    print(f"Carpeta local destino: {cierre_dir}")
    print(f"Dry-run: {args.dry_run}")
    print(f"Reemplazar cierre existente: {args.reemplazar}")
    print("-" * 100)

    if not args.dry_run and cierre_dir.exists() and not args.reemplazar:
        print(f"ERROR: ya existe el cierre mensual: {cierre_dir}")
        print("No se modifica. Usa --reemplazar solo despues de revisarlo.")
        return 4

    lock_creado = False
    if not args.dry_run:
        ok_lock, msg_lock = crear_lock()
        if not ok_lock:
            print(f"ERROR lock: {msg_lock}")
            return 2
        lock_creado = True

    try:
        audit_files = recolectar_audits(inicio, fin)
        log_files = recolectar_logs(inicio, fin)
        if CIERRE_MENSUAL_DESDE_DIARIOS:
            headers, filas_mensuales, meta_seleccion = (
                obtener_filas_mensuales_desde_diarios(
                    inicio,
                    fin,
                )
            )
        else:
            headers, filas_mensuales, meta_seleccion = (
                obtener_filas_mensuales_desde_excel(
                    inicio,
                    fin,
                    audit_files,
                )
            )

        advertencias: list[str] = []
        advertencias.extend(meta_seleccion.get("advertencias", []))
        if not audit_files:
            advertencias.append("No se encontraron archivos de auditoria dentro del rango mensual.")
        if not log_files:
            advertencias.append("No se encontraron logs especificos del mes.")
        if not filas_mensuales:
            advertencias.append("El Excel mensual no tendra filas de datos para este periodo.")

        print(f"Auditorias detectadas: {len(audit_files)}")
        print(f"Logs detectados: {len(log_files)}")
        print(f"Metodo seleccion filas: {meta_seleccion.get('metodo_seleccion')}")
        print(f"Filas mensuales detectadas: {len(filas_mensuales)}")

        if args.dry_run:
            print("-" * 100)
            print("DRY-RUN: no se generaron archivos.")
            if advertencias:
                print("Advertencias:")
                for adv in advertencias:
                    print(f"  - {adv}")
            return 0

        if not args.permitir_vacio and len(filas_mensuales) == 0:
            print("No se genero el cierre: el periodo no produjo filas mensuales.")
            print("Si el mes realmente no tuvo facturas nuevas, usa --permitir-vacio.")
            return 3

        if staging_dir.exists():
            shutil.rmtree(staging_dir)

        excel_dir = staging_dir / "01_Excel_Mensual"
        auditoria_dir = staging_dir / "02_Auditorias_Mes"
        manifest_dir = staging_dir / "04_Manifest_Mensual"
        validaciones_dir = staging_dir / "05_Validaciones"
        logs_dir = staging_dir / "06_Logs_Mes"

        excel_path = excel_dir / f"facturas_mensual_{periodo}.xlsx"
        manifest_path = manifest_dir / f"manifest_mensual_{periodo}.json"
        resumen_path = manifest_dir / f"resumen_mensual_{periodo}.txt"
        validacion_path = validaciones_dir / f"validacion_local_mensual_{periodo}.json"

        for d in (
            excel_dir,
            auditoria_dir,
            manifest_dir,
            validaciones_dir,
            logs_dir,
        ):
            d.mkdir(parents=True, exist_ok=True)

        excel_info = crear_excel_mensual(excel_path, list(headers), filas_mensuales)
        excel_info["archivo"] = f"01_Excel_Mensual/{excel_path.name}"
        auditorias_info = copiar_auditorias(audit_files, auditoria_dir)
        logs_info = copiar_logs(log_files, logs_dir)
        env_info = crear_snapshot_env_redactado(manifest_dir, periodo)

        generado_en = _dt.datetime.now(ZONA_HORARIA).isoformat(timespec="seconds")
        manifest = {
            "tipo": "cierre_mensual_local_seguro_v5",
            "version": VERSION_CIERRE,
            "trimestre": {
                "periodo_activo": trimestre["periodo_activo"],
                "nombre_carpeta": trimestre["nombre_carpeta"],
                "fecha_inicio": trimestre["fecha_inicio"],
                "fecha_fin": trimestre["fecha_fin"],
                "ruta_relativa": trimestre["ruta_relativa"],
                "state_path": trimestre["state_path"],
            },
            "anio": anio,
            "mes_carpeta": mes_nombre,
            "periodo": periodo,
            "fecha_inicio": inicio.isoformat(),
            "fecha_fin": fin.isoformat(),
            "generado_en": generado_en,
            "root": str(ROOT),
            "host": socket.gethostname(),
            "platform": platform.platform(),
            "python": sys.version.replace("\n", " "),
            "carpeta_cierre": rel(cierre_dir),
            "excel_mensual": excel_info,
            "seleccion_filas": meta_seleccion,
            "fuente_datos": {
                "modo": (
                    "cierres_diarios_fisicos"
                    if CIERRE_MENSUAL_DESDE_DIARIOS
                    else "excel_principal_mas_auditorias"
                ),
                "principal": (
                    None
                    if CIERRE_MENSUAL_DESDE_DIARIOS
                    else rel(FACTURAS_PATH)
                ),
                "cierres_diarios": (
                    meta_seleccion.get("diarios_usados", [])
                    if CIERRE_MENSUAL_DESDE_DIARIOS
                    else []
                ),
                "auditorias": [rel(p) for p in audit_files],
                "nota": (
                    "Reconstruccion mensual desde cierres diarios fisicos validados; "
                    "los cierres diarios y semanales permanecen independientes."
                    if CIERRE_MENSUAL_DESDE_DIARIOS
                    else
                    "Modo normal desde Excel principal y auditorias; "
                    "los cierres diarios y semanales permanecen independientes."
                ),
            },
            "conteos": {
                "filas_mensuales": len(filas_mensuales),
                "auditorias": len(auditorias_info),
                "logs": len(logs_info),
                "config_redactada": 1 if env_info else 0,
            },
            "validacion_local": f"05_Validaciones/{validacion_path.name}",
            "advertencias": advertencias,
            "nota": "Cierre mensual V5: evidencia propia del mes, sin copiar semanas, diarios ni generar paquetes ZIP.",
        }

        archivos_base = listar_archivos_para_manifest(
            staging_dir,
            excluir={manifest_path, resumen_path, validacion_path},
        )
        manifest["archivos"] = archivos_base
        manifest["total_archivos"] = len(archivos_base) + 1  # incluye el resumen
        manifest["total_bytes"] = sum(int(x.get("bytes", 0) or 0) for x in archivos_base)
        generar_resumen_txt(resumen_path, manifest)

        archivos = listar_archivos_para_manifest(
            staging_dir,
            excluir={manifest_path, validacion_path},
        )
        manifest["archivos"] = archivos
        manifest["total_archivos"] = len(archivos)
        manifest["total_bytes"] = sum(int(x.get("bytes", 0) or 0) for x in archivos)
        manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")

        validacion_excel = validar_excel_generado(
            excel_path,
            expected_columns=len(headers),
            expected_rows=len(filas_mensuales),
        )
        integridad_manifest = validar_archivos_manifest(staging_dir, archivos)
        copias_ok = (
            len(audit_files) == len(auditorias_info)
            and len(log_files) == len(logs_info)
        )
        nombres_sensibles = [
            p.relative_to(staging_dir).as_posix()
            for p in staging_dir.rglob("*")
            if p.is_file() and p.name in NOMBRES_EXCLUIDOS
        ]
        validacion = {
            "tipo": "validacion_local_cierre_mensual",
            "version": VERSION_CIERRE,
            "trimestre": {
                "periodo_activo": trimestre["periodo_activo"],
                "nombre_carpeta": trimestre["nombre_carpeta"],
                "fecha_inicio": trimestre["fecha_inicio"],
                "fecha_fin": trimestre["fecha_fin"],
                "ruta_relativa": trimestre["ruta_relativa"],
                "state_path": trimestre["state_path"],
            },
            "periodo": periodo,
            "fecha_inicio": inicio.isoformat(),
            "fecha_fin": fin.isoformat(),
            "generado_en": generado_en,
            "excel": validacion_excel,
            "integridad_manifest": integridad_manifest,
            "auditorias_detectadas": len(audit_files),
            "auditorias_copiadas": len(auditorias_info),
            "logs_detectados": len(log_files),
            "logs_copiados": len(logs_info),
            "copias_completas": copias_ok,
            "archivos_sensibles_detectados": nombres_sensibles,
            "ok": (
                bool(validacion_excel.get("abre_ok"))
                and not validacion_excel.get("errores")
                and bool(integridad_manifest.get("ok"))
                and copias_ok
                and not nombres_sensibles
            ),
            "advertencias": advertencias,
        }
        validacion_path.write_text(
            json.dumps(validacion, ensure_ascii=False, indent=2), encoding="utf-8"
        )
        if not validacion["ok"]:
            raise RuntimeError("La validacion local mensual no termino en OK.")

        publicar_atomico(staging_dir, cierre_dir, args.reemplazar)

        excel_final = cierre_dir / "01_Excel_Mensual" / excel_path.name
        manifest_final = cierre_dir / "04_Manifest_Mensual" / manifest_path.name
        resumen_final = cierre_dir / "04_Manifest_Mensual" / resumen_path.name
        validacion_final = cierre_dir / "05_Validaciones" / validacion_path.name
        print(f"Excel mensual: {excel_final}")
        print(f"Filas Excel mensual: {excel_info['filas_datos']}")
        print(f"Auditorias copiadas: {len(auditorias_info)}")
        print(f"Logs copiados: {len(logs_info)}")
        print(f"Manifest: {manifest_final}")
        print(f"Resumen: {resumen_final}")
        print(f"Validacion local: {validacion_final}")
        print(f"Total archivos evidencia: {manifest['total_archivos']}")
        print(f"Total bytes evidencia: {manifest['total_bytes']}")
        if advertencias:
            print("Advertencias:")
            for adv in advertencias:
                print(f"  - {adv}")

        print("=" * 100)
        print("Cierre mensual local V5 generado y publicado atomicamente.")
        print("Siguiente paso: subir/validar en OneDrive con scripts\\subir_cierre_mensual_sharepoint.py")
        print("=" * 100)
        return 0

    except Exception as exc:
        print(f"ERROR generando cierre mensual local V5: {type(exc).__name__}: {exc}")
        print("No subas ni archives nada hasta revisar el error.")
        return 1
    finally:
        if staging_dir.exists():
            shutil.rmtree(staging_dir, ignore_errors=True)
        if lock_creado:
            liberar_lock()


if __name__ == "__main__":
    raise SystemExit(main())
