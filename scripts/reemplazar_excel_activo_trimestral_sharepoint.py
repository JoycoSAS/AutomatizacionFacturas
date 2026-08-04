# -*- coding: utf-8 -*-
"""
JOYCO - Facturas Procesador
Finalización transaccional y segura del cierre trimestral.

Este script ejecuta únicamente la FASE FINAL del cierre trimestral.
Antes de modificar el Excel activo exige que:

1. Exista una preparación local válida del cierre.
2. data/facturas.xlsx conserve exactamente el SHA256 registrado al preparar.
3. El cierre histórico y el Excel limpio candidato sean íntegros.
4. Exista una validación remota OK del cierre histórico.
5. La evidencia de validación exista y coincida en los dos destinos.
6. Todos los archivos históricos vuelvan a verificarse remotamente.
7. El Excel activo de SharePoint coincida lógicamente con el Excel local.

Solo después de esas comprobaciones:

1. Respalda el Excel activo local, el estado y el Excel activo remoto.
2. Reemplaza y verifica el Excel activo de SharePoint con el candidato limpio.
3. Reemplaza de forma atómica data/facturas.xlsx con el candidato limpio.
4. Actualiza de forma atómica cierre_trimestral_state.json.
5. Publica y verifica la evidencia final en ambos destinos históricos.

Si ocurre un error después de iniciar los reemplazos, intenta restaurar:
- Excel activo de SharePoint.
- data/facturas.xlsx.
- cierre_trimestral_state.json.

Modos:
- --dry-run:
  ejecuta diagnóstico y validaciones de solo lectura. No reemplaza ni sube.
- --finalizar --confirmar FINALIZAR_CIERRE_TRIMESTRAL:
  ejecuta la finalización transaccional.
- --upload-activo:
  comando antiguo bloqueado expresamente para evitar el flujo inseguro previo.
"""

from __future__ import annotations

import argparse
import datetime
import hashlib
import json
import os
import shutil
import sys
import tempfile
import time
from collections import Counter
from contextlib import ExitStack
from pathlib import Path, PurePosixPath
from typing import Any, Iterable, Optional, Tuple
from urllib.parse import quote

import requests
from dotenv import load_dotenv
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

try:
    import config  # noqa: F401
except Exception:
    pass

from services.m365.token import get_access_token

try:
    from services.m365.sp_graph import SP_FOLDER as BASE_SP
except Exception:
    BASE_SP = ""

DATA_DIR = ROOT / "data"
STATE_PATH = DATA_DIR / "state" / "cierre_trimestral_state.json"
FACTURAS_PATH = DATA_DIR / "facturas.xlsx"
CIERRES_TRIMESTRALES_DIR = DATA_DIR / "cierres_trimestrales"
LOCKS_DIR = DATA_DIR / "state" / "locks"
LOCK_FINALIZACION_PATH = LOCKS_DIR / "cierre_trimestral_finalizacion.lock"
LOCK_APROBADAS_PATH = LOCKS_DIR / "aprobadas.lock"

VERSION = "2026-08-04-FINALIZACION-CIERRE-TRIMESTRAL-V2.2-COMPARACION-SEMANTICA"
GRAPH = "https://graph.microsoft.com/v1.0"
CONFIRMACION = "FINALIZAR_CIERRE_TRIMESTRAL"

ESTADO_PREPARADO = "PREPARADO_PENDIENTE_VALIDACION_REMOTA"
ESTADO_FINALIZADO = "FINALIZADO"
PREFIJO_ESTADO_PREPARACION = "estado_preparacion_cierre_trimestral_"
PREFIJO_VALIDACION_REMOTA = "validacion_remota_cierre_trimestral_"
PREFIJO_FINALIZACION = "estado_finalizacion_cierre_trimestral_"
PREFIJO_RESUMEN_FINAL = "RESUMEN_FINALIZACION_CIERRE_TRIMESTRAL_"

load_dotenv(ROOT / ".env")

BASE_SP = (BASE_SP or os.getenv("SP_FOLDER") or "").strip().strip("/")
SP_DRIVE_ID = (os.getenv("SP_DRIVE_ID") or "").strip()
SP_BACKUP2_HOSTNAME = (os.getenv("SP_BACKUP2_HOSTNAME") or "").strip()
SP_BACKUP2_SITE_PATH = (os.getenv("SP_BACKUP2_SITE_PATH") or "").strip()
SP_BACKUP2_DRIVE_ID = (os.getenv("SP_BACKUP2_DRIVE_ID") or "").strip()
SP_BACKUP2_FOLDER = (os.getenv("SP_BACKUP2_FOLDER") or "").strip().strip("/")

HEADERS_ESPERADOS = [
    "Radicado",
    "ProyectoProceso",
    "Archivo",
    "Empresa emisora",
    "CUFE",
    "Ciudad emisora",
    "Código ciudad",
    "NIT",
    "Cliente",
    "Número de factura",
    "Año",
    "Mes",
    "Día",
    "Tipo de contribuyente",
    "Actividad económica",
    "DESCRIPCIÓN",
    "Concepto",
    "VALOR",
    "Estado_calidad",
]

EXCEL_EXT = {".xlsx", ".xlsm"}


class LockAtomico:
    """
    Lock conservador por archivo.

    No elimina locks existentes automáticamente. Si hay un lock activo o
    huérfano, la finalización se bloquea para revisión manual.
    """

    def __init__(self, path: Path):
        self.path = path
        self.pid = os.getpid()
        self.adquirido = False

    def acquire(self) -> bool:
        self.path.parent.mkdir(parents=True, exist_ok=True)
        flags = os.O_CREAT | os.O_EXCL | os.O_WRONLY

        try:
            fd = os.open(str(self.path), flags, 0o640)
        except FileExistsError:
            return False
        except OSError:
            return False

        try:
            with os.fdopen(fd, "w", encoding="utf-8") as archivo:
                archivo.write(f"{self.pid}\n")
                archivo.flush()
                os.fsync(archivo.fileno())
        except Exception:
            try:
                self.path.unlink(missing_ok=True)
            except Exception:
                pass
            return False

        self.adquirido = True
        return True

    def release(self) -> None:
        if not self.adquirido:
            return

        try:
            contenido = self.path.read_text(encoding="utf-8").strip()
            if contenido == str(self.pid):
                self.path.unlink(missing_ok=True)
        except FileNotFoundError:
            pass
        finally:
            self.adquirido = False

    def __enter__(self) -> "LockAtomico":
        if not self.acquire():
            raise RuntimeError(
                "No fue posible adquirir el lock exclusivo: "
                f"{self.path}. No se realizará la finalización."
            )
        return self

    def __exit__(self, exc_type, exc_value, traceback) -> bool:
        self.release()
        return False


def ssl_verify() -> bool:
    return (os.getenv("SSL_VERIFY") or "true").strip().lower() not in {
        "0",
        "false",
        "no",
        "off",
    }


def headers_graph() -> dict:
    token = get_access_token()
    return {"Authorization": f"Bearer {token}"}


def headers_json() -> dict:
    resultado = headers_graph()
    resultado["Content-Type"] = "application/json"
    return resultado


def encode_path(path: str) -> str:
    return quote(str(path).strip("/"), safe="/")


def encode_drive_id(drive_id: str) -> str:
    return quote(str(drive_id), safe="!")


def resumen_drive(drive_id: str) -> str:
    if not drive_id:
        return "(vacío)"
    return f"***{drive_id[-6:]} (longitud={len(drive_id)})"


def sha256_bytes(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as archivo:
        for chunk in iter(lambda: archivo.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def leer_json(path: Path, descripcion: str) -> dict:
    if not path.exists():
        raise RuntimeError(f"No existe {descripcion}: {path}")

    try:
        datos = json.loads(path.read_text(encoding="utf-8-sig"))
    except Exception as exc:
        raise RuntimeError(
            f"No fue posible leer {descripcion}: {path}: {exc}"
        ) from exc

    if not isinstance(datos, dict):
        raise RuntimeError(f"{descripcion} no contiene un objeto JSON: {path}")

    return datos


def escribir_json_atomico(path: Path, datos: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporal = path.with_suffix(path.suffix + ".tmp")
    temporal.write_text(
        json.dumps(datos, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    temporal.replace(path)


def escribir_texto_atomico(path: Path, contenido: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    temporal = path.with_suffix(path.suffix + ".tmp")
    temporal.write_text(contenido, encoding="utf-8")
    temporal.replace(path)


def copiar_atomico(origen: Path, destino: Path) -> None:
    destino.parent.mkdir(parents=True, exist_ok=True)
    temporal = destino.with_name(destino.name + ".tmp_finalizacion")
    temporal.unlink(missing_ok=True)
    shutil.copy2(origen, temporal)

    with temporal.open("rb") as archivo:
        os.fsync(archivo.fileno())

    os.replace(temporal, destino)


def normalizar_valor_excel(valor: Any) -> str:
    if valor is None:
        return "<NULL>"
    if isinstance(
        valor,
        (datetime.datetime, datetime.date, datetime.time),
    ):
        return valor.isoformat()
    if isinstance(valor, float):
        return repr(valor)
    return f"{type(valor).__name__}:{str(valor)}"


def digest_datos_excel(path: Path) -> Tuple[str, dict]:
    h = hashlib.sha256()
    resumen = {
        "sheets": [],
        "non_empty_cells": 0,
        "max_rows_total": 0,
        "max_cols_total": 0,
    }

    wb = load_workbook(
        path,
        read_only=True,
        data_only=False,
        keep_links=False,
    )

    try:
        nombres = list(wb.sheetnames)
        h.update(json.dumps(nombres, ensure_ascii=False).encode("utf-8"))

        for ws in wb.worksheets:
            info = {
                "title": ws.title,
                "max_row": int(ws.max_row or 0),
                "max_column": int(ws.max_column or 0),
                "non_empty_cells": 0,
            }
            resumen["max_rows_total"] += info["max_row"]
            resumen["max_cols_total"] += info["max_column"]
            h.update(
                f"\n[SHEET]{ws.title}|{ws.max_row}|{ws.max_column}".encode(
                    "utf-8"
                )
            )

            for row in ws.iter_rows():
                for cell in row:
                    valor = cell.value
                    if valor is None:
                        continue
                    info["non_empty_cells"] += 1
                    resumen["non_empty_cells"] += 1
                    payload = (
                        f"{ws.title}|{cell.coordinate}|"
                        f"{normalizar_valor_excel(valor)}\n"
                    )
                    h.update(payload.encode("utf-8", errors="replace"))

            resumen["sheets"].append(info)
    finally:
        wb.close()

    return h.hexdigest(), resumen


CAMPOS_FECHA_NUMERICOS = {"Año", "Mes", "Día"}


def normalizar_campo_semantico_excel(campo: str, valor: Any) -> tuple[str, str]:
    """
    Canoniza valores para comparar dos Excel activos sin perder control.

    Reglas deliberadamente limitadas:
    - Las celdas vacías son equivalentes aunque Excel cambie su tipo interno.
    - Año, Mes y Día aceptan ceros a la izquierda ("08" == "8").
    - Los demás campos conservan valor y tipo; VALOR no se flexibiliza.
    """
    if valor is None:
        return ("VACIO", "")

    if isinstance(valor, str):
        if valor.strip() == "":
            return ("VACIO", "")

        if campo in CAMPOS_FECHA_NUMERICOS:
            texto = valor.strip()
            signo = texto[1:] if texto[:1] in {"+", "-"} else texto
            if signo.isdigit():
                return ("FECHA_ENTERA", str(int(texto)))

        return ("str", valor)

    if campo in CAMPOS_FECHA_NUMERICOS and not isinstance(valor, bool):
        if isinstance(valor, int):
            return ("FECHA_ENTERA", str(valor))
        if isinstance(valor, float) and valor.is_integer():
            return ("FECHA_ENTERA", str(int(valor)))

    if isinstance(
        valor,
        (datetime.datetime, datetime.date, datetime.time),
    ):
        return (type(valor).__name__, valor.isoformat())

    if isinstance(valor, float):
        return ("float", repr(valor))

    return (type(valor).__name__, str(valor))


def inventario_semantico_facturas(path: Path) -> dict:
    """
    Construye un multiconjunto de las 19 columnas de Facturas.

    El orden físico de las filas no afecta el resultado, pero sí se conservan
    duplicados mediante Counter. Esto permite aceptar reordenamientos de Excel
    sin ocultar cambios reales en registros o valores financieros.
    """
    wb = load_workbook(
        path,
        read_only=True,
        data_only=False,
        keep_links=False,
    )

    try:
        if "Facturas" not in wb.sheetnames:
            raise RuntimeError("El Excel no contiene la hoja Facturas.")

        ws = wb["Facturas"]
        primera_fila = next(
            ws.iter_rows(
                min_row=1,
                max_row=1,
                max_col=len(HEADERS_ESPERADOS),
                values_only=True,
            )
        )
        encabezados = tuple(primera_fila)

        if list(encabezados) != HEADERS_ESPERADOS:
            raise RuntimeError(
                "Los encabezados del Excel no coinciden con la estructura "
                "esperada para comparación semántica."
            )

        registros: Counter = Counter()

        for fila in ws.iter_rows(
            min_row=2,
            max_col=len(encabezados),
            values_only=True,
        ):
            canonica = tuple(
                normalizar_campo_semantico_excel(campo, valor)
                for campo, valor in zip(encabezados, fila)
            )
            registros[canonica] += 1

        filas_serializadas = [
            {
                "registro": registro,
                "cantidad": cantidad,
            }
            for registro, cantidad in sorted(registros.items())
        ]
        payload = {
            "encabezados": encabezados,
            "registros": filas_serializadas,
        }
        digest = hashlib.sha256(
            json.dumps(
                payload,
                ensure_ascii=False,
                separators=(",", ":"),
            ).encode("utf-8")
        ).hexdigest()

        return {
            "encabezados": encabezados,
            "registros": registros,
            "filas_datos": sum(registros.values()),
            "registros_unicos": len(registros),
            "digest_semantico": digest,
        }
    finally:
        wb.close()


def comparar_excel_facturas_semantico(local: Path, remoto: Path) -> dict:
    local_info = inventario_semantico_facturas(local)
    remoto_info = inventario_semantico_facturas(remoto)

    solo_local = local_info["registros"] - remoto_info["registros"]
    solo_remoto = remoto_info["registros"] - local_info["registros"]

    return {
        "ok": (
            local_info["encabezados"] == remoto_info["encabezados"]
            and not solo_local
            and not solo_remoto
        ),
        "filas_local": local_info["filas_datos"],
        "filas_remoto": remoto_info["filas_datos"],
        "registros_unicos_local": local_info["registros_unicos"],
        "registros_unicos_remoto": remoto_info["registros_unicos"],
        "registros_solo_local": sum(solo_local.values()),
        "registros_solo_remoto": sum(solo_remoto.values()),
        "digest_semantico_local": local_info["digest_semantico"],
        "digest_semantico_remoto": remoto_info["digest_semantico"],
    }


def validar_excel_estructura(
    path: Path,
    exigir_limpio: bool,
    tabla_requerida: Optional[str] = "TblFacturas",
) -> dict:
    if not path.exists():
        raise RuntimeError(f"No existe el Excel: {path}")

    wb = load_workbook(
        path,
        read_only=False,
        data_only=False,
        keep_links=False,
    )

    try:
        if "Facturas" not in wb.sheetnames:
            raise RuntimeError("El Excel no contiene la hoja Facturas.")

        ws = wb["Facturas"]
        headers = [cell.value for cell in ws[1]]
        if headers != HEADERS_ESPERADOS:
            raise RuntimeError(
                "Los encabezados del Excel no coinciden con la estructura "
                "esperada."
            )

        tabla_ref = (
            ws.tables[tabla_requerida].ref
            if tabla_requerida and tabla_requerida in ws.tables
            else None
        )

        info = {
            "archivo": str(path),
            "hojas": list(wb.sheetnames),
            "filas": int(ws.max_row or 0),
            "filas_datos": max(int(ws.max_row or 0) - 1, 0),
            "columnas": int(ws.max_column or 0),
            "tablas": list(ws.tables.keys()),
            "tabla_requerida": tabla_requerida,
            "tabla_requerida_ref": tabla_ref,
            "tbl_facturas_ref": (
                ws.tables["TblFacturas"].ref
                if "TblFacturas" in ws.tables
                else None
            ),
            "bytes": path.stat().st_size,
            "sha256": sha256_file(path),
        }

        if info["columnas"] != 19:
            raise RuntimeError(
                f"Columnas inválidas: {info['columnas']}. Esperadas: 19."
            )

        if tabla_requerida and tabla_ref is None:
            raise RuntimeError(
                f"No existe la tabla requerida {tabla_requerida}."
            )

        if exigir_limpio:
            if info["filas"] != 1:
                raise RuntimeError(
                    "El Excel candidato no está limpio. "
                    f"Filas detectadas: {info['filas']}. Esperadas: 1."
                )
            if tabla_ref != "A1:S1":
                raise RuntimeError(
                    "La tabla del Excel candidato no es A1:S1: "
                    f"{tabla_ref}"
                )

        return info
    finally:
        wb.close()


def excel_desde_bytes(data: bytes, prefijo: str) -> Path:
    fd, nombre = tempfile.mkstemp(prefix=prefijo, suffix=".xlsx")
    os.close(fd)
    path = Path(nombre)
    path.write_bytes(data)
    return path


def validar_excel_bytes(
    data: bytes,
    exigir_limpio: bool,
    tabla_requerida: Optional[str] = "TblFacturas",
) -> tuple[dict, str, dict]:
    temporal = excel_desde_bytes(data, "joyco_finalizacion_")
    try:
        info = validar_excel_estructura(
            temporal,
            exigir_limpio,
            tabla_requerida=tabla_requerida,
        )
        digest, resumen = digest_datos_excel(temporal)
        return info, digest, resumen
    finally:
        temporal.unlink(missing_ok=True)


def asegurar_dentro(path: Path, base: Path, descripcion: str) -> Path:
    resuelto = path.resolve()
    base_resuelta = base.resolve()

    try:
        resuelto.relative_to(base_resuelta)
    except ValueError as exc:
        raise RuntimeError(
            f"{descripcion} está fuera de la carpeta del periodo: {path}"
        ) from exc

    return resuelto


def ruta_relativa_segura(valor: str) -> str:
    ruta = PurePosixPath(str(valor).replace("\\", "/"))
    if ruta.is_absolute() or ".." in ruta.parts or not ruta.parts:
        raise RuntimeError(f"Ruta relativa insegura en evidencia: {valor}")
    return ruta.as_posix()


def validar_configuracion() -> None:
    faltantes = []
    if not SP_DRIVE_ID:
        faltantes.append("SP_DRIVE_ID")
    if not BASE_SP:
        faltantes.append("SP_FOLDER")
    if not SP_BACKUP2_FOLDER:
        faltantes.append("SP_BACKUP2_FOLDER")
    if not SP_BACKUP2_DRIVE_ID and not (
        SP_BACKUP2_HOSTNAME and SP_BACKUP2_SITE_PATH
    ):
        faltantes.append(
            "SP_BACKUP2_DRIVE_ID o SP_BACKUP2_HOSTNAME/SP_BACKUP2_SITE_PATH"
        )

    if faltantes:
        raise RuntimeError(
            "Configuración incompleta: " + ", ".join(faltantes)
        )

    if not FACTURAS_PATH.exists():
        raise RuntimeError(f"No existe el Excel activo: {FACTURAS_PATH}")
    if not STATE_PATH.exists():
        raise RuntimeError(f"No existe el estado trimestral: {STATE_PATH}")


def graph_get(url: str, *, ok=(200,), timeout=60) -> requests.Response:
    respuesta = requests.get(
        url,
        headers=headers_graph(),
        timeout=timeout,
        verify=ssl_verify(),
    )
    if respuesta.status_code not in ok:
        raise RuntimeError(
            "GET Graph falló. "
            f"HTTP={respuesta.status_code} | respuesta={respuesta.text[:400]}"
        )
    return respuesta


def validar_drive_id(drive_id: str) -> Optional[dict]:
    if not drive_id:
        return None

    url = (
        f"{GRAPH}/drives/{encode_drive_id(drive_id)}"
        "?$select=id,name,webUrl"
    )
    respuesta = requests.get(
        url,
        headers=headers_graph(),
        timeout=60,
        verify=ssl_verify(),
    )

    if respuesta.status_code == 200:
        return respuesta.json()
    return None


def listar_drives_site(hostname: str, site_path: str) -> list[dict]:
    site_path = site_path if site_path.startswith("/") else f"/{site_path}"
    url = (
        f"{GRAPH}/sites/{hostname}:{site_path}:/drives"
        "?$select=id,name,webUrl"
    )
    return graph_get(url).json().get("value", [])


def resolver_drive_secundario() -> str:
    drive = validar_drive_id(SP_BACKUP2_DRIVE_ID)
    if drive:
        return str(drive["id"])

    drives = listar_drives_site(
        SP_BACKUP2_HOSTNAME,
        SP_BACKUP2_SITE_PATH,
    )
    if not drives:
        raise RuntimeError("No se encontraron drives para el destino secundario.")

    preferidos = {"documentos", "documents", "shared documents"}
    for drive in drives:
        if str(drive.get("name") or "").strip().lower() in preferidos:
            return str(drive["id"])

    return str(drives[0]["id"])


def obtener_item_por_path(drive_id: str, remote_path: str) -> Optional[dict]:
    url = (
        f"{GRAPH}/drives/{encode_drive_id(drive_id)}/root:/"
        f"{encode_path(remote_path)}:"
        "?$select=id,name,size,file,lastModifiedDateTime"
    )
    respuesta = requests.get(
        url,
        headers=headers_graph(),
        timeout=60,
        verify=ssl_verify(),
    )

    if respuesta.status_code == 200:
        return respuesta.json()
    if respuesta.status_code == 404:
        return None

    raise RuntimeError(
        "No fue posible consultar la ruta remota. "
        f"HTTP={respuesta.status_code} | path={remote_path} | "
        f"respuesta={respuesta.text[:400]}"
    )


def descargar_item(drive_id: str, item_id: str) -> bytes:
    url = (
        f"{GRAPH}/drives/{encode_drive_id(drive_id)}/items/"
        f"{quote(item_id, safe='')}/content"
    )
    respuesta = requests.get(
        url,
        headers=headers_graph(),
        timeout=300,
        verify=ssl_verify(),
        allow_redirects=True,
    )

    if respuesta.status_code != 200:
        raise RuntimeError(
            "No fue posible descargar el archivo remoto. "
            f"HTTP={respuesta.status_code} | "
            f"respuesta={respuesta.text[:400]}"
        )

    return respuesta.content


def descargar_path(drive_id: str, remote_path: str) -> tuple[dict, bytes]:
    item = obtener_item_por_path(drive_id, remote_path)
    if not item:
        raise RuntimeError(f"No existe el archivo remoto: {remote_path}")
    return item, descargar_item(drive_id, str(item["id"]))


def graph_put_content(
    drive_id: str,
    remote_path: str,
    local_file: Path,
) -> Optional[dict]:
    url = (
        f"{GRAPH}/drives/{encode_drive_id(drive_id)}/root:/"
        f"{encode_path(remote_path)}:/content"
    )
    respuesta = requests.put(
        url,
        headers=headers_graph(),
        data=local_file.read_bytes(),
        timeout=300,
        verify=ssl_verify(),
    )

    if respuesta.status_code in (200, 201):
        return respuesta.json()

    # Un 504 puede ocurrir después de que Graph haya guardado el archivo.
    # La capa superior comprobará la ruta antes de declarar el fallo.
    if respuesta.status_code in (408, 429, 500, 502, 503, 504):
        return None

    raise RuntimeError(
        "PUT Graph falló. "
        f"HTTP={respuesta.status_code} | path={remote_path} | "
        f"respuesta={respuesta.text[:400]}"
    )


def verificar_contra_local(
    *,
    local: Path,
    drive_id: str,
    remote_path: str,
    exigir_excel_limpio: bool = False,
    tabla_excel_requerida: Optional[str] = "TblFacturas",
) -> dict:
    item, remoto = descargar_path(drive_id, remote_path)

    if local.suffix.lower() in EXCEL_EXT:
        _, digest_remoto, resumen_remoto = validar_excel_bytes(
            remoto,
            exigir_excel_limpio,
            tabla_requerida=tabla_excel_requerida,
        )
        digest_local, resumen_local = digest_datos_excel(local)

        if digest_local != digest_remoto:
            raise RuntimeError(
                "El Excel remoto no coincide lógicamente con el local: "
                f"{remote_path}\n"
                f"Digest local:  {digest_local}\n"
                f"Digest remoto: {digest_remoto}\n"
                f"Resumen local: {json.dumps(resumen_local, ensure_ascii=False)}\n"
                f"Resumen remoto: {json.dumps(resumen_remoto, ensure_ascii=False)}"
            )

        return {
            "remote_path": remote_path,
            "item_id": item.get("id"),
            "bytes_remotos": len(remoto),
            "digest_datos_excel": digest_remoto,
            "resumen_datos_excel": resumen_remoto,
        }

    hash_local = sha256_file(local)
    hash_remoto = sha256_bytes(remoto)
    if hash_local != hash_remoto:
        raise RuntimeError(
            "El archivo remoto no coincide por SHA256: "
            f"{remote_path}\n"
            f"SHA local:  {hash_local}\n"
            f"SHA remoto: {hash_remoto}"
        )

    return {
        "remote_path": remote_path,
        "item_id": item.get("id"),
        "bytes_remotos": len(remoto),
        "sha256": hash_remoto,
    }


def subir_y_verificar(
    *,
    local: Path,
    drive_id: str,
    remote_path: str,
    exigir_excel_limpio: bool = False,
    intentos: int = 4,
) -> dict:
    espera = 2
    ultimo_error: Optional[Exception] = None

    for intento in range(1, intentos + 1):
        try:
            graph_put_content(drive_id, remote_path, local)
            return verificar_contra_local(
                local=local,
                drive_id=drive_id,
                remote_path=remote_path,
                exigir_excel_limpio=exigir_excel_limpio,
            )
        except Exception as exc:
            ultimo_error = exc

        # Antes de repetir el PUT, comprobar si el archivo quedó guardado.
        try:
            item = obtener_item_por_path(drive_id, remote_path)
            if item:
                return verificar_contra_local(
                    local=local,
                    drive_id=drive_id,
                    remote_path=remote_path,
                    exigir_excel_limpio=exigir_excel_limpio,
                )
        except Exception as exc:
            ultimo_error = exc

        if intento < intentos:
            print(
                f"⚠️ Reintento {intento}/{intentos} para {remote_path} "
                f"en {espera}s..."
            )
            time.sleep(espera)
            espera *= 2

    raise RuntimeError(
        f"No fue posible subir y verificar {remote_path}: {ultimo_error}"
    )


def buscar_ultima_preparacion() -> tuple[Path, dict]:
    if not CIERRES_TRIMESTRALES_DIR.exists():
        raise RuntimeError(
            "No existe la carpeta de cierres trimestrales: "
            f"{CIERRES_TRIMESTRALES_DIR}"
        )

    candidatos = sorted(
        CIERRES_TRIMESTRALES_DIR.rglob(
            f"02_Soportes_Tecnicos/{PREFIJO_ESTADO_PREPARACION}*.json"
        ),
        key=lambda path: path.stat().st_mtime,
        reverse=True,
    )

    if not candidatos:
        raise RuntimeError("No se encontró una preparación trimestral.")

    path = candidatos[0]
    return path, leer_json(path, "estado de preparación")


def buscar_finalizacion_existente(carpeta_soportes: Path) -> Optional[dict]:
    candidatos = sorted(
        carpeta_soportes.glob(f"{PREFIJO_FINALIZACION}*.json"),
        key=lambda path: path.stat().st_mtime,
        reverse=True,
    )

    for path in candidatos:
        datos = leer_json(path, "estado de finalización")
        if datos.get("ok") is True and datos.get("estado") == ESTADO_FINALIZADO:
            datos["_path"] = str(path)
            return datos

    return None


def validar_preparacion_local() -> dict:
    estado_path, preparacion = buscar_ultima_preparacion()

    carpeta_soportes = estado_path.parent.resolve()
    carpeta_periodo = carpeta_soportes.parent.resolve()
    carpeta_excel = carpeta_periodo / "01_Excel_Cierre"

    finalizacion = buscar_finalizacion_existente(carpeta_soportes)
    if finalizacion:
        return {
            "ya_finalizado": True,
            "finalizacion": finalizacion,
            "estado_preparacion_path": estado_path,
            "carpeta_soportes": carpeta_soportes,
            "carpeta_periodo": carpeta_periodo,
        }

    if preparacion.get("tipo") != "ESTADO_PREPARACION_CIERRE_TRIMESTRAL":
        raise RuntimeError("El archivo encontrado no es un estado de preparación.")
    if preparacion.get("estado") != ESTADO_PREPARADO:
        raise RuntimeError(
            "Estado de preparación inesperado: "
            f"{preparacion.get('estado')}"
        )
    if preparacion.get("finalizacion_autorizada") is not False:
        raise RuntimeError(
            "El estado de preparación fue alterado: "
            "finalizacion_autorizada debe ser false."
        )

    manifest_path = asegurar_dentro(
        Path(str(preparacion.get("manifest") or "")),
        carpeta_periodo,
        "Manifest",
    )
    manifest = leer_json(manifest_path, "manifest de preparación")

    if manifest.get("tipo") != "PREPARACION_CIERRE_TRIMESTRAL":
        raise RuntimeError("El manifest no corresponde a una preparación trimestral.")
    if manifest.get("estado") != ESTADO_PREPARADO:
        raise RuntimeError("El manifest no está en estado PREPARADO.")
    if manifest.get("finalizacion_autorizada") is not False:
        raise RuntimeError("El manifest de preparación fue alterado.")

    periodo = str(preparacion.get("periodo") or "").strip()
    fecha_fin = str(preparacion.get("fecha_fin") or "").strip()
    if not periodo or not fecha_fin:
        raise RuntimeError("La preparación no contiene periodo y fecha_fin.")

    hoy = datetime.date.today()
    try:
        fecha_fin_date = datetime.datetime.strptime(fecha_fin, "%Y-%m-%d").date()
    except Exception as exc:
        raise RuntimeError(f"fecha_fin inválida en preparación: {fecha_fin}") from exc

    if hoy < fecha_fin_date:
        raise RuntimeError(
            "La finalización está bloqueada por fecha. "
            f"Fecha de cierre: {fecha_fin} | Fecha actual: {hoy.isoformat()}"
        )

    excel_activo_hash = str(
        preparacion.get("excel_activo_sha256_al_preparar") or ""
    )
    if not excel_activo_hash:
        raise RuntimeError("Falta el SHA256 del Excel activo al preparar.")

    info_activo = validar_excel_estructura(FACTURAS_PATH, exigir_limpio=False)
    if info_activo["sha256"] != excel_activo_hash:
        raise RuntimeError(
            "data/facturas.xlsx cambió después de la preparación.\n"
            f"SHA preparación: {excel_activo_hash}\n"
            f"SHA actual:      {info_activo['sha256']}\n"
            "Debe revisarse antes de finalizar."
        )

    historico = asegurar_dentro(
        Path(str(preparacion.get("excel_historico") or "")),
        carpeta_periodo,
        "Excel histórico",
    )
    candidato = asegurar_dentro(
        Path(str(preparacion.get("excel_limpio_candidato") or "")),
        carpeta_periodo,
        "Excel limpio candidato",
    )

    if not historico.exists() or not candidato.exists():
        raise RuntimeError("Falta el Excel histórico o el candidato limpio.")

    hash_historico = sha256_file(historico)
    if hash_historico != str(preparacion.get("excel_historico_sha256") or ""):
        raise RuntimeError("El Excel histórico no coincide con la preparación.")
    if hash_historico != excel_activo_hash:
        raise RuntimeError("El Excel histórico no coincide con el Excel activo preparado.")

    hash_candidato = sha256_file(candidato)
    if hash_candidato != str(
        preparacion.get("excel_limpio_candidato_sha256") or ""
    ):
        raise RuntimeError("El Excel limpio candidato no coincide con la preparación.")

    info_candidato = validar_excel_estructura(candidato, exigir_limpio=True)

    estado_actual = leer_json(STATE_PATH, "estado trimestral activo")
    estado_previo = manifest.get("estado_trimestral_al_preparar")
    if not isinstance(estado_previo, dict) or estado_actual != estado_previo:
        raise RuntimeError(
            "cierre_trimestral_state.json cambió después de la preparación."
        )

    respaldo_estado = asegurar_dentro(
        Path(str(manifest.get("respaldo_estado_activo") or "")),
        carpeta_periodo,
        "Respaldo del estado",
    )
    respaldo_activo = asegurar_dentro(
        Path(str(manifest.get("respaldo_excel_activo") or "")),
        carpeta_periodo,
        "Respaldo del Excel activo",
    )

    if sha256_file(respaldo_estado) != str(
        manifest.get("respaldo_estado_activo_sha256") or ""
    ):
        raise RuntimeError("El respaldo del estado no coincide con el manifest.")
    if sha256_file(respaldo_activo) != str(
        manifest.get("respaldo_excel_activo_sha256") or ""
    ):
        raise RuntimeError("El respaldo del Excel activo no coincide con el manifest.")
    if sha256_file(respaldo_activo) != excel_activo_hash:
        raise RuntimeError("El respaldo del Excel activo no coincide con el activo.")

    validacion_path = asegurar_dentro(
        Path(str(preparacion.get("validacion_remota_requerida") or "")),
        carpeta_periodo,
        "Validación remota",
    )

    return {
        "ya_finalizado": False,
        "estado_preparacion_path": estado_path,
        "preparacion": preparacion,
        "manifest_path": manifest_path,
        "manifest": manifest,
        "carpeta_periodo": carpeta_periodo,
        "carpeta_excel": carpeta_excel,
        "carpeta_soportes": carpeta_soportes,
        "periodo": periodo,
        "fecha_fin": fecha_fin,
        "excel_activo_hash": excel_activo_hash,
        "info_activo": info_activo,
        "historico": historico,
        "candidato": candidato,
        "info_candidato": info_candidato,
        "estado_actual": estado_actual,
        "validacion_path": validacion_path,
        "respaldo_estado_preparacion": respaldo_estado,
        "respaldo_activo_preparacion": respaldo_activo,
    }


def validar_evidencia_local(contexto: dict, drive_secundario: str) -> dict:
    path = contexto["validacion_path"]
    evidencia = leer_json(path, "validación remota trimestral")

    if evidencia.get("tipo") != "VALIDACION_REMOTA_CIERRE_TRIMESTRAL":
        raise RuntimeError("El archivo no es una validación remota trimestral.")
    if evidencia.get("ok") is not True:
        raise RuntimeError("La validación remota no tiene ok=true.")
    if evidencia.get("estado") != (
        "ARCHIVOS_HISTORICOS_VALIDADOS_EN_AMBOS_DESTINOS"
    ):
        raise RuntimeError(
            "Estado remoto inesperado: " f"{evidencia.get('estado')}"
        )

    principal = evidencia.get("principal")
    secundaria = evidencia.get("secundaria")
    archivos = evidencia.get("archivos_historicos")

    if not isinstance(principal, dict) or principal.get("ok") is not True:
        raise RuntimeError("La evidencia principal no está validada.")
    if not isinstance(secundaria, dict) or secundaria.get("ok") is not True:
        raise RuntimeError("La evidencia secundaria no está validada.")
    if not isinstance(archivos, list) or not archivos:
        raise RuntimeError("La evidencia no contiene archivos históricos.")

    total = int(evidencia.get("total_archivos_historicos") or 0)
    if total != len(archivos):
        raise RuntimeError(
            "El total de archivos de la evidencia no coincide con el detalle."
        )
    if int(principal.get("archivos_verificados") or 0) != total:
        raise RuntimeError("Conteo principal incompleto en la evidencia.")
    if int(secundaria.get("archivos_verificados") or 0) != total:
        raise RuntimeError("Conteo secundario incompleto en la evidencia.")

    nombre_periodo = contexto["carpeta_periodo"].name
    anio = contexto["carpeta_periodo"].parent.name

    if str(evidencia.get("periodo") or "") != nombre_periodo:
        raise RuntimeError(
            "El periodo de la evidencia no coincide con la carpeta local."
        )
    if str(evidencia.get("anio") or "") != anio:
        raise RuntimeError("El año de la evidencia no coincide con la carpeta local.")

    esperado_principal = (
        f"{BASE_SP}/Backups/03_Cierres_Trimestrales/"
        f"{anio}/{nombre_periodo}"
    ).strip("/")
    esperado_secundario = (
        f"{SP_BACKUP2_FOLDER}/03_Cierres_Trimestrales/"
        f"{anio}/{nombre_periodo}"
    ).strip("/")

    if str(principal.get("drive_id") or "") != SP_DRIVE_ID:
        raise RuntimeError("El drive principal de la evidencia no coincide.")
    if str(secundaria.get("drive_id") or "") != drive_secundario:
        raise RuntimeError("El drive secundario de la evidencia no coincide.")
    if str(principal.get("remote_base") or "").strip("/") != esperado_principal:
        raise RuntimeError("La ruta principal de la evidencia no coincide.")
    if str(secundaria.get("remote_base") or "").strip("/") != esperado_secundario:
        raise RuntimeError("La ruta secundaria de la evidencia no coincide.")

    for registro in archivos:
        if not isinstance(registro, dict):
            raise RuntimeError("Registro inválido en archivos_historicos.")

        rel = ruta_relativa_segura(str(registro.get("ruta_relativa") or ""))
        local = asegurar_dentro(
            contexto["carpeta_periodo"] / PurePosixPath(rel),
            contexto["carpeta_periodo"],
            "Archivo histórico",
        )

        if not local.exists() or not local.is_file():
            raise RuntimeError(f"Falta archivo histórico local: {rel}")

        if local.stat().st_size != int(registro.get("bytes") or -1):
            raise RuntimeError(f"Tamaño local distinto para: {rel}")

        if sha256_file(local) != str(
            registro.get("sha256_archivo_local") or ""
        ):
            raise RuntimeError(f"SHA256 local distinto para: {rel}")

        if local.suffix.lower() in EXCEL_EXT:
            digest, _ = digest_datos_excel(local)
            if digest != str(registro.get("digest_datos_excel") or ""):
                raise RuntimeError(f"Digest Excel local distinto para: {rel}")

    evidencia["_path"] = str(path)
    evidencia["_sha256"] = sha256_file(path)
    evidencia["_esperado_principal"] = esperado_principal
    evidencia["_esperado_secundario"] = esperado_secundario
    return evidencia


def verificar_evidencia_y_archivos_remotos(
    contexto: dict,
    evidencia: dict,
    drive_secundario: str,
) -> dict:
    path_evidencia = Path(evidencia["_path"])
    rel_evidencia = path_evidencia.relative_to(
        contexto["carpeta_periodo"]
    ).as_posix()

    destinos = [
        {
            "nombre": "PRINCIPAL_CONTABILIDAD",
            "drive_id": SP_DRIVE_ID,
            "remote_base": evidencia["_esperado_principal"],
        },
        {
            "nombre": "SECUNDARIA_CONTROL_INTERNO",
            "drive_id": drive_secundario,
            "remote_base": evidencia["_esperado_secundario"],
        },
    ]

    resultado_destinos = []
    for destino in destinos:
        remote_evidencia = (
            f"{destino['remote_base']}/{rel_evidencia}"
        ).strip("/")
        verificacion_evidencia = verificar_contra_local(
            local=path_evidencia,
            drive_id=destino["drive_id"],
            remote_path=remote_evidencia,
        )

        archivos_ok = []
        for registro in evidencia["archivos_historicos"]:
            rel = ruta_relativa_segura(
                str(registro.get("ruta_relativa") or "")
            )
            local = asegurar_dentro(
                contexto["carpeta_periodo"] / PurePosixPath(rel),
                contexto["carpeta_periodo"],
                "Archivo histórico",
            )
            remoto = f"{destino['remote_base']}/{rel}".strip("/")
            es_excel_controlado = local in {
                contexto["historico"],
                contexto["candidato"],
                contexto["respaldo_activo_preparacion"],
            }
            verificacion = verificar_contra_local(
                local=local,
                drive_id=destino["drive_id"],
                remote_path=remoto,
                exigir_excel_limpio=(local == contexto["candidato"]),
                tabla_excel_requerida=(
                    "TblFacturas" if es_excel_controlado else None
                ),
            )
            archivos_ok.append(
                {
                    "ruta_relativa": rel,
                    "remote_path": remoto,
                    "verificacion": verificacion,
                }
            )

        resultado_destinos.append(
            {
                "nombre": destino["nombre"],
                "drive_id_resumen": resumen_drive(destino["drive_id"]),
                "remote_base": destino["remote_base"],
                "evidencia": verificacion_evidencia,
                "archivos_verificados": len(archivos_ok),
                "detalle": archivos_ok,
            }
        )

    return {
        "ok": True,
        "total_archivos": len(evidencia["archivos_historicos"]),
        "destinos": resultado_destinos,
    }


def remote_excel_activo_path() -> str:
    return f"{BASE_SP}/excel/facturas.xlsx".strip("/")


def validar_excel_activo_remoto_con_local(contexto: dict) -> dict:
    destino = remote_excel_activo_path()
    item, remoto = descargar_path(SP_DRIVE_ID, destino)
    info_remoto, digest_remoto, resumen_remoto = validar_excel_bytes(
        remoto,
        exigir_limpio=False,
    )
    digest_local, resumen_local = digest_datos_excel(FACTURAS_PATH)

    temporal_remoto = excel_desde_bytes(
        remoto,
        "joyco_activo_sharepoint_semantico_",
    )
    try:
        comparacion = comparar_excel_facturas_semantico(
            FACTURAS_PATH,
            temporal_remoto,
        )
    finally:
        temporal_remoto.unlink(missing_ok=True)

    if not comparacion["ok"]:
        raise RuntimeError(
            "El Excel activo de SharePoint no coincide semánticamente "
            "con data/facturas.xlsx.\n"
            f"Filas local: {comparacion['filas_local']}\n"
            f"Filas remoto: {comparacion['filas_remoto']}\n"
            "Registros únicamente en local: "
            f"{comparacion['registros_solo_local']}\n"
            "Registros únicamente en SharePoint: "
            f"{comparacion['registros_solo_remoto']}\n"
            "No se reemplazará ninguno de los dos archivos."
        )

    return {
        "remote_path": destino,
        "item_id": item.get("id"),
        "bytes": len(remoto),
        "sha256_bytes_descargados": sha256_bytes(remoto),
        "digest_datos_excel": digest_remoto,
        "digest_datos_excel_local": digest_local,
        "digest_semantico": comparacion["digest_semantico_remoto"],
        "comparacion_semantica": comparacion,
        "resumen_datos_excel": resumen_remoto,
        "info_excel": info_remoto,
        "contenido": remoto,
        "resumen_local": resumen_local,
    }


def construir_nuevo_estado(contexto: dict, evidencia_final_path: Path) -> dict:
    siguiente = contexto["manifest"].get("siguiente_periodo_previsto")
    if not isinstance(siguiente, dict):
        raise RuntimeError("El manifest no contiene el siguiente periodo previsto.")

    requeridos = [
        "periodo_activo",
        "fecha_inicio_periodo_activo",
        "proximo_cierre_estimado",
    ]
    faltantes = [campo for campo in requeridos if not siguiente.get(campo)]
    if faltantes:
        raise RuntimeError(
            f"Siguiente periodo incompleto en manifest: {faltantes}"
        )

    nuevo = dict(contexto["estado_actual"])
    nuevo.update(
        {
            "ultimo_cierre_trimestral": contexto["fecha_fin"],
            "ultimo_periodo_cerrado": contexto["periodo"],
            "ultimo_archivo_generado": str(contexto["historico"]),
            "ultima_validacion_remota_trimestral": str(
                contexto["validacion_path"]
            ),
            "ultima_evidencia_finalizacion_trimestral": str(
                evidencia_final_path
            ),
            "periodo_activo": siguiente["periodo_activo"],
            "fecha_inicio_periodo_activo": siguiente[
                "fecha_inicio_periodo_activo"
            ],
            "proximo_cierre_estimado": siguiente[
                "proximo_cierre_estimado"
            ],
            "estado": "ACTIVO",
            "actualizado_en": datetime.datetime.now().isoformat(
                timespec="seconds"
            ),
            "version": (
                "2026-07-30-CIERRE-TRIMESTRAL-STATE-"
                "V3-FINALIZADO-SEGURO"
            ),
        }
    )
    return nuevo


def crear_resumen_finalizacion(datos: dict) -> str:
    lineas = [
        "FINALIZACIÓN SEGURA DEL CIERRE TRIMESTRAL",
        "=" * 80,
        f"Versión: {datos['version_script']}",
        f"Generado en: {datos['generado_en']}",
        f"Estado: {datos['estado']}",
        f"Resultado global: {datos['resultado_global']}",
        "",
        "Periodo cerrado:",
        f"- {datos['periodo']}",
        f"- Fecha de cierre: {datos['fecha_fin']}",
        "",
        "Histórico validado:",
        f"- Excel: {datos['excel_historico']}",
        f"- SHA256: {datos['excel_historico_sha256']}",
        "",
        "Excel activo después de finalizar:",
        f"- Local: {datos['excel_activo_local_despues']}",
        f"- SHA256: {datos['excel_activo_local_despues_sha256']}",
        f"- SharePoint: {datos['excel_activo_sharepoint_path']}",
        "",
        "Nuevo periodo activo:",
        f"- {datos['nuevo_estado']['periodo_activo']}",
        f"- Inicio: {datos['nuevo_estado']['fecha_inicio_periodo_activo']}",
        f"- Fin: {datos['nuevo_estado']['proximo_cierre_estimado']}",
        "",
        "Controles completados:",
        "- Histórico revalidado en ambos destinos.",
        "- Evidencia remota exacta verificada en ambos destinos.",
        "- Excel activo remoto anterior respaldado.",
        "- Excel activo de SharePoint reemplazado y verificado.",
        "- Excel activo local reemplazado de forma atómica y verificado.",
        "- Estado trimestral actualizado de forma atómica.",
        "- Evidencia final publicada y verificada en ambos destinos.",
        "=" * 80,
    ]
    return "\n".join(lineas) + "\n"


def publicar_artifactos_finales(
    *,
    contexto: dict,
    evidencia: dict,
    drive_secundario: str,
    archivos: Iterable[Path],
) -> list[dict]:
    destinos = [
        (
            "PRINCIPAL_CONTABILIDAD",
            SP_DRIVE_ID,
            evidencia["_esperado_principal"],
        ),
        (
            "SECUNDARIA_CONTROL_INTERNO",
            drive_secundario,
            evidencia["_esperado_secundario"],
        ),
    ]

    resultados = []
    for nombre, drive_id, remote_base in destinos:
        verificados = []
        for local in archivos:
            rel = local.relative_to(contexto["carpeta_periodo"]).as_posix()
            remote_path = f"{remote_base}/{rel}".strip("/")
            verificacion = subir_y_verificar(
                local=local,
                drive_id=drive_id,
                remote_path=remote_path,
                exigir_excel_limpio=False,
            )
            verificados.append(
                {
                    "ruta_relativa": rel,
                    "remote_path": remote_path,
                    "verificacion": verificacion,
                }
            )

        resultados.append(
            {
                "nombre": nombre,
                "drive_id_resumen": resumen_drive(drive_id),
                "remote_base": remote_base,
                "archivos_verificados": len(verificados),
                "detalle": verificados,
            }
        )

    return resultados


def restaurar_transaccion(
    *,
    contexto: dict,
    respaldo_local: Path,
    respaldo_estado: Path,
    respaldo_remoto: Path,
    remoto_modificado: bool,
    local_modificado: bool,
    estado_modificado: bool,
) -> dict:
    resultado = {
        "estado_restaurado": not estado_modificado,
        "excel_local_restaurado": not local_modificado,
        "excel_remoto_restaurado": not remoto_modificado,
        "errores": [],
    }

    if estado_modificado:
        try:
            copiar_atomico(respaldo_estado, STATE_PATH)
            if sha256_file(STATE_PATH) != sha256_file(respaldo_estado):
                raise RuntimeError("SHA del estado restaurado no coincide.")
            resultado["estado_restaurado"] = True
        except Exception as exc:
            resultado["errores"].append(f"Estado: {exc}")

    if local_modificado:
        try:
            copiar_atomico(respaldo_local, FACTURAS_PATH)
            if sha256_file(FACTURAS_PATH) != sha256_file(respaldo_local):
                raise RuntimeError("SHA del Excel local restaurado no coincide.")
            resultado["excel_local_restaurado"] = True
        except Exception as exc:
            resultado["errores"].append(f"Excel local: {exc}")

    if remoto_modificado:
        try:
            subir_y_verificar(
                local=respaldo_remoto,
                drive_id=SP_DRIVE_ID,
                remote_path=remote_excel_activo_path(),
                exigir_excel_limpio=False,
            )
            resultado["excel_remoto_restaurado"] = True
        except Exception as exc:
            resultado["errores"].append(f"Excel remoto: {exc}")

    resultado["ok"] = bool(
        resultado["estado_restaurado"]
        and resultado["excel_local_restaurado"]
        and resultado["excel_remoto_restaurado"]
    )
    return resultado


def ejecutar_diagnostico() -> int:
    validar_configuracion()

    principal = validar_drive_id(SP_DRIVE_ID)
    if not principal:
        raise RuntimeError("El drive principal no es válido o no es accesible.")
    drive_secundario = resolver_drive_secundario()

    contexto = validar_preparacion_local()
    if contexto["ya_finalizado"]:
        finalizacion = contexto["finalizacion"]
        print("✅ El cierre trimestral ya se encuentra finalizado.")
        print(f"Evidencia: {finalizacion.get('_path')}")
        print(f"Periodo: {finalizacion.get('periodo')}")
        return 0

    evidencia = validar_evidencia_local(contexto, drive_secundario)
    revalidacion = verificar_evidencia_y_archivos_remotos(
        contexto,
        evidencia,
        drive_secundario,
    )
    activo_remoto = validar_excel_activo_remoto_con_local(contexto)

    print("-" * 100)
    print("✅ DIAGNÓSTICO DE FINALIZACIÓN COMPLETADO.")
    print(f"Versión: {VERSION}")
    print(f"Periodo: {contexto['periodo']}")
    print(f"Fecha de cierre: {contexto['fecha_fin']}")
    print(
        "Excel activo local: "
        f"filas_datos={contexto['info_activo']['filas_datos']} | "
        f"sha256={contexto['info_activo']['sha256']}"
    )
    print(f"Excel limpio candidato: {contexto['candidato']}")
    print(
        "Archivos históricos revalidados por destino: "
        f"{revalidacion['total_archivos']}"
    )
    print(
        "Drive principal: "
        f"{principal.get('name')} | {resumen_drive(SP_DRIVE_ID)}"
    )
    print(
        "Drive secundario: "
        f"{resumen_drive(drive_secundario)}"
    )
    print(
        "Excel activo SharePoint coincide con local: "
        f"digest={activo_remoto['digest_datos_excel']}"
    )
    print("✅ No se reemplazó ni se subió ningún archivo.")
    print("=" * 100)
    return 0


def ejecutar_finalizacion() -> int:
    validar_configuracion()

    with ExitStack() as stack:
        stack.enter_context(LockAtomico(LOCK_FINALIZACION_PATH))
        stack.enter_context(LockAtomico(LOCK_APROBADAS_PATH))

        principal = validar_drive_id(SP_DRIVE_ID)
        if not principal:
            raise RuntimeError(
                "El drive principal no es válido o no es accesible."
            )
        drive_secundario = resolver_drive_secundario()

        # Repetir todas las validaciones después de adquirir los locks.
        contexto = validar_preparacion_local()
        if contexto["ya_finalizado"]:
            print("✅ El cierre ya estaba finalizado. No se modificó nada.")
            return 0

        evidencia = validar_evidencia_local(contexto, drive_secundario)
        revalidacion = verificar_evidencia_y_archivos_remotos(
            contexto,
            evidencia,
            drive_secundario,
        )
        activo_remoto = validar_excel_activo_remoto_con_local(contexto)

        stamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        nombre_seguro = contexto["carpeta_periodo"].name
        soportes = contexto["carpeta_soportes"]

        respaldo_local = soportes / (
            "facturas_ACTIVO_LOCAL_ANTES_FINALIZACION_"
            f"{nombre_seguro}_{stamp}.xlsx"
        )
        respaldo_estado = soportes / (
            "cierre_trimestral_state_ANTES_FINALIZACION_"
            f"{nombre_seguro}_{stamp}.json"
        )
        respaldo_remoto = soportes / (
            "facturas_ACTIVO_SHAREPOINT_ANTES_FINALIZACION_"
            f"{nombre_seguro}_{stamp}.xlsx"
        )
        estado_despues_path = soportes / (
            "cierre_trimestral_state_DESPUES_FINALIZACION_"
            f"{nombre_seguro}_{stamp}.json"
        )
        evidencia_final_path = soportes / (
            f"{PREFIJO_FINALIZACION}{nombre_seguro}.json"
        )
        resumen_final_path = soportes / (
            f"{PREFIJO_RESUMEN_FINAL}{nombre_seguro}.txt"
        )

        shutil.copy2(FACTURAS_PATH, respaldo_local)
        shutil.copy2(STATE_PATH, respaldo_estado)
        respaldo_remoto.write_bytes(activo_remoto["contenido"])

        if sha256_file(respaldo_local) != contexto["excel_activo_hash"]:
            raise RuntimeError("El respaldo local previo no coincide.")
        validar_excel_estructura(respaldo_remoto, exigir_limpio=False)
        digest_respaldo_remoto, _ = digest_datos_excel(respaldo_remoto)
        if digest_respaldo_remoto != activo_remoto["digest_datos_excel"]:
            raise RuntimeError("El respaldo remoto previo no coincide.")

        nuevo_estado = construir_nuevo_estado(
            contexto,
            evidencia_final_path,
        )

        remoto_modificado = False
        local_modificado = False
        estado_modificado = False

        try:
            print("☁️ Reemplazando Excel activo de SharePoint...")
            subir_y_verificar(
                local=contexto["candidato"],
                drive_id=SP_DRIVE_ID,
                remote_path=remote_excel_activo_path(),
                exigir_excel_limpio=True,
            )
            remoto_modificado = True

            print("💾 Reemplazando data/facturas.xlsx de forma atómica...")
            copiar_atomico(contexto["candidato"], FACTURAS_PATH)
            info_local_nuevo = validar_excel_estructura(
                FACTURAS_PATH,
                exigir_limpio=True,
            )
            if info_local_nuevo["sha256"] != sha256_file(contexto["candidato"]):
                raise RuntimeError("El nuevo Excel activo local no coincide.")
            local_modificado = True

            print("🧾 Actualizando el estado trimestral de forma atómica...")
            escribir_json_atomico(STATE_PATH, nuevo_estado)
            if leer_json(STATE_PATH, "nuevo estado trimestral") != nuevo_estado:
                raise RuntimeError("El estado trimestral actualizado no coincide.")
            estado_modificado = True

            # Validación final antes de emitir evidencia de éxito.
            verificar_contra_local(
                local=FACTURAS_PATH,
                drive_id=SP_DRIVE_ID,
                remote_path=remote_excel_activo_path(),
                exigir_excel_limpio=True,
            )
            if sha256_file(FACTURAS_PATH) != sha256_file(contexto["candidato"]):
                raise RuntimeError("El Excel local cambió después del reemplazo.")
            if leer_json(STATE_PATH, "estado final") != nuevo_estado:
                raise RuntimeError("El estado cambió después de actualizarlo.")

            escribir_json_atomico(estado_despues_path, nuevo_estado)

            datos_finales = {
                "tipo": "ESTADO_FINALIZACION_CIERRE_TRIMESTRAL",
                "version_script": VERSION,
                "generado_en": datetime.datetime.now().isoformat(
                    timespec="seconds"
                ),
                "ok": True,
                "estado": ESTADO_FINALIZADO,
                "resultado_global": "OK",
                "periodo": contexto["periodo"],
                "fecha_fin": contexto["fecha_fin"],
                "carpeta_periodo": str(contexto["carpeta_periodo"]),
                "excel_historico": str(contexto["historico"]),
                "excel_historico_sha256": sha256_file(
                    contexto["historico"]
                ),
                "validacion_remota_historica": str(
                    contexto["validacion_path"]
                ),
                "validacion_remota_historica_sha256": sha256_file(
                    contexto["validacion_path"]
                ),
                "archivos_historicos_revalidados_por_destino": (
                    revalidacion["total_archivos"]
                ),
                "destinos_historicos": [
                    {
                        "nombre": destino["nombre"],
                        "drive_id_resumen": destino["drive_id_resumen"],
                        "remote_base": destino["remote_base"],
                        "archivos_verificados": destino[
                            "archivos_verificados"
                        ],
                    }
                    for destino in revalidacion["destinos"]
                ],
                "excel_activo_local_antes": str(respaldo_local),
                "excel_activo_local_antes_sha256": sha256_file(
                    respaldo_local
                ),
                "excel_activo_sharepoint_antes": str(respaldo_remoto),
                "excel_activo_sharepoint_antes_sha256": sha256_file(
                    respaldo_remoto
                ),
                "excel_activo_sharepoint_antes_digest": (
                    activo_remoto["digest_datos_excel"]
                ),
                "estado_antes": str(respaldo_estado),
                "estado_antes_sha256": sha256_file(respaldo_estado),
                "excel_limpio_candidato": str(contexto["candidato"]),
                "excel_limpio_candidato_sha256": sha256_file(
                    contexto["candidato"]
                ),
                "excel_activo_local_despues": str(FACTURAS_PATH),
                "excel_activo_local_despues_sha256": sha256_file(
                    FACTURAS_PATH
                ),
                "excel_activo_sharepoint_path": remote_excel_activo_path(),
                "estado_despues": str(estado_despues_path),
                "estado_despues_sha256": sha256_file(estado_despues_path),
                "nuevo_estado": nuevo_estado,
                "rollback_requerido": False,
                "controles": {
                    "locks_exclusivos_adquiridos": True,
                    "excel_activo_original_sin_cambios_desde_preparacion": True,
                    "evidencia_remota_exacta_en_ambos_destinos": True,
                    "archivos_historicos_revalidados_en_ambos_destinos": True,
                    "excel_activo_remoto_antes_coincidia_con_local": True,
                    "excel_activo_sharepoint_reemplazado_y_verificado": True,
                    "excel_activo_local_reemplazado_atomicamente": True,
                    "estado_actualizado_atomicamente": True,
                },
            }

            escribir_texto_atomico(
                resumen_final_path,
                crear_resumen_finalizacion(datos_finales),
            )

            # Subir primero respaldos y soportes. La evidencia final se
            # genera y publica al final para que su presencia remota
            # signifique que todos los pasos anteriores terminaron bien.
            artefactos_previos = [
                respaldo_local,
                respaldo_estado,
                respaldo_remoto,
                estado_despues_path,
                resumen_final_path,
            ]
            publicacion_soportes = publicar_artifactos_finales(
                contexto=contexto,
                evidencia=evidencia,
                drive_secundario=drive_secundario,
                archivos=artefactos_previos,
            )

            datos_finales["publicacion_soportes_finales"] = (
                publicacion_soportes
            )
            datos_finales["evidencia_final_publicada_como_ultimo_paso"] = True
            escribir_json_atomico(evidencia_final_path, datos_finales)

            publicar_artifactos_finales(
                contexto=contexto,
                evidencia=evidencia,
                drive_secundario=drive_secundario,
                archivos=[evidencia_final_path],
            )

            print("-" * 100)
            print("✅ CIERRE TRIMESTRAL FINALIZADO CORRECTAMENTE.")
            print(f"Periodo cerrado: {contexto['periodo']}")
            print(f"Excel histórico: {contexto['historico']}")
            print(f"Excel activo local limpio: {FACTURAS_PATH}")
            print(f"Excel activo SharePoint: {remote_excel_activo_path()}")
            print(
                "Nuevo periodo activo: "
                f"{nuevo_estado['periodo_activo']}"
            )
            print(f"Evidencia final: {evidencia_final_path}")
            print(
                "✅ Evidencia final publicada y verificada en ambos destinos."
            )
            print("=" * 100)
            return 0

        except Exception as exc:
            print("❌ Error durante la finalización. Iniciando rollback...")
            rollback = restaurar_transaccion(
                contexto=contexto,
                respaldo_local=respaldo_local,
                respaldo_estado=respaldo_estado,
                respaldo_remoto=respaldo_remoto,
                remoto_modificado=remoto_modificado,
                local_modificado=local_modificado,
                estado_modificado=estado_modificado,
            )

            datos_fallo = {
                "tipo": "ESTADO_FINALIZACION_CIERRE_TRIMESTRAL",
                "version_script": VERSION,
                "generado_en": datetime.datetime.now().isoformat(
                    timespec="seconds"
                ),
                "ok": False,
                "estado": (
                    "REVERTIDO" if rollback["ok"] else "ROLLBACK_INCOMPLETO"
                ),
                "resultado_global": "ERROR",
                "periodo": contexto["periodo"],
                "fecha_fin": contexto["fecha_fin"],
                "error": f"{type(exc).__name__}: {exc}",
                "rollback": rollback,
                "rollback_requerido": True,
            }
            escribir_json_atomico(evidencia_final_path, datos_fallo)

            # Intento de reemplazar cualquier evidencia de éxito parcial por
            # la evidencia de rollback. No oculta un fallo si Graph no responde.
            try:
                publicar_artifactos_finales(
                    contexto=contexto,
                    evidencia=evidencia,
                    drive_secundario=drive_secundario,
                    archivos=[evidencia_final_path],
                )
            except Exception as publicacion_exc:
                rollback["errores"].append(
                    "No se pudo publicar evidencia de rollback: "
                    f"{publicacion_exc}"
                )

            print(f"Error original: {type(exc).__name__}: {exc}")
            print(f"Rollback completo: {rollback['ok']}")
            for error in rollback["errores"]:
                print(f"⚠️ {error}")
            print(f"Evidencia local del fallo: {evidencia_final_path}")
            print("=" * 100)
            return 1 if rollback["ok"] else 2


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Valida local y remoto sin reemplazar ni subir archivos.",
    )
    parser.add_argument(
        "--finalizar",
        action="store_true",
        help="Finaliza el cierre trimestral de forma transaccional.",
    )
    parser.add_argument(
        "--confirmar",
        default=None,
        help="Confirmación obligatoria para --finalizar.",
    )
    parser.add_argument(
        "--upload-activo",
        action="store_true",
        help=argparse.SUPPRESS,
    )
    args = parser.parse_args()

    if args.upload_activo:
        print("❌ El comando antiguo --upload-activo está bloqueado.")
        print(
            "Debe usarse el flujo seguro completo con --finalizar "
            f"--confirmar {CONFIRMACION}."
        )
        return 1

    if args.dry_run and args.finalizar:
        print("❌ Usa solo un modo: --dry-run o --finalizar.")
        return 1

    modo = "FINALIZAR" if args.finalizar else "DRY RUN"

    print("=" * 100)
    print(f"FINALIZACIÓN CIERRE TRIMESTRAL - {modo}")
    print("=" * 100)
    print(f"Versión: {VERSION}")
    print(f"Root: {ROOT}")
    print("-" * 100)

    try:
        if args.finalizar:
            if args.confirmar != CONFIRMACION:
                print("❌ Finalización bloqueada por falta de confirmación.")
                print("Comando requerido:")
                print(
                    "python scripts\\"
                    "reemplazar_excel_activo_trimestral_sharepoint.py "
                    f"--finalizar --confirmar {CONFIRMACION}"
                )
                return 1
            return ejecutar_finalizacion()

        return ejecutar_diagnostico()

    except Exception as exc:
        print(
            "❌ Error en la finalización trimestral: "
            f"{type(exc).__name__}: {exc}"
        )
        print("No se debe continuar hasta revisar el error.")
        print("=" * 100)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
