# -*- coding: utf-8 -*-
"""
JOYCO - Facturas Procesador
Plan de corte inicial pre-VPS.

Este script NO modifica archivos y NO llama Microsoft Graph.
Sirve para validar la lógica, calcular rutas y dejar trazabilidad del plan
antes de crear el script real de corte inicial pre-VPS.

Uso:
  python scripts/plan_corte_inicial_pre_vps_facturas.py --fecha-corte 2026-07-15
"""

from __future__ import annotations

import argparse
import datetime as dt
import json
import os
import sys
from pathlib import Path

from dotenv import load_dotenv
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

try:
    from services.m365.sp_graph import SP_FOLDER as BASE_SP_IMPORTADO
except Exception:
    BASE_SP_IMPORTADO = ""

VERSION = "2026-07-10-PLAN-CORTE-INICIAL-PRE-VPS-V1"

DATA_DIR = ROOT / "data"
FACTURAS_PATH = DATA_DIR / "facturas.xlsx"
STATE_PATH = DATA_DIR / "state" / "cierre_trimestral_state.json"
CIERRES_TRIMESTRALES_DIR = DATA_DIR / "cierres_trimestrales"

HEADERS_ESPERADOS = [
    "Radicado",
    "ProyectoProceso",
    "Archivo",
    "Empresa emisora",
    "CUFE",
    "Ciudad emisora",
    "Código ciudad",
    "NIT",
    "Cliente",
    "Número de factura",
    "Año",
    "Mes",
    "Día",
    "Tipo de contribuyente",
    "Actividad económica",
    "DESCRIPCIÓN",
    "Concepto",
    "VALOR",
    "Estado_calidad",
]

load_dotenv(ROOT / ".env")

BASE_SP = (BASE_SP_IMPORTADO or os.getenv("SP_FOLDER") or "").strip().strip("/")
SP_DRIVE_ID = (os.getenv("SP_DRIVE_ID") or "").strip()
BACKUP_DRIVE_ID = (
    os.getenv("BACKUP_DRIVE_ID")
    or os.getenv("ONEDRIVE_BACKUP_DRIVE_ID")
    or os.getenv("SP_BACKUP2_DRIVE_ID")
    or ""
).strip()
BACKUP_ROOT_FOLDER = (
    os.getenv("BACKUP_ROOT_FOLDER")
    or os.getenv("ONEDRIVE_BACKUP_FOLDER")
    or os.getenv("SP_BACKUP2_FOLDER")
    or ""
).strip().strip("/")


def parse_fecha(valor: str, campo: str) -> dt.date:
    try:
        return dt.date.fromisoformat(valor)
    except Exception as exc:
        raise RuntimeError(f"Fecha inválida para {campo}: {valor}") from exc


def ultimo_dia_mes(anio: int, mes: int) -> dt.date:
    if mes == 12:
        return dt.date(anio, 12, 31)
    return dt.date(anio, mes + 1, 1) - dt.timedelta(days=1)


def calcular_trimestre(fecha: dt.date) -> dict:
    trimestre_num = ((fecha.month - 1) // 3) + 1
    mes_inicio = ((trimestre_num - 1) * 3) + 1
    mes_fin = mes_inicio + 2
    inicio_calendario = dt.date(fecha.year, mes_inicio, 1)
    fin_calendario = ultimo_dia_mes(fecha.year, mes_fin)
    return {
        "periodo_activo": f"{fecha.year}-T{trimestre_num}",
        "trimestre": f"T{trimestre_num}",
        "anio": str(fecha.year),
        "fecha_inicio_calendario_trimestre": inicio_calendario.isoformat(),
        "proximo_cierre_estimado": fin_calendario.isoformat(),
    }


def cargar_estado() -> dict:
    if not STATE_PATH.exists():
        raise RuntimeError(f"No existe state trimestral: {STATE_PATH}")
    with STATE_PATH.open("r", encoding="utf-8-sig") as f:
        estado = json.load(f)
    requeridos = ["periodo_activo", "fecha_inicio_periodo_activo", "proximo_cierre_estimado", "estado"]
    faltantes = [k for k in requeridos if not estado.get(k)]
    if faltantes:
        raise RuntimeError(f"State incompleto. Faltan: {faltantes}")
    return estado


def validar_excel_local() -> dict:
    if not FACTURAS_PATH.exists():
        raise RuntimeError(f"No existe Excel local: {FACTURAS_PATH}")
    wb = load_workbook(FACTURAS_PATH, read_only=False, data_only=False)
    try:
        if "Facturas" not in wb.sheetnames:
            raise RuntimeError("El Excel no tiene hoja Facturas")
        ws = wb["Facturas"]
        headers = [cell.value for cell in ws[1]]
        if headers != HEADERS_ESPERADOS:
            raise RuntimeError(
                "Los encabezados no coinciden con la estructura esperada.\n"
                f"Detectados: {headers}\n"
                f"Esperados: {HEADERS_ESPERADOS}"
            )
        tablas = list(ws.tables.keys())
        tbl_ref = ws.tables["TblFacturas"].ref if "TblFacturas" in ws.tables else None
        return {
            "archivo": str(FACTURAS_PATH),
            "hojas": wb.sheetnames,
            "filas": ws.max_row,
            "filas_datos": max(ws.max_row - 1, 0),
            "columnas": ws.max_column,
            "tablas": tablas,
            "tbl_facturas_ref": tbl_ref,
            "esta_limpio": ws.max_row == 1 and ws.max_column == 19 and tbl_ref == "A1:S1",
        }
    finally:
        wb.close()


def construir_plan(fecha_corte: dt.date, estado: dict) -> dict:
    fecha_corte_s = fecha_corte.isoformat()
    fecha_inicio_base = str(estado.get("fecha_inicio_periodo_activo") or fecha_corte_s)
    tri = calcular_trimestre(fecha_corte)

    carpeta_base = f"BASE_INICIAL_PRE_VPS_{fecha_inicio_base}_a_{fecha_corte_s}"
    historico_excel = f"facturas_base_inicial_pre_vps_{fecha_inicio_base}_a_{fecha_corte_s}.xlsx"

    local_base = CIERRES_TRIMESTRALES_DIR / tri["anio"] / carpeta_base / "Corte_Inicial_Pre_VPS"

    sp_activo = f"{BASE_SP}/excel/facturas.xlsx".strip("/")
    sp_historicas_folder = f"{BASE_SP}/excel/Historicas/{carpeta_base}".strip("/")
    sp_historico = f"{sp_historicas_folder}/{historico_excel}".strip("/")

    backup_remote_base = f"{BACKUP_ROOT_FOLDER}/{tri['anio']}/{carpeta_base}/Corte_Inicial_Pre_VPS".strip("/")

    nuevo_state = dict(estado)
    nuevo_state.update({
        "version": "2026-07-10-CORTE-INICIAL-PRE-VPS-STATE-V1",
        "periodo_activo": tri["periodo_activo"],
        "fecha_inicio_periodo_activo": fecha_corte_s,
        "fecha_inicio_calendario_trimestre": tri["fecha_inicio_calendario_trimestre"],
        "proximo_cierre_estimado": tri["proximo_cierre_estimado"],
        "ultimo_corte_inicial_pre_vps": fecha_corte_s,
        "ultimo_archivo_generado": sp_historico,
        "modo_arranque": "CORTE_INICIAL_PRE_VPS",
        "estado": "ACTIVO",
        "nota": (
            "El Excel activo limpio inicia control operativo desde fecha_inicio_periodo_activo. "
            "El trimestre se controla por ciclo de vida del Excel activo, no por fecha de emisión de factura."
        ),
    })

    return {
        "version": VERSION,
        "fecha_corte": fecha_corte_s,
        "fecha_inicio_base": fecha_inicio_base,
        "carpeta_base": carpeta_base,
        "historico_excel": historico_excel,
        "trimestre": tri,
        "local_base": str(local_base),
        "local_estructura": {
            "01_Excel_Historico": str(local_base / "01_Excel_Historico"),
            "02_Excel_Activo_Limpio": str(local_base / "02_Excel_Activo_Limpio"),
            "03_State": str(local_base / "03_State"),
            "04_Manifest": str(local_base / "04_Manifest"),
            "05_Validaciones": str(local_base / "05_Validaciones"),
        },
        "sharepoint_principal": {
            "excel_activo_actual_y_futuro": sp_activo,
            "carpeta_historicas": sp_historicas_folder,
            "excel_historico_base_inicial": sp_historico,
        },
        "repositorio_backups": {
            "drive_id_detectado": BACKUP_DRIVE_ID,
            "ruta_base_remota": backup_remote_base,
        },
        "nuevo_state_preview": nuevo_state,
    }


def validar_config_minima() -> list[str]:
    advertencias = []
    if not SP_DRIVE_ID:
        advertencias.append("Falta SP_DRIVE_ID en .env")
    if not BASE_SP:
        advertencias.append("Falta SP_FOLDER en .env o services.m365.sp_graph")
    if not BACKUP_DRIVE_ID:
        advertencias.append("Falta BACKUP_DRIVE_ID / ONEDRIVE_BACKUP_DRIVE_ID / SP_BACKUP2_DRIVE_ID")
    if not BACKUP_ROOT_FOLDER:
        advertencias.append("Falta BACKUP_ROOT_FOLDER / ONEDRIVE_BACKUP_FOLDER / SP_BACKUP2_FOLDER")
    return advertencias


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--fecha-corte", default=dt.date.today().isoformat(), help="Fecha real estimada del corte inicial, formato YYYY-MM-DD")
    args = parser.parse_args()

    fecha_corte = parse_fecha(args.fecha_corte, "fecha-corte")
    estado = cargar_estado()
    info_excel = validar_excel_local()
    advertencias = validar_config_minima()
    plan = construir_plan(fecha_corte, estado)

    print("=" * 100)
    print("PLAN CORTE INICIAL PRE-VPS - FACTURAS JOYCO")
    print("=" * 100)
    print(f"Version: {VERSION}")
    print(f"Root: {ROOT}")
    print(f"Fecha corte: {plan['fecha_corte']}")
    print(f"Carpeta base: {plan['carpeta_base']}")
    print("-" * 100)
    print("Estado actual:")
    print(json.dumps(estado, ensure_ascii=False, indent=2))
    print("-" * 100)
    print("Excel local actual:")
    print(json.dumps(info_excel, ensure_ascii=False, indent=2))
    print("-" * 100)
    print("SharePoint principal limpio para contabilidad:")
    print(f"Excel activo: {plan['sharepoint_principal']['excel_activo_actual_y_futuro']}")
    print(f"Histórico:    {plan['sharepoint_principal']['excel_historico_base_inicial']}")
    print("-" * 100)
    print("Repositorio técnico de backups:")
    print(plan["repositorio_backups"]["ruta_base_remota"])
    print("-" * 100)
    print("Estructura local de evidencia técnica:")
    for nombre, ruta in plan["local_estructura"].items():
        print(f"{nombre}: {ruta}")
    print("-" * 100)
    print("Nuevo state previsto:")
    print(json.dumps(plan["nuevo_state_preview"], ensure_ascii=False, indent=2))
    print("-" * 100)
    print("Regla confirmada:")
    print("- El Excel completo representa el periodo operativo.")
    print("- NO se filtra por fecha de emisión de factura.")
    print("- Primero se guarda/verifica histórico y después se crea/reemplaza el Excel limpio.")
    print("- Este script solo diagnostica; no modifica nada y no llama Graph.")

    if advertencias:
        print("-" * 100)
        print("Advertencias de configuración:")
        for w in advertencias:
            print(f"  - {w}")

    print("=" * 100)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
