# -*- coding: utf-8 -*-
"""
JOYCO - Facturas Procesador
Corte inicial PRE-VPS de facturas.

Objetivo:
- Preparar el arranque limpio del proceso en VPS.
- Guardar el Excel acumulado actual como histórico/base inicial.
- Generar un Excel activo limpio, conservando estructura base.
- Preparar el nuevo state del trimestre operativo.
- Generar manifest, resumen y validación local.

IMPORTANTE:
- Este script NO llama Microsoft Graph.
- Este script NO toca SharePoint/OneDrive.
- En modo --local NO reemplaza data/facturas.xlsx ni actualiza state.
- El modo real queda bloqueado para una fase posterior.

Uso seguro:
  python scripts/corte_inicial_pre_vps_facturas.py --fecha-corte 2026-07-15 --dry-run
  python scripts/corte_inicial_pre_vps_facturas.py --fecha-corte 2026-07-15 --local
"""

from __future__ import annotations

import argparse
import hashlib
import json
import shutil
import sys
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "data"
FACTURAS_PATH = DATA_DIR / "facturas.xlsx"
STATE_PATH = DATA_DIR / "state" / "cierre_trimestral_state.json"
CIERRES_TRIMESTRALES_DIR = DATA_DIR / "cierres_trimestrales"

VERSION = "2026-07-10-CORTE-INICIAL-PRE-VPS-LOCAL-V1"

HEADERS_ESPERADOS = [
    "Radicado",
    "ProyectoProceso",
    "Archivo",
    "Empresa emisora",
    "CUFE",
    "Ciudad emisora",
    "Código ciudad",
    "NIT",
    "Cliente",
    "Número de factura",
    "Año",
    "Mes",
    "Día",
    "Tipo de contribuyente",
    "Actividad económica",
    "DESCRIPCIÓN",
    "Concepto",
    "VALOR",
    "Estado_calidad",
]


def parse_fecha(valor: str, campo: str) -> date:
    try:
        return datetime.strptime(str(valor).strip(), "%Y-%m-%d").date()
    except Exception as exc:
        raise RuntimeError(f"Fecha inválida en {campo}: {valor!r}. Formato esperado: YYYY-MM-DD") from exc


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def rel(path: Path) -> str:
    try:
        return path.relative_to(ROOT).as_posix()
    except Exception:
        return str(path)


def cargar_json(path: Path) -> dict:
    if not path.exists():
        raise RuntimeError(f"No existe archivo requerido: {path}")
    return json.loads(path.read_text(encoding="utf-8-sig"))


def guardar_json(path: Path, payload: dict) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def trimestre_calendario(fecha: date) -> Tuple[str, date, date]:
    q = ((fecha.month - 1) // 3) + 1
    inicio_mes = 3 * (q - 1) + 1
    fin_mes = inicio_mes + 2
    inicio = date(fecha.year, inicio_mes, 1)

    if fin_mes == 12:
        fin = date(fecha.year, 12, 31)
    else:
        fin = date(fecha.year, fin_mes + 1, 1)
        fin = date.fromordinal(fin.toordinal() - 1)

    return f"{fecha.year}-T{q}", inicio, fin


def validar_estado_actual(estado: dict) -> None:
    requeridos = [
        "periodo_activo",
        "fecha_inicio_periodo_activo",
        "proximo_cierre_estimado",
        "estado",
    ]
    faltantes = [k for k in requeridos if not estado.get(k)]
    if faltantes:
        raise RuntimeError(f"Estado trimestral incompleto. Faltan: {faltantes}")

    if str(estado.get("estado")).upper() != "ACTIVO":
        raise RuntimeError(f"Estado trimestral no está ACTIVO: {estado.get('estado')}")


def diagnosticar_excel(path: Path, exigir_limpio: Optional[bool] = None) -> dict:
    if not path.exists():
        raise RuntimeError(f"No existe Excel requerido: {path}")

    wb = load_workbook(path, read_only=False, data_only=False)
    try:
        if "Facturas" not in wb.sheetnames:
            raise RuntimeError("El Excel no tiene la hoja requerida: Facturas")

        ws = wb["Facturas"]
        headers = [cell.value for cell in ws[1]]

        if headers != HEADERS_ESPERADOS:
            raise RuntimeError(
                "Los encabezados del Excel no coinciden con la estructura esperada.\n"
                f"Detectados: {headers}\n"
                f"Esperados:  {HEADERS_ESPERADOS}"
            )

        tablas = list(ws.tables.keys())
        tbl_ref = ws.tables["TblFacturas"].ref if "TblFacturas" in ws.tables else None
        esta_limpio = ws.max_row == 1 and ws.max_column == 19 and tbl_ref == "A1:S1"

        info = {
            "archivo": rel(path),
            "existe": True,
            "abre_ok": True,
            "hojas": wb.sheetnames,
            "hoja_principal": ws.title,
            "filas": ws.max_row,
            "filas_datos": max(ws.max_row - 1, 0),
            "columnas": ws.max_column,
            "tablas": tablas,
            "tbl_facturas_ref": tbl_ref,
            "esta_limpio": esta_limpio,
            "bytes": path.stat().st_size,
            "sha256": sha256_file(path),
        }

        if ws.max_column != 19:
            raise RuntimeError(f"Columnas inválidas. Detectadas: {ws.max_column}. Esperadas: 19.")

        if "TblFacturas" not in tablas:
            raise RuntimeError("No existe la tabla requerida TblFacturas.")

        if exigir_limpio is True and not esta_limpio:
            raise RuntimeError(f"El Excel debería estar limpio y no lo está: {info}")

        if exigir_limpio is False and esta_limpio:
            raise RuntimeError("El Excel original está limpio; no parece haber base acumulada para histórico.")

        return info
    finally:
        wb.close()


def crear_excel_limpio_desde_original(origen: Path, destino: Path) -> None:
    destino.parent.mkdir(parents=True, exist_ok=True)

    if destino.exists():
        raise RuntimeError(f"Ya existe el Excel limpio destino y no se sobrescribirá: {destino}")

    shutil.copy2(origen, destino)

    wb = load_workbook(destino, read_only=False, data_only=False)
    try:
        if "Facturas" not in wb.sheetnames:
            raise RuntimeError("El Excel copiado no tiene hoja Facturas.")

        ws = wb["Facturas"]
        headers = [cell.value for cell in ws[1]]
        if headers != HEADERS_ESPERADOS:
            raise RuntimeError("Encabezados inválidos al crear Excel limpio.")

        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)

        ws.tables.clear()

        tab = Table(displayName="TblFacturas", ref="A1:S1")
        style = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        tab.tableStyleInfo = style
        ws.add_table(tab)

        wb.save(destino)
    finally:
        wb.close()


def construir_contexto(fecha_corte: date) -> dict:
    estado = cargar_json(STATE_PATH)
    validar_estado_actual(estado)

    fecha_inicio_estado = parse_fecha(estado["fecha_inicio_periodo_activo"], "fecha_inicio_periodo_activo")

    if fecha_corte < fecha_inicio_estado:
        raise RuntimeError(
            f"La fecha de corte {fecha_corte.isoformat()} no puede ser menor "
            f"al inicio registrado {fecha_inicio_estado.isoformat()}."
        )

    periodo_nuevo, inicio_calendario, fin_calendario = trimestre_calendario(fecha_corte)

    carpeta_base = f"BASE_INICIAL_PRE_VPS_{fecha_inicio_estado.isoformat()}_a_{fecha_corte.isoformat()}"
    anio = str(fecha_corte.year)

    base_local = CIERRES_TRIMESTRALES_DIR / anio / carpeta_base / "Corte_Inicial_Pre_VPS"

    carpetas = {
        "01_Excel_Historico": base_local / "01_Excel_Historico",
        "02_Excel_Activo_Limpio": base_local / "02_Excel_Activo_Limpio",
        "03_State": base_local / "03_State",
        "04_Manifest": base_local / "04_Manifest",
        "05_Validaciones": base_local / "05_Validaciones",
        "06_Logs_Pre_VPS": base_local / "06_Logs_Pre_VPS",
    }

    nombre_historico = f"facturas_base_inicial_pre_vps_{fecha_inicio_estado.isoformat()}_a_{fecha_corte.isoformat()}.xlsx"
    nombre_limpio = f"facturas_limpio_inicio_vps_{fecha_corte.isoformat()}.xlsx"

    sp_folder = ""
    try:
        # Import tardío solo para leer constante si existe; no llama Graph.
        sys.path.insert(0, str(ROOT))
        from services.m365.sp_graph import SP_FOLDER as SP_FOLDER_CONST  # type: ignore

        sp_folder = str(SP_FOLDER_CONST or "").strip().strip("/")
    except Exception:
        sp_folder = ""

    # Si no se pudo importar, dejamos ruta por .env o vacío.
    if not sp_folder:
        # Lectura manual simple del .env sin depender de python-dotenv.
        env_path = ROOT / ".env"
        if env_path.exists():
            for line in env_path.read_text(encoding="utf-8", errors="ignore").splitlines():
                if line.strip().startswith("SP_FOLDER="):
                    sp_folder = line.split("=", 1)[1].strip().strip('"').strip("'").strip("/")
                    break

    backup_root = "Backups_Facturas_Pruebas"
    env_path = ROOT / ".env"
    if env_path.exists():
        for line in env_path.read_text(encoding="utf-8", errors="ignore").splitlines():
            if line.strip().startswith("BACKUP_ROOT_FOLDER="):
                backup_root = line.split("=", 1)[1].strip().strip('"').strip("'").strip("/")
                break

    sp_excel_activo = f"{sp_folder}/excel/facturas.xlsx".strip("/")
    sp_historico = f"{sp_folder}/excel/Historicas/{carpeta_base}/{nombre_historico}".strip("/")
    backup_remoto = f"{backup_root}/{anio}/{carpeta_base}/Corte_Inicial_Pre_VPS".strip("/")

    nuevo_state = {
        "version": "2026-07-10-CORTE-INICIAL-PRE-VPS-STATE-V1",
        "periodo_activo": periodo_nuevo,
        "fecha_inicio_periodo_activo": fecha_corte.isoformat(),
        "proximo_cierre_estimado": fin_calendario.isoformat(),
        "ultimo_cierre_trimestral": None,
        "ultimo_archivo_generado": sp_historico,
        "estado": "ACTIVO",
        "nota": (
            "El Excel activo limpio inicia control operativo desde fecha_inicio_periodo_activo. "
            "El trimestre se controla por ciclo de vida del Excel activo, no por fecha de emisión de factura."
        ),
        "fecha_inicio_calendario_trimestre": inicio_calendario.isoformat(),
        "ultimo_corte_inicial_pre_vps": fecha_corte.isoformat(),
        "modo_arranque": "CORTE_INICIAL_PRE_VPS",
    }

    return {
        "estado_actual": estado,
        "fecha_inicio_estado": fecha_inicio_estado,
        "fecha_corte": fecha_corte,
        "carpeta_base": carpeta_base,
        "anio": anio,
        "base_local": base_local,
        "carpetas": carpetas,
        "nombre_historico": nombre_historico,
        "nombre_limpio": nombre_limpio,
        "archivo_historico_local": carpetas["01_Excel_Historico"] / nombre_historico,
        "archivo_limpio_local": carpetas["02_Excel_Activo_Limpio"] / nombre_limpio,
        "state_actual_backup": carpetas["03_State"] / "cierre_trimestral_state_ANTES_corte_inicial_pre_vps.json",
        "state_previsto_path": carpetas["03_State"] / "cierre_trimestral_state_PREVISTO_post_corte_inicial_pre_vps.json",
        "manifest_path": carpetas["04_Manifest"] / f"manifest_corte_inicial_pre_vps_{fecha_inicio_estado.isoformat()}_a_{fecha_corte.isoformat()}.json",
        "resumen_path": carpetas["04_Manifest"] / f"resumen_corte_inicial_pre_vps_{fecha_inicio_estado.isoformat()}_a_{fecha_corte.isoformat()}.txt",
        "validacion_path": carpetas["05_Validaciones"] / f"validacion_local_corte_inicial_pre_vps_{fecha_corte.isoformat()}.json",
        "sp_excel_activo": sp_excel_activo,
        "sp_historico": sp_historico,
        "backup_remoto": backup_remoto,
        "nuevo_state": nuevo_state,
    }


def listar_logs() -> List[Path]:
    candidatos: List[Path] = []
    for carpeta_rel in ["logs", "data/logs"]:
        carpeta = ROOT / carpeta_rel
        if carpeta.exists():
            candidatos.extend([p for p in carpeta.rglob("*") if p.is_file()])
    return sorted(set(candidatos))


def crear_manifest(ctx: dict, info_original: dict, info_historico: dict, info_limpio: dict, logs: List[Path]) -> dict:
    return {
        "tipo": "corte_inicial_pre_vps",
        "version": VERSION,
        "generado_en": datetime.now().isoformat(timespec="seconds"),
        "root": str(ROOT),
        "fecha_corte": ctx["fecha_corte"].isoformat(),
        "fecha_inicio_base": ctx["fecha_inicio_estado"].isoformat(),
        "carpeta_base": ctx["carpeta_base"],
        "regla": (
            "El Excel completo representa el periodo operativo. "
            "No se filtra por fecha de emisión de factura."
        ),
        "sharepoint_principal": {
            "excel_activo": ctx["sp_excel_activo"],
            "historico_previsto": ctx["sp_historico"],
            "nota": "En la ruta principal solo debe quedar el Excel activo y el histórico; sin logs ni auditorías técnicas.",
        },
        "repositorio_backups": {
            "ruta_remota_prevista": ctx["backup_remoto"],
            "nota": "Aquí sí se debe subir evidencia técnica, manifest, validaciones, state y soportes.",
        },
        "archivos_locales_generados": {
            "historico": rel(ctx["archivo_historico_local"]),
            "excel_limpio": rel(ctx["archivo_limpio_local"]),
            "state_actual_backup": rel(ctx["state_actual_backup"]),
            "state_previsto": rel(ctx["state_previsto_path"]),
            "manifest": rel(ctx["manifest_path"]),
            "resumen": rel(ctx["resumen_path"]),
            "validacion": rel(ctx["validacion_path"]),
        },
        "excel_original": info_original,
        "excel_historico": info_historico,
        "excel_limpio": info_limpio,
        "state_actual": ctx["estado_actual"],
        "state_previsto": ctx["nuevo_state"],
        "logs_detectados": [rel(p) for p in logs],
    }


def escribir_resumen(path: Path, manifest: dict) -> None:
    lines = [
        "CORTE INICIAL PRE-VPS - FACTURAS JOYCO",
        "=" * 80,
        f"Versión: {manifest['version']}",
        f"Generado en: {manifest['generado_en']}",
        f"Fecha corte: {manifest['fecha_corte']}",
        f"Base inicial desde: {manifest['fecha_inicio_base']}",
        f"Carpeta base: {manifest['carpeta_base']}",
        "",
        "Regla:",
        f"- {manifest['regla']}",
        "",
        "SharePoint principal previsto:",
        f"- Excel activo: {manifest['sharepoint_principal']['excel_activo']}",
        f"- Histórico: {manifest['sharepoint_principal']['historico_previsto']}",
        "",
        "Repositorio técnico de backups previsto:",
        f"- {manifest['repositorio_backups']['ruta_remota_prevista']}",
        "",
        "Excel original:",
        f"- Filas de datos: {manifest['excel_original']['filas_datos']}",
        f"- Columnas: {manifest['excel_original']['columnas']}",
        f"- Tabla: {manifest['excel_original']['tbl_facturas_ref']}",
        "",
        "Excel histórico generado:",
        f"- {manifest['archivos_locales_generados']['historico']}",
        f"- SHA256: {manifest['excel_historico']['sha256']}",
        "",
        "Excel limpio generado:",
        f"- {manifest['archivos_locales_generados']['excel_limpio']}",
        f"- Filas de datos: {manifest['excel_limpio']['filas_datos']}",
        f"- Tabla: {manifest['excel_limpio']['tbl_facturas_ref']}",
        "",
        "Nuevo state previsto:",
        f"- Periodo activo: {manifest['state_previsto']['periodo_activo']}",
        f"- Inicio operativo: {manifest['state_previsto']['fecha_inicio_periodo_activo']}",
        f"- Próximo cierre: {manifest['state_previsto']['proximo_cierre_estimado']}",
        "",
        "Importante:",
        "- Este modo local NO reemplaza data/facturas.xlsx.",
        "- Este modo local NO actualiza data/state/cierre_trimestral_state.json.",
        "- Este modo local NO llama Graph ni toca SharePoint/OneDrive.",
        "=" * 80,
    ]
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def ejecutar_dry_run(ctx: dict) -> int:
    info_original = diagnosticar_excel(FACTURAS_PATH, exigir_limpio=False)

    print("Estado actual:")
    print(json.dumps(ctx["estado_actual"], ensure_ascii=False, indent=2))
    print("-" * 100)

    print("Excel local actual:")
    print(json.dumps(info_original, ensure_ascii=False, indent=2))
    print("-" * 100)

    print("Archivos que se generarían localmente:")
    print(f"Histórico:    {ctx['archivo_historico_local']}")
    print(f"Excel limpio: {ctx['archivo_limpio_local']}")
    print(f"State antes:  {ctx['state_actual_backup']}")
    print(f"State nuevo:  {ctx['state_previsto_path']}")
    print(f"Manifest:     {ctx['manifest_path']}")
    print(f"Validación:   {ctx['validacion_path']}")
    print("-" * 100)

    print("SharePoint principal previsto:")
    print(f"Excel activo: {ctx['sp_excel_activo']}")
    print(f"Histórico:    {ctx['sp_historico']}")
    print("-" * 100)

    print("Repositorio técnico de backups previsto:")
    print(ctx["backup_remoto"])
    print("-" * 100)

    print("Nuevo state previsto:")
    print(json.dumps(ctx["nuevo_state"], ensure_ascii=False, indent=2))
    print("-" * 100)

    print("DRY-RUN: no se generaron archivos, no se reemplazó nada y no se llamó Graph.")
    return 0


def ejecutar_local(ctx: dict) -> int:
    info_original = diagnosticar_excel(FACTURAS_PATH, exigir_limpio=False)

    base_local: Path = ctx["base_local"]
    if base_local.exists() and any(base_local.rglob("*")):
        raise RuntimeError(
            "La carpeta local de corte ya existe y contiene archivos. "
            f"Para evitar sobrescritura accidental, revisa o elimina manualmente: {base_local}"
        )

    for carpeta in ctx["carpetas"].values():
        carpeta.mkdir(parents=True, exist_ok=True)

    # Copia histórico.
    shutil.copy2(FACTURAS_PATH, ctx["archivo_historico_local"])
    info_historico = diagnosticar_excel(ctx["archivo_historico_local"], exigir_limpio=False)

    # Genera Excel limpio.
    crear_excel_limpio_desde_original(FACTURAS_PATH, ctx["archivo_limpio_local"])
    info_limpio = diagnosticar_excel(ctx["archivo_limpio_local"], exigir_limpio=True)

    # Copia state actual y escribe state previsto.
    guardar_json(ctx["state_actual_backup"], ctx["estado_actual"])
    guardar_json(ctx["state_previsto_path"], ctx["nuevo_state"])

    # Copia logs si existen.
    logs = listar_logs()
    logs_copiados = []
    destino_logs = ctx["carpetas"]["06_Logs_Pre_VPS"]
    for log in logs:
        try:
            destino = destino_logs / log.relative_to(ROOT)
        except Exception:
            destino = destino_logs / log.name
        destino.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(log, destino)
        logs_copiados.append(destino)

    manifest = crear_manifest(ctx, info_original, info_historico, info_limpio, logs)
    manifest["logs_copiados"] = [rel(p) for p in logs_copiados]
    manifest["ok"] = True
    manifest["advertencias"] = []
    if not logs:
        manifest["advertencias"].append("No se encontraron logs locales para copiar.")

    guardar_json(ctx["manifest_path"], manifest)
    escribir_resumen(ctx["resumen_path"], manifest)

    validacion = {
        "tipo": "validacion_local_corte_inicial_pre_vps",
        "version": VERSION,
        "generado_en": datetime.now().isoformat(timespec="seconds"),
        "fecha_corte": ctx["fecha_corte"].isoformat(),
        "historico": {
            "archivo": rel(ctx["archivo_historico_local"]),
            "existe": ctx["archivo_historico_local"].exists(),
            "sha256": sha256_file(ctx["archivo_historico_local"]),
            "filas_datos": info_historico["filas_datos"],
            "columnas": info_historico["columnas"],
            "tbl_facturas_ref": info_historico["tbl_facturas_ref"],
        },
        "excel_limpio": {
            "archivo": rel(ctx["archivo_limpio_local"]),
            "existe": ctx["archivo_limpio_local"].exists(),
            "sha256": sha256_file(ctx["archivo_limpio_local"]),
            "filas_datos": info_limpio["filas_datos"],
            "columnas": info_limpio["columnas"],
            "tbl_facturas_ref": info_limpio["tbl_facturas_ref"],
            "esta_limpio": info_limpio["esta_limpio"],
        },
        "state": {
            "backup_actual_existe": ctx["state_actual_backup"].exists(),
            "state_previsto_existe": ctx["state_previsto_path"].exists(),
            "periodo_activo_previsto": ctx["nuevo_state"]["periodo_activo"],
            "fecha_inicio_periodo_activo_prevista": ctx["nuevo_state"]["fecha_inicio_periodo_activo"],
            "proximo_cierre_estimado_previsto": ctx["nuevo_state"]["proximo_cierre_estimado"],
        },
        "manifest_existe": ctx["manifest_path"].exists(),
        "resumen_existe": ctx["resumen_path"].exists(),
        "logs_detectados": len(logs),
        "logs_copiados": len(logs_copiados),
        "ok": True,
        "advertencias": manifest["advertencias"],
    }
    guardar_json(ctx["validacion_path"], validacion)

    print("-" * 100)
    print("✅ CORTE INICIAL PRE-VPS LOCAL GENERADO.")
    print(f"Histórico local:    {ctx['archivo_historico_local']}")
    print(f"Excel limpio local: {ctx['archivo_limpio_local']}")
    print(f"State previsto:     {ctx['state_previsto_path']}")
    print(f"Manifest:           {ctx['manifest_path']}")
    print(f"Validación local:   {ctx['validacion_path']}")
    print("-" * 100)
    print("✅ No se reemplazó data/facturas.xlsx.")
    print("✅ No se actualizó data/state/cierre_trimestral_state.json.")
    print("✅ No se llamó Graph ni se tocó SharePoint/OneDrive.")
    return 0


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--fecha-corte", required=True, help="Fecha del corte inicial PRE-VPS. Formato YYYY-MM-DD.")
    parser.add_argument("--dry-run", action="store_true", help="Solo diagnostica. No genera archivos.")
    parser.add_argument("--local", action="store_true", help="Genera evidencia local, histórico y Excel limpio local. No reemplaza activos.")
    args = parser.parse_args()

    if bool(args.dry_run) == bool(args.local):
        print("❌ Usa exactamente un modo: --dry-run o --local.")
        return 1

    fecha_corte = parse_fecha(args.fecha_corte, "fecha-corte")
    ctx = construir_contexto(fecha_corte)

    modo = "LOCAL" if args.local else "DRY-RUN"

    print("=" * 100)
    print(f"CORTE INICIAL PRE-VPS FACTURAS JOYCO - {modo}")
    print("=" * 100)
    print(f"Version: {VERSION}")
    print(f"Root: {ROOT}")
    print(f"Fecha corte: {fecha_corte.isoformat()}")
    print(f"Carpeta base: {ctx['carpeta_base']}")
    print(f"Base local: {ctx['base_local']}")
    print("-" * 100)

    try:
        if args.local:
            rc = ejecutar_local(ctx)
        else:
            rc = ejecutar_dry_run(ctx)

        print("=" * 100)
        return rc

    except Exception as exc:
        print(f"❌ Error en corte inicial PRE-VPS: {exc}")
        print("No se debe continuar hasta revisar el error.")
        print("=" * 100)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
