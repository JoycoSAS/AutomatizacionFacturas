#!/opt/joyco/facturas-procesador/venv/bin/python

from __future__ import annotations

import argparse
import ast
import grp
import hashlib
import os
import pwd
import runpy
import stat
import sys
import zipfile
from pathlib import Path

from dotenv import load_dotenv
from openpyxl import load_workbook


VERSION_LANZADOR = "2026-07-27-VPS-03-LOCK-HUERFANO"

BASE_DIR = Path("/opt/joyco/facturas-procesador")
APP_DIR = BASE_DIR / "app"
ENV_FILE = Path("/etc/joyco/facturas-procesador/facturas.env")
ENTRYPOINT = APP_DIR / "main_aprobadas_integrado.py"

EXCEL_PATH = Path(
    "/var/lib/joyco/facturas-procesador/data/facturas.xlsx"
)

HISTORIAL_PATH = Path(
    "/var/lib/joyco/facturas-procesador/data/historial_ejecuciones.xlsx"
)

STATE_PATH = Path(
    "/var/lib/joyco/facturas-procesador/state"
)

AUDIT_PATH = Path(
    "/var/lib/joyco/facturas-procesador/data/audit"
)

LOCK_PATH = STATE_PATH / "aprobadas.lock"

FACTURAS_SHEET = "Facturas"
FACTURAS_COLUMNAS_ESPERADAS = 19

HISTORIAL_SHEET = "Historial"
HISTORIAL_HEADERS_ESPERADOS = [
    "Fecha",
    "Hora",
    "Archivo ZIP",
    "Nuevos XML guardados",
    "Errores encontrados",
]


def abortar(mensaje: str) -> None:
    raise RuntimeError(mensaje)


def valor(nombre: str) -> str:
    return str(os.getenv(nombre) or "").strip()


def validar_archivo_entorno() -> None:
    if not ENV_FILE.is_file():
        abortar(f"No existe el archivo de entorno: {ENV_FILE}")

    info = ENV_FILE.stat()
    permisos = stat.S_IMODE(info.st_mode)

    propietario = pwd.getpwuid(info.st_uid).pw_name
    grupo = grp.getgrgid(info.st_gid).gr_name

    if propietario != "root":
        abortar(
            "Propietario inesperado para facturas.env: "
            f"{propietario}"
        )

    if grupo != "innovacion":
        abortar(
            "Grupo inesperado para facturas.env: "
            f"{grupo}"
        )

    if permisos != 0o640:
        abortar(
            "Permisos inesperados para facturas.env: "
            f"{permisos:o}"
        )

    contenido = ENV_FILE.read_bytes()

    if contenido.startswith(b"\xef\xbb\xbf"):
        abortar("facturas.env contiene UTF-8 BOM.")

    if b"\x00" in contenido:
        abortar("facturas.env contiene bytes NUL.")

    print(
        "OK: facturas.env protegido "
        "| propietario=root "
        "| grupo=innovacion "
        "| permisos=640"
    )


def cargar_entorno() -> None:
    os.environ["FACTURAS_ENV_FILE"] = str(ENV_FILE)

    cargado = load_dotenv(
        dotenv_path=ENV_FILE,
        override=True,
        encoding="utf-8",
    )

    if not cargado:
        abortar("No fue posible cargar facturas.env.")

    requeridas = [
        "TENANT_ID",
        "CLIENT_ID",
        "CLIENT_SECRET",
        "MAILBOX_UPN",
        "SP_HOSTNAME",
        "SP_SITE_PATH",
        "SP_DRIVE_ID",
        "SP_FOLDER",
        "SP_DRIVE_ID_RADICADOS",
        "SP_FOLDER_RADICADOS",
    ]

    faltantes = [
        nombre
        for nombre in requeridas
        if not valor(nombre)
    ]

    if faltantes:
        abortar(
            "Variables obligatorias ausentes o vacías: "
            + ", ".join(sorted(faltantes))
        )

    print(
        "OK: variables obligatorias cargadas "
        "sin mostrar credenciales."
    )


def validar_perfil_seguro() -> None:
    perfil_esperado = {
        "FACTURAS_MODO": "PRODUCCION",
        "FACTURAS_SINCE_DAYS": "6",
        "FACTURAS_MAX_MENSAJES": "1000",
        "FACTURAS_MAX_ZIP_BUSCAR": "1000",
        "FACTURAS_UNREAD_ONLY": "0",
        "FACTURAS_USE_PROCESSED_STORE": "1",
        "FACTURAS_MARCAR_LEIDO": "0",
        "FACTURAS_FORZAR_SOLO_EXCEL_SP": "1",
        "FACTURAS_PROCESAR_ANTIGUOS_PRIMERO": "0",
        "FACTURAS_HISTORICO_DRY_RUN": "0",
        "FACTURAS_DISABLE_AUTOSTOP": "0",
        "FACTURAS_RUN_APROBADAS": "1",
        "FACTURAS_RUN_NOTAS_CREDITO": "0",
        "SP_UPLOAD_DOCUMENTOS": "0",
        "SP_UPLOAD_HISTORIAL": "0",
        "SP_ENSURE_DOCUMENT_FOLDERS": "0",
        "ALERTAS_HABILITADAS": "0",
        "LOCK_TTL_SECONDS": "3600",
    }

    errores = []

    for nombre, esperado in perfil_esperado.items():
        actual = valor(nombre)

        if actual != esperado:
            errores.append(
                f"{nombre}: esperado={esperado!r}, "
                f"actual={actual!r}"
            )

    if errores:
        abortar(
            "El perfil no coincide con la configuración segura:\n- "
            + "\n- ".join(errores)
        )

    print("OK: perfil inicial de producción validado.")

    print()
    print("Perfil visible:")

    for nombre, esperado in perfil_esperado.items():
        print(f"{nombre}={esperado}")


def validar_entrypoint() -> None:
    if not ENTRYPOINT.is_file():
        abortar(
            f"No existe el punto de entrada: {ENTRYPOINT}"
        )

    source = ENTRYPOINT.read_text(
        encoding="utf-8-sig",
        errors="strict",
    )

    tree = ast.parse(
        source,
        filename=str(ENTRYPOINT),
    )

    funciones = {
        node.name
        for node in tree.body
        if isinstance(
            node,
            (
                ast.FunctionDef,
                ast.AsyncFunctionDef,
            ),
        )
    }

    if "main" not in funciones:
        abortar(
            "main_aprobadas_integrado.py no contiene "
            "una función main()."
        )

    guard_principal = False

    for node in tree.body:
        if not isinstance(node, ast.If):
            continue

        prueba = node.test

        if not isinstance(prueba, ast.Compare):
            continue

        if not isinstance(prueba.left, ast.Name):
            continue

        if prueba.left.id != "__name__":
            continue

        if not prueba.comparators:
            continue

        comparador = prueba.comparators[0]

        if (
            isinstance(comparador, ast.Constant)
            and comparador.value == "__main__"
        ):
            guard_principal = True
            break

    if not guard_principal:
        abortar(
            "No se encontró el bloque "
            "if __name__ == '__main__'."
        )

    digest = hashlib.sha256(
        ENTRYPOINT.read_bytes()
    ).hexdigest().upper()

    print("OK: punto de entrada validado.")
    print(f"ENTRYPOINT={ENTRYPOINT}")
    print(f"SHA256_ENTRYPOINT={digest}")


def _normalizar_encabezado(valor_celda: object) -> str:
    if valor_celda is None:
        return ""

    return " ".join(
        str(valor_celda).strip().split()
    )


def validar_excel_xlsx(
    path: Path,
    *,
    descripcion: str,
    hoja_requerida: str,
    columnas_esperadas: int,
    encabezados_esperados: list[str] | None = None,
) -> dict[str, int | str]:
    """
    Valida integridad y estructura mínima de un archivo XLSX.

    Un Excel con únicamente encabezados es válido. Por tanto, cero filas
    de datos no representa un error.
    """
    if not path.is_file():
        abortar(
            f"No existe {descripcion}: {path}"
        )

    try:
        bytes_archivo = path.stat().st_size
    except OSError as exc:
        abortar(
            f"No fue posible consultar {descripcion}: "
            f"{path} | {type(exc).__name__}: {exc}"
        )

    if bytes_archivo <= 0:
        abortar(
            f"{descripcion} está vacío (0 bytes): {path}"
        )

    if not zipfile.is_zipfile(path):
        abortar(
            f"{descripcion} no es un archivo XLSX válido: {path}"
        )

    workbook = None

    try:
        workbook = load_workbook(
            path,
            read_only=True,
            data_only=False,
        )

        if hoja_requerida not in workbook.sheetnames:
            abortar(
                f"{descripcion} no contiene la hoja requerida "
                f"{hoja_requerida!r}. "
                f"Hojas detectadas: {workbook.sheetnames}"
            )

        sheet = workbook[hoja_requerida]
        filas_totales = int(sheet.max_row or 0)
        columnas_detectadas = int(sheet.max_column or 0)

        if filas_totales < 1:
            abortar(
                f"{descripcion} no contiene la fila de encabezados: "
                f"{path}"
            )

        if columnas_detectadas != columnas_esperadas:
            abortar(
                f"{descripcion} tiene una cantidad de columnas "
                f"inesperada. Detectadas={columnas_detectadas}, "
                f"esperadas={columnas_esperadas}. Archivo={path}"
            )

        primera_fila = next(
            sheet.iter_rows(
                min_row=1,
                max_row=1,
                max_col=columnas_esperadas,
                values_only=True,
            ),
            None,
        )

        encabezados = [
            _normalizar_encabezado(valor_celda)
            for valor_celda in (primera_fila or ())
        ]

        if len(encabezados) != columnas_esperadas:
            abortar(
                f"{descripcion} no contiene todos los encabezados "
                f"esperados. Detectados={len(encabezados)}, "
                f"esperados={columnas_esperadas}. Archivo={path}"
            )

        posiciones_vacias = [
            str(indice)
            for indice, encabezado in enumerate(
                encabezados,
                start=1,
            )
            if not encabezado
        ]

        if posiciones_vacias:
            abortar(
                f"{descripcion} contiene encabezados vacíos en las "
                f"columnas: {', '.join(posiciones_vacias)}. "
                f"Archivo={path}"
            )

        encabezados_normalizados = [
            encabezado.casefold()
            for encabezado in encabezados
        ]

        if (
            len(set(encabezados_normalizados))
            != len(encabezados_normalizados)
        ):
            abortar(
                f"{descripcion} contiene encabezados duplicados: "
                f"{encabezados}. Archivo={path}"
            )

        if encabezados_esperados is not None:
            esperados_normalizados = [
                _normalizar_encabezado(encabezado)
                for encabezado in encabezados_esperados
            ]

            if encabezados != esperados_normalizados:
                abortar(
                    f"{descripcion} tiene encabezados inesperados. "
                    f"Detectados={encabezados}. "
                    f"Esperados={esperados_normalizados}. "
                    f"Archivo={path}"
                )

        filas_datos = max(filas_totales - 1, 0)

        print(
            f"OK: {descripcion} válido "
            f"| hoja={hoja_requerida} "
            f"| filas_datos={filas_datos} "
            f"| columnas={columnas_detectadas} "
            f"| bytes={bytes_archivo}"
        )

        if filas_datos == 0:
            print(
                f"OK: {descripcion} no tiene registros de datos, "
                "pero conserva una estructura válida con encabezados."
            )

        return {
            "bytes": bytes_archivo,
            "filas_datos": filas_datos,
            "columnas": columnas_detectadas,
            "hoja": hoja_requerida,
        }

    except RuntimeError:
        raise
    except Exception as exc:
        abortar(
            f"No fue posible abrir o validar {descripcion}: "
            f"{path} | {type(exc).__name__}: {exc}"
        )
    finally:
        if workbook is not None:
            workbook.close()


def validar_lock_aprobadas() -> bool:
    """
    Comprueba el lock usando la misma implementación del controller.

    Retorna True cuando existía un lock huérfano que fue recuperado.
    Un lock asociado a un PID activo bloquea la ejecución.
    """
    from utils.single_instance_lock import SingleInstanceLock

    lock_existia = LOCK_PATH.exists()

    try:
        ttl_seconds = int(valor("LOCK_TTL_SECONDS") or "3600")
    except (TypeError, ValueError):
        abortar(
            "LOCK_TTL_SECONDS no contiene un entero válido."
        )

    lock = SingleInstanceLock(
        str(LOCK_PATH),
        ttl_seconds=ttl_seconds,
    )

    if not lock.acquire():
        abortar(
            "Existe un lock activo o no recuperable: "
            f"{LOCK_PATH}. "
            "No se iniciará otra instancia."
        )

    lock.release()

    if LOCK_PATH.exists():
        abortar(
            "El lock de validación no pudo liberarse correctamente: "
            f"{LOCK_PATH}"
        )

    return lock_existia


def validar_rutas_config() -> None:
    if str(APP_DIR) not in sys.path:
        sys.path.insert(0, str(APP_DIR))

    os.chdir(APP_DIR)

    if (APP_DIR / ".env").exists():
        abortar(
            "Existe un .env dentro del repositorio. "
            "Debe usarse exclusivamente el archivo de /etc."
        )

    import config

    excel_config = Path(config.ARCHIVO_EXCEL)
    historial_config = Path(config.HISTORIAL_EXCEL)
    state_config = Path(config.STATE_DIR).resolve()
    audit_config = Path(config.AUDIT_DIR).resolve()
    lock_config = Path(
        config.LOCK_FILE_APROBADAS
    ).resolve()

    if excel_config != EXCEL_PATH:
        abortar(
            f"ARCHIVO_EXCEL inesperado: {excel_config}"
        )

    if historial_config != HISTORIAL_PATH:
        abortar(
            f"HISTORIAL_EXCEL inesperado: {historial_config}"
        )

    if state_config != STATE_PATH:
        abortar(
            f"STATE_DIR inesperado: {state_config}"
        )

    if audit_config != AUDIT_PATH:
        abortar(
            f"AUDIT_DIR inesperado: {audit_config}"
        )

    if lock_config != LOCK_PATH:
        abortar(
            f"LOCK_FILE_APROBADAS inesperado: {lock_config}"
        )

    validar_excel_xlsx(
        EXCEL_PATH,
        descripcion="el Excel operativo",
        hoja_requerida=FACTURAS_SHEET,
        columnas_esperadas=FACTURAS_COLUMNAS_ESPERADAS,
    )

    validar_excel_xlsx(
        HISTORIAL_PATH,
        descripcion="el historial de ejecuciones",
        hoja_requerida=HISTORIAL_SHEET,
        columnas_esperadas=len(
            HISTORIAL_HEADERS_ESPERADOS
        ),
        encabezados_esperados=HISTORIAL_HEADERS_ESPERADOS,
    )

    if not STATE_PATH.is_dir():
        abortar(
            f"No existe STATE_DIR: {STATE_PATH}"
        )

    lock_huerfano_recuperado = validar_lock_aprobadas()

    print()
    print("Rutas resueltas:")
    print(f"ARCHIVO_EXCEL={excel_config}")
    print(f"HISTORIAL_EXCEL={historial_config}")
    print(f"STATE_DIR={state_config}")
    print(f"AUDIT_DIR={audit_config}")
    print(f"LOCK_FILE={lock_config}")

    print()
    print("OK: rutas del VPS validadas.")
    print("OK: archivos XLSX operativos íntegros y estructurados.")
    print("OK: cero filas de datos se acepta cuando existen encabezados válidos.")

    if lock_huerfano_recuperado:
        print(
            "OK: se detectó y retiró un lock huérfano; "
            "no existe una instancia activa."
        )
    else:
        print("OK: no existe un lock activo.")


def validar_todo() -> None:
    print("============================================================")
    print(f"LANZADOR FACTURAS JOYCO {VERSION_LANZADOR}")
    print("MODO: VALIDACIÓN SIN EJECUCIÓN")
    print("============================================================")

    validar_archivo_entorno()
    cargar_entorno()
    validar_perfil_seguro()
    validar_entrypoint()
    validar_rutas_config()

    print()
    print("============================================================")
    print("OK: lanzador preparado correctamente.")
    print("OK: main_aprobadas_integrado.py es el entrypoint oficial.")
    print("OK: no se realizó ninguna conexión de red.")
    print("OK: no se ejecutó el procesamiento de facturas.")
    print("============================================================")


def ejecutar(confirmado: bool) -> None:
    validar_todo()

    if not confirmado:
        abortar(
            "La ejecución real requiere "
            "--confirm-production."
        )

    print()
    print("============================================================")
    print("INICIO MANUAL DEL PROCESO DE PRODUCCIÓN")
    print(f"ENTRYPOINT={ENTRYPOINT}")
    print("============================================================")
    sys.stdout.flush()

    runpy.run_path(
        str(ENTRYPOINT),
        run_name="__main__",
    )


def main() -> int:
    os.umask(0o027)

    parser = argparse.ArgumentParser(
        description=(
            "Lanzador seguro del procesador de facturas JOYCO."
        )
    )

    parser.add_argument(
        "accion",
        choices=[
            "check",
            "run",
        ],
        help=(
            "check valida sin ejecutar; "
            "run inicia el procesamiento."
        ),
    )

    parser.add_argument(
        "--confirm-production",
        action="store_true",
        help=(
            "Confirmación obligatoria para ejecutar producción."
        ),
    )

    argumentos = parser.parse_args()

    try:
        if argumentos.accion == "check":
            validar_todo()
        else:
            ejecutar(
                confirmado=argumentos.confirm_production
            )

        return 0

    except Exception as exc:
        print()
        print("ERROR DEL LANZADOR:")
        print(str(exc))
        return 1


if __name__ == "__main__":
    raise SystemExit(main())