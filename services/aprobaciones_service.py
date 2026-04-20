# services/aprobaciones_service.py
"""
Sincroniza Radicado / ProyectoProceso desde el Excel de Radicados (SharePoint)
hacia facturas.xlsx.

MEJORAS:
- Descarga el Excel de radicados al RADICADOS_LOCAL_PATH.
- Reutiliza cargar_mapa_radicados().
- Recrea la tabla TblFacturas sin usar ws._tables = [].
- Hace match robusto por múltiples variantes del "Número de factura":
    original, sin espacios, sin guiones, solo alfanumérico, etc.

Resultado:
- Si existe match por "Número de factura", llena:
    Radicado, ProyectoProceso
- Si no existe match, NO rompe, solo no actualiza.
"""

from __future__ import annotations

import os
import re
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv

from config import (
    ARCHIVO_EXCEL,
    TMP_DIR,
    RADICADOS_SP_RELATIVE_PATH,
    RADICADOS_LOCAL_PATH,
    FACT_COL_NUMERO,
    FACT_COL_RAD,
    FACT_COL_PROY,
)

from services.m365.sp_graph import download_small_file
from services.radicados_service import cargar_mapa_radicados
from utils.safe_io import safe_save_pandas
from utils.normalizacion_facturas import normalizar_texto_basico

load_dotenv()
SP_DRIVE_ID_RADICADOS = (os.getenv("SP_DRIVE_ID_RADICADOS") or "").strip()


# -------------------------------------------------
# Helpers (Excel facturas)
# -------------------------------------------------
def _find_col(ws, header_row: int, wanted_name: str) -> int | None:
    wanted = normalizar_texto_basico(wanted_name)
    for c in range(1, ws.max_column + 1):
        h = ws.cell(row=header_row, column=c).value
        if normalizar_texto_basico(h) == wanted:
            return c
    return None


def _ensure_column(ws, header_name: str) -> int:
    col = _find_col(ws, 1, header_name)
    if col:
        return col
    new_c = ws.max_column + 1
    ws.cell(row=1, column=new_c, value=header_name)
    return new_c


def _recrear_tabla_facturas(ws) -> None:
    """
    Elimina tablas existentes de forma segura y recrea TblFacturas.
    """
    try:
        for table_name in list(ws.tables.keys()):
            del ws.tables[table_name]
    except Exception:
        pass

    if ws.max_row < 1 or ws.max_column < 1:
        return

    ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    tbl = Table(displayName="TblFacturas", ref=ref)
    tbl.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tbl)
    ws.freeze_panes = "A2"


def _ordenar_y_formatear_facturas():
    """
    Mantiene el Excel bonito:
    - Radicado y ProyectoProceso al inicio
    - Orden por Radicado (si es numérico)
    - Tabla + freeze panes
    """
    if not os.path.exists(ARCHIVO_EXCEL):
        return

    df = pd.read_excel(ARCHIVO_EXCEL, sheet_name="Facturas", engine="openpyxl")

    prioridad = [c for c in [FACT_COL_RAD, FACT_COL_PROY] if c in df.columns]
    resto = [c for c in df.columns if c not in prioridad]
    df = df[prioridad + resto]

    if FACT_COL_RAD in df.columns:
        def _rad_sort(x):
            s = str(x).strip()
            return int(s) if s.isdigit() else 10**12

        df["_rad_sort"] = df[FACT_COL_RAD].apply(_rad_sort)
        df = df.sort_values("_rad_sort").drop(columns="_rad_sort").reset_index(drop=True)

    safe_save_pandas(df, ARCHIVO_EXCEL, sheet_name="Facturas")

    wb = load_workbook(ARCHIVO_EXCEL)
    try:
        ws = wb["Facturas"]
        _recrear_tabla_facturas(ws)
        wb.save(ARCHIVO_EXCEL)
    finally:
        wb.close()


# -------------------------------------------------
# Helpers (match robusto de factura)
# -------------------------------------------------
def _build_possible_factura_keys(val: str) -> list[str]:
    """
    Genera varias claves posibles para aumentar el match entre
    Número de factura en facturas.xlsx y el mapa de radicados.
    """
    raw = str(val or "").strip().upper()
    if not raw:
        return []

    keys = []
    seen = set()

    def add(x: str):
        x = str(x or "").strip().upper()
        if x and x not in seen:
            seen.add(x)
            keys.append(x)

    add(raw)
    add(raw.replace(" ", ""))
    add(raw.replace("-", ""))
    add(raw.replace("_", ""))
    add(raw.replace(" ", "").replace("-", ""))
    add(raw.replace(" ", "").replace("_", ""))
    add(raw.replace("-", "").replace("_", ""))
    add(raw.replace(" ", "").replace("-", "").replace("_", ""))

    # solo alfanumérico
    limpio = re.sub(r"[^A-Z0-9]", "", raw)
    add(limpio)

    # variantes comunes con separadores
    m = re.match(r"^([A-Z]+)(\d+)$", limpio)
    if m:
        pref, dig = m.groups()
        add(f"{pref}{dig}")
        add(f"{pref}-{dig}")
        add(f"{pref} {dig}")
        add(f"{pref}_{dig}")

    m2 = re.match(r"^([A-Z]+)-(\d+)$", raw)
    if m2:
        pref, dig = m2.groups()
        add(f"{pref}{dig}")
        add(f"{pref} {dig}")
        add(f"{pref}_{dig}")

    return keys


def _find_match_in_mapa(val: str, mapa: dict) -> tuple[str, str] | None:
    """
    Busca coincidencia del número de factura contra el mapa de radicados
    usando múltiples variantes.
    """
    for key in _build_possible_factura_keys(val):
        if key in mapa:
            return mapa[key]
    return None


# -------------------------------------------------
# FUNCIÓN PRINCIPAL
# -------------------------------------------------
def sincronizar_aprobaciones_en_facturas(force_reload_radicados: bool = True) -> int:
    """
    1) Descarga el Excel de radicados desde SharePoint a RADICADOS_LOCAL_PATH
    2) Construye mapa con cargar_mapa_radicados()
    3) Aplica a facturas.xlsx (solo llena si está vacío)
    """
    if not os.path.exists(ARCHIVO_EXCEL):
        return 0

    if not SP_DRIVE_ID_RADICADOS:
        print("[RAD] Falta SP_DRIVE_ID_RADICADOS en .env")
        return 0

    Path(TMP_DIR).mkdir(parents=True, exist_ok=True)
    Path(os.path.dirname(RADICADOS_LOCAL_PATH)).mkdir(parents=True, exist_ok=True)

    ok = download_small_file(
        sp_relative_path=RADICADOS_SP_RELATIVE_PATH,
        local_path=RADICADOS_LOCAL_PATH,
        drive_id=SP_DRIVE_ID_RADICADOS,
    )
    if not ok:
        print("[RAD] No se pudo descargar el Excel de radicados.")
        return 0

    try:
        mapa = cargar_mapa_radicados(force_reload=force_reload_radicados)
    except Exception as e:
        print(f"[RAD] Error cargando mapa de radicados: {e}")
        return 0

    if not mapa:
        print("[RAD] Mapa de radicados vacío (0).")
        return 0

    wb_f = load_workbook(ARCHIVO_EXCEL)
    try:
        ws_f = wb_f["Facturas"]

        col_num = _find_col(ws_f, 1, FACT_COL_NUMERO)
        if not col_num:
            print(f"[RAD] No existe columna '{FACT_COL_NUMERO}' en facturas.xlsx")
            return 0

        col_rad_f = _ensure_column(ws_f, FACT_COL_RAD)
        col_proy_f = _ensure_column(ws_f, FACT_COL_PROY)

        updated = 0

        for r in range(2, ws_f.max_row + 1):
            val = ws_f.cell(row=r, column=col_num).value
            if not val:
                continue

            match = _find_match_in_mapa(str(val), mapa)
            if not match:
                continue

            rad, proy = match
            changed = False

            cell_rad = ws_f.cell(row=r, column=col_rad_f)
            cell_proy = ws_f.cell(row=r, column=col_proy_f)

            if (cell_rad.value is None or str(cell_rad.value).strip() == "") and rad:
                cell_rad.value = rad
                changed = True

            if (cell_proy.value is None or str(cell_proy.value).strip() == "") and proy:
                cell_proy.value = proy
                changed = True

            if changed:
                updated += 1

        wb_f.save(ARCHIVO_EXCEL)

    finally:
        wb_f.close()

    if updated > 0:
        _ordenar_y_formatear_facturas()
        print(f"[RAD] ✅ Filas actualizadas en facturas.xlsx: {updated}")
    else:
        print("[RAD] ℹ️ No hubo cambios (sin match o ya estaban llenos).")

    return updated