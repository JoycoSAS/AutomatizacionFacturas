# services/excel_service.py
import os
import re
import unicodedata
from pathlib import Path
from typing import Any, Dict, List, Set, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from config import ARCHIVO_EXCEL, HISTORIAL_EXCEL
from utils.safe_io import safe_save_pandas

print("🔥 EXCEL_SERVICE VERSION ACTIVA: 2026-06-10-H3-NAN-CALIDAD-ANTIDUP-RADICADO-MINIMA")


CONCEPTOS_BASE_FIJOS = [
    "Subtotal",
    "IVA 5%",
    "IVA 19%",
    "Retención de IVA",
    "Retención de ICA",
    "Retención en la fuente",
    "Total",
]

COLUMNAS_BASE_LARGO = [
    "Archivo",
    "Empresa emisora",
    "CUFE",
    "Ciudad emisora",
    "Código ciudad",
    "NIT",
    "Cliente",
    "Número de factura",
    "Año",
    "Mes",
    "Día",
    "Tipo de contribuyente",
    "Actividad económica",
    "DESCRIPCIÓN",
    "Concepto",
    "VALOR",
]

COLUMNAS_EXTRA_PERMITIDAS = [
    "Radicado",
    "ProyectoProceso",
]

# Estado_calidad debe quedar de última para formato condicional en Excel Web.
COLUMNA_ESTADO_CALIDAD = "Estado_calidad"

# Valores que NO deben aparecer escritos en Excel.
# En pandas/openpyxl pueden llegar como float NaN, pd.NA, None o como texto "nan".
# La regla oficial desde 2026-06-10 es: estos valores se guardan como celda vacía.
VALORES_TEXTO_VACIOS_EXCEL = {
    "", "NAN", "NONE", "NULL", "N/A", "NA", "SIN_DATO", "SIN DATO", "<NA>", "NAT"
}

COLUMNAS_VALIDAS_FINALES = COLUMNAS_EXTRA_PERMITIDAS + COLUMNAS_BASE_LARGO + [COLUMNA_ESTADO_CALIDAD]
CONCEPTO_ORDEN = {nombre: idx for idx, nombre in enumerate(CONCEPTOS_BASE_FIJOS, start=1)}

# Campos obligatorios para considerar una factura COMPLETA.
# Regla aprobada:
# - Actividad económica NO cuenta.
# - En conceptos, solo Total debe ser > 0.
# - IVA y retenciones pueden estar en cero.
CAMPOS_CALIDAD_COMPLETA_EXCEL = [
    "Radicado",
    "ProyectoProceso",
    "Archivo",
    "Empresa emisora",
    "CUFE",
    "Ciudad emisora",
    "Código ciudad",
    "NIT",
    "Cliente",
    "Número de factura",
    "Año",
    "Mes",
    "Día",
    "DESCRIPCIÓN",
]

# Campos que definen si una factura tiene una base real de información.
# Si varios de estos faltan, la factura no debe quedar PARCIAL: debe quedar MINIMA.
CAMPOS_CALIDAD_BASE_REAL_EXCEL = [
    "Empresa emisora",
    "NIT",
    "Cliente",
    "Número de factura",
    "Año",
    "Mes",
    "Día",
    "DESCRIPCIÓN",
]

# Campos técnicos/operativos que por sí solos NO son suficientes para considerar
# una factura como PARCIAL: pueden existir incluso en un registro mínimo.
CAMPOS_CALIDAD_OPERATIVOS_EXCEL = [
    "Radicado",
    "ProyectoProceso",
    "Archivo",
    "CUFE",
]


def _es_valor_vacio_excel(valor: Any) -> bool:
    """
    Detecta vacíos reales y vacíos escritos como texto.

    Evita que valores como "nan" queden visibles en Excel o cuenten como dato
    presente para Estado_calidad.
    """
    if valor is None:
        return True

    try:
        if pd.isna(valor):
            return True
    except Exception:
        pass

    try:
        s = str(valor).strip()
    except Exception:
        return True

    if s.upper() in VALORES_TEXTO_VACIOS_EXCEL:
        return True

    return False


def _sanitizar_valor_excel(valor: Any) -> Any:
    """
    Normaliza valores antes de escribirlos al Excel.
    Los vacíos técnicos quedan como celda vacía.
    """
    if _es_valor_vacio_excel(valor):
        return ""
    return valor


def _valor_presente_calidad_excel(valor: Any) -> bool:
    return not _es_valor_vacio_excel(valor)


ANCHO_COLUMNAS = {
    "Radicado": 13,
    "ProyectoProceso": 30,
    "Archivo": 32,
    "Empresa emisora": 34,
    "CUFE": 34,
    "Ciudad emisora": 18,
    "Código ciudad": 13,
    "NIT": 16,
    "Cliente": 32,
    "Número de factura": 20,
    "Año": 9,
    "Mes": 8,
    "Día": 8,
    "Tipo de contribuyente": 22,
    "Actividad económica": 24,
    "DESCRIPCIÓN": 55,
    "Concepto": 22,
    "VALOR": 16,
    "Estado_calidad": 18,
}

COLUMNAS_WRAP = {
    "ProyectoProceso",
    "Archivo",
    "Empresa emisora",
    "CUFE",
    "Cliente",
    "DESCRIPCIÓN",
}

COLUMNAS_TEXTO = {
    "Radicado",
    "ProyectoProceso",
    "Archivo",
    "Empresa emisora",
    "CUFE",
    "Ciudad emisora",
    "Código ciudad",
    "NIT",
    "Cliente",
    "Número de factura",
    "Año",
    "Mes",
    "Día",
    "Tipo de contribuyente",
    "Actividad económica",
    "DESCRIPCIÓN",
    "Concepto",
    "Estado_calidad",
}


def _read_facturas_df() -> pd.DataFrame:
    if not os.path.exists(ARCHIVO_EXCEL):
        return pd.DataFrame()

    try:
        return pd.read_excel(ARCHIVO_EXCEL, sheet_name="Facturas", engine="openpyxl")
    except Exception:
        return pd.read_excel(ARCHIVO_EXCEL, engine="openpyxl")


def _to_str(v: Any) -> str:
    if _es_valor_vacio_excel(v):
        return ""
    return str(v).strip()


def _money_float(v: Any) -> float:
    if v is None:
        return 0.0

    if isinstance(v, (int, float)) and not isinstance(v, bool):
        try:
            if pd.isna(v):
                return 0.0
        except Exception:
            pass
        try:
            return float(v)
        except Exception:
            return 0.0

    s = str(v).strip()
    if not s:
        return 0.0

    s = s.replace("\xa0", " ")
    s = s.replace("$", "").replace("COP", "").replace(" ", "")

    if "," in s and "." in s:
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    elif "," in s:
        s = s.replace(".", "").replace(",", ".")

    try:
        return float(s)
    except Exception:
        return 0.0


def _forzar_texto_excel(valor: Any) -> str:
    if _es_valor_vacio_excel(valor):
        return ""

    if isinstance(valor, str):
        s = valor.strip()
    else:
        s = str(valor).strip()

    if _es_valor_vacio_excel(s):
        return ""

    s_upper = s.upper().replace(",", ".")

    if "E+" in s_upper or "E-" in s_upper:
        try:
            return "{:.0f}".format(float(s_upper))
        except Exception:
            return s

    if re.fullmatch(r"\d+\.0", s):
        try:
            return str(int(float(s)))
        except Exception:
            return s

    return s


def _limpiar_descripcion(s: Any) -> str:
    if _es_valor_vacio_excel(s):
        return ""

    txt = str(s)
    txt = txt.replace("\r\n", "\n").replace("\r", "\n")
    txt = txt.replace("\n", "; ")
    txt = re.sub(r"[ \t]+", " ", txt).strip()
    txt = re.sub(r"(;\s*){2,}", "; ", txt).strip(" ;")
    return txt


def _descripcion_por_concepto(d: Dict[str, Any], concepto: str) -> str:
    full = d.get("DescripcionLineas", "") or ""

    if concepto == "IVA 19%":
        return d.get("DescripcionIVA19") or full
    if concepto == "IVA 5%":
        return d.get("DescripcionIVA5") or full

    return full


def _normalizar_valor_excel(v: Any) -> Any:
    return _sanitizar_valor_excel(v)


def _limpiar_dataframe_a_formato_largo(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame(columns=COLUMNAS_VALIDAS_FINALES)

    work = df.copy()

    for col in COLUMNAS_VALIDAS_FINALES:
        if col not in work.columns:
            work[col] = ""

    work = work[[c for c in COLUMNAS_VALIDAS_FINALES if c in work.columns]].copy()

    # Sanitización global: ningún "nan" textual o NaN técnico debe llegar al Excel.
    for col in work.columns:
        work[col] = work[col].apply(_sanitizar_valor_excel)

    for col in COLUMNAS_TEXTO:
        if col in work.columns:
            work[col] = work[col].apply(_forzar_texto_excel)

    if "DESCRIPCIÓN" in work.columns:
        work["DESCRIPCIÓN"] = work["DESCRIPCIÓN"].apply(_limpiar_descripcion)

    if "Concepto" in work.columns:
        work["Concepto"] = work["Concepto"].astype(str).str.strip()

    if "VALOR" in work.columns:
        work["VALOR"] = work["VALOR"].apply(_normalizar_valor_excel)

    if COLUMNA_ESTADO_CALIDAD in work.columns:
        work[COLUMNA_ESTADO_CALIDAD] = work[COLUMNA_ESTADO_CALIDAD].apply(_forzar_texto_excel).str.strip().str.upper()
        work.loc[work[COLUMNA_ESTADO_CALIDAD].isin(VALORES_TEXTO_VACIOS_EXCEL), COLUMNA_ESTADO_CALIDAD] = ""

    # Segunda pasada defensiva para evitar que pandas convierta algo a "nan" textual.
    for col in work.columns:
        work[col] = work[col].apply(_sanitizar_valor_excel)

    return work


def _radicado_sort_key(v: Any) -> int:
    s = _to_str(v)
    if not s:
        return 10**12
    try:
        return int(float(s))
    except Exception:
        return 10**12


def _fecha_sort_key(row: pd.Series) -> str:
    y = _to_str(row.get("Año"))
    m = _to_str(row.get("Mes"))
    d = _to_str(row.get("Día"))

    try:
        yy = int(float(y)) if y else 9999
        mm = int(float(m)) if m else 99
        dd = int(float(d)) if d else 99
        return f"{yy:04d}-{mm:02d}-{dd:02d}"
    except Exception:
        return "9999-99-99"


def _concepto_sort_key(v: Any) -> int:
    s = str(v or "").strip()
    return CONCEPTO_ORDEN.get(s, 999)


def _ordenar_dataframe_facturas(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return df

    work = df.copy()

    work["__rad_sort__"] = work["Radicado"].apply(_radicado_sort_key) if "Radicado" in work.columns else 10**12
    work["__fecha_sort__"] = work.apply(_fecha_sort_key, axis=1)
    work["__concepto_sort__"] = work["Concepto"].apply(_concepto_sort_key) if "Concepto" in work.columns else 999

    sort_cols = ["__rad_sort__", "__fecha_sort__"]
    if "Número de factura" in work.columns:
        sort_cols.append("Número de factura")
    sort_cols.append("__concepto_sort__")

    work = (
        work.sort_values(sort_cols, kind="mergesort")
        .drop(columns=["__rad_sort__", "__fecha_sort__", "__concepto_sort__"], errors="ignore")
        .reset_index(drop=True)
    )

    return work


def _es_recibo_caja_ph_sin_cufe_aceptable_excel_20260521(base: Dict[str, Any], filas: list[dict]) -> bool:
    """
    Excepción controlada para documentos tipo recibo de caja/propiedad horizontal.

    Caso validado H8:
    - RC 601 Abril26 -MP.pdf / EDIFICIO MURO DE PIEDRA - PROPIEDAD HORIZONTAL.
    - No trae CUFE porque no es factura electrónica DIAN.
    - Sí trae emisor, NIT, cliente, ciudad, número, fecha, descripción y total.

    No se aplica de forma genérica a cualquier PDF sin CUFE. Debe parecer recibo
    de caja/propiedad horizontal y tener el resto de campos operativos completos.
    """
    try:
        archivo = _to_str(base.get("Archivo")).upper()
        empresa = _to_str(base.get("Empresa emisora")).upper()
        numero = _to_str(base.get("Número de factura")).upper()
        descripcion = " ".join(
            _to_str((r or {}).get("DESCRIPCIÓN"))
            for r in (filas or [])
            if isinstance(r, dict)
        ).upper()

        texto = f"{archivo} {empresa} {numero} {descripcion}"

        parece_recibo = (
            "RC " in texto
            or "RECIBO" in texto
            or "COMPROBANTE" in texto
            or numero.startswith("R-")
        )

        parece_ph = (
            "PROPIEDAD HORIZONTAL" in texto
            or "EDIFICIO MURO DE PIEDRA" in texto
            or "ADMON" in texto
            or "ADMINISTRACION" in texto
            or "ADMINISTRACIÓN" in texto
        )

        if not (parece_recibo and parece_ph):
            return False

        campos_necesarios = [
            "Radicado",
            "ProyectoProceso",
            "Archivo",
            "Empresa emisora",
            "Ciudad emisora",
            "Código ciudad",
            "NIT",
            "Cliente",
            "Número de factura",
            "Año",
            "Mes",
            "Día",
            "DESCRIPCIÓN",
        ]

        return all(_valor_presente_calidad_excel(base.get(c)) for c in campos_necesarios)
    except Exception:
        return False


def _calidad_visual_por_grupo(filas: list[dict]) -> str:
    """
    Clasificación REAL por factura para Estado_calidad.

    REGLAS FINALES:
    - COMPLETA:
        * Todos los campos obligatorios llenos.
        * Total > 0.

    - PARCIAL:
        * Falta algún campo obligatorio.
        * O Total <= 0.
        * Pero existe información útil.

    - MINIMA:
        * Registro mínimo obligatorio.
        * O factura casi vacía.
        * O faltan varios datos base reales aunque exista Radicado/Archivo/CUFE.

    IMPORTANTE:
    - Actividad económica NO cuenta.
    - Tipo de contribuyente NO cuenta.
    - En conceptos, SOLO se valida Total.
    - IVA 5%, IVA 19%, Retención de IVA, Retención de ICA y Retención en la fuente
      pueden estar en cero sin afectar la calidad.
    """
    if not filas:
        return "MINIMA"

    base: Dict[str, Any] = {}

    # Consolidar la factura completa, no evaluar fila por fila.
    for r in filas:
        if not isinstance(r, dict):
            continue

        for k, v in r.items():
            if (
                not _valor_presente_calidad_excel(base.get(k))
                and _valor_presente_calidad_excel(v)
            ):
                base[k] = v

    # Detectar registro mínimo por descripción.
    descripcion_total = " ".join(
        _to_str((r or {}).get("DESCRIPCIÓN"))
        for r in filas
        if isinstance(r, dict)
    ).upper()

    if "REGISTRO MINIMO" in descripcion_total or "REGISTRO MÍNIMO" in descripcion_total:
        return "MINIMA"

    # Obtener el Total real únicamente desde la fila Concepto == Total.
    total = 0.0
    total_encontrado = False

    for r in filas:
        if not isinstance(r, dict):
            continue

        concepto = _to_str(r.get("Concepto")).strip().lower()

        if concepto == "total":
            total = _money_float(r.get("VALOR"))
            total_encontrado = True
            break

    # Validar campos obligatorios.
    faltantes = []

    for campo in CAMPOS_CALIDAD_COMPLETA_EXCEL:
        if not _valor_presente_calidad_excel(base.get(campo)):
            faltantes.append(campo)

    # Excepción controlada H8/G5:
    # algunos recibos de caja/propiedad horizontal no tienen CUFE por naturaleza,
    # pero sí pueden quedar operativamente completos si todos los demás campos
    # relevantes y el total están presentes.
    if "CUFE" in faltantes and _es_recibo_caja_ph_sin_cufe_aceptable_excel_20260521(base, filas):
        faltantes = [f for f in faltantes if f != "CUFE"]

    # Validar Total.
    if not total_encontrado or total <= 0:
        faltantes.append("Total")

    if not faltantes:
        return "COMPLETA"

    # Si tiene muy pocos datos útiles, se considera mínima.
    # Importante 2026-06-10:
    # Radicado/Proyecto/Archivo/CUFE pueden existir por el correo o por fallback,
    # pero no significan que la factura esté realmente identificada. Para evitar
    # falsos PARCIAL, se evalúa una base real: emisor, NIT, cliente, número,
    # fecha y descripción.
    campos_presentes = sum(
        1
        for campo in CAMPOS_CALIDAD_COMPLETA_EXCEL
        if _valor_presente_calidad_excel(base.get(campo))
    )

    campos_base_real_presentes = sum(
        1
        for campo in CAMPOS_CALIDAD_BASE_REAL_EXCEL
        if _valor_presente_calidad_excel(base.get(campo))
    )

    faltantes_base_real = [
        campo
        for campo in CAMPOS_CALIDAD_BASE_REAL_EXCEL
        if not _valor_presente_calidad_excel(base.get(campo))
    ]

    if total_encontrado and total > 0:
        campos_presentes += 1

    # Caso casi vacío: antes podía quedar PARCIAL si tenía Radicado/Archivo/CUFE.
    if campos_presentes <= 6:
        return "MINIMA"

    # Si no existe una base real mínima de identificación, debe ser MINIMA.
    if campos_base_real_presentes <= 4:
        return "MINIMA"

    # Si faltan campos de identidad fuertes al mismo tiempo, también es MINIMA.
    identidad_fuerte_faltante = (
        not _valor_presente_calidad_excel(base.get("Empresa emisora"))
        and not _valor_presente_calidad_excel(base.get("NIT"))
    )
    descripcion_faltante = not _valor_presente_calidad_excel(base.get("DESCRIPCIÓN"))
    fecha_incompleta = any(
        not _valor_presente_calidad_excel(base.get(c))
        for c in ["Año", "Mes", "Día"]
    )

    if identidad_fuerte_faltante and (descripcion_faltante or fecha_incompleta):
        return "MINIMA"

    # Si faltan muchos datos de la base real, no se considera parcial útil.
    if len(faltantes_base_real) >= 4:
        return "MINIMA"

    return "PARCIAL"


def _calcular_estado_calidad_para_filas(filas: List[Dict[str, Any]]) -> str:
    """
    Wrapper seguro para evitar estados inválidos.
    """
    try:
        estado = _calidad_visual_por_grupo(filas)
    except Exception:
        return "PARCIAL"

    if estado not in {"COMPLETA", "PARCIAL", "MINIMA"}:
        return "PARCIAL"

    return estado


def _recalcular_estado_calidad_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """
    Recalcula Estado_calidad para TODO el Excel antes de guardar.

    Esto evita que queden valores viejos como PARCIAL cuando la factura ya está completa.
    Se calcula por factura completa, no por fila individual.

    Reglas:
    - Actividad económica NO cuenta.
    - Tipo de contribuyente NO cuenta.
    - En conceptos, solo Total debe ser > 0.
    - IVA y retenciones pueden estar en cero.
    """
    if df is None or df.empty:
        return df

    work = df.copy()

    for col in COLUMNAS_VALIDAS_FINALES:
        if col not in work.columns:
            work[col] = ""

    work = work[COLUMNAS_VALIDAS_FINALES].copy()

    grupos: Dict[tuple, List[Dict[str, Any]]] = {}

    for _, row in work.iterrows():
        row_dict = row.to_dict()
        key = _factura_group_key_excel_20260515(row_dict)
        grupos.setdefault(key, []).append(row_dict)

    estado_por_key = {
        key: _calcular_estado_calidad_para_filas(filas)
        for key, filas in grupos.items()
    }

    estados = []
    for _, row in work.iterrows():
        key = _factura_group_key_excel_20260515(row.to_dict())
        estados.append(estado_por_key.get(key, "PARCIAL"))

    work[COLUMNA_ESTADO_CALIDAD] = estados
    return work



def _aplicar_formato_visual_facturas() -> None:
    """
    Versión segura:
    - NO usa auto_filter manual adicional.
    - NO aplica bordes celda por celda.
    - NO combina tabla + estilos agresivos.
    - Sí deja tabla, encabezado, anchos, wrap, colores y freeze panes.
    """
    wb = load_workbook(ARCHIVO_EXCEL)
    ws = wb["Facturas"] if "Facturas" in wb.sheetnames else wb.active
    ws.title = "Facturas"

    max_row = ws.max_row
    max_col = ws.max_column

    if max_row < 1 or max_col < 1:
        wb.save(ARCHIVO_EXCEL)
        wb.close()
        return

    # Encabezados como texto válido.
    for c in range(1, max_col + 1):
        value = ws.cell(row=1, column=c).value
        ws.cell(row=1, column=c).value = _to_str(value) or f"Columna{c}"

    last_col = get_column_letter(max_col)
    table_ref = f"A1:{last_col}{max_row}"

    # Borrar tablas previas y reconstruir TblFacturas.
    try:
        for table_name in list(ws.tables.keys()):
            del ws.tables[table_name]
    except Exception:
        pass

    tbl = Table(displayName="TblFacturas", ref=table_ref)
    tbl.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tbl)

    # IMPORTANTE: no asignar ws.auto_filter.ref aparte.
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False

    header_idx = {}
    for c in range(1, max_col + 1):
        header = _to_str(ws.cell(row=1, column=c).value)
        header_idx[header] = c

    # Encabezado.
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for c in range(1, max_col + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        header = _to_str(cell.value)
        ws.column_dimensions[get_column_letter(c)].width = ANCHO_COLUMNAS.get(header, 18)

    ws.row_dimensions[1].height = 28

    # Calcular calidad por factura.
    grupos = {}

    for r in range(2, max_row + 1):
        row_dict = {}
        for header, idx in header_idx.items():
            row_dict[header] = ws.cell(r, idx).value

        key = _factura_group_key_excel_20260515(row_dict)
        grupos.setdefault(key, []).append(row_dict)

    calidad_por_key = {k: _calidad_visual_por_grupo(v) for k, v in grupos.items()}

    fill_normal = PatternFill("solid", fgColor="FFFFFF")
    fill_parcial = PatternFill("solid", fgColor="FFF2CC")
    fill_minima = PatternFill("solid", fgColor="FCE4D6")
    fill_total = PatternFill("solid", fgColor="E2F0D9")

    for r in range(2, max_row + 1):
        row_dict_for_key = {}
        for header, idx in header_idx.items():
            row_dict_for_key[header] = ws.cell(r, idx).value

        key = _factura_group_key_excel_20260515(row_dict_for_key)

        calidad = calidad_por_key.get(key, "PARCIAL")

        # IMPORTANTE:
        # La calidad recalculada no solo debe usarse para pintar.
        # También debe escribirse en la celda Estado_calidad para no dejar valores viejos.
        col_estado = header_idx.get(COLUMNA_ESTADO_CALIDAD)
        if col_estado:
            estado_cell = ws.cell(row=r, column=col_estado)
            estado_cell.value = calidad
            estado_cell.number_format = "@"

        concepto = _to_str(ws.cell(r, header_idx.get("Concepto", 1)).value)

        row_fill = fill_normal
        if calidad == "PARCIAL":
            row_fill = fill_parcial
        elif calidad == "MINIMA":
            row_fill = fill_minima

        if concepto == "Total":
            row_fill = fill_total if calidad == "COMPLETA" else row_fill

        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            header = _to_str(ws.cell(row=1, column=c).value)

            cell.fill = row_fill

            wrap = header in COLUMNAS_WRAP
            shrink = header in {"CUFE", "Número de factura", "NIT", "Radicado"}

            if header == "VALOR":
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right", vertical="top")
            else:
                if header in COLUMNAS_TEXTO:
                    cell.number_format = "@"

                cell.alignment = Alignment(
                    horizontal="left",
                    vertical="top",
                    wrap_text=wrap,
                    shrink_to_fit=shrink,
                )

        if concepto == "Total":
            for c in range(1, max_col + 1):
                ws.cell(r, c).font = Font(bold=True)

        ws.row_dimensions[r].height = 32

    # Formato condicional amplio para que Excel Web pinte nuevas filas insertadas por Graph.
    try:
        ws.conditional_formatting._cf_rules.clear()
    except Exception:
        pass

    col_estado = header_idx.get(COLUMNA_ESTADO_CALIDAD)
    col_concepto = header_idx.get("Concepto")
    if col_estado and col_concepto:
        data_range = f"A1:{last_col}100000"
        estado_letter = get_column_letter(col_estado)
        concepto_letter = get_column_letter(col_concepto)

        ws.conditional_formatting.add(
            data_range,
            FormulaRule(formula=[f'=${estado_letter}1="MINIMA"'], fill=fill_minima),
        )
        ws.conditional_formatting.add(
            data_range,
            FormulaRule(formula=[f'=${estado_letter}1="PARCIAL"'], fill=fill_parcial),
        )
        ws.conditional_formatting.add(
            data_range,
            FormulaRule(formula=[f'=${concepto_letter}1="Total"'], font=Font(bold=True)),
        )

    wb.save(ARCHIVO_EXCEL)
    wb.close()


def obtener_cufes_existentes() -> Set[str]:
    if not os.path.exists(ARCHIVO_EXCEL):
        return set()

    try:
        df = _read_facturas_df()
    except Exception as e:
        print(f"[Excel] No se pudo leer facturas.xlsx para índice de CUFEs: {e}")
        return set()

    if df.empty or "CUFE" not in df.columns:
        return set()

    cufes: Set[str] = set()
    for v in df["CUFE"]:
        s = _to_str(v)
        if s:
            cufes.add(s)

    return cufes



# ============================================================
# PATCH 2026-05-12 - UPSERT real para facturas.xlsx
# ============================================================
# Problema detectado:
# - El parser puede mejorar una factura ya registrada, pero el Excel/flujo
#   puede conservar filas viejas o reportar 0 nuevos.
# - Se requiere actualizar filas existentes cuando la misma factura + concepto
#   vuelve con datos mejores.
#
# Criterio de identidad 2026-05-15:
# 1. Radicado + CUFE + Concepto cuando existe Radicado y CUFE.
# 2. Radicado + Archivo + Número + Concepto cuando existe Radicado.
# 3. CUFE + Archivo + Número + Concepto solo si no existe Radicado.
# 4. Últimos recursos con Archivo/Número, evitando colapsar aprobaciones distintas.
#
# Además:
# - Se cuentan filas nuevas y filas realmente actualizadas.
# - El return de guardar_en_excel será filas afectadas = nuevas + actualizadas,
#   para que el controller no interprete una corrección como "0 cambios".
# ============================================================

def _normalizar_texto_key_excel_20260512(v: Any) -> str:
    s = _to_str(v)
    if not s:
        return ""

    s = s.replace("\\", "/")
    s = os.path.basename(s)
    s = s.strip().lower()
    s = re.sub(r"\s+", " ", s)
    return s


def _normalizar_cufe_key_excel_20260512(v: Any) -> str:
    s = _to_str(v).strip().lower()
    if not s:
        return ""

    c = re.sub(r"[^0-9a-f]", "", s)
    if len(c) >= 64:
        return c

    return ""


def _normalizar_numero_key_excel_20260512(v: Any) -> str:
    s = _to_str(v).upper().strip()
    if not s:
        return ""

    s = re.sub(r"[^A-Z0-9]+", "", s)
    return s


def _normalizar_radicado_key_excel_20260512(v: Any) -> str:
    s = _to_str(v).strip()
    if not s:
        return ""

    if re.fullmatch(r"\d+\.0", s):
        try:
            return str(int(float(s)))
        except Exception:
            return s

    return s


def _normalizar_concepto_key_excel_20260512(v: Any) -> str:
    s = _to_str(v).strip()
    if not s:
        return ""

    # Mantener nombres oficiales para evitar diferencias por mayúsculas.
    for c in CONCEPTOS_BASE_FIJOS:
        if s.lower() == c.lower():
            return c

    return s


def _normalizar_valor_para_comparar_excel_20260512(v: Any) -> str:
    if v is None:
        return ""

    try:
        if isinstance(v, float) and pd.isna(v):
            return ""
    except Exception:
        pass

    # Valores monetarios se comparan como número redondeado.
    try:
        if isinstance(v, (int, float)) and not isinstance(v, bool):
            return f"{float(v):.6f}"
    except Exception:
        pass

    s = _to_str(v)
    if not s:
        return ""

    # Si parece número monetario, normalizarlo.
    if re.search(r"\d", s) and re.fullmatch(r"[\s$COPcop0-9,.\-]+", s):
        try:
            return f"{_money_float(s):.6f}"
        except Exception:
            pass

    return re.sub(r"\s+", " ", s.strip())


def _factura_group_key_excel_20260515(row: Dict[str, Any]) -> Tuple[str, str]:
    """
    Llave de agrupación por factura completa, sin Concepto.

    Regla principal: el Radicado separa aprobaciones distintas.
    Esto evita que el Excel colapse facturas cuando llega un mismo nombre de archivo,
    un mismo número o un CUFE reutilizado por error operativo en correos/reenvíos.
    """
    radicado = _normalizar_radicado_key_excel_20260512(row.get("Radicado"))
    cufe = _normalizar_cufe_key_excel_20260512(row.get("CUFE"))
    archivo = _normalizar_texto_key_excel_20260512(row.get("Archivo"))
    numero = _normalizar_numero_key_excel_20260512(row.get("Número de factura"))

    if radicado and cufe:
        return ("RAD_CUFE", f"{radicado}|{cufe}")

    if radicado:
        return ("RAD_ARCH_NUM", f"{radicado}|{archivo}|{numero}")

    if cufe:
        return ("CUFE_ARCH_NUM", f"{cufe}|{archivo}|{numero}")

    if archivo and numero:
        return ("ARCH_NUM", f"{archivo}|{numero}")

    if archivo:
        return ("ARCH", archivo)

    return ("ROW", str(id(row)))


def _texto_generico_malo_excel_20260515(v: Any) -> bool:
    s = _to_str(v).upper()
    if not s:
        return True

    malos = {
        "NAN", "NONE", "NULL", "N/A", "NA", "SIN_DATO", "SIN DATO",
        "REGISTRO MINIMO OBLIGATORIO", "REGISTRO MÍNIMO OBLIGATORIO",
    }
    return s in malos


def _descripcion_contaminada_excel_20260521(v: Any) -> bool:
    """
    Detecta descripciones que vienen de columnas/encabezados mezclados por pdfminer
    o de textos técnicos de la representación gráfica, no del producto/servicio real.
    """
    s = _to_str(v)
    if not s:
        return True

    su = s.upper()
    su_compacto = re.sub(r"\s+", " ", su).strip()

    patrones_malos = [
        "REGISTRO MINIMO",
        "REGISTRO MÍNIMO",
        "U MEDIDA UND. VALOR UNITARIO",
        "VALOR UNITARIO IVA",
        "FECHA VALIDACION DIAN",
        "FECHA VALIDACIÓN DIAN",
        "I N A D",
        "PROVEEDOR TECNOLÓGICO",
        "PROVEEDOR TECNOLOGICO",
        "SIMBA SOFTWARE",
        "WORLD OFFICE",
        "FABRICANTE Y PROVEEDOR",
        "ABOUT:BLANK",
        "HOJA 1 DE",
        "UNITARIO DE HOJA",
        "WSD WSD",
        "LIQUIDACION 20",
        "LIQUIDACIÓN 20",
        "DETALLE; CONCEPTO/FUNCIONES",
        "E-MAIL FACT. ELECT.",
        "CONTACTO:",
        "LIQUIDADO POR:",
        "CANT DCTO",
        "VALOR IVA",
        "TOTAL LINEAS O ITEMS",
        "TOTAL LÍNEAS O ÍTEMS",
        "NOTAS: SON",
        "NETO FACTURA",
        "CARGOS:",
        "DESCUENTOS:",
        "SUBTOTAL:",
    ]

    if any(p in su_compacto for p in patrones_malos):
        return True

    # Texto típico de extracción invertida/letra por letra:
    # "T; : I N A D n ó c a d i i l a V..."
    letras_sueltas = re.findall(r"(?:^|[ ;:])(?:[A-ZÁÉÍÓÚÑ])(?:[ ;:])", su)
    if len(letras_sueltas) >= 8:
        return True

    # Demasiados separadores suele indicar tabla/encabezados contaminados.
    if su.count(";") >= 8 and len(su) > 120:
        return True

    return False


def _score_texto_excel_20260515(v: Any) -> int:
    """
    Puntúa texto útil para decidir si una descripción nueva debe reemplazar
    una descripción vieja.

    Antes el puntaje dependía casi solo del largo. Eso hacía que textos basura
    largos ganaran sobre descripciones nuevas limpias, por ejemplo:
    - "U Medida Und. Valor Unitario IVA..."
    - "Fecha Validación DIAN..."
    - "LIQUIDACION ... Contacto..."
    """
    s = _limpiar_descripcion(v)
    if not s:
        return 0

    if _texto_generico_malo_excel_20260515(s):
        return 1

    contaminada = _descripcion_contaminada_excel_20260521(s)

    # Base alta para texto limpio, base baja para texto contaminado.
    score = 20 if contaminada else 120

    # Premiar longitud útil, pero sin dejar que el texto enorme gane por tamaño.
    palabras = re.findall(r"[A-ZÁÉÍÓÚÑa-záéíóúñ0-9]{3,}", s)
    score += min(60, len(palabras) * 3)
    score += min(40, len(s) // 8)

    # Castigo fuerte para contaminación.
    if contaminada:
        score -= 100

    return max(1, score)


def _score_fila_para_merge_excel_20260521(row: Dict[str, Any]) -> int:
    """
    Puntaje operativo de una fila para decidir cuál dato gana cuando dos filas
    representan el mismo Radicado+Archivo+Concepto.

    Evita que un duplicado viejo/parcial sobrescriba una mejora completa cuando
    el número de factura o el CUFE cambian durante la corrección del parser.
    """
    if not isinstance(row, dict):
        return 0

    score = 0

    campos_fuertes = [
        "Radicado",
        "ProyectoProceso",
        "Archivo",
        "Empresa emisora",
        "CUFE",
        "Ciudad emisora",
        "Código ciudad",
        "NIT",
        "Cliente",
        "Número de factura",
        "Año",
        "Mes",
        "Día",
        "DESCRIPCIÓN",
    ]

    for campo in campos_fuertes:
        if _valor_presente_calidad_excel(row.get(campo)):
            score += 10

    if _valor_presente_calidad_excel(row.get("CUFE")):
        score += 25

    estado = _to_str(row.get(COLUMNA_ESTADO_CALIDAD)).upper()
    if estado == "COMPLETA":
        score += 40
    elif estado == "PARCIAL":
        score += 15
    elif estado == "MINIMA":
        score -= 20

    desc = _limpiar_descripcion(row.get("DESCRIPCIÓN"))
    if desc:
        score += min(30, _score_texto_excel_20260515(desc) // 6)
        if _descripcion_contaminada_excel_20260521(desc):
            score -= 30

    valor = abs(_money_float(row.get("VALOR")))
    if valor > 0:
        score += 5

    return score


def _merge_rows_prefer_better_excel_20260515(
    viejo: Dict[str, Any],
    nuevo: Dict[str, Any],
) -> Dict[str, Any]:
    """
    Fusiona dos filas con la misma llave de UPSERT sin perder datos buenos.

    - Si nuevo trae un campo vacío, conserva el viejo.
    - Si viejo está vacío y nuevo trae dato, usa nuevo.
    - Para VALOR: conserva valor viejo si nuevo viene en cero y viejo tenía valor.
    - Para DESCRIPCIÓN: evita que un registro mínimo pise una descripción real.
    - Estado_calidad se recalcula después sobre todo el grupo.
    """
    out = {col: viejo.get(col, "") for col in COLUMNAS_VALIDAS_FINALES}

    for col in COLUMNAS_VALIDAS_FINALES:
        if col == COLUMNA_ESTADO_CALIDAD:
            continue

        old_val = out.get(col, "")
        new_val = nuevo.get(col, "")

        old_pres = _valor_presente_calidad_excel(old_val)
        new_pres = _valor_presente_calidad_excel(new_val)

        if col == "VALOR":
            old_num = _money_float(old_val)
            new_num = _money_float(new_val)
            old_score = _score_fila_para_merge_excel_20260521(viejo)
            new_score = _score_fila_para_merge_excel_20260521(nuevo)

            if abs(new_num) > 0.000001 and abs(old_num) <= 0.000001:
                out[col] = new_val
            elif abs(old_num) > 0.000001 and abs(new_num) <= 0.000001:
                out[col] = old_val
            elif abs(new_num) > 0.000001 and abs(old_num) > 0.000001:
                # Si una fila es claramente más completa, gana su valor aunque sea
                # menor. Esto corrige casos como Roberto Car, donde el valor viejo
                # era un NIT/monto mal escalado y la mejora real es menor.
                if new_score >= old_score + 15:
                    out[col] = new_val
                elif old_score >= new_score + 15:
                    out[col] = old_val
                else:
                    # Si la calidad es similar, preferimos el valor entrante.
                    # En un upsert normal el entrante es el parser corregido.
                    out[col] = new_val
            elif new_pres:
                out[col] = new_val
            else:
                out[col] = old_val
            continue

        if col == "DESCRIPCIÓN":
            old_clean = _limpiar_descripcion(old_val)
            new_clean = _limpiar_descripcion(new_val)

            old_score = _score_texto_excel_20260515(old_clean)
            new_score = _score_texto_excel_20260515(new_clean)

            old_bad = _descripcion_contaminada_excel_20260521(old_clean)
            new_bad = _descripcion_contaminada_excel_20260521(new_clean)

            if new_pres and not old_pres:
                out[col] = new_clean
            elif old_pres and not new_pres:
                out[col] = old_clean
            elif new_pres and old_pres:
                # Regla 2026-05-21-G3:
                # Si la descripción vieja está contaminada y la nueva está limpia,
                # la nueva debe reemplazarla aunque sea más corta.
                if old_bad and not new_bad:
                    out[col] = new_clean
                elif new_bad and not old_bad:
                    out[col] = old_clean
                elif new_score >= old_score:
                    out[col] = new_clean
                else:
                    out[col] = old_clean
            else:
                out[col] = ""
            continue

        if new_pres and not old_pres:
            out[col] = new_val
        elif old_pres and not new_pres:
            out[col] = old_val
        elif new_pres and old_pres:
            old_score = _score_fila_para_merge_excel_20260521(viejo)
            new_score = _score_fila_para_merge_excel_20260521(nuevo)

            # Si la fila entrante es claramente peor, no debe pisar campos buenos
            # de una fila ya completa. Esto repara duplicados existentes creados
            # por cambios de CUFE/número sin perder la mejora buena.
            if old_score >= new_score + 20:
                out[col] = old_val
            else:
                out[col] = new_val
        else:
            out[col] = ""

    # Se recalcula más adelante, pero dejamos algo consistente temporalmente.
    out[COLUMNA_ESTADO_CALIDAD] = nuevo.get(COLUMNA_ESTADO_CALIDAD) or viejo.get(COLUMNA_ESTADO_CALIDAD) or ""
    return out


def _validar_integridad_7_conceptos_excel_20260515(df: pd.DataFrame) -> Tuple[int, int, List[Tuple[Tuple[str, str], int, str]]]:
    """
    Retorna: total_grupos, grupos_incompletos, muestra_incompletos.
    Solo diagnostica; no bloquea el guardado.
    """
    if df is None or df.empty:
        return 0, 0, []

    grupos: Dict[Tuple[str, str], set] = {}
    for _, row in df.iterrows():
        row_dict = row.to_dict()
        key = _factura_group_key_excel_20260515(row_dict)
        concepto = _normalizar_concepto_key_excel_20260512(row_dict.get("Concepto"))
        if concepto:
            grupos.setdefault(key, set()).add(concepto)

    incompletos = []
    requeridos = set(CONCEPTOS_BASE_FIJOS)
    for key, conceptos in grupos.items():
        faltan = sorted(requeridos - conceptos, key=lambda x: CONCEPTO_ORDEN.get(x, 999))
        if faltan or len(conceptos) != len(requeridos):
            incompletos.append((key, len(conceptos), ", ".join(faltan)))

    return len(grupos), len(incompletos), incompletos[:12]


def _legacy_rad_arch_num_key_excel_20260521(row: Dict[str, Any]) -> Tuple[str, str] | None:
    """
    Llave secundaria segura para reparar/actualizar registros antiguos.

    Caso detectado H6:
    - La fila vieja no tenía CUFE/CUDE, por lo que su llave era Radicado+Archivo+Número+Concepto.
    - La fila nueva sí trae CUFE/CUDE, por lo que su llave primaria pasa a Radicado+CUFE+Concepto.
    - Si no usamos esta llave secundaria, el upsert inserta 7 filas nuevas y duplica la factura.

    Esta llave solo se usa cuando existen Radicado + Archivo + Número + Concepto,
    por lo que es suficientemente segura para NO mezclar aprobaciones distintas.
    """
    radicado = _normalizar_radicado_key_excel_20260512(row.get("Radicado"))
    archivo = _normalizar_texto_key_excel_20260512(row.get("Archivo"))
    numero = _normalizar_numero_key_excel_20260512(row.get("Número de factura"))
    concepto = _normalizar_concepto_key_excel_20260512(row.get("Concepto"))

    if radicado and archivo and numero and concepto:
        return ("LEGACY_RAD_ARCH_NUM", f"{radicado}|{archivo}|{numero}|{concepto}")

    return None


def _legacy_rad_arch_key_excel_20260521_g5(row: Dict[str, Any]) -> Tuple[str, str] | None:
    """
    Llave de respaldo G5: Radicado + Archivo + Concepto.

    Se usa para evitar duplicados cuando una mejora cambia el número de factura
    o agrega CUFE/CUDE a una fila vieja que antes no lo tenía.

    Caso H8: Claro/COMCEL tenía número viejo 586455300000013255 y la mejora
    correcta llega como 3-292586455 con CUFE real. La llave Radicado+Archivo+
    Concepto permite actualizar la misma factura en vez de crear otras 7 filas.
    """
    radicado = _normalizar_radicado_key_excel_20260512(row.get("Radicado"))
    archivo = _normalizar_texto_key_excel_20260512(row.get("Archivo"))
    concepto = _normalizar_concepto_key_excel_20260512(row.get("Concepto"))

    if radicado and archivo and concepto:
        return ("LEGACY_RAD_ARCH", f"{radicado}|{archivo}|{concepto}")

    return None


def _dedupe_key_row_excel_20260512(row: Dict[str, Any]) -> Tuple[str, str]:
    concepto = _normalizar_concepto_key_excel_20260512(row.get("Concepto"))
    tipo_factura, key_factura = _factura_group_key_excel_20260515(row)
    return (tipo_factura, f"{key_factura}|{concepto}")


def _upsert_keys_row_excel_20260521(row: Dict[str, Any]) -> List[Tuple[str, str]]:
    """
    Devuelve llaves de búsqueda para una fila.

    Orden:
    1. Llave primaria actual, por ejemplo Radicado+CUFE+Concepto.
    2. Llave legacy segura Radicado+Archivo+Número+Concepto.
    3. Llave G5 Radicado+Archivo+Concepto.

    Así una mejora que ahora trae CUFE o corrige el número de factura puede
    actualizar la fila vieja, sin crear un segundo grupo de 7 filas.
    """
    keys: List[Tuple[str, str]] = []

    primary = _dedupe_key_row_excel_20260512(row)
    if primary and primary[1].strip("|"):
        keys.append(primary)

    legacy = _legacy_rad_arch_num_key_excel_20260521(row)
    if legacy and legacy not in keys:
        keys.append(legacy)

    legacy_arch = _legacy_rad_arch_key_excel_20260521_g5(row)
    if legacy_arch and legacy_arch not in keys:
        keys.append(legacy_arch)

    return keys


# ============================================================
# PATCH 2026-06-10-H3 - Anti-duplicado estricto de registros mínimos por Radicado
# ============================================================
# Problema detectado en primera corrida de producción normal:
# - Ya existía una factura COMPLETA por XML.
# - Producción volvió a tomar el PDF de la misma aprobación/radicado.
# - El PDF cayó a REGISTRO MÍNIMO y se insertó como si fuera otra factura,
#   porque su llave era Radicado+Archivo+Número y no coincidía con Radicado+CUFE.
#
# Regla H2:
# - Bloqueaba MINIMA / REGISTRO MÍNIMO si encontraba una fila mejor compatible
#   para el mismo Radicado + Concepto usando número/empresa/archivo.
#
# Ajuste H3:
# - Si una fila nueva es MINIMA / REGISTRO MÍNIMO y ya existe una fila mejor
#   para el mismo Radicado + Concepto, NO se inserta, aunque el PDF tenga otro
#   nombre, otra empresa mal extraída o un número incompleto.
# - Si el Radicado no existe con una fila mejor, el mínimo se conserva para no
#   perder casos legítimos sin XML o documentos no electrónicos.
# ============================================================

EMPRESA_TOKENS_IGNORAR_EXCEL = {
    "S", "A", "SAS", "SA", "S A S", "LTDA", "LIMITADA", "CIA", "COMPANIA", "COMPAÑIA",
    "COLOMBIA", "SUCURSAL", "NIT", "DE", "DEL", "LA", "EL", "Y", "EN",
}


def _normalizar_ascii_excel_20260610(v: Any) -> str:
    s = _to_str(v).upper()
    if not s:
        return ""
    try:
        s = unicodedata.normalize("NFKD", s)
        s = "".join(ch for ch in s if not unicodedata.combining(ch))
    except Exception:
        pass
    s = re.sub(r"[^A-Z0-9]+", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _normalizar_empresa_para_match_excel_20260610(v: Any) -> str:
    s = _normalizar_ascii_excel_20260610(v)
    if not s:
        return ""
    tokens = [t for t in s.split() if t and t not in EMPRESA_TOKENS_IGNORAR_EXCEL]
    return " ".join(tokens)


def _empresas_compatibles_excel_20260610(a: Any, b: Any) -> bool:
    """
    True si las empresas parecen la misma.
    Si una de las dos viene vacía, no bloquea el match porque muchos registros
    mínimos no tienen NIT/empresa completa.
    """
    ea = _normalizar_empresa_para_match_excel_20260610(a)
    eb = _normalizar_empresa_para_match_excel_20260610(b)

    if not ea or not eb:
        return True

    if ea == eb:
        return True

    if len(ea) >= 4 and len(eb) >= 4 and (ea in eb or eb in ea):
        return True

    ta = {t for t in ea.split() if len(t) >= 3}
    tb = {t for t in eb.split() if len(t) >= 3}

    if not ta or not tb:
        return False

    inter = ta & tb
    return len(inter) >= max(1, min(len(ta), len(tb)) // 2)


def _numeros_factura_compatibles_excel_20260610(a: Any, b: Any) -> bool:
    na = _normalizar_numero_key_excel_20260512(a)
    nb = _normalizar_numero_key_excel_20260512(b)

    if not na or not nb:
        return False

    if na == nb:
        return True

    # Casos reales: PDF extrae "K2" y XML trae "K2B6164633".
    # Se acepta prefijo/subcadena cuando hay mismo radicado y empresa compatible.
    if len(na) >= 2 and nb.startswith(na):
        return True
    if len(nb) >= 2 and na.startswith(nb):
        return True

    if len(na) >= 4 and len(nb) >= 4 and (na in nb or nb in na):
        return True

    return False


def _stems_archivo_compatibles_excel_20260610(a: Any, b: Any) -> bool:
    aa = Path(_normalizar_texto_key_excel_20260512(a)).stem.upper()
    bb = Path(_normalizar_texto_key_excel_20260512(b)).stem.upper()

    aa = re.sub(r"[^A-Z0-9]+", "", aa)
    bb = re.sub(r"[^A-Z0-9]+", "", bb)

    if not aa or not bb:
        return False

    if aa == bb:
        return True

    if len(aa) >= 8 and len(bb) >= 8 and (aa in bb or bb in aa):
        return True

    return False


def _row_es_minima_o_registro_minimo_excel_20260610(row: Dict[str, Any]) -> bool:
    if not isinstance(row, dict):
        return False

    estado = _to_str(row.get(COLUMNA_ESTADO_CALIDAD)).upper()
    descripcion = _to_str(row.get("DESCRIPCIÓN")).upper()
    fuente_archivo = _to_str(row.get("Archivo")).upper()

    if estado == "MINIMA":
        return True

    if "REGISTRO MINIMO" in descripcion or "REGISTRO MÍNIMO" in descripcion:
        return True

    # Algunos mínimos llegan sin descripción, pero con muchos campos base vacíos.
    base_real_presentes = sum(
        1
        for c in CAMPOS_CALIDAD_BASE_REAL_EXCEL
        if _valor_presente_calidad_excel(row.get(c))
    )
    if base_real_presentes <= 4 and not _valor_presente_calidad_excel(row.get("CUFE")):
        return True

    return False


def _row_es_mejor_que_minima_excel_20260610(row: Dict[str, Any]) -> bool:
    if not isinstance(row, dict):
        return False

    if _row_es_minima_o_registro_minimo_excel_20260610(row):
        return False

    estado = _to_str(row.get(COLUMNA_ESTADO_CALIDAD)).upper()
    if estado in {"COMPLETA", "PARCIAL"}:
        return True

    if _valor_presente_calidad_excel(row.get("CUFE")) and _valor_presente_calidad_excel(row.get("NIT")):
        return True

    if _score_fila_para_merge_excel_20260521(row) >= 100:
        return True

    return False


def _radicado_concepto_key_excel_20260610(row: Dict[str, Any]) -> Tuple[str, str] | None:
    radicado = _normalizar_radicado_key_excel_20260512(row.get("Radicado"))
    concepto = _normalizar_concepto_key_excel_20260512(row.get("Concepto"))
    if not radicado or not concepto:
        return None
    return (radicado, concepto)


def _row_minima_compatible_con_mejor_excel_20260610(minima: Dict[str, Any], mejor: Dict[str, Any]) -> bool:
    """
    Decide si una fila MINIMA parece duplicado de una fila mejor ya existente.
    Se exige mismo Radicado+Concepto por índice externo y compatibilidad adicional.
    """
    if not _empresas_compatibles_excel_20260610(minima.get("Empresa emisora"), mejor.get("Empresa emisora")):
        return False

    if _numeros_factura_compatibles_excel_20260610(minima.get("Número de factura"), mejor.get("Número de factura")):
        return True

    if _stems_archivo_compatibles_excel_20260610(minima.get("Archivo"), mejor.get("Archivo")):
        return True

    # Si el mínimo no trae número confiable, pero la empresa coincide y el mejor tiene CUFE/NIT,
    # se considera duplicado solo cuando el mínimo es claramente registro mínimo obligatorio.
    desc = _to_str(minima.get("DESCRIPCIÓN")).upper()
    if (
        ("REGISTRO MINIMO" in desc or "REGISTRO MÍNIMO" in desc)
        and _valor_presente_calidad_excel(mejor.get("CUFE"))
        and _valor_presente_calidad_excel(mejor.get("NIT"))
        and _empresas_compatibles_excel_20260610(minima.get("Empresa emisora"), mejor.get("Empresa emisora"))
    ):
        return True

    return False

def _row_signature_excel_20260512(row: Dict[str, Any]) -> str:
    parts = []

    for col in COLUMNAS_VALIDAS_FINALES:
        parts.append(f"{col}={_normalizar_valor_para_comparar_excel_20260512(row.get(col))}")

    return "|".join(parts)


def _consolidar_filas_por_upsert_excel_20260512(
    antiguo: pd.DataFrame,
    nuevo: pd.DataFrame,
) -> Tuple[pd.DataFrame, int, int]:
    """
    Devuelve:
    - DataFrame consolidado.
    - filas_nuevas.
    - filas_actualizadas.

    Versión 2026-05-21-G5:
    - Preserva filas históricas.
    - Repara duplicados seguros generados cuando una mejora nueva trae CUFE/CUDE
      o corrige el número de factura.
    - Para buscar una fila existente usa:
        1) llave primaria actual: Radicado+CUFE+Concepto, si existe CUFE.
        2) llave secundaria segura: Radicado+Archivo+Número+Concepto.
        3) llave de respaldo G5: Radicado+Archivo+Concepto.
    - Esto evita que Claro/COMCEL u otros casos pasen de 385 a 386 facturas por duplicado.
    """
    antiguo = _limpiar_dataframe_a_formato_largo(antiguo)
    nuevo = _limpiar_dataframe_a_formato_largo(nuevo)

    if antiguo is None or antiguo.empty:
        return nuevo.copy(), len(nuevo), 0

    if nuevo is None or nuevo.empty:
        return antiguo.copy(), 0, 0

    filas_ordenadas: List[Dict[str, Any]] = []
    index_por_key: Dict[Tuple[str, str], int] = {}
    firma_por_idx: Dict[int, str] = {}
    mejores_por_radicado_concepto: Dict[Tuple[str, str], List[Dict[str, Any]]] = {}

    nuevas = 0
    actualizadas = 0
    minimas_bloqueadas = 0

    def registrar_keys(idx: int, row: Dict[str, Any]) -> None:
        """Registra todas las llaves seguras de una fila hacia el índice físico."""
        for k in _upsert_keys_row_excel_20260521(row):
            if k not in index_por_key:
                index_por_key[k] = idx

    def registrar_mejor(row: Dict[str, Any]) -> None:
        """Indexa filas mejores que MINIMA por Radicado+Concepto para bloquear mínimos duplicados."""
        if not _row_es_mejor_que_minima_excel_20260610(row):
            return
        k = _radicado_concepto_key_excel_20260610(row)
        if not k:
            return
        bucket = mejores_por_radicado_concepto.setdefault(k, [])
        bucket.append(row)

    def es_minima_duplicada_de_mejor(row: Dict[str, Any]) -> bool:
        """
        Evita insertar PDF/registro mínimo cuando ya hay una fila mejor
        para el mismo Radicado + Concepto.

        H3 endurece la regla:
        - No exige compatibilidad por empresa, número o archivo.
        - Motivo: en registros mínimos esos campos pueden venir mal extraídos
          o vacíos, como ocurrió con PDF mínimo frente a XML completo.
        - Solo bloquea cuando ya existe una fila mejor indexada para el mismo
          Radicado + Concepto. Si no existe mejor, el mínimo se deja pasar.
        """
        if not _row_es_minima_o_registro_minimo_excel_20260610(row):
            return False

        k = _radicado_concepto_key_excel_20260610(row)
        if not k:
            return False

        return bool(mejores_por_radicado_concepto.get(k))

    def buscar_idx(row: Dict[str, Any]) -> int | None:
        """Busca por llave primaria y luego por llave legacy segura."""
        for k in _upsert_keys_row_excel_20260521(row):
            if k in index_por_key:
                return index_por_key[k]
        return None

    # 1) Cargar filas antiguas.
    #    Regla G4: preservar histórico, pero si ya existen duplicados seguros con
    #    mismo Radicado+Archivo+Número+Concepto, se fusionan en la primera fila.
    #    Eso corrige los duplicados creados por una mejora anterior sin volver a
    #    colapsar aprobaciones distintas.
    for row in antiguo.to_dict(orient="records"):
        clean_row = {col: row.get(col, "") for col in COLUMNAS_VALIDAS_FINALES}
        idx_existente = buscar_idx(clean_row)

        if idx_existente is not None:
            firma_anterior = firma_por_idx.get(idx_existente, _row_signature_excel_20260512(filas_ordenadas[idx_existente]))
            merged_row = _merge_rows_prefer_better_excel_20260515(
                filas_ordenadas[idx_existente],
                clean_row,
            )
            filas_ordenadas[idx_existente] = merged_row
            firma_por_idx[idx_existente] = _row_signature_excel_20260512(merged_row)
            registrar_keys(idx_existente, merged_row)
            registrar_keys(idx_existente, clean_row)
            registrar_mejor(merged_row)
        else:
            idx = len(filas_ordenadas)
            filas_ordenadas.append(clean_row)
            firma_por_idx[idx] = _row_signature_excel_20260512(clean_row)
            registrar_keys(idx, clean_row)
            registrar_mejor(clean_row)

    # 2) Upsert con filas nuevas.
    for row in nuevo.to_dict(orient="records"):
        clean_row = {col: row.get(col, "") for col in COLUMNAS_VALIDAS_FINALES}

        if es_minima_duplicada_de_mejor(clean_row):
            minimas_bloqueadas += 1
            continue

        idx = buscar_idx(clean_row)

        if idx is not None:
            firma_anterior = firma_por_idx.get(idx, _row_signature_excel_20260512(filas_ordenadas[idx]))
            merged_row = _merge_rows_prefer_better_excel_20260515(
                filas_ordenadas[idx],
                clean_row,
            )
            firma_merged = _row_signature_excel_20260512(merged_row)

            if firma_merged != firma_anterior:
                filas_ordenadas[idx] = merged_row
                firma_por_idx[idx] = firma_merged
                actualizadas += 1

            registrar_keys(idx, merged_row)
            registrar_keys(idx, clean_row)
            registrar_mejor(merged_row)
        else:
            idx_nuevo = len(filas_ordenadas)
            filas_ordenadas.append(clean_row)
            firma_por_idx[idx_nuevo] = _row_signature_excel_20260512(clean_row)
            registrar_keys(idx_nuevo, clean_row)
            registrar_mejor(clean_row)
            nuevas += 1

    if minimas_bloqueadas:
        facturas_aprox = minimas_bloqueadas // max(1, len(CONCEPTOS_BASE_FIJOS))
        print(
            "🛡️ [Excel anti-duplicado mínimo] "
            f"filas_minimas_ignoradas={minimas_bloqueadas} | "
            f"facturas_aprox={facturas_aprox}"
        )

    final_df = pd.DataFrame(filas_ordenadas)

    for col in COLUMNAS_VALIDAS_FINALES:
        if col not in final_df.columns:
            final_df[col] = ""

    final_df = final_df[COLUMNAS_VALIDAS_FINALES].copy()

    return final_df, nuevas, actualizadas

def guardar_en_excel(datos: List[Dict[str, Any]]) -> int:
    """
    Guarda facturas en formato largo, 7 filas por factura.

    Versión 2026-05-15:
    - Hace UPSERT real usando Radicado como separador principal:
        * Si la factura/concepto ya existe, fusiona datos buenos.
        * Si no existe, inserta.
        * CUFE solo no colapsa aprobaciones/radicados distintos.
    - Recalcula Estado_calidad para todo el archivo.
    - Devuelve filas afectadas = insertadas + actualizadas.

    Esto permite que mejoras del parser sí se reflejen en facturas.xlsx,
    incluso cuando la factura ya estaba registrada con datos parciales/mínimos.
    """
    registros_transformados: List[Dict[str, Any]] = []

    for d in (datos or []):
        if not isinstance(d, dict):
            continue

        numero_factura_texto = _forzar_texto_excel(d.get("Número de factura", ""))

        extras = {
            "Radicado": d.get("Radicado", ""),
            "ProyectoProceso": d.get("ProyectoProceso", ""),
        }

        filas_temp: List[Dict[str, Any]] = []

        for medida in CONCEPTOS_BASE_FIJOS:
            base = {
                "Archivo":               d.get("Archivo", ""),
                "Empresa emisora":       d.get("Empresa emisora", ""),
                "CUFE":                  d.get("CUFE", ""),
                "Ciudad emisora":        d.get("Ciudad emisora", ""),
                "Código ciudad":         d.get("Código ciudad", ""),
                "NIT":                   d.get("NIT", ""),
                "Cliente":               d.get("Cliente", ""),
                "Número de factura":     numero_factura_texto,
                "Año":                   d.get("Año", ""),
                "Mes":                   d.get("Mes", ""),
                "Día":                   d.get("Día", ""),
                "Tipo de contribuyente": d.get("Tipo de contribuyente", ""),
                "Actividad económica":   d.get("Actividad económica", ""),
                "DESCRIPCIÓN":           _limpiar_descripcion(_descripcion_por_concepto(d, medida)),
                "Concepto":              medida,
                "VALOR":                 d.get(medida, 0),
                **extras,
            }
            filas_temp.append(base)

        estado = _calcular_estado_calidad_para_filas(filas_temp)

        for fila in filas_temp:
            fila[COLUMNA_ESTADO_CALIDAD] = estado

        registros_transformados.extend(filas_temp)

    df_nuevo = pd.DataFrame(registros_transformados)
    df_nuevo = _limpiar_dataframe_a_formato_largo(df_nuevo)

    if os.path.exists(ARCHIVO_EXCEL):
        antiguo = _read_facturas_df()
        antiguo = _limpiar_dataframe_a_formato_largo(antiguo)

        final_df, filas_nuevas, filas_actualizadas = _consolidar_filas_por_upsert_excel_20260512(
            antiguo,
            df_nuevo,
        )
    else:
        final_df = df_nuevo
        filas_nuevas = len(df_nuevo)
        filas_actualizadas = 0

    final_df = _limpiar_dataframe_a_formato_largo(final_df)

    if "Concepto" in final_df.columns:
        final_df = final_df[final_df["Concepto"].astype(str).str.strip() != ""].copy()

    for col in COLUMNAS_VALIDAS_FINALES:
        if col not in final_df.columns:
            final_df[col] = ""

    final_df = final_df[COLUMNAS_VALIDAS_FINALES].copy()

    final_df = _ordenar_dataframe_facturas(final_df)

    # Recalcular Estado_calidad sobre TODO el archivo antes de guardar.
    final_df = _recalcular_estado_calidad_dataframe(final_df)

    grupos_total, grupos_incompletos, muestra_incompletos = _validar_integridad_7_conceptos_excel_20260515(final_df)
    if grupos_incompletos:
        print(
            f"⚠️ [Excel Integridad] grupos_factura={grupos_total} | "
            f"grupos_incompletos={grupos_incompletos}"
        )
        for key, cant, faltan in muestra_incompletos:
            print(f"   - key={key} | conceptos={cant} | faltan={faltan}")

    safe_save_pandas(
        final_df,
        ARCHIVO_EXCEL,
        sheet_name="Facturas",
        header=True,
        index=False,
    )

    _aplicar_formato_visual_facturas()

    filas_afectadas = int(filas_nuevas or 0) + int(filas_actualizadas or 0)

    print(f"✅ Excel formateado y actualizado: {ARCHIVO_EXCEL}")
    print(
        "[Excel UPSERT] "
        f"filas_nuevas={filas_nuevas} | "
        f"filas_actualizadas={filas_actualizadas} | "
        f"filas_afectadas={filas_afectadas} | "
        f"filas_finales={len(final_df)} | "
        f"facturas_grupo={grupos_total}"
    )

    return filas_afectadas

def registrar_historial_por_zip(filas: List[Dict[str, Any]]) -> None:
    df_h = pd.DataFrame(filas)

    if os.path.exists(HISTORIAL_EXCEL):
        try:
            antiguo = pd.read_excel(HISTORIAL_EXCEL, engine="openpyxl")
            unido = pd.concat([antiguo, df_h], ignore_index=True)
        except Exception:
            unido = df_h
    else:
        unido = df_h

    safe_save_pandas(
        unido,
        HISTORIAL_EXCEL,
        sheet_name="Historial",
        header=True,
        index=False,
    )

    print(f"📁 Historial actualizado: {HISTORIAL_EXCEL}")


def _normalizar_ref_archivo(s: Any) -> str:
    if s is None:
        return ""
    txt = str(s).strip().replace("\\", "/")
    if not txt:
        return ""
    return os.path.basename(txt).strip().lower()


def _stem_archivo(s: Any) -> str:
    norm = _normalizar_ref_archivo(s)
    if not norm:
        return ""
    return Path(norm).stem.lower().strip()


def obtener_filas_por_archivos(archivos: Set[str]) -> List[Dict[str, Any]]:
    if not archivos:
        return []

    df = _read_facturas_df()
    df = _limpiar_dataframe_a_formato_largo(df)

    if df.empty or "Archivo" not in df.columns:
        return []

    archivos_norm = {_normalizar_ref_archivo(a) for a in archivos if _normalizar_ref_archivo(a)}
    if not archivos_norm:
        return []

    stems_ref = {_stem_archivo(a) for a in archivos_norm if _stem_archivo(a)}

    df = df.copy()
    df["__archivo_norm__"] = df["Archivo"].apply(_normalizar_ref_archivo)
    df["__archivo_stem__"] = df["Archivo"].apply(_stem_archivo)

    mask_exacta = df["__archivo_norm__"].isin(archivos_norm)
    mask_stem = df["__archivo_stem__"].isin(stems_ref) if stems_ref else False

    df2 = df[mask_exacta | mask_stem].copy()
    if df2.empty:
        return []

    for col in COLUMNAS_TEXTO:
        if col in df2.columns:
            df2[col] = df2[col].apply(_forzar_texto_excel)

    df2 = df2.drop(columns=["__archivo_norm__", "__archivo_stem__"], errors="ignore")
    df2 = df2.fillna("")
    for col in df2.columns:
        df2[col] = df2[col].apply(_sanitizar_valor_excel)
    return df2.to_dict(orient="records")
            